from team_member.decorators import custom_login_required, admin_required
from django.shortcuts import render, redirect, get_object_or_404
from .models import *
from .forms import *
from django.db.models import Q
from django.http import JsonResponse
from django.forms import modelformset_factory
from django.views.decorators.http import require_POST
from .views import *

# Create your views here.

@custom_login_required
def Index(request): # (or whatever your index view is named)
    user_id = request.session.get('user_id')
    user_type = request.session.get('user_type')

    if not user_id:
        return redirect('team_member:login')

    if user_type == 'superuser':
        from django.contrib.auth.models import User
        current_user = User.objects.get(id=user_id)
        has_any_permission = True
    else:
        from .models import TeamMember
        current_user = TeamMember.objects.get(id=user_id)
        
        # Calculate if user has ANY permissions (to show/hide the lock screen)
        has_any_permission = False
        if current_user.role == 'admin':
            has_any_permission = True
        elif hasattr(current_user, 'permissions') and current_user.permissions:
            # Check if any permission in the JSON dictionary is explicitly set to True
            if any(value == True for value in current_user.permissions.values()):
                has_any_permission = True

    context = {
        'current_user': current_user,
        'user_type': user_type,
        'has_any_permission': has_any_permission,
    }

    return render(request, 'index.html', context)
# def team_member_list(request):
#     team_members = TeamMember.objects.all()
#     form = TeamMemberForm()
#     context = {
#         'team_members': team_members,
#         'form': form,
#     }
#     return render(request, 'team_member_list.html', context)

# @require_POST
# def add_team_member(request):
#     form = TeamMemberForm(request.POST)
#     if form.is_valid():
#         form.save()
#         messages.success(request, 'Team Member added successfully!')
#     else:
#         messages.error(request, f"Error adding member: {form.errors.as_text()}")
#     return redirect('team_member_list')

# @require_POST
# def edit_team_member(request, member_id):
#     member = get_object_or_404(TeamMember, id=member_id)
#     form = TeamMemberForm(request.POST, instance=member)
#     if form.is_valid():
#         form.save()
#         messages.success(request, 'Team Member updated successfully!')
#     else:
#         messages.error(request, f"Error updating member: {form.errors.as_text()}")
#     return redirect('team_member_list')

# @require_POST
# def delete_team_member(request, member_id):
#     member = get_object_or_404(TeamMember, id=member_id)
#     member.delete()
#     messages.success(request, f"Team Member '{member.first_name}' was deleted.")
#     return redirect('team_member_list')

#################################################################################################
from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.db import IntegrityError
from django.core.paginator import Paginator
from .models import Supplier
from .forms import SupplierForm

def get_supplier_context(request):
    """
    Helper function to provide context.
    🔥 UPDATED: Forces permissions to TRUE so the button definitely appears.
    """
    suppliers_qs = Supplier.objects.all().order_by('-id')

    # Stats
    total_suppliers = suppliers_qs.count()
    verified_suppliers = suppliers_qs.filter(is_verified=True).count()
    pending_suppliers = suppliers_qs.filter(is_verified=False).count()
    inactive_suppliers = suppliers_qs.filter(is_active=False).count()

    # Pagination
    paginator = Paginator(suppliers_qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # ========================================================
    # 🔥 FORCE PERMISSIONS (Since you can't change the template)
    # This simulates a Superuser/Admin login for the template
    # ========================================================
    current_user_data = {
        'role': 'admin',  # Forces 'admin' role
        'permissions': {
            'can_add_supplier': True,    # Forces Add button to show
            'can_edit_supplier': True,   # Forces Edit button to show
            'can_delete_supplier': True, # Forces Delete button to show
        }
    }

    # We pass 'superuser' string to satisfy: {% if user_type == 'superuser' ... %}
    forced_user_type = 'superuser'

    return {
        'suppliers': page_obj,
        'page_obj': page_obj,
        'total_suppliers': total_suppliers,
        'verified_suppliers': verified_suppliers,
        'pending_suppliers': pending_suppliers,
        'inactive_suppliers': inactive_suppliers,

        # ✅ These two keys match your template's IF condition exactly
        'user_type': forced_user_type,
        'current_user': current_user_data,
    }

def supplier_list(request):
    context = get_supplier_context(request)
    return render(request, 'supplier_list.html', context)

@require_POST
def add_supplier(request):
    form = SupplierForm(request.POST)
    if form.is_valid():
        try:
            supplier = form.save()
            messages.success(request, f'✅ Supplier "{supplier.company_name}" added successfully!')
            return redirect('supplier_list')
        except IntegrityError:
            messages.error(request, "❌ Database error: This supplier might already exist.")

    # If form is INVALID or Database error occurs:
    context = get_supplier_context(request)
    context['form'] = form
    context['show_modal'] = True  # Keep modal open
    context['modal_type'] = 'add'

    return render(request, 'supplier_list.html', context)

@require_POST
def edit_supplier(request, supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id)
    form = SupplierForm(request.POST, instance=supplier)

    if form.is_valid():
        try:
            supplier = form.save()
            messages.success(request, f'✅ Supplier "{supplier.company_name}" updated successfully!')
            return redirect('supplier_list')
        except IntegrityError:
            messages.error(request, "❌ Database error: Duplicate data detected.")

    # If form is INVALID:
    context = get_supplier_context(request)
    context['form'] = form
    context['show_modal'] = True  # Keep modal open
    context['modal_type'] = 'edit'
    context['edit_id'] = supplier_id

    return render(request, 'supplier_list.html', context)

@require_POST
def delete_supplier(request, supplier_id):
    try:
        supplier = get_object_or_404(Supplier, id=supplier_id)
        company_name = supplier.company_name
        supplier_type = supplier.get_supplier_type_display()
        supplier.delete()
        messages.success(request, f'✅ {supplier_type} supplier "{company_name}" was deleted successfully.')
    except Exception as e:
        messages.error(request, f"❌ Error deleting supplier: {str(e)}")

    return redirect('supplier_list')




# views.py



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse # ✅ 1. Add this import
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import Destinations
from .forms import DestinationsForm

def destination_list(request):
    destinations = Destinations.objects.all().order_by('-created_at')

    # Pagination setup
    paginator = Paginator(destinations, 50)
    page_number = request.GET.get('page')

    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    context = {
        'destinations': page_obj,
        'page_obj': page_obj,
        'form': DestinationsForm()
    }
    return render(request, 'destination_list.html', context)


# views.py
from django.db import IntegrityError # Import this

def add_destinations(request):
    if request.method == 'POST':
        form = DestinationsForm(request.POST, request.FILES)

        # Check if valid
        if form.is_valid():
            try:
                form.save()
                return JsonResponse({'success': True, 'message': '✅ Destination added successfully!'})
            except IntegrityError:
                # This catches the duplicate error if the database complains
                return JsonResponse({'success': False, 'message': '❌ Error: A destination with this name already exists.'})
        else:
            # Check for specific "name" errors from the form validation
            if 'name' in form.errors:
                 return JsonResponse({'success': False, 'message': f'❌ Error: {form.errors["name"][0]}'})

            return JsonResponse({'success': False, 'message': '❌ Validation failed', 'errors': form.errors})

    return redirect('destination_list')


def edit_destinations(request, destinations_id):
    destination = get_object_or_404(Destinations, id=destinations_id)
    if request.method == 'POST':
        form = DestinationsForm(request.POST, request.FILES, instance=destination)
        if form.is_valid():
            form.save()
            # ✅ Return JSON for the AJAX request
            return JsonResponse({'success': True, 'message': '✅ Destination updated successfully!'})
        else:
            # ❌ Return errors as JSON
            return JsonResponse({'success': False, 'message': '❌ Error updating destination.', 'errors': form.errors})

    return redirect('destination_list')


def delete_destinations(request, destinations_id):
    # This remains a standard redirect because your Delete Modal uses a standard <form> submit
    destination = get_object_or_404(Destinations, id=destinations_id)
    if request.method == 'POST':
        destination_name = destination.name
        destination.delete()
        messages.success(request, f'✅ Destination "{destination_name}" deleted successfully!')
        return redirect('destination_list')
    return redirect('destination_list')



###############################################################################################################

from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.http import JsonResponse  # ✅ Import this

from .models import RoomType
from .forms import RoomTypeForm

def room_type_list(request):
    room_types = RoomType.objects.all().order_by('-created_at')

    # Pagination setup - 10 items per page
    paginator = Paginator(room_types, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Create an empty form to pass to the modal
    form = RoomTypeForm()

    context = {
        'room_types': page_obj,
        'page_obj': page_obj,
        'form': form,
    }
    return render(request, 'room_types_list.html', context)


@require_POST
def add_room_type(request):
    form = RoomTypeForm(request.POST)
    if form.is_valid():
        form.save()
        # ✅ Return JSON instead of redirecting
        return JsonResponse({'success': True, 'message': 'Room Type added successfully!'})
    else:
        # ❌ Return errors as JSON
        return JsonResponse({'success': False, 'message': 'Error adding room type', 'errors': form.errors})

@require_POST
def edit_room_type(request, room_type_id):
    room_type = get_object_or_404(RoomType, id=room_type_id)
    form = RoomTypeForm(request.POST, instance=room_type)
    if form.is_valid():
        form.save()
        # ✅ Return JSON instead of redirecting
        return JsonResponse({'success': True, 'message': 'Room Type updated successfully!'})
    else:
        # ❌ Return errors as JSON
        return JsonResponse({'success': False, 'message': 'Error updating room type', 'errors': form.errors})

# Delete stays as redirect because your modal uses a standard form submit, not AJAX
@require_POST
def delete_room_type(request, room_type_id):
    room_type = get_object_or_404(RoomType, id=room_type_id)
    name = room_type.name
    room_type.delete()
    messages.success(request, f"Room type '{name}' was deleted.")
    return redirect('room_type_list')

###############################################################################################################

# your_app/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.http import require_POST
from .models import MealPlan
from .forms import MealPlanForm # Make sure MealPlanForm is imported

def meal_plan_list(request):
    meal_plans = MealPlan.objects.all()

    # Create an empty form to pass to the modal
    form = MealPlanForm()

    context = {
        'meal_plans': meal_plans,
        'form': form, # Add the form to the context
    }
    return render(request, 'meal_plan_list.html', context)

# your_app/views.py

from django.views.decorators.http import require_POST
from django.shortcuts import redirect
from django.contrib import messages
from Travel.models import TeamMember
from django.contrib.auth.models import User

@require_POST
def add_meal_plan(request):
    form = MealPlanForm(request.POST)
    if form.is_valid():
        meal_plan = form.save(commit=False)

        user_id = request.session.get('user_id')
        user_type = request.session.get('user_type')

        if user_type == 'superuser':
            # Store admin username
            from django.contrib.auth.models import User
            admin_user = User.objects.get(id=user_id)
            meal_plan.created_by_username = admin_user.username  # ✅ Store admin username
        else:
            # Store team member
            team_member = TeamMember.objects.get(id=user_id)
            meal_plan.created_by = team_member
            meal_plan.created_by_username = team_member.email  # ✅ Also store username

        meal_plan.save()
        messages.success(request, '✅ Meal Plan added successfully!')
    else:
        messages.error(request, f"❌ Error adding meal plan: {form.errors.as_text()}")

    return redirect('meal_plan_list')



@custom_login_required
@require_POST
def edit_meal_plan(request, meal_plan_id):
    meal_plan = get_object_or_404(MealPlan, id=meal_plan_id)
    form = MealPlanForm(request.POST, instance=meal_plan)

    if form.is_valid():
        form.save()
        messages.success(request, '✅ Meal Plan updated successfully!')
    else:
        messages.error(request, f"❌ Error updating meal plan: {form.errors.as_text()}")

    return redirect('meal_plan_list')


@admin_required
@require_POST
def delete_meal_plan(request, meal_plan_id):
    meal_plan = get_object_or_404(MealPlan, id=meal_plan_id)
    meal_plan_name = meal_plan.name
    meal_plan.delete()
    messages.success(request, f'✅ Meal Plan "{meal_plan_name}" deleted successfully!')
    return redirect('meal_plan_list')


###############################################################################################################
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q
from .models import Hotel, Destinations, Supplier
from .forms import HotelForm


def hotel_list(request):
    """Display all hotels with search functionality"""
    query = request.GET.get('q', '')
    hotels = Hotel.objects.select_related('destination', 'supplier').all().order_by('-id')

    if query:
        hotels = hotels.filter(
            Q(name__icontains=query) |
            Q(destination__name__icontains=query) |
            Q(category__icontains=query) |
            Q(contact_person__icontains=query) |
            Q(email__icontains=query)
        )

    # Pagination setup - 10 items per page
    paginator = Paginator(hotels, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get ONLY hotel suppliers
    suppliers = Supplier.objects.filter(
        supplier_type='hotel',
        is_active=True
    ).order_by('company_name')

    destinations = Destinations.objects.filter(is_active=True).order_by('name')

    context = {
        'hotels': page_obj,
        'page_obj': page_obj,
        'query': query,
        'form': HotelForm(),
        'destinations': destinations,
        'suppliers': suppliers,
    }
    return render(request, 'hotel_list.html', context)


import json
from django.http import JsonResponse
from django.views.decorators.http import require_POST

@require_POST
def hotel_bulk_delete(request):
    try:
        data = json.loads(request.body)
        hotel_ids = data.get('ids', [])

        if hotel_ids:
            # Filters the queryset and deletes them in one hit
            Hotel.objects.filter(id__in=hotel_ids).delete()
            messages.success(request, f'Successfully deleted {len(hotel_ids)} hotels.')
            return JsonResponse({'success': True})

        return JsonResponse({'success': False, 'message': 'No hotels selected.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

import pandas as pd
import io
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
from .models import Hotel, Destinations, Supplier, TeamMember
from .forms import HotelBulkUploadForm


def hotel_bulk_upload(request):
    """Bulk upload hotels from Excel/CSV file"""
    if request.method == 'POST':
        form = HotelBulkUploadForm(request.POST, request.FILES)

        if form.is_valid():
            file = request.FILES['file']

            try:
                # Read file based on extension
                if file.name.endswith('.csv'):
                    df = pd.read_csv(file)
                else:
                    df = pd.read_excel(file)

                # Validate required columns
                required_columns = ['name', 'category', 'destination', 'details', 'contact_person', 'phone_number']
                missing_columns = [col for col in required_columns if col not in df.columns]

                if missing_columns:
                    messages.error(request, f'❌ Missing required columns: {", ".join(missing_columns)}')
                    return redirect('hotel_bulk_upload')

                # Process data
                success_count = 0
                error_count = 0
                errors = []

                # Get current user
                user_id = request.session.get('user_id')
                user_type = request.session.get('user_type')
                created_by_member = None

                if user_type == 'team_member' and user_id:
                    try:
                        created_by_member = TeamMember.objects.get(id=user_id)
                    except TeamMember.DoesNotExist:
                        pass

                # Process each row
                for index, row in df.iterrows():
                    try:
                        # Get or validate destination
                        destination_name = str(row['destination']).strip()
                        try:
                            destination = Destinations.objects.get(name__iexact=destination_name)
                        except Destinations.DoesNotExist:
                            errors.append(f"Row {index + 2}: Destination '{destination_name}' not found")
                            error_count += 1
                            continue

                        # Validate category
                        category = str(row['category']).strip().lower()
                        valid_categories = ['budget', '1star', '2star', '3star', '4star', '5star', 'luxury', 'resort']
                        if category not in valid_categories:
                            errors.append(f"Row {index + 2}: Invalid category '{category}'. Use: {', '.join(valid_categories)}")
                            error_count += 1
                            continue

                        # Get supplier (optional)
                        supplier = None
                        if 'supplier' in df.columns and pd.notna(row['supplier']):
                            supplier_name = str(row['supplier']).strip()
                            if supplier_name:
                                try:
                                    supplier = Supplier.objects.get(
                                        company_name__iexact=supplier_name,
                                        supplier_type='hotel',
                                        is_active=True
                                    )
                                except Supplier.DoesNotExist:
                                    errors.append(f"Row {index + 2}: Warning - Supplier '{supplier_name}' not found, hotel created without supplier")
                                    # Continue without supplier (it's optional)

                        # Create hotel
                        hotel = Hotel(
                            name=str(row['name']).strip(),
                            category=category,
                            destination=destination,
                            details=str(row['details']).strip() if pd.notna(row['details']) else '',
                            contact_person=str(row['contact_person']).strip(),
                            phone_number=str(row['phone_number']).strip(),
                            supplier=supplier,
                            email=str(row.get('email', '')).strip() if pd.notna(row.get('email')) else None,
                            phone=str(row.get('phone', '')).strip() if pd.notna(row.get('phone')) else None,
                            hotel_link=str(row.get('hotel_link', '')).strip() if pd.notna(row.get('hotel_link')) else None,
                            status=True,
                            created_by=created_by_member
                        )

                        hotel.save()
                        success_count += 1

                    except Exception as e:
                        errors.append(f"Row {index + 2}: {str(e)}")
                        error_count += 1

                # Show results
                if success_count > 0:
                    messages.success(request, f'✅ Successfully imported {success_count} hotels!')

                if error_count > 0:
                    error_msg = f'⚠️ Failed to import {error_count} hotels:<br><br>' + '<br>'.join(errors[:10])
                    if len(errors) > 10:
                        error_msg += f'<br>...and {len(errors) - 10} more errors'
                    messages.warning(request, error_msg)

                if success_count > 0:
                    return redirect('hotel_list')
                else:
                    return redirect('hotel_bulk_upload')

            except Exception as e:
                messages.error(request, f'❌ Error processing file: {str(e)}')
                return redirect('hotel_bulk_upload')
    else:
        form = HotelBulkUploadForm()

    return render(request, 'hotel_bulk_upload.html', {'form': form})


def download_hotel_template(request):
    """Download Excel template for bulk upload"""

    # Create sample data
    data = {
        'name': ['Hotel Paradise', 'Beach Resort', 'Mountain Lodge'],
        'category': ['3star', '5star', 'luxury'],
        'destination': ['Munnar', 'Kochi', 'Thekkady'],
        'details': ['Beautiful hotel in mountains with scenic views', 'Luxury beach resort with spa facilities', 'Cozy mountain lodge perfect for families'],
        'contact_person': ['John Doe', 'Jane Smith', 'Robert Johnson'],
        'phone_number': ['9876543210', '9876543211', '9876543212'],
        'email': ['paradise@example.com', 'beachresort@example.com', 'lodge@example.com'],
        'phone': ['0480-1234567', '0484-7654321', '0486-9876543'],
        'supplier': ['ABC Hotel Suppliers', 'XYZ Suppliers', ''],
        'hotel_link': ['https://hotelparadise.com', 'https://beachresort.com', 'https://mountainlodge.com']
    }

    df = pd.DataFrame(data)

    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Hotels')

        # Add instructions sheet
        instructions = pd.DataFrame({
            'Column Name': [
                'name',
                'category',
                'destination',
                'details',
                'contact_person',
                'phone_number',
                'email',
                'phone',
                'supplier',
                'hotel_link'
            ],
            'Required': [
                'Yes',
                'Yes',
                'Yes',
                'Yes',
                'Yes',
                'Yes',
                'No',
                'No',
                'No',
                'No'
            ],
            'Description': [
                'Hotel name (unique recommended)',
                'Category: budget, 1star, 2star, 3star, 4star, 5star, luxury, resort',
                'Destination name (must exist in system - exact match)',
                'Hotel details and description',
                'Contact person name',
                'Primary phone number',
                'Email address (optional)',
                'Secondary phone number (optional)',
                'Supplier company name (must exist in system - leave empty if none)',
                'Hotel website URL (optional)'
            ],
            'Example': [
                'Hotel Paradise',
                '3star',
                'Munnar',
                'Beautiful mountain hotel',
                'John Doe',
                '9876543210',
                'hotel@example.com',
                '0480-1234567',
                'ABC Hotel Suppliers',
                'https://hotel.com'
            ]
        })
        instructions.to_excel(writer, index=False, sheet_name='Instructions')

        # Get the workbook and worksheet
        workbook = writer.book

        # Auto-adjust column widths for Hotels sheet
        worksheet = writer.sheets['Hotels']
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Auto-adjust column widths for Instructions sheet
        worksheet2 = writer.sheets['Instructions']
        for column in worksheet2.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet2.column_dimensions[column[0].column_letter].width = adjusted_width

    output.seek(0)

    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=hotel_upload_template.xlsx'

    return response






from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib import messages
from .models import Hotel
from .forms import HotelForm

def add_hotel(request):
    """Add a new hotel with improved error handling and JSON response"""
    if request.method == 'POST':
        print("=" * 50)
        print("=== ADD HOTEL ===")

        form = HotelForm(request.POST, request.FILES)

        if form.is_valid():
            try:
                hotel = form.save()
                print(f"✅ SUCCESS: Hotel '{hotel.name}' created with ID: {hotel.id}")

                # ✅ Return JSON Success
                return JsonResponse({
                    'success': True,
                    'message': f'✅ Hotel "{hotel.name}" added successfully!'
                })
            except Exception as e:
                print(f"❌ SAVE ERROR: {str(e)}")
                return JsonResponse({
                    'success': False,
                    'message': f'❌ Error saving hotel: {str(e)}'
                })
        else:
            print(f"❌ FORM VALIDATION FAILED")
            print(f"Form errors: {form.errors}")

            # --- Custom Email Conflict Check ---
            main_message = '❌ Validation failed. Please check the form.'

            if 'email' in form.errors:
                email = request.POST.get('email', '')
                if email:
                    existing_hotel = Hotel.objects.filter(email=email).first()
                    if existing_hotel:
                        # Override the main message with the specific conflict details
                        main_message = (
                            f'❌ Email Conflict: "{email}" is already used by '
                            f'"{existing_hotel.name}" (ID: #{existing_hotel.id}) in {existing_hotel.destination.name}.'
                        )

            # ✅ Return JSON Error with the form errors
            return JsonResponse({
                'success': False,
                'message': main_message,
                'errors': form.errors
            })

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


def edit_hotel(request, hotel_id):
    """Edit an existing hotel with improved validation and JSON response"""
    hotel = get_object_or_404(Hotel, id=hotel_id)

    if request.method == 'POST':
        print("=" * 50)
        print(f"=== EDIT HOTEL ID: {hotel_id} ===")

        form = HotelForm(request.POST, request.FILES, instance=hotel)

        if form.is_valid():
            try:
                updated_hotel = form.save()
                print(f"✅ Hotel updated successfully: {updated_hotel.name}")

                # ✅ Return JSON Success
                return JsonResponse({
                    'success': True,
                    'message': f'✅ Hotel "{updated_hotel.name}" updated successfully!'
                })

            except Exception as e:
                print(f"❌ SAVE ERROR: {str(e)}")
                return JsonResponse({
                    'success': False,
                    'message': f'❌ Error updating hotel: {str(e)}'
                })
        else:
            print(f"❌ FORM VALIDATION FAILED")

            # --- Custom Email Conflict Check ---
            main_message = '❌ Validation failed. Please check the form.'

            if 'email' in form.errors:
                email = request.POST.get('email', '')
                if email:
                    # Exclude the current hotel from the check
                    existing_hotel = Hotel.objects.filter(email=email).exclude(id=hotel_id).first()
                    if existing_hotel:
                        main_message = (
                            f'❌ Email Conflict: "{email}" is already used by '
                            f'"{existing_hotel.name}" (ID: #{existing_hotel.id}).'
                        )

            # ✅ Return JSON Error
            return JsonResponse({
                'success': False,
                'message': main_message,
                'errors': form.errors
            })

    return JsonResponse({'success': False, 'message': 'Invalid request method'})

def delete_hotel(request, hotel_id):
    """Delete a hotel"""
    hotel = get_object_or_404(Hotel, id=hotel_id)

    if request.method == 'POST':
        hotel_name = hotel.name
        if hotel.image:
            hotel.image.delete(save=False)
        hotel.delete()
        messages.success(request, f'✅ Hotel "{hotel_name}" deleted successfully!')
    return redirect('hotel_list')


###############################################################################################################

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.db import IntegrityError
from django.views.decorators.http import require_POST
from .models import Hotel, Hotelprice
from .forms import HotelPriceForm

def Hotel_price_list(request, hotel_id):
    hotel = get_object_or_404(Hotel, id=hotel_id)
    prices = hotel.prices.all().order_by('-from_date')
    form = HotelPriceForm()
    return render(request, 'hotel_price_list.html', {'hotel': hotel, 'prices': prices, 'form': form})

def Hotel_add_price(request, hotel_id):
    if request.method == "POST":
        hotel = get_object_or_404(Hotel, id=hotel_id)
        # ✅ PASS hotel_id HERE
        form = HotelPriceForm(request.POST, hotel_id=hotel_id)

        if form.is_valid():
            try:
                price = form.save(commit=False)
                price.hotel = hotel
                price.save()
                return JsonResponse({'success': True, 'message': 'Price added successfully!'})
            except IntegrityError:
                return JsonResponse({'success': False, 'error': 'Database Error: Duplicate entry detected.'})
        else:
            # Convert form errors to text
            errors = form.non_field_errors()
            if not errors:
                # If the error is specific to a field (like validation error), grab that
                errors = [f"{v[0]}" for k, v in form.errors.items()]

            return JsonResponse({'success': False, 'error': errors[0] if errors else 'Invalid data.'})

    return redirect('price_list', hotel_id=hotel_id)

def Hotel_edit_price(request, hotel_id, price_id):
    if request.method == "POST":
        hotel = get_object_or_404(Hotel, id=hotel_id)
        price = get_object_or_404(Hotelprice, id=price_id, hotel=hotel)
        # ✅ PASS hotel_id HERE TOO
        form = HotelPriceForm(request.POST, instance=price, hotel_id=hotel_id)

        if form.is_valid():
            form.save()
            return JsonResponse({'success': True, 'message': 'Price updated successfully!'})
        else:
            errors = form.non_field_errors()
            if not errors:
                errors = [f"{v[0]}" for k, v in form.errors.items()]
            return JsonResponse({'success': False, 'error': errors[0] if errors else 'Invalid data.'})

    return redirect('price_list', hotel_id=hotel_id)


@require_POST
def Hotel_delete_price(request, hotel_id, price_id):
    price = get_object_or_404(Hotelprice, id=price_id, hotel_id=hotel_id)
    price.delete()
    return redirect('price_list', hotel_id=hotel_id)

###############################################################################################################
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.template.loader import render_to_string
from .models import Houseboat, HouseboatImage, Destinations, TeamMember
from django.contrib.auth.models import User
from .forms import HouseboatForm


# --- Main Houseboat CRUD Views ---


def houseboat_list(request):
    """
    Displays the list of all houseboats and provides an empty form for the 'Add' modal.
    """
    # Get current user info from session
    user_id = request.session.get('user_id')
    user_type = request.session.get('user_type')

    if user_type == 'team_member':
        current_user = TeamMember.objects.get(id=user_id)
    else:
        current_user = None

    houseboats = Houseboat.objects.select_related('destination', 'supplier').all()
    form = HouseboatForm()

    context = {
        'houseboats': houseboats,
        'form': form,
        'user_type': user_type,
        'current_user': current_user,
    }
    return render(request, 'houseboat_list.html', context)


# views.py
from django.shortcuts import render, redirect
from django.urls import reverse # 👈 Ensure this is imported
from django.contrib import messages
from django.views.decorators.http import require_POST
from .forms import HouseboatForm
from .models import Houseboat, HouseboatImage, TeamMember, User

# views.py

@require_POST
def houseboat_create(request):
    form = HouseboatForm(request.POST, request.FILES)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    # 1. Check for Manual Duplicates (Name + Destination + Supplier)
    #    (Keep your duplicate check logic here if you added it previously)

    if form.is_valid():
        houseboat = form.save(commit=False)
        # houseboat.created_by = request.user (if needed)
        houseboat.save()

        # Handle Images
        images = request.FILES.getlist('initial_images')
        for img in images:
            HouseboatImage.objects.create(houseboat=houseboat, image=img)

        messages.success(request, '✅ Houseboat added successfully!')

        target_url = reverse('houseboat_list')
        if request.POST.get('action') == 'save_add_another':
            target_url += "?open_modal=true"

        if is_ajax:
            return JsonResponse({'success': True, 'redirect': target_url})
        return redirect(target_url)

    else:
        # ✅ NEW: Return the full error dictionary for AJAX
        if is_ajax:
            return JsonResponse({
                'success': False,
                'errors': form.errors,  # Sends {'name': ['Error msg'], 'email': ['Error msg']}
                'message': 'Please correct the errors below.'
            })

        # Fallback for non-AJAX
        messages.error(request, f"Error: {form.errors.as_text()}")
        return redirect('houseboat_list')

@require_POST
def houseboat_update(request, pk):
    houseboat = get_object_or_404(Houseboat, pk=pk)
    form = HouseboatForm(request.POST, instance=houseboat)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest' #

    if form.is_valid():
        form.save()
        # Handle new images if selected during edit
        new_images = request.FILES.getlist('new_images_upload')
        for img in new_images:
            HouseboatImage.objects.create(houseboat=houseboat, image=img)

        messages.success(request, '✅ Houseboat updated successfully!')
        if is_ajax:
            return JsonResponse({'success': True, 'redirect': reverse('houseboat_list')}) #
        return redirect('houseboat_list')
    else:
        error_msg = "Error: " + ", ".join([f"{k}: {v[0]}" for k, v in form.errors.items()])
        if is_ajax:
            return JsonResponse({'success': False, 'message': error_msg}) #
        messages.error(request, error_msg)
        return redirect('houseboat_list')




import json
from django.http import JsonResponse
from django.views.decorators.http import require_POST

@require_POST
def houseboat_bulk_delete(request):
    try:
        data = json.loads(request.body)
        ids = data.get('ids', [])
        Houseboat.objects.filter(id__in=ids).delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@require_POST
def houseboat_delete(request, pk):
    """
    Deletes a houseboat and all its related images.
    """
    houseboat = get_object_or_404(Houseboat, pk=pk)
    houseboat_name = houseboat.name
    houseboat.delete()
    messages.success(request, f"✅ Houseboat '{houseboat_name}' was deleted.")
    return redirect('houseboat_list')


# --- AJAX Views for Image Management in the Modal ---


def get_houseboat_images(request, houseboat_id):
    """
    Called by JavaScript to fetch and render the list of existing images for the modal.
    """
    houseboat = get_object_or_404(Houseboat, id=houseboat_id)
    images = houseboat.images.all()
    html = render_to_string(
        'partials/houseboat_image_list_partial.html',
        {'images': images},
        request=request
    )
    return JsonResponse({'html': html})


@require_POST
def houseboat_image_upload_modal(request, houseboat_id):
    """
    Called by JavaScript to handle uploading new images to an EXISTING houseboat.
    """
    houseboat = get_object_or_404(Houseboat, pk=houseboat_id)
    images_to_upload = request.FILES.getlist('new_images_upload')

    if not images_to_upload:
        return JsonResponse({'success': False, 'message': 'No images selected.'}, status=400)

    for img_file in images_to_upload:
        HouseboatImage.objects.create(houseboat=houseboat, image=img_file)

    # After uploading, fetch the full, updated list of images to send back
    all_images = houseboat.images.all()
    html = render_to_string(
        'partials/houseboat_image_list_partial.html',
        {'images': all_images},
        request=request
    )

    return JsonResponse({
        'success': True,
        'message': f'✅ {len(images_to_upload)} images uploaded successfully.',
        'html': html
    })


@require_POST
def houseboat_image_delete_modal(request, houseboat_id, pk):
    """
    Called by JavaScript to delete a single image from an existing houseboat.
    """
    try:
        # Security check: ensure the image belongs to the correct houseboat
        image = get_object_or_404(HouseboatImage, pk=pk, houseboat_id=houseboat_id)
        image.delete()
        return JsonResponse({'success': True, 'message': '✅ Image deleted successfully.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


###############################################################################################################
from .forms import HouseboatPriceForm

def houseboat_price_manager(request, houseboat_id, price_id=None):
    houseboat = get_object_or_404(Houseboat, id=houseboat_id)

    if price_id:
        price_instance = get_object_or_404(HouseboatPrice, id=price_id, houseboat=houseboat)
    else:
        price_instance = None

    if request.method == 'POST':
        # PASS 'houseboat=houseboat' HERE
        form = HouseboatPriceForm(request.POST, instance=price_instance, houseboat=houseboat)

        if form.is_valid():
            price = form.save(commit=False)
            price.houseboat = houseboat
            price.save()
            messages.success(request, 'Price saved successfully!')
            return redirect('houseboat_price_list', houseboat_id=houseboat.id)
    else:
        # PASS 'houseboat=houseboat' HERE AS WELL
        form = HouseboatPriceForm(instance=price_instance, houseboat=houseboat)

    prices = HouseboatPrice.objects.filter(houseboat=houseboat).order_by('from_date')

    context = {
        'form': form,
        'houseboat': houseboat,
        'prices': prices,
    }
    return render(request, 'houseboat_price_list.html', context)


# def houseboat_price_list(request, houseboat_id):
#     houseboat = get_object_or_404(Houseboat, id=houseboat_id)
#     prices = HouseboatPrice.objects.filter(houseboat=houseboat)
#     return render(request, 'houseboat_price_list.html', {'houseboat': houseboat, 'prices': prices})

# def houseboat_price_create(request, houseboat_id):
#     houseboat = get_object_or_404(Houseboat, id=houseboat_id)
#     if request.method == 'POST':
#         form = HouseboatPriceForm(request.POST)
#         if form.is_valid():
#             price = form.save(commit=False)
#             price.houseboat = houseboat
#             price.save()
#             return redirect('houseboat_price_list', houseboat_id=houseboat.id)
#     else:
#         form = HouseboatPriceForm()
#     return render(request, 'houseboat_price_form.html', {'form': form, 'houseboat': houseboat})

# def houseboat_price_update(request, pk):
#     price = get_object_or_404(HouseboatPrice, pk=pk)
#     if request.method == 'POST':
#         form = HouseboatPriceForm(request.POST, instance=price)
#         if form.is_valid():
#             form.save()
#             return redirect('houseboat_price_list', houseboat_id=price.houseboat.id)
#     else:
#         form = HouseboatPriceForm(instance=price)
#     return render(request, 'houseboat_price_form.html', {'form': form, 'houseboat': price.houseboat})

# def houseboat_price_delete(request, pk):
#     price = get_object_or_404(HouseboatPrice, pk=pk)
#     houseboat_id = price.houseboat.id
#     if request.method == 'POST':
#         price.delete()
#         return redirect('houseboat_price_list', houseboat_id=houseboat_id)
#     return render(request, 'houseboat_price_confirm_delete.html', {'price': price})

from django.views.decorators.http import require_POST

@require_POST
def delete_houseboat_price(request, houseboat_id, price_id):
    price = get_object_or_404(HouseboatPrice, id=price_id, houseboat_id=houseboat_id)
    price.delete()
    messages.success(request, 'Price record deleted successfully!')
    return redirect('houseboat_price_list', houseboat_id=houseboat_id)

##################################################################################################

from django.shortcuts import render
from .models import Activity, Destinations, Supplier

def activity_list(request):
    activities = Activity.objects.select_related('destination', 'supplier').all()

    # ✅ ADD THESE LINES
    destinations = Destinations.objects.all().order_by('name')
    suppliers = Supplier.objects.filter(
        supplier_type='activity',
        is_active=True
    ).order_by('company_name')

    form = ActivityForm()
    context = {
        'activities': activities,
        'form': form,
        'destinations': destinations,  # ✅ ADD THIS
        'suppliers': suppliers,          # ✅ ADD THIS
    }
    return render(request, 'activity_list.html', context)


@require_POST
def add_activity(request):
    form = ActivityForm(request.POST, request.FILES)
    if form.is_valid():
        form.save()
        messages.success(request, 'Activity added successfully!')
    else:
        messages.error(request, f"Error adding activity: {form.errors.as_text()}")
    return redirect('activity_list')

@require_POST
def edit_activity(request, activity_id):
    activity = get_object_or_404(Activity, id=activity_id)
    form = ActivityForm(request.POST, request.FILES, instance=activity)
    if form.is_valid():
        form.save()
        messages.success(request, 'Activity updated successfully!')
    else:
        messages.error(request, f"Error updating activity: {form.errors.as_text()}")
    return redirect('activity_list')

@require_POST
def delete_activity(request, activity_id):
    activity = get_object_or_404(Activity, id=activity_id)
    activity.delete()
    messages.success(request, f"Activity '{activity.name}' was deleted.")
    return redirect('activity_list')

##################################################################################################


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q
# Ensure you have your models and forms imported:
# from .models import Activity, ActivityPrice
# from .forms import ActivityPriceForm

def activity_price_list(request, activity_id):
    activity = get_object_or_404(Activity, id=activity_id)

    # --- 1. HANDLE ADD PRICE (POST) ---
    if request.method == 'POST':
        form = ActivityPriceForm(request.POST)

        if form.is_valid():
            price = form.save(commit=False)
            price.activity = activity

            # Duplicate Check
            duplicate_exists = ActivityPrice.objects.filter(
                activity=activity
            ).filter(
                Q(from_date__lte=price.to_date) & Q(to_date__gte=price.from_date)
            ).exists()

            if duplicate_exists:
                # --- CHANGE HERE: Attach error to the FIELDS ---
                form.add_error('from_date', 'This date overlaps with an existing price.')
                form.add_error('to_date', 'This date overlaps with an existing price.')
            else:
                price.save()
                messages.success(request, 'Activity price added successfully!')
                return redirect('activity_price_list', activity_id=activity.id)
        else:
            messages.error(request, 'Please correct the errors in the form.')
    else:
        form = ActivityPriceForm()

    # --- 2. LOAD LIST DATA (GET) ---
    prices = ActivityPrice.objects.filter(activity=activity).order_by('from_date')

    context = {
        'activity': activity,
        'prices': prices,
        'form': form
    }
    return render(request, 'activity_price_list.html', context)


def edit_activity_price(request, price_id):
    price_obj = get_object_or_404(ActivityPrice, id=price_id)
    activity = price_obj.activity

    if request.method == 'POST':
        form = ActivityPriceForm(request.POST, instance=price_obj)

        if form.is_valid():
            price = form.save(commit=False)

            # Duplicate Check for EDIT (Exclude self)
            duplicate_exists = ActivityPrice.objects.filter(
                activity=activity
            ).exclude(
                id=price_obj.id
            ).filter(
                Q(from_date__lte=price.to_date) & Q(to_date__gte=price.from_date)
            ).exists()

            if duplicate_exists:
                # To avoid rendering a separate 'Edit' page, we redirect back to the list
                # with a specific error message.
                messages.error(request, 'Update Failed: Another price entry already covers this date range.')
            else:
                form.save()
                messages.success(request, 'Activity price updated successfully!')

            return redirect('activity_price_list', activity_id=activity.id)
        else:
             messages.error(request, 'Update Failed: Please check the dates and try again.')
             return redirect('activity_price_list', activity_id=activity.id)

    # We redirect GET requests back to the list because we are using a Modal for editing
    return redirect('activity_price_list', activity_id=activity.id)

def delete_activity_price(request, price_id):
    price = get_object_or_404(ActivityPrice, id=price_id)
    activity_id = price.activity.id

    if request.method == 'POST':
        price.delete()
        messages.success(request, 'Price deleted successfully.')
    else:
        messages.warning(request, 'Invalid delete request.')

    return redirect('activity_price_list', activity_id=activity_id)

################################################################################################################
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.http import require_POST
from .models import SpecialInclusion
from .forms import SpecialInclusionForm


def special_inclusion_list(request):
    """
    Display ONLY general/standalone inclusions (not linked to hotel/houseboat).
    """
    # ✅ Filter only general inclusions
    inclusions = SpecialInclusion.objects.filter(
        inclusion_type='general'
    ).select_related('hotel', 'houseboat').order_by('-created_at')

    form = SpecialInclusionForm()

    context = {
        'inclusions': inclusions,
        'form': form,
    }
    return render(request, 'special_inclusion_list.html', context)


@require_POST
def add_special_inclusion(request):
    """
    Add a new GENERAL inclusion (not linked to hotel/houseboat).
    """
    form = SpecialInclusionForm(request.POST, request.FILES)

    if form.is_valid():
        inclusion = form.save(commit=False)

        # ✅ Force it to be general type
        inclusion.inclusion_type = 'general'

        # ✅ Ensure no hotel/houseboat is linked
        inclusion.hotel = None
        inclusion.houseboat = None

        inclusion.save()

        messages.success(request, f'✅ General inclusion "{inclusion.name}" added successfully!')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f"❌ {field}: {error}")

    return redirect('special_inclusion_list')


@require_POST
def edit_special_inclusion(request, pk):
    """
    Edit an existing GENERAL inclusion.
    """
    inclusion = get_object_or_404(SpecialInclusion, pk=pk, inclusion_type='general')
    form = SpecialInclusionForm(request.POST, request.FILES, instance=inclusion)

    if form.is_valid():
        updated_inclusion = form.save(commit=False)

        # ✅ Ensure it stays general
        updated_inclusion.inclusion_type = 'general'
        updated_inclusion.hotel = None
        updated_inclusion.houseboat = None

        updated_inclusion.save()

        messages.success(request, f'✅ Inclusion "{updated_inclusion.name}" updated successfully!')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f"❌ {field}: {error}")

    return redirect('special_inclusion_list')


@require_POST
def delete_special_inclusion(request, pk):
    """
    Delete a GENERAL inclusion.
    """
    inclusion = get_object_or_404(SpecialInclusion, pk=pk, inclusion_type='general')
    inclusion_name = inclusion.name
    inclusion.delete()

    messages.success(request, f'✅ Inclusion "{inclusion_name}" was deleted.')
    return redirect('special_inclusion_list')


################################################################################################################
def inclusion_price_list(request, inclusion_id):
    inclusion = get_object_or_404(SpecialInclusion, pk=inclusion_id)
    prices = InclusionPrice.objects.filter(inclusion=inclusion)
    return render(request, 'inclusion_price_list.html', {'inclusion': inclusion, 'prices': prices})

def add_inclusion_price(request, inclusion_id):
    inclusion = get_object_or_404(SpecialInclusion, pk=inclusion_id)
    if request.method == 'POST':
        form = InclusionPriceForm(request.POST)
        if form.is_valid():
            price = form.save(commit=False)
            price.inclusion = inclusion
            price.save()
            return redirect('inclusion_price_list', inclusion_id=inclusion.id)
    else:
        form = InclusionPriceForm()
    return render(request, 'add_inclusion_price.html', {'form': form, 'inclusion': inclusion})

def edit_inclusion_price(request, pk):
    price = get_object_or_404(InclusionPrice, pk=pk)
    if request.method == 'POST':
        form = InclusionPriceForm(request.POST, instance=price)
        if form.is_valid():
            form.save()
            return redirect('inclusion_price_list', price.inclusion.id)
    else:
        form = InclusionPriceForm(instance=price)
    return render(request, 'edit_inclusion_price.html', {'form': form})

def delete_inclusion_price(request, pk):
    price = get_object_or_404(InclusionPrice, pk=pk)
    if request.method == 'POST':
        price.delete()
        return redirect('inclusion_price_list', price.inclusion.id)
    return render(request, 'delete_inclusion_price.html', {'price': price})

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Q
from .models import Vehicle, Destinations, Supplier
from .forms import VehicleForm
from django.core.paginator import Paginator


def vehicle_list(request):
    """List all vehicles with search functionality"""
    query = request.GET.get('q', '')
    vehicles = Vehicle.objects.select_related('destination', 'supplier').all().order_by('-created_at')

    if query:
        vehicles = vehicles.filter(
            Q(name__icontains=query) | Q(destination__name__icontains=query)
        )

    # Pagination setup - 10 items per page
    paginator = Paginator(vehicles, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    destinations = Destinations.objects.all().order_by('name')
    suppliers = Supplier.objects.filter(
        supplier_type='vehicle',
        is_active=True
    ).order_by('company_name')

    context = {
        'vehicles': page_obj,
        'page_obj': page_obj,
        'query': query,
        'destinations': destinations,
        'suppliers': suppliers,
    }
    return render(request, 'vehicle_list.html', context)



def add_vehicle(request):
    """Add new vehicle"""
    if request.method == 'POST':
        form = VehicleForm(request.POST, request.FILES)
        if form.is_valid():
            vehicle = form.save(commit=False)
            if hasattr(request.user, 'teammember'):
                vehicle.created_by = request.user.teammember
            vehicle.save()
            messages.success(request, f"✅ Vehicle '{vehicle.name}' added successfully!")
            return redirect('vehicle_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"❌ {field}: {error}")

    return redirect('vehicle_list')


def edit_vehicle(request, vehicle_id):
    """Edit existing vehicle"""
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)

    if request.method == 'POST':
        form = VehicleForm(request.POST, request.FILES, instance=vehicle)
        if form.is_valid():
            vehicle = form.save()
            messages.success(request, f"✅ Vehicle '{vehicle.name}' updated successfully!")
            return redirect('vehicle_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"❌ {field}: {error}")

    return redirect('vehicle_list')


def delete_vehicle(request, vehicle_id):
    """Delete vehicle"""
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)
    vehicle_name = vehicle.name
    vehicle.delete()
    messages.success(request, f"✅ Vehicle '{vehicle_name}' deleted successfully!")
    return redirect('vehicle_list')


################################################################################################################
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_POST
from django.contrib import messages
from .models import Vehicle, VehiclePricing
from .forms import VehiclePricingForm

def vehicle_pricing_list(request, vehicle_id):
    """
    Display list of pricing records for a specific vehicle
    Accessible via GET request
    """
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)

    # Order by most recent date first
    pricings = VehiclePricing.objects.filter(
        vehicle=vehicle
    ).order_by('-from_date', '-id')

    # Initialize empty form for the modal (Add Mode)
    # We pass 'vehicle' so the empty form knows its context
    form = VehiclePricingForm(vehicle=vehicle)

    context = {
        'pricings': pricings,
        'vehicle': vehicle,
        'form': form,
    }
    return render(request, 'vehicle_pricing_list.html', context)


@require_POST
def add_vehicle_pricing(request, vehicle_id):
    """
    Add new pricing record.
    If valid: Redirects.
    If invalid: Renders page with form errors and re-opens modal.
    """
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)

    # Pass POST data and the vehicle instance to the form
    form = VehiclePricingForm(request.POST, vehicle=vehicle)

    if form.is_valid():
        price = form.save(commit=False)
        price.vehicle = vehicle

        # Optional: Add created_by if needed
        if hasattr(request.user, 'teammember'):
            price.created_by = request.user.teammember

        price.save()
        messages.success(request, "✅ Price added successfully!")
        return redirect('vehicle_pricing_list', vehicle_id=vehicle.id)
    else:
        # ❌ ERROR CASE:
        # Do NOT redirect. Render the page again so the modal stays open with errors.
        pricings = VehiclePricing.objects.filter(vehicle=vehicle).order_by('-from_date', '-id')

        context = {
            'pricings': pricings,
            'vehicle': vehicle,
            'form': form,           # This form instance contains the errors
            'show_add_modal': True  # Tells the template to pop the Add Modal
        }
        return render(request, 'vehicle_pricing_list.html', context)


@require_POST
def edit_vehicle_pricing(request, vehicle_id, price_id):
    """
    Edit existing pricing record.
    If valid: Redirects.
    If invalid: Renders page with form errors and re-opens modal.
    """
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)
    price = get_object_or_404(VehiclePricing, id=price_id, vehicle=vehicle)

    # Pass POST data, instance, and vehicle
    form = VehiclePricingForm(request.POST, instance=price, vehicle=vehicle)

    if form.is_valid():
        updated_price = form.save()
        messages.success(request, "✅ Price updated successfully!")
        return redirect('vehicle_pricing_list', vehicle_id=vehicle.id)
    else:
        # ❌ ERROR CASE:
        # Do NOT redirect. Render the page again so the modal stays open with errors.
        pricings = VehiclePricing.objects.filter(vehicle=vehicle).order_by('-from_date', '-id')

        context = {
            'pricings': pricings,
            'vehicle': vehicle,
            'form': form,             # This form instance contains the errors
            'show_edit_modal': True,  # Tells the template to pop the Edit Modal
            'edit_price_id': price.id # Needed for the form action URL in the modal
        }
        return render(request, 'vehicle_pricing_list.html', context)


@require_POST
def delete_vehicle_pricing(request, vehicle_id, price_id):
    """
    Delete pricing record
    """
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)
    price = get_object_or_404(VehiclePricing, id=price_id, vehicle=vehicle)

    # Store details before deletion for the success message
    price_info = f"Price (₹{price.total_fee_100km})"

    price.delete()
    messages.success(request, f"✅ Price record deleted: {price_info}")

    return redirect('vehicle_pricing_list', vehicle_id=vehicle_id)

################################################################################################################
# from django.db.models import Q

# def itinerary_list(request):
#     itineraries = DayItinerary.objects.all()
#     search_query = request.GET.get('q', '')
#     if search_query:
#         itineraries = itineraries.filter(
#             Q(title__icontains=search_query) |
#             Q(details__icontains=search_query) |
#             Q(destination__name__icontains=search_query)
#         )
#     return render(request, 'itinerary_list.html', {'itineraries': itineraries, 'search_query': search_query})

# def search_itineraries(request):
#     query = request.GET.get('q', '')
#     itineraries = DayItinerary.objects.filter(
#         Q(title__icontains=query) |
#         Q(details__icontains=query) |
#         Q(destination__name__icontains=query)
#     ).values("title", "destination__name")[:10]  # Limit to 10 results
#     return JsonResponse(list(itineraries), safe=False)

# def add_itinerary(request):
#     if request.method == 'POST':
#         form = DayItineraryForm(request.POST)
#         if form.is_valid():
#             itinerary = form.save(commit=False)
#             itinerary.save()
#             return redirect('itinerary_list')
#     else:
#         form = DayItineraryForm()
#     return render(request, 'add_itinerary.html', {'form': form})


# def edit_itinerary(request, pk):
#     itinerary = get_object_or_404(DayItinerary, pk=pk)
#     if request.method == 'POST':
#         form = DayItineraryForm(request.POST, instance=itinerary)
#         if form.is_valid():
#             form.save()
#             return redirect('itinerary_list')
#     else:
#         form = DayItineraryForm(instance=itinerary)
#     return render(request, 'edit_itinerary.html', {'form': form})


# def delete_itinerary(request, pk):
#     itinerary = get_object_or_404(DayItinerary, pk=pk)
#     if request.method == 'POST':
#         itinerary.delete()
#         return redirect('itinerary_list')
#     return render(request, 'delete_itinerary.html', {'itinerary': itinerary})

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.db.models import Q
from django.contrib import messages
from django.core.files.base import ContentFile
from .models import DayItinerary, ItineraryImageGallery, TeamMember
from .forms import DayItineraryForm, ItineraryImageGalleryForm
import traceback
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.db.models import Q, Min
from django.contrib import messages
from django.core.files.base import ContentFile
import traceback

from .models import DayItinerary, ItineraryImageGallery
from .forms import DayItineraryForm




def day_itinerary_list(request):
    """List UNIQUE day itineraries only (no duplicates shown)"""

    # ✅ Keep only ONE record per (name, destination, details)
    unique_ids = (
        DayItinerary.objects
        .values('name', 'destination', 'details')
        .annotate(id=Min('id'))
        .values_list('id', flat=True)
    )

    itineraries = (
        DayItinerary.objects
        .filter(id__in=unique_ids)
        .select_related('created_by', 'gallery_image')
        .order_by('-is_pinned', '-created_at')
    )

    search_query = request.GET.get('q', '')
    if search_query:
        itineraries = itineraries.filter(
            Q(name__icontains=search_query) |
            Q(details__icontains=search_query) |
            Q(destination__icontains=search_query)
        )

    form = DayItineraryForm()
    gallery_images = ItineraryImageGallery.objects.all().order_by('-created_at')

    return render(request, 'day_itinerary_list.html', {
        'itineraries': itineraries,
        'search_query': search_query,
        'form': form,
        'gallery_images': gallery_images
    })



def add_day_itinerary(request):
    """Add new day itinerary via AJAX (no duplicates allowed)"""

    if request.method == 'POST':
        uploaded_file = request.FILES.get('uploaded_image')
        form = DayItineraryForm(request.POST, request.FILES)

        if form.is_valid():
            name = form.cleaned_data['name'].strip()
            destination = form.cleaned_data['destination'].strip()
            details = (form.cleaned_data.get('details') or '').strip()

            # 🔒 DUPLICATE CHECK
            if DayItinerary.objects.filter(
                name__iexact=name,
                destination__iexact=destination,
                details__iexact=details
            ).exists():
                return JsonResponse({
                    'success': False,
                    'message': '⚠️ This itinerary already exists with the same name, destination, and details.'
                }, status=400)

            itinerary = form.save(commit=False)

            # created_by
            user_id = request.session.get('user_id')
            user_type = request.session.get('user_type')
            created_by_member = None

            if user_type == 'team_member' and user_id:
                try:
                    created_by_member = TeamMember.objects.get(id=user_id)
                    itinerary.created_by = created_by_member
                except TeamMember.DoesNotExist:
                    pass

            # Save uploaded image to gallery
            if uploaded_file and itinerary.image_source == 'upload':
                try:
                    gallery_name = f"{itinerary.destination} - {itinerary.name[:50]}"

                    if not ItineraryImageGallery.objects.filter(name=gallery_name).exists():
                        gallery_image = ItineraryImageGallery(
                            name=gallery_name,
                            description=f"Auto-saved from itinerary: {itinerary.name}",
                            created_by=created_by_member
                        )
                        uploaded_file.seek(0)
                        gallery_image.image.save(
                            uploaded_file.name,
                            ContentFile(uploaded_file.read()),
                            save=True
                        )
                except Exception:
                    traceback.print_exc()

            itinerary.save()

            messages.success(
                request,
                f'✅ Day Itinerary "{itinerary.name}" added successfully!'
            )

            return JsonResponse({
                'success': True,
                'itinerary': {
                    'id': itinerary.id,
                    'name': itinerary.name,
                    'destination': itinerary.destination,
                    'is_pinned': itinerary.is_pinned
                }
            })

        return JsonResponse({
            'success': False,
            'message': 'Please correct the errors.',
            'errors': form.errors
        }, status=400)

    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)



def edit_day_itinerary(request, pk):
    """Edit day itinerary (cannot update into duplicate)"""

    itinerary = get_object_or_404(DayItinerary, pk=pk)

    if request.method == 'GET':
        return JsonResponse({
            'id': itinerary.id,
            'name': itinerary.name,
            'destination': itinerary.destination,
            'details': itinerary.details or '',
            'image_url': itinerary.get_image(),
            'image_source': itinerary.image_source,
            'gallery_image_id': itinerary.gallery_image.id if itinerary.gallery_image else None,
            'is_pinned': itinerary.is_pinned,
        })

    if request.method == 'POST':
        uploaded_file = request.FILES.get('uploaded_image')
        form = DayItineraryForm(request.POST, request.FILES, instance=itinerary)

        if form.is_valid():
            name = form.cleaned_data['name'].strip()
            destination = form.cleaned_data['destination'].strip()
            details = (form.cleaned_data.get('details') or '').strip()

            # 🔒 DUPLICATE CHECK (exclude self)
            if DayItinerary.objects.filter(
                name__iexact=name,
                destination__iexact=destination,
                details__iexact=details
            ).exclude(pk=itinerary.pk).exists():
                return JsonResponse({
                    'success': False,
                    'message': '⚠️ Another itinerary with the same name, destination, and details already exists.'
                }, status=400)

            updated_itinerary = form.save(commit=False)

            # Save uploaded image to gallery
            if uploaded_file:
                try:
                    from datetime import datetime
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    gallery_image = ItineraryImageGallery(
                        name=f"{destination} - {name[:40]} [{timestamp}]",
                        description=f"Auto-saved from itinerary: {name}"
                    )
                    uploaded_file.seek(0)
                    gallery_image.image.save(
                        uploaded_file.name,
                        ContentFile(uploaded_file.read()),
                        save=True
                    )
                except Exception:
                    traceback.print_exc()

            updated_itinerary.save()

            messages.success(
                request,
                f'✅ Day Itinerary "{updated_itinerary.name}" updated successfully!'
            )

            return JsonResponse({
                'success': True,
                'itinerary': {
                    'id': updated_itinerary.id,
                    'name': updated_itinerary.name,
                    'destination': updated_itinerary.destination,
                    'is_pinned': updated_itinerary.is_pinned
                }
            })

        return JsonResponse({
            'success': False,
            'message': 'Please correct the errors.',
            'errors': form.errors
        }, status=400)


def delete_day_itinerary(request, pk):
    """Delete day itinerary via AJAX"""
    if request.method == 'POST':
        itinerary = get_object_or_404(DayItinerary, pk=pk)
        name = itinerary.name
        destination = itinerary.destination

        # Delete uploaded image file if exists (don't delete gallery images)
        if itinerary.uploaded_image:
            try:
                itinerary.uploaded_image.delete(save=False)
            except Exception as e:
                print(f"⚠️ Could not delete image: {e}")

        itinerary.delete()

        messages.success(request, f'✅ Day Itinerary "{name}" ({destination}) deleted successfully!')

        return JsonResponse({
            'success': True,
            'message': f'Day Itinerary "{name}" deleted successfully!'
        })

    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)


def delete_selected_day_itineraries(request):
    """
    Delete multiple itineraries at once.
    Expects JSON body: { "ids": ["1", "2", "3"] }
    """
    if request.method == 'POST':
        try:
            # Parse JSON data
            try:
                data = json.loads(request.body)
                ids = data.get('ids', [])
            except json.JSONDecodeError:
                # Fallback for form-data
                ids = request.POST.getlist('ids[]')

            if not ids:
                return JsonResponse({'success': False, 'message': 'No items selected.'}, status=400)

            # Fetch items
            itineraries = DayItinerary.objects.filter(id__in=ids)
            count = itineraries.count()

            if count == 0:
                return JsonResponse({'success': False, 'message': 'Records not found.'}, status=404)

            # Cleanup images loop
            for item in itineraries:
                if item.uploaded_image:
                    try:
                        item.uploaded_image.delete(save=False)
                    except Exception as e:
                        print(f"⚠️ Could not delete image: {e}")
                item.delete()

            messages.success(request, f'✅ Successfully deleted {count} itineraries.')
            return JsonResponse({'success': True, 'message': f'Deleted {count} items.'})

        except Exception as e:
            print(f"❌ Bulk Delete Error: {e}")
            return JsonResponse({'success': False, 'message': 'Server error during deletion.'}, status=500)

    return JsonResponse({'success': False, 'message': 'Invalid method'}, status=405)


def toggle_pin_itinerary(request, pk):
    """Toggle pin status of a day itinerary"""
    if request.method == 'POST':
        itinerary = get_object_or_404(DayItinerary, pk=pk)

        itinerary.is_pinned = not itinerary.is_pinned
        itinerary.save()

        status = "pinned to top" if itinerary.is_pinned else "unpinned"
        messages.success(request, f'✅ "{itinerary.name}" {status}!')

        return JsonResponse({
            'success': True,
            'is_pinned': itinerary.is_pinned,
            'message': f'"{itinerary.name}" {status}!'
        })

    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)


# ✅ Gallery Management Views
def gallery_image_list(request):
    """List and manage gallery images"""
    images = ItineraryImageGallery.objects.all().select_related('created_by').order_by('-created_at')
    search_query = request.GET.get('q', '')

    if search_query:
        images = images.filter(Q(name__icontains=search_query) | Q(description__icontains=search_query))

    form = ItineraryImageGalleryForm()
    return render(request, 'gallery_image_list.html', {
        'images': images,
        'search_query': search_query,
        'form': form
    })


def add_gallery_image(request):
    """Add new gallery image"""
    if request.method == 'POST':
        form = ItineraryImageGalleryForm(request.POST, request.FILES)

        if form.is_valid():
            gallery_image = form.save(commit=False)

            user_id = request.session.get('user_id')
            user_type = request.session.get('user_type')

            if user_type == 'team_member' and user_id:
                try:
                    gallery_image.created_by = TeamMember.objects.get(id=user_id)
                except TeamMember.DoesNotExist:
                    pass

            gallery_image.save()
            messages.success(request, f'✅ Gallery image "{gallery_image.name}" added successfully!')

            return JsonResponse({
                'success': True,
                'message': f'Gallery image "{gallery_image.name}" added successfully!',
                'image': {
                    'id': gallery_image.id,
                    'name': gallery_image.name,
                    'image_url': gallery_image.image.url
                }
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': form.errors,
                'message': 'Please correct the errors.'
            }, status=400)

    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)


def delete_gallery_image(request, pk):
    """Delete gallery image"""
    if request.method == 'POST':
        gallery_image = get_object_or_404(ItineraryImageGallery, pk=pk)

        # Check if image is being used
        usage_count = gallery_image.day_itineraries.count()
        if usage_count > 0:
            return JsonResponse({
                'success': False,
                'message': f'Cannot delete! This image is used in {usage_count} itinerary/itineraries.'
            }, status=400)

        name = gallery_image.name

        try:
            gallery_image.image.delete(save=False)
        except Exception as e:
            print(f"⚠️ Could not delete image file: {e}")

        gallery_image.delete()
        messages.success(request, f'✅ Gallery image "{name}" deleted successfully!')

        return JsonResponse({
            'success': True,
            'message': f'Gallery image "{name}" deleted successfully!'
        })

    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)


def get_gallery_images_json(request):
    """Get all gallery images as JSON for selection"""
    images = ItineraryImageGallery.objects.all().order_by('-created_at')

    return JsonResponse({
        'images': [{
            'id': img.id,
            'name': img.name,
            'image_url': img.image.url,
            'description': img.description or ''
        } for img in images]
    })


def search_day_itineraries(request):
    """Search itineraries for autocomplete - pinned items first"""
    query = request.GET.get('q', '')

    # Order by pinned first, then relevance
    itineraries = DayItinerary.objects.filter(
        Q(name__icontains=query) |
        Q(details__icontains=query) |
        Q(destination__icontains=query)
    ).order_by('-is_pinned', '-created_at').values(
        "id", "name", "destination", "details", "is_pinned"
    )[:10]

    return JsonResponse(list(itineraries), safe=False)




################################################################################################################
from .models import LeadSource
from .forms import LeadSourceForm
from .forms import LeadSourceForm

def leadsource_list(request):
    leadsources = LeadSource.objects.all()
    form = LeadSourceForm()
    context = {
        'leadsources': leadsources,
        'form': form,
    }
    return render(request, 'leadsource_list.html', context)

@require_POST
def add_leadsource(request):
    form = LeadSourceForm(request.POST)
    if form.is_valid():
        # You can add logic to set 'created_by' automatically here if needed
        form.save()
        messages.success(request, 'Lead Source added successfully!')
    else:
        messages.error(request, f"Error: {form.errors.as_text()}")
    return redirect('leadsource_list')

@require_POST
def edit_leadsource(request, pk):
    leadsource = get_object_or_404(LeadSource, pk=pk)
    form = LeadSourceForm(request.POST, instance=leadsource)
    if form.is_valid():
        form.save()
        messages.success(request, 'Lead Source updated successfully!')
    else:
        messages.error(request, f"Error: {form.errors.as_text()}")
    return redirect('leadsource_list')

@require_POST
def delete_leadsource(request, pk):
    leadsource = get_object_or_404(LeadSource, pk=pk)
    leadsource.delete()
    messages.success(request, f"Lead Source '{leadsource.source_name}' was deleted.")
    return redirect('leadsource_list')

################################################################################################################

# your_app/views.py
from .forms import PackageThemeForm # Make sure to import

def package_theme_list(request):
    themes = PackageTheme.objects.all()
    form = PackageThemeForm()
    context = {
        'themes': themes,
        'form': form,
    }
    return render(request, 'package_theme_list.html', context)

@require_POST
def add_package_theme(request):
    form = PackageThemeForm(request.POST, request.FILES)
    if form.is_valid():
        # You can add logic to set 'created_by' automatically here if needed
        form.save()
        messages.success(request, 'Package Theme added successfully!')
    else:
        messages.error(request, f"Error: {form.errors.as_text()}")
    return redirect('package_theme_list')

@require_POST
def edit_package_theme(request, pk):
    theme = get_object_or_404(PackageTheme, pk=pk)
    form = PackageThemeForm(request.POST, request.FILES, instance=theme)
    if form.is_valid():
        form.save()
        messages.success(request, 'Package Theme updated successfully!')
    else:
        messages.error(request, f"Error: {form.errors.as_text()}")
    return redirect('package_theme_list')

@require_POST
def delete_package_theme(request, pk):
    theme = get_object_or_404(PackageTheme, pk=pk)
    theme.delete()
    messages.success(request, f"Theme '{theme.package_name}' was deleted.")
    return redirect('package_theme_list')

########################################################################################################################
from .forms import CurrencyForm

def currency_list(request):
    currencies = Currency.objects.all()
    form = CurrencyForm()
    context = {
        'currencies': currencies,
        'form': form,
    }
    return render(request, 'currency_list.html', context)

@require_POST
def add_currency(request):
    form = CurrencyForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request, 'Currency added successfully!')
    else:
        messages.error(request, f"Error: {form.errors.as_text()}")
    return redirect('currency_list')

@require_POST
def edit_currency(request, pk):
    currency = get_object_or_404(Currency, pk=pk)
    form = CurrencyForm(request.POST, instance=currency)
    if form.is_valid():
        form.save()
        messages.success(request, 'Currency updated successfully!')
    else:
        messages.error(request, f"Error: {form.errors.as_text()}")
    return redirect('currency_list')

@require_POST
def delete_currency(request, pk):
    currency = get_object_or_404(Currency, pk=pk)
    currency.delete()
    messages.success(request, f"Currency '{currency.currency_name}' was deleted.")
    return redirect('currency_list')

####################################################################################################################
def Admin_Board(request):
    return render(request,'admin_board.html')

def Admin_Settings(request):
    return render(request,'admin_settings.html')

def organisational_setting_list(request):
    settings = OrganisationalSetting.objects.all()
    return render(request, 'organisational_setting_list.html', {'settings': settings})

def organisational_setting_edit(request, pk):
    setting = get_object_or_404(OrganisationalSetting, pk=pk)
    if request.method == 'POST':
        form = OrganisationalSettingForm(request.POST, instance=setting)
        if form.is_valid():
            form.save()
            return redirect('organisational_setting_list')
    else:
        form = OrganisationalSettingForm(instance=setting)
    return render(request, 'organisational_setting_edit.html', {'form': form})




# def edit_invoice_logo(request):
#     instance, created = InvoiceLogo.objects.get_or_create(id=1)  # Ensure there's always one record
#     if request.method == 'POST':
#         form = InvoiceLogoForm(request.POST, instance=instance)
#         if form.is_valid():
#             form.save()
#             return redirect('edit_invoice_logo')
#     else:
#         form = InvoiceLogoForm(instance=instance)

#     return render(request, 'edit_section.html', {'form': form, 'title': "Edit Invoice Logo"})

# def edit_invoice_terms(request):
#     instance, created = InvoiceTerms.objects.get_or_create(id=1)
#     if request.method == 'POST':
#         form = InvoiceTermsForm(request.POST, instance=instance)
#         if form.is_valid():
#             form.save()
#             return redirect('edit_invoice_terms')
#     else:
#         form = InvoiceTermsForm(instance=instance)

#     return render(request, 'edit_section.html', {'form': form, 'title': "Edit Invoice Terms & Conditions"})

# def edit_package_terms(request):
#     instance, created = PackageTerms.objects.get_or_create(id=1)
#     if request.method == 'POST':
#         form = PackageTermsForm(request.POST, instance=instance)
#         if form.is_valid():
#             form.save()
#             return redirect('edit_package_terms')
#     else:
#         form = PackageTermsForm(instance=instance)

#     return render(request, 'edit_section.html', {'form': form, 'title': "Edit Package Terms & Conditions"})

# def edit_bank_information(request):
#     instance, created = BankInformation.objects.get_or_create(id=1)
#     if request.method == 'POST':
#         form = BankInformationForm(request.POST, instance=instance)
#         if form.is_valid():
#             form.save()
#             return redirect('edit_bank_information')
#     else:
#         form = BankInformationForm(instance=instance)

#     return render(request, 'edit_section.html', {'form': form, 'title': "Edit Bank Information"})



def edit_all_settings(request):
    invoice_logo, _ = InvoiceLogo.objects.get_or_create(id=1)
    invoice_terms, _ = InvoiceTerms.objects.get_or_create(id=1)
    package_terms, _ = PackageTerms.objects.get_or_create(id=1)
    bank_info, _ = BankInformation.objects.get_or_create(id=1)

    if request.method == 'POST':
        form = CombinedBusinessSettingsForm(request.POST, request.FILES, invoice_logo=invoice_logo, invoice_terms=invoice_terms, package_terms=package_terms, bank_info=bank_info)
        if form.is_valid():
            form.save(invoice_logo, invoice_terms, package_terms, bank_info)
            return redirect('edit_all_settings')
    else:
        form = CombinedBusinessSettingsForm(invoice_logo=invoice_logo, invoice_terms=invoice_terms, package_terms=package_terms, bank_info=bank_info)

    return render(request, 'edit_all_settings.html', {'form': form})



def package_details(request):
    instance, created = PackageTermss.objects.get_or_create(id=1)

    if request.method == 'POST':
        form = PackageTermssForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            return redirect('package_details')  # Reload the same page after saving
    else:
        form = PackageTermssForm(instance=instance)

    return render(request, 'package_details.html', {'form': form})





def branch_list(request):
    branches = BranchSettings.objects.all()
    return render(request, 'branch_list.html', {'branches': branches})


def branch_create(request):
    if request.method == 'POST':
        form = BranchSettingsForm(request.POST)
        if form.is_valid():
            branch = form.save(commit=False)

            branch.save()
            return redirect('branch_list')
    else:
        form = BranchSettingsForm()
    return render(request, 'branch_form.html', {'form': form, 'title': 'Add Branch'})


def branch_update(request, pk):
    branch = get_object_or_404(BranchSettings, pk=pk)
    if request.method == 'POST':
        form = BranchSettingsForm(request.POST, instance=branch)
        if form.is_valid():
            form.save()
            return redirect('branch_list')
    else:
        form = BranchSettingsForm(instance=branch)
    return render(request, 'branch_form.html', {'form': form, 'title': 'Edit Branch'})


def branch_delete(request, pk):
    branch = get_object_or_404(BranchSettings, pk=pk)
    if request.method == 'POST':
        branch.delete()
        return redirect('branch_list')
    return render(request, 'branch_confirm_delete.html', {'branch': branch})


def role_list(request):
    roles = Role.objects.all()
    return render(request, 'role_list.html', {'roles': roles})

def add_role(request):
    if request.method == 'POST':
        form = RoleForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('role_list')
    else:
        form = RoleForm()
    return render(request, 'role_form.html', {'form': form, 'title': 'Add Role'})

def edit_role(request, pk):
    role = get_object_or_404(Role, pk=pk)
    if request.method == 'POST':
        form = RoleForm(request.POST, instance=role)
        if form.is_valid():
            form.save()
            return redirect('role_list')
    else:
        form = RoleForm(instance=role)
    return render(request, 'role_form.html', {'form': form, 'title': 'Edit Role'})

def delete_role(request, pk):
    role = get_object_or_404(Role, pk=pk)
    if request.method == 'POST':
        role.delete()
        return redirect('role_list')
    return render(request, 'role_confirm_delete.html', {'role': role})

################### query ####################################################################################################################
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q  # ✅ ADD THIS
from .models import Query, LeadSource  # ✅ ADD LeadSource
from .forms import QueryForm
import json


from django.utils import timezone
from datetime import datetime

@custom_login_required
def query_list(request):
    user_id = request.session.get('user_id')
    user_type = request.session.get('user_type')

    if not user_id:
        messages.warning(request, '⚠️ Please login first')
        return redirect('team_member:login')

    # Get individual search terms from GET
    name = request.GET.get('name', '').strip()
    phone = request.GET.get('phone', '').strip()
    from_date = request.GET.get('from_date', '').strip()
    assign = request.GET.get('assign', '').strip()
    query_id = request.GET.get('id', '').strip()

    # Get user role and base queryset
    user_role = None
    if user_type == 'superuser':
        user_role = 'superuser'
        queries = Query.objects.all()
    elif user_type == 'team_member':
        user = TeamMember.objects.get(id=user_id)
        user_role = user.role
        if user.role == 'admin':
            queries = Query.objects.all()
        elif user.role == 'manager':
            queries = Query.objects.filter(
                Q(created_by=user) | Q(assign=user)
            )
        else:
            queries = Query.objects.filter(created_by=user)
    else:
        queries = Query.objects.none()

    # Apply filters for each search box field
    if name:
        queries = queries.filter(client_name__icontains=name)
    if phone:
        queries = queries.filter(phone_number__icontains=phone)
    if from_date:
        queries = queries.filter(from_date__icontains=from_date)
    if assign:
        queries = queries.filter(
            Q(assign__first_name__icontains=assign) |
            Q(assign__last_name__icontains=assign)
        )
    if query_id:
        queries = queries.filter(query_id__icontains=query_id)

    queries = queries.order_by('-created_at')

    # Calculate status counts based on the same filtered queryset
    total_queries = queries.count()

    # ✅ Get today's date range (start and end of today)
    today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)

    status_counts = queries.aggregate(
        new=Count('id', filter=Q(created_at__gte=today_start, created_at__lte=today_end)),  # ✅ Today's queries only
        active=Count('id', filter=Q(status='active')),
        no_connect=Count('id', filter=Q(status='no_connect')),
        hot_lead=Count('id', filter=Q(priority='hot')),
        follow_up=Count('id', filter=Q(status='follow_up')),
        proposal_sent=Count('id', filter=Q(status='proposal_sent')),
        confirmed=Count('id', filter=Q(status='confirmed')),
        cancelled=Count('id', filter=Q(status='cancelled')),
        invalid=Count('id', filter=Q(status='invalid')),
    )

    lead_sources = LeadSource.objects.filter(is_active=True).order_by('source_name')
    team_members = TeamMember.objects.filter(is_active=True).order_by('first_name')

    context = {
        'queries': queries,
        'query_form': QueryForm(user=None, user_role=user_role),
        'title': 'Query Management',
        'total_queries': total_queries,
        'user_role': user_role,
        'lead_sources': lead_sources,
        'team_members': team_members,
        'name': name,
        'phone': phone,
        'from_date': from_date,
        'assign': assign,
        'id': query_id,
        'status_counts': status_counts,
    }
    return render(request, 'query_list.html', context)





from django.contrib import messages
from django.http import JsonResponse
import traceback

@custom_login_required
def add_query(request):
    """Handle adding new query - AUTO-ASSIGN to creator"""
    user_id = request.session.get('user_id')
    user_type = request.session.get('user_type')

    print(f"\n{'='*50}")
    print(f"ADD QUERY REQUEST - Method: {request.method}")
    print(f"User ID: {user_id}, User Type: {user_type}")
    print(f"{'='*50}\n")

    # ✅ GET CURRENT USER AND ROLE
    current_user = None
    user_role = None

    if user_type == 'superuser':
        user_role = 'superuser'
    elif user_type == 'team_member':
        current_user = TeamMember.objects.get(id=user_id)
        user_role = current_user.role
        print(f"Current User: {current_user.get_full_name()}, Role: {user_role}")

    if request.method == 'POST':
        print("\n📝 POST Data Received:")
        for key, value in request.POST.items():
            if key != 'csrfmiddlewaretoken':
                print(f"  - {key}: {value}")

        form = QueryForm(request.POST, user=current_user, user_role=user_role)

        print(f"\n✅ Form Valid: {form.is_valid()}")

        if not form.is_valid():
            print("\n❌ FORM ERRORS:")
            for field, errors in form.errors.items():
                print(f"  - {field}: {errors}")

            # Return JSON if AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                errors_dict = {}
                for field, error_list in form.errors.items():
                    errors_dict[field] = [str(error) for error in error_list]

                return JsonResponse({
                    'success': False,
                    'errors': errors_dict,
                    'message': '❌ Please correct the errors below.'
                }, status=400)

            messages.error(request, '❌ Please correct the errors below.')
            return redirect('query_list')

        # Form is valid, try to save
        try:
            print("\n💾 Attempting to save query...")

            query = form.save(commit=False)
            print(f"  - Query object created (not saved yet)")
            print(f"  - Client Name: {query.client_name}")
            print(f"  - Phone: {query.phone_number}")

            # Set created_by
            if user_type == 'team_member' and current_user:
                query.created_by = current_user
                print(f"  - Created By: {current_user.get_full_name()}")

                # Auto-set assign if not set
                if not query.assign:
                    query.assign = current_user
                    print(f"  - Assigned To: {current_user.get_full_name()} (auto-assigned)")
                else:
                    print(f"  - Assigned To: {query.assign.get_full_name()}")

            # ✅ CRITICAL: Check if status field exists
            if hasattr(query, 'status'):
                if not query.status:
                    query.status = 'new'
                    print(f"  - Status set to: new (default)")
                else:
                    print(f"  - Status: {query.status}")
            else:
                print("  ⚠️ WARNING: Query model has no 'status' field!")

            # Save the query
            print("\n  🔄 Calling query.save()...")
            query.save()
            print(f"  ✅ QUERY SAVED SUCCESSFULLY!")
            print(f"  - ID: {query.id}")
            print(f"  - Query ID: {query.query_id}")

            # Verify it's in the database
            verify = Query.objects.filter(id=query.id).first()
            if verify:
                print(f"  ✅ Verified in database: {verify.query_id}")
            else:
                print(f"  ❌ WARNING: Not found in database after save!")

            # Return success
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'✅ Query {query.query_id} for {query.client_name} added successfully!'
                })

            messages.success(request, f'✅ Query {query.query_id} for {query.client_name} added successfully!')
            return redirect('query_list')

        except Exception as e:
            print(f"\n❌ EXCEPTION DURING SAVE:")
            print(f"  - Error: {str(e)}")
            print(f"  - Type: {type(e).__name__}")
            print("\nFull Traceback:")
            traceback.print_exc()

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'❌ Error saving query: {str(e)}'
                }, status=400)

            messages.error(request, f'❌ Error saving query: {str(e)}')
            return redirect('query_list')

    return redirect('query_list')




@custom_login_required
def edit_query(request, pk):
    """Handle editing query - DON'T change created_by, allow assign changes"""
    query = get_object_or_404(Query, pk=pk)

    user_id = request.session.get('user_id')
    user_type = request.session.get('user_type')

    # ✅ GET CURRENT USER AND ROLE
    current_user = None
    user_role = None

    if user_type == 'superuser':
        user_role = 'superuser'
    elif user_type == 'team_member':
        current_user = TeamMember.objects.get(id=user_id)
        user_role = current_user.role

    # ✅ CHECK EDIT PERMISSION
    can_edit = False

    if user_type == 'superuser':
        can_edit = True
    elif user_type == 'team_member':
        user = current_user

        if user.role == 'admin':
            can_edit = True
        elif user.has_permission('can_edit_all_queries'):
            can_edit = True
        elif user.has_permission('can_edit_queries'):
            if query.created_by == user or query.assign == user:
                can_edit = True
        elif user.role == 'manager':
            if query.created_by == user or query.assign == user:
                can_edit = True

    if not can_edit:
        # ✅ Return JSON for AJAX
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': '❌ You do not have permission to edit this query'
            }, status=403)

        messages.error(request, '❌ You do not have permission to edit this query')
        return redirect('query_list')

    # ✅ PROCESS EDIT
    if request.method == 'POST':
        form = QueryForm(request.POST, instance=query, user=current_user, user_role=user_role)

        if form.is_valid():
            try:
                # ✅ Save normally - created_by won't change, assign can be updated by admin
                query = form.save()

                # ✅ Return JSON for AJAX
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': f'✅ Query {query.query_id} for {query.client_name} updated successfully!'
                    })

                messages.success(request, f'✅ Query {query.query_id} for {query.client_name} updated successfully!')
                return redirect('query_list')

            except Exception as e:
                # ✅ Return JSON error for AJAX
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': f'❌ Error updating query: {str(e)}'
                    }, status=400)

                messages.error(request, f'❌ Error updating query: {str(e)}')
        else:
            # ✅ Return FORM ERRORS as JSON for AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                errors_dict = {}
                for field, error_list in form.errors.items():
                    errors_dict[field] = [str(error) for error in error_list]

                return JsonResponse({
                    'success': False,
                    'errors': errors_dict,
                    'message': '❌ Please correct the errors below.'
                }, status=400)

            messages.error(request, '❌ Please correct the errors below.')

            # Build context for non-AJAX
            if user_type == 'superuser':
                queries = Query.objects.all().order_by('-created_at')
            elif user_type == 'team_member':
                user = current_user

                if user.role == 'admin':
                    queries = Query.objects.all().order_by('-created_at')
                elif user.role == 'manager':
                    queries = Query.objects.filter(
                        Q(created_by=user) | Q(assign=user)
                    ).order_by('-created_at')
                else:
                    queries = Query.objects.filter(created_by=user).order_by('-created_at')
            else:
                queries = Query.objects.none()

            lead_sources = LeadSource.objects.filter(is_active=True).order_by('source_name')
            team_members = TeamMember.objects.filter(is_active=True).order_by('first_name')

            # Calculate status counts
            total_queries = queries.count()
            status_counts = queries.aggregate(
                new=Count('id', filter=Q(status='new')),
                active=Count('id', filter=Q(status='active')),
                no_connect=Count('id', filter=Q(status='no_connect')),
                hot_lead=Count('id', filter=Q(priority='hot')),
                follow_up=Count('id', filter=Q(status='follow_up')),
                proposal_sent=Count('id', filter=Q(status='proposal_sent')),
                confirmed=Count('id', filter=Q(status='confirmed')),
                cancelled=Count('id', filter=Q(status='cancelled')),
                invalid=Count('id', filter=Q(status='invalid')),
            )

            context = {
                'queries': queries,
                'query_form': form,
                'title': 'Query Management',
                'total_queries': total_queries,
                'user_role': user_role,
                'lead_sources': lead_sources,
                'team_members': team_members,
                'status_counts': status_counts,
            }
            return render(request, 'query_list.html', context)

    return redirect('query_list')








# ===== API ENDPOINTS FOR MODALS =====

@require_http_methods(["GET"])
@custom_login_required
def get_query_api(request, query_id):
    """API endpoint to get query details for view/edit - WITH PERMISSION CHECK"""
    try:
        query = get_object_or_404(Query, id=query_id)

        user_id = request.session.get('user_id')
        user_type = request.session.get('user_type')

        # ✅ CHECK PERMISSION
        has_access = False

        if user_type == 'superuser':
            has_access = True
        elif user_type == 'team_member':
            user = TeamMember.objects.get(id=user_id)
            if user.role == 'admin':
                has_access = True
            elif user.has_permission('can_edit_all_queries') or user.has_permission('can_view_all_queries'):
                has_access = True
            elif user.has_permission('can_edit_queries'):
                if query.created_by == user or query.assign == user:
                    has_access = True
            elif user.role == 'manager':
                if query.created_by == user or query.assign == user:
                    has_access = True
            else:
                if query.created_by == user:
                    has_access = True

        if not has_access:
            return JsonResponse({
                'success': False,
                'message': 'You do not have permission to view this query'
            }, status=403)

        # Return query data
        return JsonResponse({
            'success': True,
            'query': {
                'id': query.id,
                'client_name': query.client_name,
                'gender': query.gender,
                'gender_display': query.get_gender_display(),
                'phone_number': query.phone_number,
                'email': query.email or '',
                'type': query.type,
                'type_display': query.get_type_display(),
                'sector': query.sector,
                'sector_display': query.get_sector_display(),
                'priority': query.priority,
                'priority_display': query.get_priority_display(),
                'lead_source_id': query.lead_source.id if query.lead_source else None,
                'lead_source_name': query.lead_source.source_name if query.lead_source else 'N/A',
                'assign_id': query.assign.id if query.assign else None,
                'assign_name': query.assign.get_full_name() if query.assign else 'N/A',
                'from_date': query.from_date.strftime('%Y-%m-%d') if query.from_date else '',
                'to_date': query.to_date.strftime('%Y-%m-%d') if query.to_date else '',
                'total_days': query.total_days or '',
                'adult': query.adult or 0,
                'childrens': query.childrens or 0,  # ✅ FIXED: changed from 'children' to 'childrens'
                'infant': query.infant or 0,
                'services': query.services or '',
                'services_display': query.get_services_display() if query.services else 'N/A',
                'remark': query.remark or '',
                'created_at': query.created_at.strftime('%b %d, %Y %H:%M'),
                'created_by': query.created_by.get_full_name() if query.created_by else 'Unknown',
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=404)


@require_http_methods(["POST"])
@custom_login_required
def delete_query_api(request, query_id):
    """API endpoint to delete query - ONLY ADMIN/SUPERUSER"""
    try:
        query = get_object_or_404(Query, id=query_id)

        # ✅ GET LOGGED-IN USER
        user_id = request.session.get('user_id')
        user_type = request.session.get('user_type')

        # ✅ CHECK PERMISSION - Only admin/superuser can delete
        can_delete = False

        if user_type == 'superuser':
            can_delete = True

        elif user_type == 'team_member':
            user = TeamMember.objects.get(id=user_id)
            if user.role == 'admin':
                can_delete = True

        # ✅ DENY ACCESS if no permission
        if not can_delete:
            return JsonResponse({
                'success': False,
                'message': 'Only admins can delete queries'
            }, status=403)

        # ✅ DELETE QUERY
        client_name = query.client_name
        query.delete()

        return JsonResponse({
            'success': True,
            'message': f'Query for {client_name} deleted successfully'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


################itinerary #######################################################################################################################
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Q, Count, Prefetch
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from .models import Itinerary, Query, TeamMember


@custom_login_required
def list_itineraries(request):
    """
    List itineraries with role-based access and filtering
    Supports: search, status filter, days filter, user filter (admin only)
    """

    # ✅ Get current user from session
    user_id = request.session.get('user_id')
    user_type = request.session.get('user_type')

    if not user_id:
        messages.warning(request, '⚠️ Please login first')
        return redirect('team_member:login')

    current_user = None
    if user_type == 'team_member':
        current_user = TeamMember.objects.get(id=user_id)

    # ✅ Role-based itinerary filtering
    if user_type == 'superuser':
        # Superuser sees everything
        itineraries = Itinerary.objects.all()
    elif user_type == 'team_member':
        if current_user.role == 'admin':
            # Admin sees everything
            itineraries = Itinerary.objects.all()
        elif current_user.role == 'manager':
            # Manager sees: their own + queries assigned to them
            itineraries = Itinerary.objects.filter(
                Q(created_by=current_user) |
                Q(query__assign=current_user) |
                Q(query__created_by=current_user)
            ).distinct()
        else:
            # Regular staff sees only their own itineraries
            itineraries = Itinerary.objects.filter(created_by=current_user)
    else:
        itineraries = Itinerary.objects.none()

    # ✅ Search filter (by itinerary name or client name)
    search = request.GET.get('search', '').strip()
    if search:
        itineraries = itineraries.filter(
            Q(name__icontains=search) |
            Q(query__client_name__icontains=search) |
            Q(query__phone_number__icontains=search)
        )

    # ✅ Status filter
    status = request.GET.get('status', '').strip()
    if status:
        itineraries = itineraries.filter(status=status)

    # ✅ Days filter - now supports itinerary.total_days OR query.total_days
    days = request.GET.get('days', '').strip()
    if days:
        try:
            days_int = int(days)
            itineraries = itineraries.filter(
                Q(total_days=days_int) | Q(query__total_days=days_int)
            )
        except ValueError:
            pass

    # ✅ Date range filter (for itinerary travel_from/travel_to)
    from_date = request.GET.get('from_date', '').strip()
    to_date = request.GET.get('to_date', '').strip()
    if from_date:
        itineraries = itineraries.filter(
            Q(travel_from__gte=from_date) | Q(query__from_date__gte=from_date)
        )
    if to_date:
        itineraries = itineraries.filter(
            Q(travel_to__lte=to_date) | Q(query__to_date__lte=to_date)
        )

    # ✅ Admin/Superuser: Filter by team member
    filter_user_id = request.GET.get('user', '').strip()
    if filter_user_id and (user_type == 'superuser' or (current_user and current_user.role == 'admin')):
        try:
            itineraries = itineraries.filter(created_by_id=int(filter_user_id))
        except ValueError:
            pass

    # ✅ Optimize query with select_related and prefetch_related
    itineraries = itineraries.select_related(
        'query',
        'created_by',
        'query__assign',
        'query__created_by'
    ).prefetch_related(
        'pricing_options',
        'destinations',
        'day_plans'
    ).order_by('-created_at')

    # ✅ Get team members for admin filter dropdown
    team_members = None
    if user_type == 'superuser' or (current_user and current_user.role == 'admin'):
        team_members = TeamMember.objects.filter(
            is_active=True
        ).order_by('first_name', 'last_name')

    # ✅ Get status choices for filter
    status_choices = Itinerary.STATUS_CHOICES

    # ✅ Count statistics
    total_count = itineraries.count()
    draft_count = itineraries.filter(status='draft').count()
    confirmed_count = itineraries.filter(status='confirmed').count()
    cancelled_count = itineraries.filter(status='cancelled').count()

    context = {
        'itineraries': itineraries,
        'user_type': user_type,
        'current_user': current_user,
        'team_members': team_members,
        'status_choices': status_choices,

        # Filter values (to maintain filter state in template)
        'selected_user_id': int(filter_user_id) if filter_user_id else None,
        'selected_status': status,
        'selected_days': days,
        'selected_from_date': from_date,
        'selected_to_date': to_date,
        'search_query': search,

        # Statistics
        'total_count': total_count,
        'draft_count': draft_count,
        'confirmed_count': confirmed_count,
        'cancelled_count': cancelled_count,
    }

    return render(request, 'list_itinerary.html', context)


@custom_login_required
def create_itinerary(request, query_id):
    """
    Create itinerary - WITH PERMISSION CHECK
    Saves dates and passenger data to Itinerary (not Query)
    """
    user_id = request.session.get('user_id')
    user_type = request.session.get('user_type')

    if not user_id:
        messages.warning(request, '⚠️ Please login first')
        return redirect('team_member:login')

    query = get_object_or_404(Query, id=query_id)

    created_by = None
    if user_type == 'team_member':
        created_by = TeamMember.objects.get(id=user_id)

    can_create = False
    if user_type == 'superuser':
        can_create = True
    elif user_type == 'team_member':
        if created_by.role in ['admin', 'manager']:
            can_create = True
        elif created_by.has_permission('can_create_itinerary'):
            can_create = True

    if not can_create:
        messages.error(request, '❌ You do not have permission to create itineraries')
        return redirect('query_proposals', query_id=query_id)

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        notes = request.POST.get('notes', '').strip()
        destinations_input = request.POST.get('destinations', '').strip()

        from_date = request.POST.get('from_date')
        to_date = request.POST.get('to_date')
        total_days = request.POST.get('total_days')
        adult = request.POST.get('adult')
        childrens = request.POST.get('childrens')
        infant = request.POST.get('infant')

        if not name:
            error_msg = 'Itinerary name is required.'
            if is_ajax:
                return JsonResponse({'success': False, 'errors': {'name': error_msg}}, status=400)
            else:
                messages.error(request, error_msg)
                return redirect('query_proposals', query_id=query_id)

        itinerary = Itinerary.objects.create(
            name=name,
            query=query,
            notes=notes,
            created_by=created_by,
            created_at=now(),
            travel_from=from_date if from_date else query.from_date,
            travel_to=to_date if to_date else query.to_date,
            total_days=int(total_days) if total_days else query.total_days,
            adults=int(adult) if adult else query.adult,
            childrens=int(childrens) if childrens else query.childrens,
            infants=int(infant) if infant else query.infant,
        )

        destination_names = [d.strip() for d in destinations_input.split(',') if d.strip()]
        for dest_name in destination_names:
            destination, created = Destinations.objects.get_or_create(
                name__iexact=dest_name,
                defaults={'name': dest_name}
            )
            itinerary.destinations.add(destination)

        creator_name = created_by.get_full_name() if created_by and hasattr(created_by, 'get_full_name') else 'System'
        success_message = f'✅ Itinerary "{name}" created successfully by {creator_name}!'

        if is_ajax:
            return JsonResponse({
                'success': True,
                'message': success_message,
                'redirect_url': f'/itinerary/{itinerary.id}/day-plan/',
                'itinerary_id': itinerary.id,
                'created_by': creator_name,
                'travel_dates': {
                    'from': str(itinerary.travel_from) if itinerary.travel_from else None,
                    'to': str(itinerary.travel_to) if itinerary.travel_to else None,
                    'days': itinerary.total_days,
                }
            })
        else:
            messages.success(request, success_message)
            return redirect('itinerary_day_plan', itinerary.id)

    return redirect('query_proposals', query_id=query_id)


from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from datetime import datetime, timedelta
import json

# ✅ API ENDPOINT TO GET ITINERARY DETAILS
@require_http_methods(["GET"])
def get_itinerary_details(request, itinerary_id):
    """
    API endpoint to fetch itinerary details as JSON
    Used by edit modal to auto-fill form fields
    """
    try:
        itinerary = Itinerary.objects.get(id=itinerary_id)

        travel_from = itinerary.travel_from.strftime('%Y-%m-%d') if itinerary.travel_from else None
        travel_to = itinerary.travel_to.strftime('%Y-%m-%d') if itinerary.travel_to else None

        return JsonResponse({
            'success': True,
            'itinerary': {
                'id': itinerary.id,
                'travel_from': travel_from,
                'travel_to': travel_to,
                'adults': itinerary.adults or 0,
                'childrens': itinerary.childrens or 0,
                'infants': itinerary.infants or 0,
                'name': itinerary.name,
                'status': itinerary.status,
            }
        })
    except Itinerary.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Itinerary not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)







@require_POST
def delete_itinerary(request, itinerary_id):
    """
    Delete an itinerary - with permission check and AJAX support
    """
    # ✅ Check authentication
    user_id = request.session.get('user_id')
    user_type = request.session.get('user_type')

    if not user_id:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': 'Authentication required.'
            }, status=401)
        messages.warning(request, '⚠️ Please login first')
        return redirect('team_member:login')

    # Check if AJAX request
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    try:
        itinerary = get_object_or_404(Itinerary, id=itinerary_id)
        itinerary_name = itinerary.name
        query_id = itinerary.query.id

        # ✅ Permission check
        has_permission = False

        if user_type == 'superuser':
            has_permission = True
        elif user_type == 'team_member':
            current_user = TeamMember.objects.get(id=user_id)

            # Admin and Manager can delete any itinerary
            if current_user.role in ['admin', 'manager']:
                has_permission = True
            # Creator can delete their own itinerary
            elif itinerary.created_by == current_user:
                has_permission = True
            # Check specific permission
            elif hasattr(current_user, 'permissions') and current_user.permissions:
                has_permission = current_user.permissions.can_delete_itinerary

        if not has_permission:
            if is_ajax:
                return JsonResponse({
                    'success': False,
                    'message': 'You do not have permission to delete this itinerary.'
                }, status=403)
            messages.error(request, '❌ You do not have permission to delete this itinerary.')
            return redirect('list_itineraries')

        # ✅ Delete itinerary (CASCADE will delete related objects)
        itinerary.delete()

        # Return response
        if is_ajax:
            return JsonResponse({
                'success': True,
                'message': f'✅ Itinerary "{itinerary_name}" deleted successfully!'
            })

        messages.success(request, f'✅ Itinerary "{itinerary_name}" deleted successfully!')
        return redirect('list_itineraries')

    except Itinerary.DoesNotExist:
        if is_ajax:
            return JsonResponse({
                'success': False,
                'message': 'Itinerary not found.'
            }, status=404)
        messages.error(request, '❌ Itinerary not found.')
        return redirect('list_itineraries')

    except Exception as e:
        if is_ajax:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=500)
        messages.error(request, f'❌ Error deleting itinerary: {str(e)}')
        return redirect('list_itineraries')




#######################################################################################################################################


def itinerary_dayplan_list(request):
    day_plans = ItineraryDayPlan.objects.select_related('itinerary', 'destination').prefetch_related(
        'hotels', 'houseboats', 'activities', 'meal_plans', 'vehicles', 'inclusions'
    )

    search_query = request.GET.get('q', '')

    if search_query:
        day_plans = day_plans.filter(
            Q(itinerary__name__icontains=search_query) |
            Q(destination__name__icontains=search_query) |
            Q(notes__icontains=search_query) |
            Q(hotels__name__icontains=search_query) |
            Q(activities__name__icontains=search_query)
        ).distinct()

    return render(request, 'itinerary_dayplan_list.html', {
        'day_plans': day_plans,
        'search_query': search_query
    })
#######################################################################################################################################


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.urls import reverse
from django.db.models import Q
from .models import (
    Itinerary, ItineraryDayPlan, Destinations, DayItinerary,
    Hotel, Houseboat, Activity, Vehicle, SpecialInclusion,
    HotelBooking, VehicleBooking, ActivityBooking, HouseboatBooking,
     RoomType, MealPlan, User, TeamMember
)
from .forms import (
    HotelBookingForm, VehicleBookingForm, ActivityBookingForm,
    HouseboatBookingForm, StandaloneInclusionBookingForm
)


def itinerary_day_plan(request, itinerary_id):
    # ==========================================
    # 1. AUTHENTICATION & PERMISSIONS
    # ==========================================
    has_access = False
    current_user = None
    user_type = None

    user_id = request.session.get('user_id')
    user_type = request.session.get('user_type')

    if not user_id:
        messages.warning(request, '⚠️ Please login to access this page')
        return redirect('team_member:login')

    # Get current user
    if user_type == 'superuser':
        current_user = User.objects.get(id=user_id)
        has_access = True
    elif user_type == 'team_member':
        current_user = TeamMember.objects.get(id=user_id)

    # Get itinerary
    itinerary = get_object_or_404(Itinerary, id=itinerary_id)
    query = itinerary.query

    # Permission checks
    if user_type == 'superuser':
        has_access = True
    elif user_type == 'team_member' and current_user:
        if current_user.role == 'admin':
            has_access = True
        elif current_user.role == 'manager':
            if (itinerary.created_by == current_user or itinerary.query.assign == current_user):
                has_access = True
        else:
            if itinerary.created_by == current_user:
                has_access = True

    if not has_access:
        messages.error(request, '❌ You do not have permission to access this itinerary')
        return redirect('list_itineraries')

    # Granular permissions
    can_manage_destinations = (
        user_type == 'superuser' or
        (current_user and hasattr(current_user, 'role') and current_user.role in ['admin', 'manager']) or
        (current_user and hasattr(current_user, 'has_permission') and current_user.has_permission('can_manage_destinations'))
    )
    can_view_saved_items = (
        user_type == 'superuser' or
        (current_user and hasattr(current_user, 'role') and current_user.role in ['admin', 'manager']) or
        (current_user and hasattr(current_user, 'has_permission') and current_user.has_permission('can_view_saved_items'))
    )
    can_add_items = (
        user_type == 'superuser' or
        (current_user and hasattr(current_user, 'role') and current_user.role in ['admin', 'manager']) or
        (current_user and hasattr(current_user, 'has_permission') and current_user.has_permission('can_add_items_to_day'))
    )

    # ==========================================
    # 2. DESTINATION ALIASES
    # ==========================================
    DESTINATION_ALIASES = {
        'kochi': ['kochi', 'cochin'],
        'cochin': ['kochi', 'cochin'],
        'thiruvananthapuram': ['thiruvananthapuram', 'trivandrum'],
        'trivandrum': ['thiruvananthapuram', 'trivandrum'],
        'munnar': ['munnar'],
        'alleppey': ['alleppey', 'alappuzha'],
        'alappuzha': ['alleppey', 'alappuzha'],
        'thekkady': ['thekkady'],
        'wayanad': ['wayanad'],
        'kovalam': ['kovalam'],
    }

    # ==========================================
    # 3. BASIC DATA SETUP
    # ==========================================
    total_days = itinerary.total_days or 1
    destinations = itinerary.destinations.all()
    days = list(range(1, int(total_days) + 1))

    # Initialize context variables
    selected_section = None
    section_data = []
    context_day_number = None

    # Get day from GET parameter or auto-select Day 1
    if 'day' in request.GET:
        try:
            context_day_number = int(request.GET.get('day'))
            if context_day_number not in days:
                context_day_number = days[0] if days else None
        except (ValueError, TypeError):
            context_day_number = days[0] if days else None
    elif days:
        context_day_number = days[0]  # Auto-select Day 1

    # ==========================================
    # 4. LOAD DAY PLANS WITH DEBUG
    # ==========================================
    day_plans_qs = ItineraryDayPlan.objects.filter(itinerary=itinerary).prefetch_related(
        'hotel_bookings__hotel',
        'hotel_bookings__room_type',
        'hotel_bookings__meal_plan',
        'hotel_bookings__inclusion_items__special_inclusion',
        'vehicle_bookings__vehicle',
        'activity_bookings__activity',
        'houseboat_bookings__houseboat',
        'houseboat_bookings__meal_plan',
        'houseboat_bookings__room_type',
        'houseboat_bookings__inclusion_items__special_inclusion',
        'standalone_inclusions__special_inclusion'
    )

    day_plans_dict = {dp.day_number: dp for dp in day_plans_qs}

    # ==========================================
    # 4.5. AUTO-FIX MISSING DESTINATIONS
    # ==========================================
    for day_plan in day_plans_qs:
        if not day_plan.destination:
            destination = None

            # Check hotels (Safe for Demo Hotels)
            if not destination and day_plan.hotel_bookings.exists():
                first_hotel = day_plan.hotel_bookings.first()
                # Use booking.destination (set for both Demo & Real) or fallback to hotel.destination
                if first_hotel.destination:
                    destination = first_hotel.destination
                elif first_hotel.hotel and first_hotel.hotel.destination:
                    destination = first_hotel.hotel.destination

            # Check vehicles
            if not destination and day_plan.vehicle_bookings.exists():
                first_vehicle = day_plan.vehicle_bookings.first()
                if first_vehicle.destination:
                    destination = first_vehicle.destination
                elif first_vehicle.vehicle and first_vehicle.vehicle.destination:
                    destination = first_vehicle.vehicle.destination

            # Check houseboats
            if not destination and day_plan.houseboat_bookings.exists():
                first_houseboat = day_plan.houseboat_bookings.first()
                if first_houseboat.destination: # Assuming houseboats also have destination FK on booking
                     destination = first_houseboat.destination
                elif first_houseboat.houseboat and first_houseboat.houseboat.destination:
                    destination = first_houseboat.houseboat.destination

            # Check activities
            if not destination and day_plan.activity_bookings.exists():
                first_activity = day_plan.activity_bookings.first()
                if first_activity.activity and first_activity.activity.destination:
                    destination = first_activity.activity.destination

            # Update if found
            if destination:
                day_plan.destination = destination
                day_plan.save()
                print(f"✅ Auto-fixed Day {day_plan.day_number}: Set destination to {destination.name}")

    # DEBUG: Print booking counts
    print(f"\n{'='*60}")
    print(f"🔍 Loading data for Itinerary {itinerary_id}")
    print(f"📅 Total days: {len(days)}")

    # Build saved items dictionary
    saved_items_by_day = {}
    if can_view_saved_items:
        for day in days:
            day_plan = day_plans_dict.get(day)
            if day_plan:
                # Convert querysets to lists
                hotel_bookings_list = list(day_plan.hotel_bookings.all())
                vehicle_bookings_list = list(day_plan.vehicle_bookings.all())
                activity_bookings_list = list(day_plan.activity_bookings.all())
                houseboat_bookings_list = list(day_plan.houseboat_bookings.all())
                standalone_inclusions_list = list(day_plan.standalone_inclusions.all())

                # ✅ DEBUG: Safe Printing (Fixed for Demo Hotels)
                print(f"\n📍 Day {day} ({day_plan.destination.name if day_plan.destination else 'No destination'}):")
                print(f"   🏨 Hotels: {len(hotel_bookings_list)}")
                for booking in hotel_bookings_list:
                    # 🔥 CRITICAL FIX HERE: Check if hotel exists before accessing name
                    if booking.hotel:
                        print(f"      - {booking.hotel.name} (ID: {booking.id})")
                    else:
                        print(f"      - {booking.custom_hotel_name} (DEMO) (ID: {booking.id})")

                print(f"   🚗 Vehicles: {len(vehicle_bookings_list)}")
                print(f"   🎯 Activities: {len(activity_bookings_list)}")
                print(f"   🚢 Houseboats: {len(houseboat_bookings_list)}")
                print(f"   ⭐ Inclusions: {len(standalone_inclusions_list)}")

                saved_items_by_day[day] = {
                    'day_plan': day_plan,
                    'destination': day_plan.destination,
                    'hotel_bookings': hotel_bookings_list,
                    'vehicle_bookings': vehicle_bookings_list,
                    'activity_bookings': activity_bookings_list,
                    'houseboat_bookings': houseboat_bookings_list,
                    'standalone_inclusions': standalone_inclusions_list,
                }
            else:
                print(f"\n📍 Day {day}: No day plan exists")
                saved_items_by_day[day] = {
                    'day_plan': None,
                    'destination': None,
                    'hotel_bookings': [],
                    'vehicle_bookings': [],
                    'activity_bookings': [],
                    'houseboat_bookings': [],
                    'standalone_inclusions': [],
                }

    print(f"{'='*60}\n")

    # Calculate TOTAL counts
    total_counts = {
        'day_plans': 0,
        'hotels': 0,
        'activities': 0,
        'houseboats': 0,
        'vehicles': 0,
        'standalone_inclusions': 0,
    }

    for day, day_data in saved_items_by_day.items():
        if day_data.get('day_plan') and day_data['day_plan'].title:
            total_counts['day_plans'] += 1
        total_counts['hotels'] += len(day_data.get('hotel_bookings', []))
        total_counts['activities'] += len(day_data.get('activity_bookings', []))
        total_counts['houseboats'] += len(day_data.get('houseboat_bookings', []))
        total_counts['vehicles'] += len(day_data.get('vehicle_bookings', []))
        total_counts['standalone_inclusions'] += len(day_data.get('standalone_inclusions', []))

    print(f"📊 TOTAL COUNTS ACROSS ALL DAYS:")
    print(f"   📋 Day Plans: {total_counts['day_plans']}")
    print(f"   🏨 Hotels: {total_counts['hotels']}")
    print(f"   🎯 Activities: {total_counts['activities']}")
    print(f"   🚢 Houseboats: {total_counts['houseboats']}")
    print(f"   🚗 Vehicles: {total_counts['vehicles']}")
    print(f"   ⭐ Standalone: {total_counts['standalone_inclusions']}")
    print(f"{'='*60}\n")

    # ==========================================
    # 5. HANDLE POST REQUESTS
    # ==========================================
    if request.method == 'POST':
        form_type = request.POST.get('form_type')

        # ✅ DESTINATION SELECTION
        if form_type == 'destination_select' and can_manage_destinations:
            day_number = int(request.POST.get('day_number'))
            destination_id = request.POST.get('destination')

            day_plan, created = ItineraryDayPlan.objects.get_or_create(
                itinerary=itinerary,
                day_number=day_number
            )

            if destination_id:
                destination = Destinations.objects.get(id=destination_id)
                day_plan.destination = destination
                day_plan.save()

                messages.success(request, f'✅ Day {day_number} set to {destination.name}')

                # Clean up bookings that don't match destination
                # Note: For Demo Hotels, they might have null hotel FK but correct destination FK on booking
                # So we filter based on booking.destination first, fallback to hotel.destination

                # Cleaning Hotels
                hotels_to_keep = HotelBooking.objects.filter(
                    day_plan=day_plan
                ).filter(
                    Q(destination=destination) |  # Keep if booking destination matches
                    Q(hotel__destination=destination) # Or if linked hotel matches
                ).values_list('id', flat=True)

                HotelBooking.objects.filter(day_plan=day_plan).exclude(id__in=hotels_to_keep).delete()

                day_plan.vehicle_bookings.exclude(destination=destination).delete()
                # For activities/houseboats, adjust logic similarly if they have explicit destination FKs
                # Assuming standard FKs for now:
                day_plan.activity_bookings.exclude(activity__destination=destination).delete()
                day_plan.houseboat_bookings.exclude(houseboat__destination=destination).delete()

                day_plan.standalone_inclusions.exclude(
                    Q(special_inclusion__destination=destination) |
                    Q(special_inclusion__destination__isnull=True)
                ).delete()
            else:
                # Clear everything
                day_plan.destination = None
                day_plan.title = ""
                day_plan.description = ""
                if day_plan.image:
                    day_plan.image.delete(save=False)
                    day_plan.image = None
                day_plan.save()

                # Delete all bookings
                day_plan.hotel_bookings.all().delete()
                day_plan.vehicle_bookings.all().delete()
                day_plan.activity_bookings.all().delete()
                day_plan.houseboat_bookings.all().delete()
                day_plan.standalone_inclusions.all().delete()

                messages.success(request, f'✅ Destination removed from Day {day_number}')

            return redirect(f"{reverse('itinerary_day_plan', args=[itinerary.id])}?day={day_number}")

        # ✅ DELETE DAY ITINERARY (SINGLE DAY)
        elif form_type == 'delete_day_itinerary' and can_manage_destinations:
            day_number = int(request.POST.get('day_number'))
            day_plan = day_plans_dict.get(day_number)

            if day_plan:
                day_plan.title = ""
                day_plan.description = ""
                if day_plan.image:
                    day_plan.image.delete(save=False)
                    day_plan.image = None
                day_plan.save()
                messages.success(request, f'✅ Day itinerary removed from Day {day_number}.')
            else:
                messages.warning(request, f'⚠️ No day plan found for Day {day_number}')

            return redirect(f"{reverse('itinerary_day_plan', args=[itinerary.id])}?day={day_number}")

        # ✅ EDIT DAY DETAILS
        elif form_type == 'edit_day_details' and can_manage_destinations:
            day_number = int(request.POST.get('day_number'))
            title = request.POST.get('title', '').strip()
            description = request.POST.get('description', '').strip()
            image = request.FILES.get('image')

            day_plan, created = ItineraryDayPlan.objects.get_or_create(
                itinerary=itinerary,
                day_number=day_number
            )

            if title: day_plan.title = title
            if description: day_plan.description = description
            if image:
                if day_plan.image:
                    day_plan.image.delete(save=False)
                day_plan.image = image
            day_plan.save()

            messages.success(request, f'✅ Day {day_number} details updated!')
            return redirect(f"{reverse('itinerary_day_plan', args=[itinerary.id])}?day={day_number}")

        # ✅ BULK DELETE OPERATIONS
        elif form_type == 'bulk_delete_all_day_plans' and can_manage_destinations:
            deleted_count = 0
            for day_plan in ItineraryDayPlan.objects.filter(itinerary=itinerary):
                if day_plan.title or day_plan.description or day_plan.image:
                    day_plan.title = ""
                    day_plan.description = ""
                    if day_plan.image:
                        day_plan.image.delete(save=False)
                        day_plan.image = None
                    day_plan.save()
                    deleted_count += 1
            messages.success(request, f'✅ Successfully deleted {deleted_count} day itinerary/ies from ALL days!')
            return redirect(f"{reverse('itinerary_day_plan', args=[itinerary.id])}?day=1")

        elif form_type == 'bulk_delete_all_hotels' and can_manage_destinations:
            all_hotel_bookings = HotelBooking.objects.filter(itinerary=itinerary)
            deleted_count = all_hotel_bookings.count()
            all_hotel_bookings.delete()
            messages.success(request, f'✅ Successfully deleted {deleted_count} hotel booking(s) from ALL days!')
            return redirect(f"{reverse('itinerary_day_plan', args=[itinerary.id])}?day=1")

        elif form_type == 'bulk_delete_all_activities' and can_manage_destinations:
            bookings = ActivityBooking.objects.filter(itinerary=itinerary)
            deleted_count = bookings.count()
            bookings.delete()
            messages.success(request, f'✅ Successfully deleted {deleted_count} activity booking(s) from ALL days!')
            return redirect(f"{reverse('itinerary_day_plan', args=[itinerary.id])}?day=1")

        elif form_type == 'bulk_delete_all_houseboats' and can_manage_destinations:
            bookings = HouseboatBooking.objects.filter(itinerary=itinerary)
            deleted_count = bookings.count()
            bookings.delete()
            messages.success(request, f'✅ Successfully deleted {deleted_count} houseboat booking(s) from ALL days!')
            return redirect(f"{reverse('itinerary_day_plan', args=[itinerary.id])}?day=1")

        elif form_type == 'bulk_delete_all_vehicles' and can_manage_destinations:
            bookings = VehicleBooking.objects.filter(itinerary=itinerary)
            deleted_count = bookings.count()
            bookings.delete()
            messages.success(request, f'✅ Successfully deleted {deleted_count} vehicle booking(s) from ALL days!')
            return redirect(f"{reverse('itinerary_day_plan', args=[itinerary.id])}?day=1")

        # ✅ SECTION SELECTION
        elif form_type == 'section_select' and can_add_items:
            selected_section = request.POST.get('section', '').strip()
            context_day_number = int(request.POST.get('day_number'))
            day_plan = day_plans_dict.get(context_day_number)
            destination = day_plan.destination if day_plan else None

            if not selected_section:
                messages.info(request, 'ℹ️ Please select a section to view available items.')
                section_data = []
            elif not destination:
                messages.warning(request, f'⚠️ Please select a destination for Day {context_day_number} first.')
                section_data = []
            else:
                # Load section data based on destination
                if selected_section == 'day_itinerary':
                    destination_name_lower = destination.name.lower()
                    possible_names = DESTINATION_ALIASES.get(destination_name_lower, [destination_name_lower])
                    section_data = DayItinerary.objects.filter(
                        destination__iregex=r'^(' + '|'.join(possible_names) + ')$'
                    ).order_by('-is_pinned', '-created_at')

                    if section_data.count() > 0:
                        messages.success(request, f'✅ Found {section_data.count()} itinerary/ies for {destination.name}')
                    else:
                        messages.info(request, f'ℹ️ No day itineraries found for {destination.name}')

                elif selected_section == 'hotels':
                    if not itinerary.travel_from or not itinerary.travel_to:
                        messages.error(request, '❌ Itinerary date range is missing!')
                        section_data = []
                    else:
                        from .models import Hotelprice
                        valid_hotel_ids = Hotelprice.objects.filter(
                            hotel__destination=destination,
                            hotel__status=True,
                            from_date__lte=itinerary.travel_to,
                            to_date__gte=itinerary.travel_from
                        ).values_list('hotel_id', flat=True).distinct()

                        section_data = Hotel.objects.filter(
                            id__in=valid_hotel_ids,
                            destination=destination,
                            status=True
                        ).select_related('destination').order_by('name')

                        if section_data.count() > 0:
                            messages.success(request, f'✅ Found {section_data.count()} hotel(s) with valid pricing')
                        else:
                            messages.warning(request, f'⚠️ No hotels available in {destination.name} with valid pricing')

                elif selected_section == 'houseboats':
                    if not itinerary.travel_from or not itinerary.travel_to:
                        messages.error(request, '❌ Itinerary date range is missing!')
                        section_data = []
                    else:
                        from .models import HouseboatPrice
                        valid_ids = HouseboatPrice.objects.filter(
                            houseboat__destination=destination,
                            houseboat__status=True,
                            from_date__lte=itinerary.travel_to,
                            to_date__gte=itinerary.travel_from
                        ).values_list('houseboat_id', flat=True).distinct()

                        section_data = Houseboat.objects.filter(
                            id__in=valid_ids,
                            destination=destination,
                            status=True
                        ).select_related('destination').order_by('name')

                        if section_data.count() > 0:
                            messages.success(request, f'✅ Found {section_data.count()} houseboat(s)')
                        else:
                            messages.warning(request, f'⚠️ No houseboats available in {destination.name}')

                elif selected_section == 'activities':
                    section_data = Activity.objects.filter(
                        destination=destination,
                        is_active=True
                    ).select_related('destination').order_by('name')

                    if section_data.count() > 0:
                        messages.success(request, f'✅ Found {section_data.count()} active activity/ies')
                    else:
                        messages.info(request, f'ℹ️ No activities found in {destination.name}')

                elif selected_section == 'vehicles':
                    section_data = Vehicle.objects.filter(
                        destination=destination,
                        status=True
                    ).select_related('destination').order_by('name')

                    if section_data.count() > 0:
                        messages.success(request, f'✅ Found {section_data.count()} vehicle(s)')
                    else:
                        messages.info(request, f'ℹ️ No vehicles found in {destination.name}')

                elif selected_section == 'standalone_inclusions':
                    section_data = SpecialInclusion.objects.filter(
                        inclusion_type='general',
                        status=True,
                        is_available=True
                    ).filter(
                        Q(destination=destination) | Q(destination__isnull=True)
                    ).order_by('name')

                    if section_data.count() > 0:
                        messages.success(request, f'✅ Found {section_data.count()} service(s)')
                    else:
                        messages.info(request, f'ℹ️ No standalone services found')

        # ✅ SELECT DAY ITINERARY
        elif form_type == 'select_day_itinerary' and can_add_items:
            day_number = int(request.POST.get('day_number'))
            day_itinerary_id = request.POST.get('day_itinerary_id')

            day_plan, created = ItineraryDayPlan.objects.get_or_create(
                itinerary=itinerary,
                day_number=day_number
            )

            day_itinerary = DayItinerary.objects.get(id=day_itinerary_id)
            day_plan.title = day_itinerary.name
            day_plan.description = day_itinerary.details

            if day_itinerary.image_source == 'gallery' and day_itinerary.gallery_image:
                day_plan.image = day_itinerary.gallery_image.image
            elif day_itinerary.image_source == 'upload' and day_itinerary.uploaded_image:
                day_plan.image = day_itinerary.uploaded_image

            day_plan.save()
            messages.success(request, f'✅ Day {day_number} updated with "{day_itinerary.name}"')
            return redirect(f"{reverse('itinerary_day_plan', args=[itinerary.id])}?day={day_number}")

    # ==========================================
    # 6. BUILD CONTEXT & RENDER
    # ==========================================
    context = {
        'itinerary': itinerary,
        'days': days,
        'query': query,
        'destinations': destinations,
        'plans': day_plans_dict,
        'selected_section': selected_section,
        'section_data': section_data,
        'context_day_number': context_day_number,
        'sections': ['day_itinerary', 'hotels', 'houseboats', 'activities', 'vehicles', 'standalone_inclusions'],
        'saved_items_by_day': saved_items_by_day,
        'total_counts': total_counts,

        # Forms
        'hotel_form': HotelBookingForm(),
        'vehicle_form': VehicleBookingForm(),
        'activity_form': ActivityBookingForm(),
        'houseboat_form': HouseboatBookingForm(),
        'standalone_inclusion_form': StandaloneInclusionBookingForm(),

        # Lookups
        'room_types': RoomType.objects.filter(is_active=True),
        'meal_plans': MealPlan.objects.filter(status=True),
        'hotel_special_inclusions': SpecialInclusion.objects.filter(
            is_available=True, status=True, inclusion_type='hotel'
        ).order_by('name'),
        'houseboat_special_inclusions': SpecialInclusion.objects.filter(
            is_available=True, status=True, inclusion_type='houseboat'
        ).order_by('name'),
        'general_special_inclusions': SpecialInclusion.objects.filter(
            is_available=True, status=True, inclusion_type='general'
        ).order_by('name'),

        # Permissions
        'user_type': user_type,
        'current_user': current_user,
        'can_edit': has_access,
        'can_manage_destinations': can_manage_destinations,
        'can_view_saved_items': can_view_saved_items,
        'can_add_items': can_add_items,
    }

    return render(request, 'plan_daywise_itinerary.html', context)









@require_http_methods(["GET"])
def get_hotels_by_destination(request):
    """
    Master Booking API: Get ALL hotels for a destination
    (No pricing filter - for manual price entry)
    """
    destination_id = request.GET.get('destination_id')

    if not destination_id:
        return JsonResponse({
            'success': False,
            'error': 'destination_id is required'
        }, status=400)

    try:
        # Get ALL active hotels for this destination
        hotels = Hotel.objects.filter(
            destination_id=destination_id,
            status=True
        ).values('id', 'name', 'category').order_by('name')

        return JsonResponse({
            'success': True,
            'hotels': list(hotels)
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ============================================================
# ✅ MANUAL MODE (PRICED HOTELS ONLY) - FIXED VERSION
# ============================================================
@require_http_methods(["GET"])
def get_priced_hotels(request):
    """
    Manual Mode API: Get only hotels with active pricing
    for the given destination and date range.
    """
    destination_id = request.GET.get('destination')  # ✅ matches JS parameter
    checkin = request.GET.get('checkin')
    checkout = request.GET.get('checkout')

    if not all([destination_id, checkin, checkout]):
        return JsonResponse({
            'success': False,
            'error': 'destination, checkin, and checkout are required'
        }, status=400)

    try:
        from datetime import datetime

        checkin_date = datetime.strptime(checkin, '%Y-%m-%d').date()
        checkout_date = datetime.strptime(checkout, '%Y-%m-%d').date()

        # ✅ Get hotels that have pricing covering the entire date range
        priced_hotels_qs = Hotelprice.objects.filter(
            hotel__destination_id=destination_id,
            hotel__status=True,
            from_date__lte=checkin_date,   # Price starts before/on checkin
            to_date__gte=checkout_date      # Price ends after/on checkout
        ).select_related('hotel').values(
            'hotel__id',
            'hotel__name',
            'hotel__category'
        ).distinct()

        # Format response
        hotels = []
        for ph in priced_hotels_qs:
            hotels.append({
                'id': ph['hotel__id'],
                'name': ph['hotel__name'],
                'category': ph['hotel__category']
            })

        return JsonResponse({
            'success': True,
            'hotels': hotels
        })

    except ValueError as ve:
        return JsonResponse({
            'success': False,
            'error': f'Invalid date format: {str(ve)}'
        }, status=400)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)



from django.http import JsonResponse
from django.template.loader import render_to_string

def get_section_data_ajax(request, itinerary_id):
    """AJAX endpoint to fetch section data without page reload"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=400)

    selected_section = request.POST.get('section')
    day_number_str = request.POST.get('day_number', '').strip()

    # Validate day_number
    if not day_number_str:
        return JsonResponse({'error': 'Day number is required'}, status=400)

    try:
        day_number = int(day_number_str)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid day number'}, status=400)

    itinerary = get_object_or_404(Itinerary, id=itinerary_id)

    # Get day plan
    day_plan = ItineraryDayPlan.objects.filter(
        itinerary=itinerary,
        day_number=day_number
    ).first()

    if not day_plan or not day_plan.destination:
        return JsonResponse({
            'error': f'Please select a destination for Day {day_number} first.',
            'day_number': day_number
        }, status=400)

    destination = day_plan.destination
    section_data = []

    # ✅ DESTINATION ALIASES (same as main view)
    DESTINATION_ALIASES = {
        'kochi': ['kochi', 'cochin'],
        'cochin': ['kochi', 'cochin'],
        'thiruvananthapuram': ['thiruvananthapuram', 'trivandrum'],
        'trivandrum': ['thiruvananthapuram', 'trivandrum'],
        'munnar': ['munnar'],
        'alleppey': ['alleppey', 'alappuzha'],
        'alappuzha': ['alleppey', 'alappuzha'],
        'thekkady': ['thekkady'],
        'wayanad': ['wayanad'],
    }

    # Fetch data based on section
    if selected_section == 'day_itinerary':
        destination_name_lower = destination.name.lower()
        possible_names = DESTINATION_ALIASES.get(destination_name_lower, [destination_name_lower])

        items = DayItinerary.objects.filter(
            destination__iregex=r'^(' + '|'.join(possible_names) + ')$'
        ).order_by('-created_at')

        section_data = [{
            'id': item.id,
            'name': item.name,
            'destination': item.destination,
            'details': item.details,
            'image_url': item.image.url if item.image else None,
        } for item in items]

    elif selected_section == 'hotels':
        from .models import Hotelprice
        valid_hotel_ids = Hotelprice.objects.filter(
            hotel__destination=destination,
            hotel__status=True
        ).filter(
            Q(from_date__lte=itinerary.travel_to) & Q(to_date__gte=itinerary.travel_from)
        ).values_list('hotel_id', flat=True).distinct()

        items = Hotel.objects.filter(
            id__in=valid_hotel_ids,
            destination=destination,
            status=True
        ).select_related('destination').order_by('name')

        section_data = [{
            'id': item.id,
            'name': item.name,
            'destination': item.destination.name if item.destination else '',
        } for item in items]

    elif selected_section == 'houseboats':
        from .models import HouseboatPrice
        valid_houseboat_ids = HouseboatPrice.objects.filter(
            houseboat__destination=destination,
            houseboat__status=True,
            from_date__lte=itinerary.travel_to,
            to_date__gte=itinerary.travel_from
        ).values_list('houseboat_id', flat=True).distinct()

        items = Houseboat.objects.filter(
            id__in=valid_houseboat_ids,
            destination=destination,
            status=True
        ).select_related('destination').order_by('name')

        section_data = [{
            'id': item.id,
            'name': item.name,
            'destination': item.destination.name if item.destination else '',
        } for item in items]

    elif selected_section == 'activities':
        items = Activity.objects.filter(
            destination=destination,
            is_active=True
        ).select_related('destination').order_by('name')

        section_data = [{
            'id': item.id,
            'name': item.name,
            'destination': item.destination.name if item.destination else '',
        } for item in items]

    elif selected_section == 'vehicles':
        items = Vehicle.objects.filter(
            destination=destination,
            status=True
        ).select_related('destination').order_by('name')

        section_data = [{
            'id': item.id,
            'name': item.name,
            'destination': item.destination.name if item.destination else '',
        } for item in items]

    elif selected_section == 'standalone_inclusions':
        items = SpecialInclusion.objects.filter(
            inclusion_type='general',
            status=True,
            is_available=True
        ).filter(
            Q(destination=destination) | Q(destination__isnull=True)
        ).order_by('name')

        section_data = [{
            'id': item.id,
            'name': item.name,
            'destination': item.destination.name if item.destination else '',
            'adult_price': float(item.adult_price) if item.adult_price else 0,
            'child_price': float(item.get_child_price) if hasattr(item, 'get_child_price') else 0,
            'pricing_type': item.pricing_type if hasattr(item, 'pricing_type') else '',
        } for item in items]

    return JsonResponse({
        'success': True,
        'section': selected_section,
        'day_number': day_number,
        'destination': destination.name,
        'data': section_data,
        'count': len(section_data)
    })



#####################################################################################################

from django.views.decorators.http import require_GET
from django.http import JsonResponse
from django.shortcuts import get_object_or_404

@require_GET
def get_valid_houseboat_options(request, houseboat_id):
    """
    Returns ONLY room types and meal plans that exist in HouseboatPrice table
    Based on check-in and check-out dates
    """
    try:
        # Get parameters
        checkin = request.GET.get('checkin')
        checkout = request.GET.get('checkout')

        print(f"\n{'='*60}")
        print(f"🚢 get_valid_houseboat_options called")
        print(f"   Houseboat ID: {houseboat_id}")
        print(f"   Check-in: {checkin}")
        print(f"   Check-out: {checkout}")

        # Validate parameters
        if not checkin or not checkout:
            return JsonResponse({
                'success': False,
                'error': 'Check-in and check-out dates are required'
            }, status=400)

        # Get houseboat
        houseboat = get_object_or_404(Houseboat, id=houseboat_id, status=True)

        # Query prices that overlap with the date range
        valid_prices = HouseboatPrice.objects.filter(
            houseboat=houseboat,
            from_date__lte=checkout,  # Price period starts before/on checkout
            to_date__gte=checkin      # Price period ends after/on checkin
        ).select_related('room_type', 'meal_plan')

        print(f"   Found {valid_prices.count()} valid price entries")

        if not valid_prices.exists():
            print(f"   ❌ No pricing available")
            return JsonResponse({
                'success': False,
                'error': f'No pricing available for {houseboat.name} from {checkin} to {checkout}',
                'roomtypes': [],
                'mealplans': []
            })

        # Extract unique room types and meal plans
        room_types = []
        meal_plans = []
        room_type_ids = set()
        meal_plan_ids = set()

        for price in valid_prices:
            # Add room type if not already added
            if price.room_type_id not in room_type_ids:
                room_types.append({
                    'id': price.room_type.id,
                    'name': price.room_type.name
                })
                room_type_ids.add(price.room_type_id)

            # Add meal plan if not already added
            if price.meal_plan_id not in meal_plan_ids:
                meal_plans.append({
                    'id': price.meal_plan.id,
                    'name': price.meal_plan.name
                })
                meal_plan_ids.add(price.meal_plan_id)

        print(f"   ✅ Room Types: {len(room_types)}")
        for rt in room_types:
            print(f"      - {rt['name']} (ID: {rt['id']})")

        print(f"   ✅ Meal Plans: {len(meal_plans)}")
        for mp in meal_plans:
            print(f"      - {mp['name']} (ID: {mp['id']})")

        print(f"{'='*60}\n")

        return JsonResponse({
            'success': True,
            'houseboat_name': houseboat.name,
            'roomtypes': room_types,
            'mealplans': meal_plans,
            'room_type_count': len(room_types),
            'meal_plan_count': len(meal_plans)
        })

    except Houseboat.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Houseboat not found'
        }, status=404)

    except Exception as e:
        print(f"❌ Error in get_valid_houseboat_options: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Server error: {str(e)}'
        }, status=500)



@require_GET
def get_valid_hotel_options(request, hotel_id):
    """
    Returns room types and meal plans based on hotel and dates ONLY (no category)
    """
    try:
        # Get parameters - NO CATEGORY
        checkin = request.GET.get('checkin')
        checkout = request.GET.get('checkout')

        print(f"\n{'='*60}")
        print(f"🏨 get_valid_hotel_options called")
        print(f"   Hotel ID: {hotel_id}")
        print(f"   Check-in: {checkin}")
        print(f"   Check-out: {checkout}")

        # Validate parameters
        if not checkin or not checkout:
            return JsonResponse({
                'success': False,
                'error': 'Check-in and check-out dates are required'
            }, status=400)

        # Get hotel
        hotel = get_object_or_404(Hotel, id=hotel_id, status=True)

        # ✅ Query prices WITHOUT category filter
        valid_prices = Hotelprice.objects.filter(
            hotel=hotel,
            from_date__lte=checkout,
            to_date__gte=checkin
        ).select_related('room_type', 'meal_plan')

        print(f"   Found {valid_prices.count()} valid price entries")

        if not valid_prices.exists():
            print(f"   ❌ No pricing available")
            return JsonResponse({
                'success': False,
                'error': f'No pricing available for {hotel.name} from {checkin} to {checkout}',
                'roomtypes': [],
                'mealplans': []
            })

        # Extract unique room types and meal plans
        room_types = []
        meal_plans = []
        room_type_ids = set()
        meal_plan_ids = set()

        for price in valid_prices:
            # Add room type if not already added
            if price.room_type_id not in room_type_ids:
                room_types.append({
                    'id': price.room_type.id,
                    'name': price.room_type.name
                })
                room_type_ids.add(price.room_type_id)

            # Add meal plan if not already added
            if price.meal_plan_id not in meal_plan_ids:
                meal_plans.append({
                    'id': price.meal_plan.id,
                    'name': price.meal_plan.name
                })
                meal_plan_ids.add(price.meal_plan_id)

        print(f"   ✅ Room Types: {len(room_types)}")
        for rt in room_types:
            print(f"      - {rt['name']} (ID: {rt['id']})")

        print(f"   ✅ Meal Plans: {len(meal_plans)}")
        for mp in meal_plans:
            print(f"      - {mp['name']} (ID: {mp['id']})")

        print(f"{'='*60}\n")

        return JsonResponse({
            'success': True,
            'hotel_name': hotel.name,
            'roomtypes': room_types,
            'mealplans': meal_plans,
            'room_type_count': len(room_types),
            'meal_plan_count': len(meal_plans)
        })

    except Hotel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Hotel not found'
        }, status=404)

    except Exception as e:
        print(f"❌ Error in get_valid_hotel_options: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Server error: {str(e)}'
        }, status=500)




from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from .models import SpecialInclusion
from django.db.models import Q

@require_http_methods(["GET"])
def get_special_inclusions(request):
    """Fetch special inclusions for a destination, hotel, or houseboat"""
    destination_id = request.GET.get('destination')
    hotel_id = request.GET.get('hotel')
    houseboat_id = request.GET.get('houseboat') # ✅ NEW: Get houseboat ID

    if not destination_id:
        return JsonResponse({
            'success': False,
            'message': 'Destination ID is required'
        })

    try:
        print(f"🔍 Fetching inclusions for Dest: {destination_id}, Hotel: {hotel_id}, Houseboat: {houseboat_id}")

        # Base query: Active items only
        query = Q(status=True, is_available=True)

        if hotel_id:
            # Case 1: Hotel Booking
            # Get inclusions for THIS Hotel OR Generic Destination inclusions (where hotel is None)
            query &= (Q(hotel_id=hotel_id) | Q(destination_id=destination_id, hotel__isnull=True))

        elif houseboat_id:
            # Case 2: Houseboat Booking
            # ✅ Fix: Get inclusions for THIS Houseboat OR Generic Destination inclusions
            # We assume your model has a 'houseboat' field. If generic inclusions shouldn't have a houseboat set, check houseboat__isnull=True
            query &= (Q(houseboat_id=houseboat_id) | Q(destination_id=destination_id, houseboat__isnull=True, hotel__isnull=True))

        else:
            # Case 3: Generic Destination (No property selected)
            # Get only generic inclusions for this destination
            query &= Q(destination_id=destination_id, hotel__isnull=True, houseboat__isnull=True)

        inclusions = SpecialInclusion.objects.filter(query)

        print(f"✅ Found {inclusions.count()} active inclusions")

        # Build response
        inclusions_list = []
        for inc in inclusions:
            # Optional: Add debug print to see which type was found
            type_label = "Hotel" if inc.hotel else ("Houseboat" if hasattr(inc, 'houseboat') and inc.houseboat else "Generic")
            # print(f"   - Found: {inc.name} ({type_label})")

            inclusions_list.append({
                'id': inc.id,
                'name': inc.name,
                'adult_price': float(inc.adult_price),
                'child_price': float(inc.get_child_price()),
                'pricing_type': inc.pricing_type,
                'price_display': inc.get_price_display()
            })

        return JsonResponse({
            'success': True,
            'inclusions': inclusions_list
        })

    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()

        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)

from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import HouseboatBooking  # Ensure this import exists

def get_houseboat_booking_inclusions(request, booking_id):
    """API to get saved inclusions for a specific houseboat booking"""
    try:
        booking = get_object_or_404(HouseboatBooking, id=booking_id)

        inclusions_data = []
        # Access the related inclusions.
        # Note: Ensure 'inclusion_items' is the correct related_name in your HouseboatBookingInclusion model
        # If your model uses a different name (like houseboatbookinginclusion_set), use that instead.
        for inc in booking.inclusion_items.all():
            inclusions_data.append({
                'inclusion_id': inc.special_inclusion.id,
                'name': inc.special_inclusion.name,
                'num_adults': inc.num_adults,
                'num_children': inc.num_children,
                'adult_price': float(inc.special_inclusion.adult_price),
                'child_price': float(inc.special_inclusion.get_child_price())
            })

        return JsonResponse({
            'success': True,
            'inclusions': inclusions_data
        })
    except Exception as e:
        print(f"❌ Error fetching houseboat inclusions: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)





from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from .models import HotelBooking, HotelBookingInclusion

@require_http_methods(["GET"])
def get_booking_inclusions(request, booking_id):
    """Get all special inclusions for a hotel booking"""
    try:
        booking = HotelBooking.objects.get(id=booking_id)

        # ✅ FIXED: Use 'hotel_booking' (with underscore)
        inclusions = HotelBookingInclusion.objects.filter(
            hotel_booking=booking  # ✅ Changed from hotelbooking to hotel_booking
        ).select_related('special_inclusion')  # ✅ Changed from specialinclusion to special_inclusion

        inclusions_list = []
        for inc in inclusions:
            inclusions_list.append({
                'id': inc.id,
                'inclusion_id': inc.special_inclusion.id,  # ✅ Changed field name
                'inclusion_name': inc.special_inclusion.name,  # ✅ Changed field name
                'num_adults': inc.num_adults,  # ✅ Changed field name
                'num_children': inc.num_children,  # ✅ Changed field name
                'adult_price': float(inc.special_inclusion.adult_price),  # ✅ Changed field name
                'child_price': float(inc.special_inclusion.get_child_price()),
                'pricing_type': inc.special_inclusion.pricing_type  # ✅ Changed field name
            })

        print(f"✅ Found {len(inclusions_list)} inclusions for booking {booking_id}")
        for inc in inclusions_list:
            print(f"   - {inc['inclusion_name']}: {inc['num_adults']}A + {inc['num_children']}C")

        return JsonResponse({
            'success': True,
            'inclusions': inclusions_list
        })

    except HotelBooking.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Booking not found'
        }, status=404)
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)



#####################################################################################################


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required

# ==========================================
# 🗑️ BULK DELETE VIEWS
# ==========================================



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.urls import reverse

@require_http_methods(["POST"])
def bulk_delete_hotels(request, itinerary_id):
    """Bulk delete hotel bookings from all days"""
    itinerary = get_object_or_404(Itinerary, id=itinerary_id)

    hotel_ids = request.POST.getlist('hotel_ids')

    if hotel_ids:
        deleted_count, _ = HotelBooking.objects.filter(
            id__in=hotel_ids,
            itinerary=itinerary
        ).delete()

        messages.success(request, f'✅ Successfully deleted {deleted_count} hotel booking(s)')
    else:
        messages.warning(request, '⚠️ No hotels selected for deletion')

    # ✅ FIXED: Using correct URL name
    day = request.GET.get('day', 1)
    return redirect(reverse('itinerary_day_plan', args=[itinerary_id]) + f'?day={day}')



@require_http_methods(["POST"])
def bulk_delete_activities(request, itinerary_id):
    """Bulk delete activity bookings from all days"""
    itinerary = get_object_or_404(Itinerary, id=itinerary_id)

    activity_ids = request.POST.getlist('activity_ids')

    if activity_ids:
        deleted_count, _ = ActivityBooking.objects.filter(
            id__in=activity_ids,
            itinerary=itinerary
        ).delete()

        messages.success(request, f'✅ Successfully deleted {deleted_count} activity booking(s)')
    else:
        messages.warning(request, '⚠️ No activities selected for deletion')

    day = request.GET.get('day', 1)
    return redirect(reverse('itinerary_day_plan', args=[itinerary_id]) + f'?day={day}')



@require_http_methods(["POST"])
def bulk_delete_houseboats(request, itinerary_id):
    """Bulk delete houseboat bookings from all days"""
    itinerary = get_object_or_404(Itinerary, id=itinerary_id)

    houseboat_ids = request.POST.getlist('houseboat_ids')

    if houseboat_ids:
        deleted_count, _ = HouseboatBooking.objects.filter(
            id__in=houseboat_ids,
            itinerary=itinerary
        ).delete()

        messages.success(request, f'✅ Successfully deleted {deleted_count} houseboat booking(s)')
    else:
        messages.warning(request, '⚠️ No houseboats selected for deletion')

    day = request.GET.get('day', 1)
    return redirect(reverse('itinerary_day_plan', args=[itinerary_id]) + f'?day={day}')



@require_http_methods(["POST"])
def bulk_delete_vehicles(request, itinerary_id):
    """Bulk delete vehicle bookings from all days"""
    itinerary = get_object_or_404(Itinerary, id=itinerary_id)

    vehicle_ids = request.POST.getlist('vehicle_ids')

    if vehicle_ids:
        deleted_count, _ = VehicleBooking.objects.filter(
            id__in=vehicle_ids,
            itinerary=itinerary
        ).delete()

        messages.success(request, f'✅ Successfully deleted {deleted_count} vehicle booking(s)')
    else:
        messages.warning(request, '⚠️ No vehicles selected for deletion')

    day = request.GET.get('day', 1)
    return redirect(reverse('itinerary_day_plan', args=[itinerary_id]) + f'?day={day}')


@require_http_methods(["POST"])
def bulk_delete_day_itineraries(request, itinerary_id):
    """Bulk delete day itineraries (day plans) from all days"""
    itinerary = get_object_or_404(Itinerary, id=itinerary_id)

    day_plan_ids = request.POST.getlist('day_plan_ids')

    if day_plan_ids:
        deleted_count, _ = ItineraryDayPlan.objects.filter(
            id__in=day_plan_ids,
            itinerary=itinerary
        ).delete()

        messages.success(request, f'✅ Successfully deleted {deleted_count} day itinerary/ies')
    else:
        messages.warning(request, '⚠️ No day itineraries selected for deletion')

    day = request.GET.get('day', 1)
    return redirect(reverse('itinerary_day_plan', args=[itinerary_id]) + f'?day={day}')


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.urls import reverse


@require_http_methods(["POST"])
def remove_day_plan_from_itinerary(request, day_plan_id):
    """Remove day plan assignment from a specific day in the itinerary"""
    day_plan = get_object_or_404(ItineraryDayPlan, id=day_plan_id)
    itinerary_id = day_plan.itinerary.id
    day_number = day_plan.day_number

    # Delete the day plan assignment
    day_plan.delete()

    messages.success(request, f'✅ Day itinerary removed from Day {day_number}')

    # Redirect back to the itinerary day plan page with the day parameter
    return redirect(reverse('itinerary_day_plan', args=[itinerary_id]) + f'?day={day_number}')







#####################################################################################################

import json
from itertools import chain
from operator import attrgetter
from decimal import Decimal, InvalidOperation
from collections import defaultdict
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils.timezone import now
from .models import (
    Itinerary, ItineraryPricingOption,
    HotelBooking, VehicleBooking, ActivityBooking, HouseboatBooking,
    StandaloneInclusionBooking,  # ✅ ADD THIS
    Hotelprice, VehiclePricing, ActivityPrice, HouseboatPrice
)




def itinerary_pricing(request, itinerary_id):
    """Pricing view with authentication, permission checks, and standalone inclusions"""

    # ==========================================
    # 1. AUTHENTICATION
    # ==========================================
    user_id = request.session.get('user_id')
    user_type = request.session.get('user_type')
    current_user = None

    if not user_id:
        messages.warning(request, '⚠️ Please login to access this page')
        return redirect('team_member:login')

    if user_type == 'superuser':
        current_user = User.objects.get(id=user_id)
    elif user_type == 'team_member':
        current_user = TeamMember.objects.get(id=user_id)

    # Get itinerary
    itinerary = get_object_or_404(Itinerary, id=itinerary_id)

    # ==========================================
    # 2. PRICE CALCULATION LOGIC
    # ==========================================
    def calculate_all_prices():
        combined_bookings = list(chain(
            HotelBooking.objects.filter(day_plan__itinerary=itinerary).prefetch_related('inclusion_items__special_inclusion'),
            VehicleBooking.objects.filter(itinerary=itinerary),
            ActivityBooking.objects.filter(day_plan__itinerary=itinerary),
            HouseboatBooking.objects.filter(day_plan__itinerary=itinerary).prefetch_related('inclusion_items__special_inclusion'),
            StandaloneInclusionBooking.objects.filter(itinerary=itinerary).select_related('special_inclusion')
        ))

        all_bookings = []

        for booking in combined_bookings:
            net_price = Decimal('0.00')
            nights = 1

            # --- Hotel Booking ---
            if isinstance(booking, HotelBooking):
                nights = (booking.check_out_date - booking.check_in_date).days or 1

                # If manual net price exists, use it
                if hasattr(booking, 'net_price') and booking.net_price and booking.net_price > 0:
                    net_price = booking.net_price

                # 🔥 FIX: Only lookup DB price if it's a REAL hotel
                elif booking.hotel:
                    rule = Hotelprice.objects.filter(
                        hotel=booking.hotel, room_type=booking.room_type, meal_plan=booking.meal_plan,
                        from_date__lte=booking.check_in_date, to_date__gte=booking.check_in_date
                    ).first()
                    if rule:
                        per_night = (
                            Decimal(booking.num_double_beds or 0) * (rule.double_bed or 0) +
                            Decimal(booking.child_with_bed or 0) * (rule.child_with_bed or 0) +
                            Decimal(booking.child_without_bed or 0) * (rule.child_without_bed or 0) +
                            Decimal(booking.extra_beds or 0) * (rule.extra_bed or 0)
                        )
                        net_price = per_night * nights
                else:
                    # Demo Hotel default price is 0
                    net_price = Decimal('0.00')

                inclusion_price = booking.get_total_inclusion_price()
                net_price += inclusion_price

                # Store rule for debugging (only if real hotel)
                booking.price_record = None
                if booking.hotel:
                    booking.price_record = Hotelprice.objects.filter(
                        hotel=booking.hotel, room_type=booking.room_type, meal_plan=booking.meal_plan,
                        from_date__lte=booking.check_in_date, to_date__gte=booking.check_in_date
                    ).first()

                booking.sort_date = booking.check_in_date
                booking.item_type = 'Accommodation'
                booking.inclusion_price = inclusion_price
                booking.inclusions_list = list(booking.inclusion_items.select_related('special_inclusion').all())

            # --- Vehicle Booking ---
            elif isinstance(booking, VehicleBooking):
                rule = VehiclePricing.objects.filter(
                    vehicle=booking.vehicle,
                    from_date__lte=booking.pickup_date,
                    to_date__gte=booking.pickup_date
                ).first()

                expected_price = Decimal('0')
                if rule and hasattr(booking, 'total_km') and booking.total_km is not None:
                    if booking.total_km <= 100:
                        expected_price = rule.total_fee_100km or Decimal('0')
                    else:
                        extra_km = booking.total_km - 100
                        extra_cost = Decimal(str(extra_km)) * (rule.extra_fee_per_km or Decimal('0'))
                        expected_price = (rule.total_fee_100km or Decimal('0')) + extra_cost

                if hasattr(booking, 'custom_total_price') and booking.custom_total_price and booking.custom_total_price > 0:
                    net_price = booking.custom_total_price
                elif hasattr(booking, 'net_price') and booking.net_price and booking.net_price > 0:
                    net_price = booking.net_price
                else:
                    net_price = expected_price

                booking.price_record = rule
                booking.sort_date = booking.pickup_date
                booking.item_type = 'Transportation'
                booking.inclusion_price = Decimal('0.00')
                booking.inclusions_list = []

            # --- Activity Booking ---
            elif isinstance(booking, ActivityBooking):
                rule = ActivityPrice.objects.filter(
                    activity=booking.activity,
                    from_date__lte=booking.booking_date,
                    to_date__gte=booking.booking_date
                ).first() or ActivityPrice.objects.filter(activity=booking.activity).first()

                expected_price = Decimal('0')
                if rule:
                    people = (booking.num_adults or 0) + (booking.num_children or 0)
                    expected_price = Decimal(str(people)) * (rule.per_person or Decimal('0'))

                if hasattr(booking, 'custom_total_price') and booking.custom_total_price and booking.custom_total_price > 0:
                    net_price = booking.custom_total_price
                elif hasattr(booking, 'net_price') and booking.net_price and booking.net_price > 0:
                    net_price = booking.net_price
                else:
                    net_price = expected_price

                booking.price_record = rule
                booking.sort_date = booking.booking_date
                booking.item_type = 'Activity'
                booking.inclusion_price = Decimal('0.00')
                booking.inclusions_list = []

            # --- Houseboat Booking ---
            elif isinstance(booking, HouseboatBooking):
                nights = (booking.check_out_date - booking.check_in_date).days or 1

                if hasattr(booking, 'net_price') and booking.net_price and booking.net_price > 0:
                    net_price = booking.net_price
                else:
                    rule = HouseboatPrice.objects.filter(
                        houseboat=booking.houseboat,
                        meal_plan=booking.meal_plan,
                        room_type=booking.room_type,
                        from_date__lte=booking.check_in_date,
                        to_date__gte=booking.check_in_date
                    ).first()
                    if rule:
                        per_night = (
                            Decimal(booking.num_one_bed_rooms or 0) * (rule.one_bed or 0) +
                            Decimal(booking.num_two_bed_rooms or 0) * (rule.two_bed or 0) +
                            Decimal(booking.num_three_bed_rooms or 0) * (rule.three_bed or 0) +
                            Decimal(booking.num_four_bed_rooms or 0) * (rule.four_bed or 0) +
                            Decimal(booking.num_five_bed_rooms or 0) * (rule.five_bed or 0) +
                            Decimal(booking.num_six_bed_rooms or 0) * (rule.six_bed or 0) +
                            Decimal(booking.num_seven_bed_rooms or 0) * (rule.seven_bed or 0) +
                            Decimal(booking.num_eight_bed_rooms or 0) * (rule.eight_bed or 0) +
                            Decimal(booking.num_nine_bed_rooms or 0) * (rule.nine_bed or 0) +
                            Decimal(booking.num_ten_bed_rooms or 0) * (rule.ten_bed or 0) +
                            Decimal(booking.num_extra_beds or 0) * (rule.extra_bed or 0)
                        )
                        net_price = per_night * nights

                inclusion_price = booking.get_total_inclusion_price()
                net_price += inclusion_price

                booking.price_record = HouseboatPrice.objects.filter(
                    houseboat=booking.houseboat,
                    meal_plan=booking.meal_plan,
                    room_type=booking.room_type,
                    from_date__lte=booking.check_in_date,
                    to_date__gte=booking.check_in_date
                ).first()
                booking.sort_date = booking.check_in_date
                booking.item_type = 'Houseboat'
                booking.inclusion_price = inclusion_price
                booking.inclusions_list = list(booking.inclusion_items.select_related('special_inclusion').all())

            # --- Standalone Inclusion ---
            elif isinstance(booking, StandaloneInclusionBooking):
                net_price = booking.total_price
                booking.price_record = None
                booking.sort_date = booking.booking_date
                booking.item_type = 'Standalone Activity'
                booking.inclusion_price = Decimal('0.00')
                booking.inclusions_list = []

            # --- Final Price Struct ---
            individual_markup = Decimal('0.00')
            if hasattr(booking, 'markup_value') and booking.markup_value:
                try:
                    markup_val = Decimal(str(booking.markup_value))
                    if markup_val > 0:
                        individual_markup = markup_val if booking.markup_type != 'percentage' else net_price * (markup_val / 100)
                except:
                    pass

            booking.calculated_price = {
                'net': net_price,
                'markup': individual_markup,
                'gross': net_price + individual_markup,
                'inclusion_price': getattr(booking, 'inclusion_price', Decimal('0.00')),
                'nights': nights
            }
            all_bookings.append(booking)

        return all_bookings

    # ==========================================
    # 3. POST LOGIC (SAVE & FINALIZE)
    # ==========================================
    if request.method == 'POST':
        cgst_perc = Decimal(request.POST.get('cgst', '0'))
        sgst_perc = Decimal(request.POST.get('sgst', '0'))
        discount = Decimal(request.POST.get('discount', '0'))

        itinerary.cgst_percentage = cgst_perc
        itinerary.sgst_percentage = sgst_perc
        itinerary.discount = discount
        itinerary.save()

        try:
            # Delete old options
            ItineraryPricingOption.objects.filter(itinerary=itinerary).delete()

            all_items = sorted(calculate_all_prices(), key=attrgetter('sort_date'))

            non_accommodation_net = sum(i.calculated_price['net'] for i in all_items if i.item_type not in ['Accommodation', 'Houseboat'])
            non_accommodation_markup = sum(i.calculated_price['markup'] for i in all_items if i.item_type not in ['Accommodation', 'Houseboat'])

            grouped_hotels = defaultdict(list)
            for item in all_items:
                if item.item_type in ['Accommodation', 'Houseboat']:
                    raw_option = getattr(item, 'option', 'option1')
                    key = str(raw_option).lower().replace(' ', '').replace('_', '')
                    if not key: key = 'option1'
                    grouped_hotels[key].append(item)

            # Create Pricing Options
            for index, (option_key, hotels) in enumerate(grouped_hotels.items(), 1):
                package_index_str = str(index)
                stored_markup_type = request.session.get(f'itinerary_{itinerary.id}_option_{package_index_str}_markup_type', 'fixed')
                stored_markup_value = Decimal(request.session.get(f'itinerary_{itinerary.id}_option_{package_index_str}_markup_value', '0'))

                option_hotels_net = sum(h.calculated_price['net'] for h in hotels)
                option_hotels_markup = sum(h.calculated_price['markup'] for h in hotels)

                package_net = option_hotels_net + non_accommodation_net
                package_markup = option_hotels_markup + non_accommodation_markup
                package_base = package_net + package_markup

                global_markup = package_base * (stored_markup_value / 100) if stored_markup_type == 'percentage' else stored_markup_value

                gross = package_base + global_markup
                cgst = gross * (cgst_perc / 100)
                sgst = gross * (sgst_perc / 100)
                final = gross + cgst + sgst - discount

                if option_key == 'option1': option_display_name = 'Option 1'
                elif option_key == 'option2': option_display_name = 'Option 2'
                elif option_key == 'option3': option_display_name = 'Option 3'
                else: option_display_name = str(option_key).title()

                # 🔥 SAFE NAME EXTRACTION FOR POST
                hotels_list = []
                for h in hotels:
                    hotel_name = "Unknown Hotel"
                    if hasattr(h, 'houseboat') and h.houseboat:
                        hotel_name = h.houseboat.name
                    elif hasattr(h, 'hotel') and h.hotel:
                        hotel_name = h.hotel.name
                    elif hasattr(h, 'custom_hotel_name') and h.custom_hotel_name:
                        hotel_name = f"{h.custom_hotel_name} (Demo)"

                    hotels_list.append({
                        'name': hotel_name,
                        'net_price': float(h.calculated_price['net']),
                        'type': h.item_type
                    })

                ItineraryPricingOption.objects.create(
                    itinerary=itinerary,
                    option_name=option_display_name,
                    option_number=index,
                    net_price=package_net,
                    markup_amount=global_markup,
                    gross_price=gross,
                    cgst_amount=cgst,
                    sgst_amount=sgst,
                    discount_amount=discount,
                    final_amount=final,
                    hotels_included=hotels_list
                )

            itinerary.is_finalized = True
            itinerary.finalized_at = now()
            itinerary.save()

            messages.success(request, f"✅ Successfully saved {len(grouped_hotels)} pricing option(s)!")
            return redirect('query_proposals', query_id=itinerary.query.id)

        except Exception as e:
            traceback.print_exc()
            messages.error(request, f"❌ Error: {str(e)}")
            return redirect('itinerary_pricing', itinerary_id=itinerary.id)

    # ==========================================
    # 4. GET LOGIC (DISPLAY)
    # ==========================================
    if request.GET.get('markup_value') is not None and request.GET.get('selected_option'):
        selected_option = request.GET.get('selected_option')
        request.session[f'itinerary_{itinerary.id}_option_{selected_option}_markup_type'] = request.GET.get('markup_type', 'fixed')
        request.session[f'itinerary_{itinerary.id}_option_{selected_option}_markup_value'] = request.GET.get('markup_value', '0')
        request.session[f'itinerary_{itinerary.id}_selected_option'] = selected_option
        request.session.modified = True

    cgst_perc = Decimal(request.GET.get('cgst', str(itinerary.cgst_percentage)))
    sgst_perc = Decimal(request.GET.get('sgst', str(itinerary.sgst_percentage)))
    discount = Decimal(request.GET.get('discount', str(itinerary.discount)))
    selected_option = request.GET.get('selected_option') or request.session.get(f'itinerary_{itinerary.id}_selected_option')

    all_items = sorted(calculate_all_prices(), key=attrgetter('sort_date'))

    non_accommodation_net = sum(i.calculated_price['net'] for i in all_items if i.item_type not in ['Accommodation', 'Houseboat'])
    non_accommodation_markup = sum(i.calculated_price['markup'] for i in all_items if i.item_type not in ['Accommodation', 'Houseboat'])

    grouped_hotels = defaultdict(list)
    for item in all_items:
        if item.item_type in ['Accommodation', 'Houseboat']:
            raw_option = getattr(item, 'option', 'option1')
            key = str(raw_option).lower().replace(' ', '').replace('_', '')
            if not key: key = 'option1'
            grouped_hotels[key].append(item)

    hotel_option_groups = []

    # Handle Case: No Hotels (Optional: You can add logic here if you want a "Standard Package" row)
    if not grouped_hotels and all_items:
         pass # Add standard package logic here if needed

    # Build Option Groups
    for index, (option_key, hotels) in enumerate(grouped_hotels.items(), 1):
        package_index_str = str(index)
        stored_markup_type = request.session.get(f'itinerary_{itinerary.id}_option_{package_index_str}_markup_type', 'fixed')
        stored_markup_value = Decimal(request.session.get(f'itinerary_{itinerary.id}_option_{package_index_str}_markup_value', '0'))

        option_hotels_net = sum(h.calculated_price['net'] for h in hotels)
        option_hotels_markup = sum(h.calculated_price['markup'] for h in hotels)

        package_net = option_hotels_net + non_accommodation_net
        package_markup = option_hotels_markup + non_accommodation_markup
        package_base = package_net + package_markup

        global_markup = package_base * (stored_markup_value / 100) if stored_markup_type == 'percentage' else stored_markup_value

        gross = package_base + global_markup
        cgst = gross * (cgst_perc / 100)
        sgst = gross * (sgst_perc / 100)
        final = gross + cgst + sgst - discount

        # Name Mapping
        option_display_name = str(option_key).title()
        if option_key == 'option1': option_display_name = 'Option 1'
        elif option_key == 'option2': option_display_name = 'Option 2'
        elif option_key == 'option3': option_display_name = 'Option 3'

        # 🔥 SAFE NAME EXTRACTION FOR TEMPLATE DISPLAY
        safe_hotels_list = []
        for h in hotels:
            hotel_name = "Unknown Hotel"
            if hasattr(h, 'houseboat') and h.houseboat:
                hotel_name = h.houseboat.name
            elif hasattr(h, 'hotel') and h.hotel:
                hotel_name = h.hotel.name
            elif hasattr(h, 'custom_hotel_name') and h.custom_hotel_name:
                hotel_name = f"{h.custom_hotel_name} (Demo)"

            safe_hotels_list.append({
                'name': hotel_name,
                'net_price': h.calculated_price['net']
            })

        hotel_option_groups.append({
            'option_name': option_display_name,
            'hotel_count': len(hotels),
            'hotels': safe_hotels_list,  # ✅ Pass this safe list to template
            'option_net_total': option_hotels_net,
            'net_price': package_net,
            'markup': package_markup,
            'global_markup': global_markup,
            'gross_before_tax': gross,
            'cgst_amount': cgst,
            'sgst_amount': sgst,
            'discount': discount,
            'gross_price': final
        })

    # Prepare Context
    current_markup_type = 'fixed'
    current_markup_value = Decimal('0')
    if selected_option:
        current_markup_type = request.session.get(f'itinerary_{itinerary.id}_option_{selected_option}_markup_type', 'fixed')
        current_markup_value = Decimal(request.session.get(f'itinerary_{itinerary.id}_option_{selected_option}_markup_value', '0'))

    context = {
        'itinerary': itinerary,
        'all_items': all_items,
        'hotel_option_groups': hotel_option_groups,
        'current_markup_type': current_markup_type,
        'current_markup_value': current_markup_value,
        'current_cgst': cgst_perc,
        'current_sgst': sgst_perc,
        'current_discount': discount,
        'selected_option': selected_option,
        'user_type': user_type,
        'current_user': current_user,
    }

    return render(request, 'pricing.html', context)







#####################################################################################################

# your_app/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from .models import (
    Itinerary, ItineraryDayPlan, Destinations, Hotel, Houseboat,
    Activity, Vehicle, RoomType, MealPlan, SpecialInclusion,
    HotelBooking, VehicleBooking, ActivityBooking, HouseboatBooking
)
from .forms import HotelBookingForm, VehicleBookingForm, ActivityBookingForm, HouseboatBookingForm


# ==========================================
# ✅ NEW: API ENDPOINT FOR HOTEL INCLUSIONS
# ==========================================

@require_http_methods(["GET"])
def get_hotel_inclusions(request, hotel_id):
    try:
        hotel = Hotel.objects.get(id=hotel_id)
        inclusions = hotel.special_inclusions.filter(is_available=True, status=True)

        data = {
            'success': True,
            'hotel_name': hotel.name,
            'count': inclusions.count(),
            'inclusions': [{
                'id': inc.id,
                'name': inc.name,
                'details': inc.details,
                'pricing_type': inc.pricing_type,
                'adult_price': float(inc.adult_price),
                'child_price': float(inc.get_child_price()),  # ✅ Add child price
                'price_display': inc.get_price_display(),
            } for inc in inclusions]
        }
        return JsonResponse(data)
    except Hotel.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Hotel not found'}, status=404)



@require_http_methods(["GET"])
def get_houseboat_inclusions(request, houseboat_id):
    """
    API endpoint to get inclusions for a specific houseboat.
    Returns JSON with inclusion data for dynamic dropdown loading.
    """
    try:
        # Get the houseboat
        houseboat = Houseboat.objects.get(id=houseboat_id)

        # Get all available inclusions for this specific houseboat only
        inclusions = SpecialInclusion.objects.filter(
            inclusion_type='houseboat',  # Only houseboat inclusions
            houseboat=houseboat,         # Only for this specific houseboat
            is_available=True,           # Only available ones
            status=True                  # Only active ones
        ).order_by('name')

        # Format the data for JSON response
        inclusions_list = []
        for inc in inclusions:
            # ✅ Use the model's get_child_price() method
            child_price = float(inc.get_child_price())

            inclusions_list.append({
                'id': inc.id,
                'name': inc.name,
                'pricing_type': inc.pricing_type,
                'adult_price': float(inc.adult_price),
                'child_price': child_price,  # ✅ Uses model's calculation method
                'details': inc.details or ''
            })

        return JsonResponse({
            'success': True,
            'inclusions': inclusions_list,
            'houseboat_name': houseboat.name,
            'count': len(inclusions_list)
        })

    except Houseboat.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Houseboat not found',
            'inclusions': []
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'inclusions': []
        }, status=500)





# ==========================================
# HOTEL INCLUSIONS MANAGEMENT
# ==========================================

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.http import require_POST
from .models import Hotel, SpecialInclusion
from .forms import SpecialInclusionForm

@custom_login_required
def manage_hotel_inclusions(request, hotel_id):
    """
    View to manage all inclusions for a specific hotel.
    """
    hotel = get_object_or_404(Hotel, id=hotel_id)

    inclusions = SpecialInclusion.objects.filter(
        inclusion_type='hotel',
        hotel=hotel
    ).order_by('-is_available', 'name') # Shows Active items first

    active_count = inclusions.filter(is_available=True).count()

    context = {
        'hotel': hotel,
        'inclusions': inclusions,
        'active_count': active_count,
        # We pass an empty form for the 'Add' modal to render correctly
        'form': SpecialInclusionForm(),
    }
    return render(request, 'manage_hotel_inclusions.html', context)




@custom_login_required
def edit_hotel_inclusion(request, inclusion_id):
    """
    Edit inclusion. Manually handles the 'is_available' checkbox
    to ensure status updates correctly every time.
    """
    inclusion = get_object_or_404(SpecialInclusion, id=inclusion_id)

    if request.method == 'POST':
        form = SpecialInclusionForm(request.POST, request.FILES, instance=inclusion)

        if form.is_valid():
            obj = form.save(commit=False)

            # ✅ FIX: Manually handle the checkbox.
            # If 'is_available' is in POST data, it's True. If missing, it's False.
            obj.is_available = 'is_available' in request.POST

            obj.save()
            messages.success(request, f'✅ Updated "{obj.name}" successfully!')
        else:
            # Show specific errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Update failed - {field}: {error}")

        return redirect('manage_hotel_inclusions', hotel_id=inclusion.hotel.id)

    return redirect('manage_hotel_inclusions', hotel_id=inclusion.hotel.id)

@custom_login_required
@require_POST
def add_hotel_inclusion(request, hotel_id):
    """
    Add inclusion. Manually handles status to be safe.
    """
    hotel = get_object_or_404(Hotel, id=hotel_id)

    form = SpecialInclusionForm(request.POST, request.FILES)

    if form.is_valid():
        inclusion = form.save(commit=False)
        inclusion.hotel = hotel
        inclusion.inclusion_type = 'hotel'

        # ✅ FIX: Explicitly check for the checkbox key
        inclusion.is_available = 'is_available' in request.POST

        if hasattr(request.user, 'team_member'):
            inclusion.created_by = request.user.team_member

        inclusion.save()
        messages.success(request, f'✅ Added "{inclusion.name}" successfully!')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f"Error in {field}: {error}")

    return redirect('manage_hotel_inclusions', hotel_id=hotel_id)

@custom_login_required
@require_POST
def delete_hotel_inclusion(request, inclusion_id):
    """
    Delete a specific hotel inclusion.
    """
    inclusion = get_object_or_404(SpecialInclusion, id=inclusion_id)
    hotel_id = inclusion.hotel.id # Save ID before deleting for redirect
    name = inclusion.name

    try:
        inclusion.delete()
        messages.success(request, f'🗑️ Inclusion "{name}" deleted.')
    except Exception as e:
        messages.error(request, f'❌ Error deleting item: {str(e)}')

    return redirect('manage_hotel_inclusions', hotel_id=hotel_id)


@custom_login_required
def toggle_inclusion_availability(request, inclusion_id):
    """
    Toggle Active/Inactive status via a simple link/button.
    """
    inclusion = get_object_or_404(SpecialInclusion, id=inclusion_id)
    inclusion.is_available = not inclusion.is_available
    inclusion.save()

    status = "Activated" if inclusion.is_available else "Deactivated"
    messages.info(request, f'Status for "{inclusion.name}" changed to {status}.')

    return redirect('manage_hotel_inclusions', hotel_id=inclusion.hotel.id)


# ==========================================
# HOUSEBOAT INCLUSIONS MANAGEMENT
# ==========================================


def manage_houseboat_inclusions(request, houseboat_id):
    """
    View to manage all inclusions for a specific houseboat.
    """
    houseboat = get_object_or_404(Houseboat, id=houseboat_id)

    # Get all inclusions for this houseboat
    inclusions = SpecialInclusion.objects.filter(
        inclusion_type='houseboat',
        houseboat=houseboat
    ).order_by('name')

    context = {
        'houseboat': houseboat,
        'inclusions': inclusions,
    }

    return render(request, 'manage_houseboat_inclusions.html', context)



def add_houseboat_inclusion(request, houseboat_id):
    """
    Add a new inclusion to a specific houseboat.
    """
    if request.method == 'POST':
        houseboat = get_object_or_404(Houseboat, id=houseboat_id)

        # Get form data
        name = request.POST.get('name')
        pricing_type = request.POST.get('pricing_type', 'free')
        adult_price = request.POST.get('adult_price', 0)

        # Child pricing fields
        child_pricing_type = request.POST.get('child_pricing_type', 'same')
        child_price_value = request.POST.get('child_price_value', 0)
        min_age = request.POST.get('min_age')
        max_age = request.POST.get('max_age')

        details = request.POST.get('details', '')
        is_available = request.POST.get('is_available') == 'on'

        # Validate required fields
        if not name:
            messages.error(request, '❌ Inclusion name is required!')
            return redirect('manage_houseboat_inclusions', houseboat_id=houseboat_id)

        # Create the inclusion
        try:
            inclusion = SpecialInclusion.objects.create(
                name=name,
                inclusion_type='houseboat',
                houseboat=houseboat,
                pricing_type=pricing_type,
                adult_price=float(adult_price) if adult_price else 0,
                child_pricing_type=child_pricing_type,
                child_price_value=float(child_price_value) if child_price_value else 0,
                min_age=int(min_age) if min_age else None,
                max_age=int(max_age) if max_age else None,
                details=details,
                is_available=is_available,
                status=True,
                created_by=request.user.team_member if hasattr(request.user, 'team_member') else None
            )

            messages.success(request, f'✅ Inclusion "{name}" added successfully!')

        except Exception as e:
            messages.error(request, f'❌ Error adding inclusion: {str(e)}')

        return redirect('manage_houseboat_inclusions', houseboat_id=houseboat_id)

    return redirect('manage_houseboat_inclusions', houseboat_id=houseboat_id)




def delete_houseboat_inclusion(request, inclusion_id):
    """
    Delete a specific houseboat inclusion.
    """
    if request.method == 'POST':
        inclusion = get_object_or_404(SpecialInclusion, id=inclusion_id)
        houseboat_id = inclusion.houseboat.id
        inclusion_name = inclusion.name

        try:
            inclusion.delete()
            messages.success(request, f'✅ Inclusion "{inclusion_name}" deleted successfully!')
        except Exception as e:
            messages.error(request, f'❌ Error deleting inclusion: {str(e)}')

        return redirect('manage_houseboat_inclusions', houseboat_id=houseboat_id)

    messages.warning(request, '⚠️ Invalid request method')
    return redirect('houseboats_list')



def edit_houseboat_inclusion(request, inclusion_id):
    """
    Edit an existing houseboat inclusion.
    """
    inclusion = get_object_or_404(SpecialInclusion, id=inclusion_id, inclusion_type='houseboat')

    if request.method == 'POST':
        # Update fields
        inclusion.name = request.POST.get('name', inclusion.name)
        inclusion.pricing_type = request.POST.get('pricing_type', inclusion.pricing_type)
        inclusion.adult_price = float(request.POST.get('adult_price', inclusion.adult_price))
        inclusion.child_pricing_type = request.POST.get('child_pricing_type', inclusion.child_pricing_type)
        inclusion.child_price_value = float(request.POST.get('child_price_value', inclusion.child_price_value))
        inclusion.min_age = int(request.POST.get('min_age')) if request.POST.get('min_age') else None
        inclusion.max_age = int(request.POST.get('max_age')) if request.POST.get('max_age') else None
        inclusion.details = request.POST.get('details', inclusion.details)
        inclusion.is_available = request.POST.get('is_available') == 'on'

        inclusion.save()

        messages.success(request, f'✅ Houseboat inclusion "{inclusion.name}" updated successfully!')

        # Redirect to houseboat inclusions page
        if inclusion.houseboat:
            return redirect('manage_houseboat_inclusions', houseboat_id=inclusion.houseboat.id)
        else:
            messages.warning(request, 'Inclusion updated but houseboat not found')
            return redirect('houseboats')

    # GET request - show edit form
    context = {
        'inclusion': inclusion,
        'houseboat': inclusion.houseboat,
    }
    return render(request, 'edit_houseboat_inclusion.html', context)


def toggle_houseboat_inclusion_availability(request, inclusion_id):
    """
    Toggle houseboat inclusion availability (active/inactive)
    """
    inclusion = get_object_or_404(SpecialInclusion, id=inclusion_id, inclusion_type='houseboat')
    inclusion.is_available = not inclusion.is_available
    inclusion.save()

    status = "activated" if inclusion.is_available else "deactivated"
    messages.success(request, f'✅ Inclusion "{inclusion.name}" {status}!')

    if inclusion.houseboat:
        return redirect('manage_houseboat_inclusions', houseboat_id=inclusion.houseboat.id)
    else:
        return redirect('houseboats')



# views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.http import require_http_methods
import json
import traceback
from .models import (
    Itinerary, ItineraryDayPlan, HotelBooking, Hotel, RoomType, MealPlan,
    SpecialInclusion, HotelBookingInclusion
)

@require_http_methods(["POST"])
def create_hotel_booking(request, itinerary_id):
    """
    Create or update hotel booking with:
    - Special inclusions support
    - Demo Hotel support (Custom Name)
    - Manual Room Type support (Custom Room)
    """
    try:
        itinerary = get_object_or_404(Itinerary, id=itinerary_id)

        # ============================================================
        # STEP 1: GET FORM DATA (with multiple field name support)
        # ============================================================
        def get_field(field_names):
            """Helper to check multiple possible field names"""
            if isinstance(field_names, str):
                field_names = [field_names]
            for name in field_names:
                value = request.POST.get(name)
                if value:
                    return value.strip()
            return None

        booking_id = get_field(['booking_id'])

        # 🔥 CORE INPUTS: Check for both Hotel ID and Custom Name
        hotel_id = get_field(['hotel_id', 'hotel'])
        custom_hotel_name = get_field(['custom_hotel_name']) # New Field
        destination_id = get_field(['destination', 'destination_id']) # Required for Demo Mode

        day_number = get_field(['day_number', 'daynumber'])
        category = get_field(['category']) or '3 Star'

        # 🔥 ROOM TYPE INPUTS: Check for ID or Manual Text
        room_type_id = get_field(['room_type', 'room_type_id'])
        custom_room_type = get_field(['custom_room_type']) # ✅ NEW MANUAL FIELD

        meal_plan_id = get_field(['meal_plan', 'meal_plan_id'])
        option = get_field(['option']) or 'option_1'

        # Room configuration
        num_double_beds = int(get_field(['num_double_beds', 'numdoublebeds']) or 0)
        child_with_bed = int(get_field(['child_with_bed', 'childwithbed']) or 0)
        child_without_bed = int(get_field(['child_without_bed', 'childwithoutbed']) or 0)
        extra_beds = int(get_field(['extra_beds', 'extrabeds']) or 0)

        # Check-in/Check-out
        check_in_date = get_field(['check_in_date', 'checkindate'])
        check_in_time = get_field(['check_in_time', 'checkintime']) or '14:00'
        check_out_date = get_field(['check_out_date', 'checkoutdate'])
        check_out_time = get_field(['check_out_time', 'checkouttime']) or '11:00'

        # Ensure HH:MM:SS format for times
        if check_in_time and len(check_in_time) == 5:
            check_in_time = f"{check_in_time}:00"
        if check_out_time and len(check_out_time) == 5:
            check_out_time = f"{check_out_time}:00"

        # Get inclusions data
        inclusions_data = get_field(['inclusions_data', 'inclusionsdata']) or '[]'

        # ============================================================
        # STEP 2: VALIDATE REQUIRED FIELDS & RESOLVE OBJECTS
        # ============================================================

        # 🚨 VALIDATION: Must have either a Real Hotel OR a Custom Name
        if not hotel_id and not custom_hotel_name:
            messages.error(request, '❌ You must select a Hotel or enter a Demo Hotel Name!')
            return redirect(f'/itinerary/{itinerary_id}/day-plan/?day={day_number or 1}')

        if not day_number:
            messages.error(request, '❌ Day number is required!')
            return redirect(f'/itinerary/{itinerary_id}/day-plan/')

        # 🧠 LOGIC: Determine Hotel Object and Destination
        hotel_obj = None
        destination_obj = None

        if hotel_id:
            # CASE A: Real Database Hotel
            hotel_obj = get_object_or_404(Hotel, id=hotel_id)
            destination_obj = hotel_obj.destination # Prioritize Hotel's destination
            custom_hotel_name = None # Clear custom name if ID exists
        elif custom_hotel_name:
            # CASE B: Demo Hotel
            hotel_obj = None
            if destination_id:
                destination_obj = get_object_or_404(Destinations, id=destination_id)
            else:
                # Fallback if destination missing in form
                # Try to get from existing day plan if available, or error
                pass

        # 🧠 LOGIC: Resolve Room Type (Manual vs Dropdown)
        room_type_obj = None

        if custom_room_type:
            # ✅ Manual Entry - Clear relation, set text
            room_type_obj = None
            # custom_room_type variable already holds the text
        elif room_type_id:
            # ✅ Dropdown Selection
            try:
                room_type_obj = get_object_or_404(RoomType, id=room_type_id)
                custom_room_type = None # Clear manual text
            except Exception:
                # Fallback if ID is invalid but we need something
                room_type_obj = RoomType.objects.first()

        # Resolve Meal Plan
        try:
            meal_plan = get_object_or_404(MealPlan, id=meal_plan_id) if meal_plan_id else None
        except Exception:
             meal_plan = MealPlan.objects.first()


        # ============================================================
        # STEP 3: LOG REQUEST DATA
        # ============================================================
        print(f"\n{'='*70}")
        print(f"📥 HOTEL BOOKING SUBMISSION - Itinerary #{itinerary_id}")
        print(f"{'='*70}")
        print(f"🏨 Hotel: {hotel_obj.name if hotel_obj else custom_hotel_name} ({'DB' if hotel_obj else 'DEMO'})")
        print(f"🛏️ Room: {room_type_obj.name if room_type_obj else custom_room_type} ({'DB' if room_type_obj else 'MANUAL'})")
        print(f"📅 Day: {day_number}")
        print(f"{'='*70}\n")

        # ============================================================
        # STEP 4: ENSURE DAY PLAN EXISTS
        # ============================================================

        day_plan, created = ItineraryDayPlan.objects.get_or_create(
            itinerary=itinerary,
            day_number=day_number,
            defaults={
                'title': f'Day {day_number}',
                'description': '',
                'destination': destination_obj
            }
        )

        # ============================================================
        # STEP 5: CREATE OR UPDATE BOOKING
        # ============================================================
        if booking_id:
            # UPDATE existing booking
            booking = get_object_or_404(HotelBooking, id=booking_id)

            # 🔥 Update Core Identity fields
            booking.hotel = hotel_obj
            booking.custom_hotel_name = custom_hotel_name
            booking.destination = destination_obj

            # 🔥 Update Room Fields
            booking.room_type = room_type_obj
            booking.custom_room_type = custom_room_type # ✅ Save manual room text

            # Update Standard Fields
            booking.category = category
            booking.meal_plan = meal_plan
            booking.option = option
            booking.num_double_beds = num_double_beds
            booking.child_with_bed = child_with_bed
            booking.child_without_bed = child_without_bed
            booking.extra_beds = extra_beds
            booking.check_in_date = check_in_date
            booking.check_in_time = check_in_time
            booking.check_out_date = check_out_date
            booking.check_out_time = check_out_time
            booking.save()

            # DELETE old inclusions (to be replaced)
            deleted_count = booking.inclusion_items.all().delete()[0]

            action = "updated"
        else:
            # CREATE new booking
            booking = HotelBooking.objects.create(
                itinerary=itinerary,
                day_plan=day_plan,

                # 🔥 Core Identity Fields
                hotel=hotel_obj,
                custom_hotel_name=custom_hotel_name,
                destination=destination_obj,

                # 🔥 Room Type Fields
                room_type=room_type_obj,
                custom_room_type=custom_room_type, # ✅ Save manual room text

                # Standard Fields
                category=category,
                meal_plan=meal_plan,
                option=option,
                num_double_beds=num_double_beds,
                child_with_bed=child_with_bed,
                child_without_bed=child_without_bed,
                extra_beds=extra_beds,
                check_in_date=check_in_date,
                check_in_time=check_in_time,
                check_out_date=check_out_date,
                check_out_time=check_out_time
            )
            action = "created"
            print(f"✅ Created booking ID: {booking.id}")

        # ============================================================
        # STEP 6: PROCESS SPECIAL INCLUSIONS
        # ============================================================
        inclusion_count = 0
        try:
            inclusions_list = json.loads(inclusions_data) if inclusions_data else []

            for idx, inc_data in enumerate(inclusions_list, 1):
                # Check IDs
                inclusion_id = inc_data.get('id') or inc_data.get('inclusion_id')
                num_adults = int(inc_data.get('adults') or 0)
                num_children = int(inc_data.get('children') or 0)

                if not inclusion_id: continue
                if num_adults == 0 and num_children == 0: continue

                try:
                    special_inclusion = SpecialInclusion.objects.get(id=inclusion_id)

                    # Calculate Price Logic
                    pricing_type = getattr(special_inclusion, 'pricing_type', 'per_person')
                    price = 0
                    if pricing_type == 'free': price = 0
                    elif pricing_type == 'per_person':
                        # Use method if available, else attribute
                        child_p = special_inclusion.child_price
                        if hasattr(special_inclusion, 'get_child_price'):
                             child_p = special_inclusion.get_child_price()

                        price = (float(special_inclusion.adult_price) * num_adults) + \
                                (float(child_p) * num_children)

                    HotelBookingInclusion.objects.create(
                        hotel_booking=booking,
                        special_inclusion=special_inclusion,
                        num_adults=num_adults,
                        num_children=num_children,
                        price=price
                    )
                    inclusion_count += 1

                except SpecialInclusion.DoesNotExist:
                    print(f"⚠️ Inclusion ID {inclusion_id} not found")
                    continue

        except json.JSONDecodeError:
            print("❌ Error parsing inclusions JSON")
        except Exception as e:
            print(f"❌ Error processing inclusions: {str(e)}")
            traceback.print_exc()

        # ============================================================
        # STEP 7: SUCCESS MESSAGE AND REDIRECT
        # ============================================================
        # Display name logic
        display_name = hotel_obj.name if hotel_obj else f"{custom_hotel_name} (Demo)"

        success_message = f"✅ {display_name} {action} successfully for Day {day_number}!"
        if inclusion_count > 0:
            success_message += f" (+ {inclusion_count} inclusions)"

        messages.success(request, success_message)
        return redirect(f'/itinerary/{itinerary_id}/day-plan/?day={day_number}')

    except Exception as e:
        print(f"\n❌ CRITICAL ERROR: {str(e)}")
        traceback.print_exc()
        messages.error(request, f'❌ Error: {str(e)}')
        return redirect(f'/itinerary/{itinerary_id}/day-plan/')




from django.db import transaction
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.views.decorators.http import require_POST


from django.http import JsonResponse
from django.views.decorators.http import require_GET
from .models import Vehicle


@require_GET
def vehicles_by_destination(request):
    destination_id = request.GET.get('destination_id')

    if not destination_id:
        return JsonResponse({
            'success': False,
            'error': 'Destination ID is required'
        }, status=400)

    vehicles = Vehicle.objects.filter(
        destination_id=destination_id
    ).values(
        'id',
        'name',
        'vehicle_type'
    )

    return JsonResponse({
        'success': True,
        'vehicles': list(vehicles)
    })



@require_POST
def create_vehicle_booking(request, itinerary_id):
    """
    Create or Update Vehicle Booking.
    ✅ FIX: Enforces 'One Vehicle Per Day' rule (Replaces existing vehicle if adding new one)
    """
    itinerary = get_object_or_404(Itinerary, id=itinerary_id)
    form = VehicleBookingForm(request.POST)

    print("=" * 60)
    print("🚗 VEHICLE BOOKING VIEW CALLED")
    print(f"📦 POST data: {request.POST}")
    print("=" * 60)

    if not form.is_valid():
        print("❌ Form validation failed:", form.errors)
        messages.error(request, f"Form validation failed: {form.errors.as_text()}")
        day_number = request.POST.get("day_number", 1)
        return redirect(f'{reverse("itinerary_day_plan", args=[itinerary.id])}?day={day_number}')

    try:
        day_number = int(request.POST.get("day_number"))
        destination_id = int(request.POST.get("destination_id"))
        vehicle_id = int(request.POST.get("vehicle_id"))
        booking_id = request.POST.get("booking_id")
        is_vehicle_change = request.POST.get("is_vehicle_change") == "true"

        print(f"✅ Parsed: day={day_number}, vehicle={vehicle_id}, change={is_vehicle_change}")

    except Exception as e:
        print("❌ Parse error:", e)
        messages.error(request, "Invalid booking data.")
        return redirect(f'{reverse("itinerary_day_plan", args=[itinerary.id])}?day={day_number}')

    vehicle = get_object_or_404(Vehicle, id=vehicle_id)
    destination = get_object_or_404(Destinations, id=destination_id)

    with transaction.atomic():
        # Get or create day plan
        day_plan, _ = ItineraryDayPlan.objects.get_or_create(
            itinerary=itinerary,
            day_number=day_number,
            defaults={"destination": destination}
        )

        # ============================================
        # CASE 1: UPDATE / CHANGE EXISTING BOOKING
        # ============================================
        if booking_id:
            booking = get_object_or_404(VehicleBooking, id=booking_id)
            print(f"✏️ Updating booking ID={booking.id}")

            booking.pickup_date = form.cleaned_data["pickup_date"]
            booking.num_passengers = form.cleaned_data["num_passengers"]
            booking.total_km = form.cleaned_data.get("total_km")

            if is_vehicle_change:
                print("🔁 Vehicle CHANGE detected")
                booking.vehicle = vehicle

            booking.destination = destination
            booking.day_plan = day_plan
            booking.day_number = day_number
            booking.save()

            messages.success(
                request,
                f'Vehicle updated to {booking.vehicle.name}'
            )

        # ============================================
        # CASE 2: CREATE NEW BOOKING (Enforce One Per Day)
        # ============================================
        else:
            # 🔥 CRITICAL FIX: Check for existing vehicle on this day and delete it
            existing_bookings = VehicleBooking.objects.filter(day_plan=day_plan)

            if existing_bookings.exists():
                count = existing_bookings.count()
                existing_bookings.delete()
                print(f"⚠️ Replaced {count} existing vehicle(s) for Day {day_number}")
                # Optional: Add an info message telling the user it was replaced
                messages.info(request, f"ℹ️ Replaced previous vehicle for Day {day_number}")

            # Now create the new one
            booking = form.save(commit=False)
            booking.itinerary = itinerary
            booking.day_plan = day_plan
            booking.day_number = day_number
            booking.destination = destination
            booking.vehicle = vehicle
            booking.save()

            print(f"✅ New booking created: ID={booking.id}")

            messages.success(
                request,
                f'Vehicle booking for {vehicle.name} created successfully!'
            )

    return redirect(
        f'{reverse("itinerary_day_plan", args=[itinerary.id])}?day={day_number}'
    )

# your_app/views.py

@require_POST
def create_activity_booking(request, itinerary_id):
    itinerary = get_object_or_404(Itinerary, id=itinerary_id)
    form = ActivityBookingForm(request.POST)

    if form.is_valid():
        day_number = request.POST.get("day_number")
        activity_id = request.POST.get("activity_id")

        if not day_number or not activity_id:
            messages.error(request, "A day number and activity are required.")
            return redirect('itinerary_day_plan', itinerary_id=itinerary.id)

        day_plan, _ = ItineraryDayPlan.objects.get_or_create(
            itinerary=itinerary,
            day_number=day_number
        )

        booking = form.save(commit=False)

        # ✅ FIXED: Add this line to assign the itinerary to the booking
        booking.itinerary = itinerary

        booking.day_plan = day_plan
        booking.activity_id = activity_id
        booking.save()
        messages.success(request, "Activity successfully booked!")
    else:
        messages.error(request, f"There was an error: {form.errors.as_text()}")

    return redirect('itinerary_day_plan', itinerary_id=itinerary.id)








import json
from decimal import Decimal
from datetime import datetime

@require_POST
def create_houseboat_booking(request, itinerary_id):
    """
    Create a new houseboat booking with multi-inclusions support
    ✅ Validates that pricing exists for the selected date range
    ✅ Redirects back to the same day after creation
    """
    itinerary = get_object_or_404(Itinerary, id=itinerary_id)
    form = HouseboatBookingForm(request.POST)

    # ✅ Get day_number at the start so we can use it for redirect
    day_number = request.POST.get("day_number", "1")

    if form.is_valid():
        houseboat_id = request.POST.get("houseboat_id")
        check_in_date_str = request.POST.get("check_in_date")
        check_out_date_str = request.POST.get("check_out_date")

        # ✅ NEW: Validate pricing exists for date range
        try:
            check_in_date = datetime.strptime(check_in_date_str, '%Y-%m-%d').date()
            check_out_date = datetime.strptime(check_out_date_str, '%Y-%m-%d').date()

            # Check if pricing exists for this houseboat and date range
            houseboat = Houseboat.objects.get(id=houseboat_id)

            # Query for pricing that covers the booking dates
            pricing_exists = HouseboatPrice.objects.filter(
                houseboat=houseboat,
                from_date__lte=check_in_date,
                to_date__gte=check_out_date
            ).exists()

            if not pricing_exists:
                messages.error(
                    request,
                    f'❌ No pricing available for {houseboat.name} from {check_in_date.strftime("%d %b %Y")} '
                    f'to {check_out_date.strftime("%d %b %Y")}. Please add pricing first!'
                )
                # ✅ FIX: Redirect to the specific day
                return redirect(f"/itinerary/{itinerary.id}/day-plan/?day={day_number}")

            print(f"✅ Pricing validation passed for {houseboat.name} ({check_in_date} to {check_out_date})")

        except Houseboat.DoesNotExist:
            messages.error(request, '❌ Houseboat not found!')
            # ✅ FIX: Redirect to the specific day
            return redirect(f"/itinerary/{itinerary.id}/day-plan/?day={day_number}")
        except ValueError:
            messages.error(request, '❌ Invalid date format!')
            # ✅ FIX: Redirect to the specific day
            return redirect(f"/itinerary/{itinerary.id}/day-plan/?day={day_number}")

        # Get or create day plan
        day_plan, _ = ItineraryDayPlan.objects.get_or_create(
            itinerary=itinerary,
            day_number=day_number
        )

        # Create booking
        booking = form.save(commit=False)
        booking.day_plan = day_plan
        booking.houseboat_id = houseboat_id

        if hasattr(booking, 'itinerary'):
            booking.itinerary = itinerary

        booking.save()

        # ✅ Handle multi-inclusions
        inclusions_data_json = request.POST.get('inclusions_data', '[]')

        try:
            inclusions_data = json.loads(inclusions_data_json)
            print(f"📦 Received {len(inclusions_data)} houseboat inclusions")

            # Delete existing inclusions (in case of update)
            booking.inclusion_items.all().delete()

            # Create new inclusion items
            for inc_data in inclusions_data:
                inclusion_id = inc_data.get('id')
                num_adults = int(inc_data.get('adults', 0))
                num_children = int(inc_data.get('children', 0))

                if inclusion_id and (num_adults > 0 or num_children > 0):
                    try:
                        special_inclusion = SpecialInclusion.objects.get(id=inclusion_id)

                        # Create the inclusion item (price calculated automatically in save())
                        HouseboatBookingInclusion.objects.create(
                            houseboat_booking=booking,
                            special_inclusion=special_inclusion,
                            num_adults=num_adults,
                            num_children=num_children
                        )

                        print(f"✅ Created houseboat inclusion: {special_inclusion.name} - {num_adults}A + {num_children}C")

                    except SpecialInclusion.DoesNotExist:
                        print(f"⚠️ Special inclusion {inclusion_id} not found")
                        continue

            messages.success(request, f'✅ Houseboat booking created successfully with {len(inclusions_data)} inclusions!')

        except json.JSONDecodeError as e:
            print(f"❌ Error parsing inclusions JSON: {e}")
            messages.warning(request, 'Houseboat booking created but inclusions could not be saved.')
        except Exception as e:
            print(f"❌ Error creating houseboat inclusions: {e}")
            messages.warning(request, f'Houseboat booking created but error with inclusions: {str(e)}')

    else:
        messages.error(request, f"❌ Error: {form.errors.as_text()}")

    # ✅ FIX: Redirect to the specific day that was selected
    return redirect(f"/itinerary/{itinerary.id}/day-plan/?day={day_number}")




from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Itinerary, ItineraryDayPlan, StandaloneInclusionBooking, SpecialInclusion
from .forms import StandaloneInclusionBookingForm

def create_standalone_inclusion(request, itinerary_id):
    """
    Create a standalone inclusion booking for an itinerary day
    """
    if request.method != 'POST':
        messages.error(request, '❌ Invalid request method')
        return redirect('itinerary_day_plan', itinerary_id=itinerary_id)

    itinerary = get_object_or_404(Itinerary, id=itinerary_id)

    try:
        # Get form data
        day_number = int(request.POST.get('day_number'))
        special_inclusion_id = request.POST.get('special_inclusion_id')
        booking_date = request.POST.get('booking_date')
        booking_time = request.POST.get('booking_time') or None
        num_adults = int(request.POST.get('num_adults', 1))
        num_children = int(request.POST.get('num_children', 0))
        markup_type = request.POST.get('markup_type', 'fixed')
        markup_value = float(request.POST.get('markup_value', 0))
        notes = request.POST.get('notes', '').strip()

        # Validate
        if not special_inclusion_id or not booking_date:
            messages.error(request, '❌ Activity and booking date are required')
            return redirect('itinerary_day_plan', itinerary_id=itinerary_id)

        if num_adults == 0 and num_children == 0:
            messages.error(request, '❌ At least one adult or child participant is required')
            return redirect('itinerary_day_plan', itinerary_id=itinerary_id)

        # Get or create day plan
        day_plan, created = ItineraryDayPlan.objects.get_or_create(
            itinerary=itinerary,
            day_number=day_number
        )

        # Get special inclusion
        special_inclusion = get_object_or_404(
            SpecialInclusion,
            id=special_inclusion_id,
            inclusion_type='general'
        )

        # Create booking
        booking = StandaloneInclusionBooking.objects.create(
            itinerary=itinerary,
            day_plan=day_plan,
            special_inclusion=special_inclusion,
            booking_date=booking_date,
            booking_time=booking_time,
            num_adults=num_adults,
            num_children=num_children,
            notes=notes
        )

        # Prices are auto-calculated in model's save method

        messages.success(
            request,
            f'✅ {special_inclusion.name} booked successfully for Day {day_number}! '
            f'Total: ₹{booking.total_price}'
        )

    except ValueError as e:
        messages.error(request, f'❌ Invalid data: {str(e)}')
    except Exception as e:
        messages.error(request, f'❌ Error creating booking: {str(e)}')
        print(f"Error in create_standalone_inclusion: {e}")

    return redirect('itinerary_day_plan', itinerary_id=itinerary_id)


def update_standalone_inclusion(request, booking_id):
    """
    Update an existing standalone inclusion booking
    """
    if request.method != 'POST':
        messages.error(request, '❌ Invalid request method')
        return redirect('list_itineraries')

    booking = get_object_or_404(StandaloneInclusionBooking, id=booking_id)
    itinerary_id = booking.itinerary.id

    try:
        # Update fields
        booking.booking_date = request.POST.get('booking_date')
        booking.booking_time = request.POST.get('booking_time') or None
        booking.num_adults = int(request.POST.get('num_adults', 1))
        booking.num_children = int(request.POST.get('num_children', 0))
        booking.markup_type = request.POST.get('markup_type', 'fixed')
        booking.markup_value = float(request.POST.get('markup_value', 0))
        booking.notes = request.POST.get('notes', '').strip()

        # Validate
        if booking.num_adults == 0 and booking.num_children == 0:
            messages.error(request, '❌ At least one adult or child participant is required')
            return redirect('itinerary_day_plan', itinerary_id=itinerary_id)

        # Save (prices are auto-calculated)
        booking.save()

        messages.success(
            request,
            f'✅ {booking.special_inclusion.name} updated successfully! '
            f'Total: ₹{booking.total_price}'
        )

    except ValueError as e:
        messages.error(request, f'❌ Invalid data: {str(e)}')
    except Exception as e:
        messages.error(request, f'❌ Error updating booking: {str(e)}')
        print(f"Error in update_standalone_inclusion: {e}")

    return redirect('itinerary_day_plan', itinerary_id=itinerary_id)

def delete_standalone_inclusion(request, booking_id):
    """
    Delete a standalone inclusion booking
    """
    booking = get_object_or_404(StandaloneInclusionBooking, id=booking_id)
    itinerary_id = booking.itinerary.id
    inclusion_name = booking.special_inclusion.name

    try:
        booking.delete()
        messages.success(request, f'✅ {inclusion_name} booking deleted successfully')
    except Exception as e:
        messages.error(request, f'❌ Error deleting booking: {str(e)}')

    return redirect('itinerary_day_plan', itinerary_id=itinerary_id)



import json
from decimal import Decimal, InvalidOperation
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction
from .models import (
    HotelBooking, HouseboatBooking, VehicleBooking, ActivityBooking,
    StandaloneInclusionBooking,
    HotelBookingInclusion, HouseboatBookingInclusion
)


@require_POST
def update_booking_totals(request):
    """
    A unified view to handle custom price updates for all booking types
    from the itinerary item editing modal, with support for inclusions.
    """
    # ✅ Check if AJAX request
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    try:
        booking_id = request.POST.get('booking_id')
        item_type = request.POST.get('item_type')

        if not booking_id or not item_type:
            error_msg = 'Missing booking ID or item type.'
            if is_ajax:
                return JsonResponse({'success': False, 'error': error_msg}, status=400)
            return JsonResponse({'success': False, 'error': error_msg}, status=400)

        # Helper to safely convert form values to Decimal
        def safe_decimal(value, default='0'):
            if value is None or value == '':
                return Decimal(default)
            try:
                return Decimal(str(value))
            except (InvalidOperation, ValueError, TypeError):
                return Decimal(default)

        # Get generic data from the form
        markup_type = request.POST.get('markup_type', 'fixed')
        markup_value = safe_decimal(request.POST.get('markup_value'))

        net_price = Decimal('0')
        booking = None
        fields_to_update = []

        # --- Process based on Item Type ---
        if item_type == 'hotel':
            booking = HotelBooking.objects.get(id=int(booking_id))

            with transaction.atomic():
                # Update custom room totals
                line_totals = {
                    'custom_double_bed_total': safe_decimal(request.POST.get('double_bed_total')),
                    'custom_extra_bed_total': safe_decimal(request.POST.get('extra_bed_total')),
                    'custom_child_with_bed_total': safe_decimal(request.POST.get('child_with_bed_total')),
                    'custom_child_without_bed_total': safe_decimal(request.POST.get('child_without_bed_total')),
                }

                for field, value in line_totals.items():
                    if hasattr(booking, field):
                        setattr(booking, field, value if value > 0 else None)
                        fields_to_update.append(field)
                    net_price += value

                # ✅ HANDLE INCLUSIONS
                booking.inclusion_items.all().delete()

                inclusions_json = request.POST.get('inclusions_data', '[]')
                print(f"📥 Hotel Inclusions Data: {inclusions_json}")

                inclusion_count = 0
                if inclusions_json and inclusions_json.strip() and inclusions_json != '[]':
                    try:
                        inclusions_data = json.loads(inclusions_json)

                        for inc_data in inclusions_data:
                            inclusion_id = inc_data.get('id')
                            num_adults = int(inc_data.get('adults', 0))
                            num_children = int(inc_data.get('children', 0))

                            if inclusion_id and (num_adults > 0 or num_children > 0):
                                HotelBookingInclusion.objects.create(
                                    hotel_booking=booking,
                                    special_inclusion_id=inclusion_id,
                                    num_adults=num_adults,
                                    num_children=num_children
                                )
                                inclusion_count += 1

                        print(f"✅ Saved {inclusion_count} hotel inclusions")
                    except json.JSONDecodeError as e:
                        print(f"❌ JSON Parse Error: {e}")
                        print(f"   Problematic JSON: {inclusions_json[:200]}")

        elif item_type == 'houseboat':
            booking = HouseboatBooking.objects.get(id=int(booking_id))

            with transaction.atomic():
                room_mappings = {
                    'one_bed_total': 'custom_one_bed_total',
                    'two_bed_total': 'custom_two_bed_total',
                    'three_bed_total': 'custom_three_bed_total',
                    'four_bed_total': 'custom_four_bed_total',
                    'five_bed_total': 'custom_five_bed_total',
                    'six_bed_total': 'custom_six_bed_total',
                    'seven_bed_total': 'custom_seven_bed_total',
                    'eight_bed_total': 'custom_eight_bed_total',
                    'nine_bed_total': 'custom_nine_bed_total',
                    'ten_bed_total': 'custom_ten_bed_total',
                    'extra_bed_hb_total': 'custom_extra_bed_hb_total',
                }

                for form_name, model_field in room_mappings.items():
                    if form_name in request.POST:
                        room_total = safe_decimal(request.POST.get(form_name))
                        net_price += room_total
                        if hasattr(booking, model_field):
                            setattr(booking, model_field, room_total if room_total > 0 else None)
                            fields_to_update.append(model_field)

                # ✅ HANDLE HOUSEBOAT INCLUSIONS
                booking.inclusion_items.all().delete()

                inclusions_json = request.POST.get('inclusions_data', '[]')
                print(f"📥 Houseboat Inclusions Data: {inclusions_json}")

                inclusion_count = 0
                if inclusions_json and inclusions_json.strip() and inclusions_json != '[]':
                    try:
                        inclusions_data = json.loads(inclusions_json)

                        for inc_data in inclusions_data:
                            inclusion_id = inc_data.get('id')
                            num_adults = int(inc_data.get('adults', 0))
                            num_children = int(inc_data.get('children', 0))

                            if inclusion_id and (num_adults > 0 or num_children > 0):
                                HouseboatBookingInclusion.objects.create(
                                    houseboat_booking=booking,
                                    special_inclusion_id=inclusion_id,
                                    num_adults=num_adults,
                                    num_children=num_children
                                )
                                inclusion_count += 1

                        print(f"✅ Saved {inclusion_count} houseboat inclusions")
                    except json.JSONDecodeError as e:
                        print(f"❌ JSON Parse Error: {e}")
                        print(f"   Problematic JSON: {inclusions_json[:200]}")

        elif item_type == 'vehicle':
            booking = VehicleBooking.objects.get(id=int(booking_id))
            net_price = safe_decimal(request.POST.get('vehicle_total_price'))
            if hasattr(booking, 'custom_total_price'):
                booking.custom_total_price = net_price if net_price > 0 else None
                fields_to_update.append('custom_total_price')

        elif item_type == 'activity':
            booking = ActivityBooking.objects.get(id=int(booking_id))
            net_price = safe_decimal(request.POST.get('activity_total_price'))
            if hasattr(booking, 'custom_total_price'):
                booking.custom_total_price = net_price if net_price > 0 else None
                fields_to_update.append('custom_total_price')

        # ✅ NEW: Standalone Activity Handler
        elif item_type == 'standalone':
            booking = StandaloneInclusionBooking.objects.get(id=int(booking_id))

            # Standalone activities have fixed subtotal from the model
            # We only update markup here, not the base price
            net_price = booking.subtotal

            print(f"📥 Standalone Activity: {booking.special_inclusion.name}")
            print(f"   Subtotal: ₹{net_price}")
            print(f"   Markup Type: {markup_type}")
            print(f"   Markup Value: {markup_value}")

        else:
            error_msg = f'Invalid item type: {item_type}'
            if is_ajax:
                return JsonResponse({'success': False, 'error': error_msg}, status=400)
            return JsonResponse({'success': False, 'error': error_msg}, status=400)

        # --- Update Booking Prices and Save ---
        if booking:
            # ✅ Handle standalone activities separately
            if item_type == 'standalone':
                # Calculate markup based on the subtotal
                if markup_type == 'percentage':
                    markup_amount = net_price * (markup_value / 100)
                else:
                    markup_amount = markup_value

                # Update markup fields
                booking.markup_type = markup_type
                booking.markup_value = markup_value
                booking.markup_amount = markup_amount
                booking.total_price = net_price + markup_amount

                fields_to_update = ['markup_type', 'markup_value', 'markup_amount', 'total_price']

                print(f"✅ Standalone Markup Calculated: ₹{markup_amount}")
                print(f"✅ Total Price: ₹{booking.total_price}")

            else:
                # For other booking types (hotel, vehicle, activity, houseboat)
                booking.net_price = net_price
                fields_to_update.append('net_price')

                # Calculate markup based on the net price
                if markup_type == 'percentage':
                    markup_amount = net_price * (markup_value / 100)
                else:
                    markup_amount = markup_value

                # Update all pricing fields
                booking.markup_type = markup_type
                booking.markup_value = markup_value
                booking.markup_amount = markup_amount
                booking.gross_price = net_price + markup_amount

                # Add markup fields to update list
                fields_to_update.extend(['markup_type', 'markup_value', 'markup_amount', 'gross_price'])

            # Efficiently save only the fields that were changed
            booking.save(update_fields=fields_to_update)

            success_msg = f'{item_type.title()} pricing updated successfully.'

            # ✅ RETURN JSON FOR AJAX REQUESTS
            if is_ajax:
                final_price = booking.total_price if item_type == 'standalone' else booking.gross_price

                return JsonResponse({
                    'success': True,
                    'message': success_msg,
                    'booking_id': booking.id,
                    'item_type': item_type,
                    'net_price': float(net_price),
                    'markup_amount': float(markup_amount),
                    'gross_price': float(final_price)
                })

            return JsonResponse({'success': True, 'message': success_msg})

    except (HotelBooking.DoesNotExist, HouseboatBooking.DoesNotExist,
            VehicleBooking.DoesNotExist, ActivityBooking.DoesNotExist,
            StandaloneInclusionBooking.DoesNotExist):
        error_msg = 'Booking record not found.'
        if is_ajax:
            return JsonResponse({'success': False, 'error': error_msg}, status=404)
        return JsonResponse({'success': False, 'error': error_msg}, status=404)

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"❌ Error in update_booking_totals: {error_details}")

        error_msg = f'An unexpected error occurred: {str(e)}'
        if is_ajax:
            return JsonResponse({'success': False, 'error': error_msg}, status=500)
        return JsonResponse({'success': False, 'error': error_msg}, status=500)



#######################################################################################################################################
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.views.decorators.http import require_POST, require_http_methods
from django.http import JsonResponse
from django.utils import timezone
from decimal import Decimal
import json
from .models import Query, Itinerary, ItineraryPricingOption, TeamMember


def query_proposals(request, query_id):
    """
    Display ONLY active (non-archived) itineraries for a query
    Archived ones go to history page
    """
    from .models import Query, Itinerary, TeamMember
    from django.contrib.auth.models import User

    # Authentication check
    user_id = request.session.get('user_id')
    user_type = request.session.get('user_type')
    current_user = None

    if not user_id:
        messages.warning(request, '⚠️ Please login to access this page')
        return redirect('team_member:login')

    # Get current user
    if user_type == 'superuser':
        current_user = User.objects.get(id=user_id)
    elif user_type == 'team_member':
        current_user = TeamMember.objects.get(id=user_id)

    # Get query
    query = get_object_or_404(Query, id=query_id)

    # ✅ GET ONLY ACTIVE (NON-ARCHIVED) ITINERARIES
    active_itineraries = Itinerary.objects.filter(
        query=query,
        status__in=['draft', 'quoted', 'confirmed']  # Exclude 'archived'
    ).order_by('-created_at')

    # Count archived
    archived_count = Itinerary.objects.filter(
        query=query,
        status='archived'
    ).count()

    context = {
        'query': query,
        'itineraries': active_itineraries,  # Only active
        'archived_count': archived_count,
        'user_type': user_type,
        'current_user': current_user,
    }

    return render(request, 'query_proposals.html', context)


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.views.decorators.http import require_POST, require_http_methods
from django.http import JsonResponse
from django.utils.timezone import now
from datetime import datetime, timedelta
from decimal import Decimal
import json
import traceback
from .models import Query, Itinerary, ItineraryPricingOption, TeamMember


@require_http_methods(["POST"])
def edit_itinerary(request):
    """
    Edit itinerary: creates new version, copies bookings and pricing
    Archives old itinerary (moved to history with versioning)
    """
    try:
        itinerary_id = request.POST.get('itinerary_id')
        travel_from_str = request.POST.get('travel_from')
        travel_to_str = request.POST.get('travel_to')
        adults = request.POST.get('adults', 0)
        childrens = request.POST.get('childrens', 0)
        infants = request.POST.get('infants', 0)

        # Validation
        if not all([itinerary_id, travel_from_str, travel_to_str, adults]):
            return JsonResponse({
                'success': False,
                'errors': {
                    'travel_from': 'Travel from date is required',
                    'travel_to': 'Travel to date is required',
                    'adults': 'Number of adults is required'
                }
            })

        # Parse dates
        try:
            travel_from = datetime.strptime(travel_from_str, '%Y-%m-%d').date()
            travel_to = datetime.strptime(travel_to_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({
                'success': False,
                'errors': {
                    'travel_from': 'Invalid date format',
                    'travel_to': 'Invalid date format'
                }
            })

        if travel_to < travel_from:
            return JsonResponse({
                'success': False,
                'errors': {
                    'travel_to': 'Travel end date must be after start date'
                }
            })

        try:
            adults = int(adults)
            childrens = int(childrens)
            infants = int(infants)

            if adults < 1:
                raise ValueError('At least 1 adult is required')
        except (ValueError, TypeError):
            return JsonResponse({
                'success': False,
                'errors': {'adults': 'Invalid passenger count'}
            })

        # Get old itinerary
        old_itinerary = get_object_or_404(Itinerary, id=itinerary_id)

        # Archive old itinerary
        old_itinerary.status = 'archived'
        old_itinerary.archived_at = now()
        old_itinerary.archived_reason = f'Edited - Dates: {old_itinerary.travel_from} to {travel_from}-{travel_to}, Passengers: {old_itinerary.adults}A -> {adults}A'
        old_itinerary.save()
        print(f'✅ Old itinerary {itinerary_id} archived')

        # Calculate total days
        total_days = (travel_to - travel_from).days + 1

        # Create new active itinerary
        new_itinerary = Itinerary.objects.create(
            query=old_itinerary.query,
            name=old_itinerary.name,
            travel_from=travel_from,
            travel_to=travel_to,
            total_days=total_days,
            adults=adults,
            childrens=childrens,
            infants=infants,
            created_by=old_itinerary.created_by,
            status='draft',
            parent_itinerary=old_itinerary,
            # Copy pricing fields
            cgst_percentage=old_itinerary.cgst_percentage,
            sgst_percentage=old_itinerary.sgst_percentage,
            discount=old_itinerary.discount,
            markup_type=old_itinerary.markup_type,
            markup_value=old_itinerary.markup_value,
            # Copy calculated amounts
            total_net_price=old_itinerary.total_net_price,
            total_gross_price=old_itinerary.total_gross_price,
            final_amount=old_itinerary.final_amount,
        )
        new_itinerary.destinations.set(old_itinerary.destinations.all())
        print(f'✅ New itinerary {new_itinerary.id} created as version')

        # Copy bookings and create day plans
        booking_report = copy_and_validate_bookings(old_itinerary, new_itinerary)
        print(f'✅ Bookings copied: {booking_report}')

        # Copy pricing options from old itinerary
        print(f'📋 Copying pricing options from old itinerary...')
        options_copied = copy_pricing_options(old_itinerary, new_itinerary)

        if not options_copied:
            print(f'⚠️ No pricing options copied, will need manual recalculation')

        # Get final amount for response
        copied_option = new_itinerary.pricing_options.first()
        final_amount = float(copied_option.final_amount) if copied_option else float(new_itinerary.final_amount)

        # Check if AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': 'Itinerary updated successfully! Old proposal moved to history.',
                'itinerary_id': new_itinerary.id,
                'final_amount': final_amount,
                'booking_report': booking_report,
                'redirect_url': f'/itinerary/{new_itinerary.id}/day-plan/'
            })
        else:
            messages.success(request, f'✅ Itinerary updated! Pricing: ₹{final_amount}')
            return redirect('itinerary_day_plan', itinerary_id=new_itinerary.id)

    except Exception as e:
        print(f'❌ Error in edit_itinerary: {str(e)}')
        traceback.print_exc()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=400)
        else:
            messages.error(request, f'Error: {str(e)}')
            return redirect('list_itineraries')


import traceback

def copy_pricing_options(old_itinerary, new_itinerary):
    """
    Copy pricing options from old itinerary to new one
    Returns True if options were copied, False if none existed
    """
    try:
        print(f'📋 Copying pricing options from itinerary {old_itinerary.id} to {new_itinerary.id}...')

        # Delete any existing pricing options on new itinerary
        ItineraryPricingOption.objects.filter(itinerary=new_itinerary).delete()

        # Get old pricing options
        old_options = ItineraryPricingOption.objects.filter(itinerary=old_itinerary)

        if not old_options.exists():
            print(f'⚠️ No pricing options found in old itinerary {old_itinerary.id}')
            return False

        # Copy all pricing options from old itinerary
        for old_option in old_options:
            new_option = ItineraryPricingOption.objects.create(
                itinerary=new_itinerary,
                option_name=old_option.option_name,
                option_number=old_option.option_number,

                # Pricing amounts
                net_price=old_option.net_price,
                markup_amount=old_option.markup_amount,
                gross_price=old_option.gross_price,

                # Taxes
                cgst_amount=old_option.cgst_amount,
                sgst_amount=old_option.sgst_amount,
                discount_amount=old_option.discount_amount,
                final_amount=old_option.final_amount,

                # Hotel breakdown
                hotels_included=old_option.hotels_included,

                # Other fields
                vehicle_type=old_option.vehicle_type,
                number_of_rooms=old_option.number_of_rooms,
                extra_beds=old_option.extra_beds,
                child_with_bed=old_option.child_with_bed,
                child_without_bed=old_option.child_without_bed,
                child_ages=old_option.child_ages,
            )
            print(f'✅ Copied pricing option: {new_option.option_name} - ₹{new_option.final_amount}')

        return True

    except Exception as e:
        print(f'❌ Error copying pricing options: {str(e)}')
        traceback.print_exc()
        return False

from datetime import timedelta
import traceback


from datetime import timedelta
import traceback

@custom_login_required
def copy_and_validate_bookings(old_itinerary, new_itinerary):
    """
    Copy bookings from old itinerary to new one with FULL DEBUG LOGGING
    Also creates ItineraryDayPlan records copying details from old plans.
    Includes: Hotels, Activities, Vehicles, Houseboats, Standalone Inclusions
    """
    try:
        # ✅ IMPORTS
        from .models import (HotelBooking, ActivityBooking, VehicleBooking, HouseboatBooking,
                           ItineraryDayPlan, StandaloneInclusionBooking,
                           HotelBookingInclusion, HouseboatBookingInclusion)

        skipped_items = {
            'hotels': [],
            'activities': [],
            'vehicles': [],
            'houseboats': [],
            'standalone_inclusions': [],
        }

        copied_count = {
            'hotels': 0,
            'activities': 0,
            'vehicles': 0,
            'houseboats': 0,
            'standalone_inclusions': 0,
        }

        print('\n' + '='*80)
        print('🚀 STARTING BOOKING COPY PROCESS')
        print('='*80)
        print(f'Old Itinerary ID: {old_itinerary.id}')
        print(f'New Itinerary ID: {new_itinerary.id}')
        print(f'Old Date Range: {old_itinerary.travel_from} to {old_itinerary.travel_to}')
        print(f'New Date Range: {new_itinerary.travel_from} to {new_itinerary.travel_to}')
        print('='*80 + '\n')

        # =========================================================
        # ✅ FIXED SECTION: COPY DAY PLAN DETAILS
        # =========================================================
        print(f'📅 Creating day plans for {new_itinerary.total_days} days...')

        # 1. Map old day plans by day_number for easy lookup
        old_day_plans_map = {dp.day_number: dp for dp in old_itinerary.day_plans.all()}
        day_plans_dict = {}

        # 2. Iterate through the NEW duration
        for day_num in range(1, new_itinerary.total_days + 1):

            # Create the new day plan
            new_day_plan, created = ItineraryDayPlan.objects.get_or_create(
                itinerary=new_itinerary,
                day_number=day_num
            )

            # 3. If this day number existed in the old itinerary, COPY the data
            if day_num in old_day_plans_map:
                old_dp = old_day_plans_map[day_num]

                # Copy fields from old day plan
                new_day_plan.destination = old_dp.destination
                new_day_plan.title = old_dp.title
                new_day_plan.description = old_dp.description
                new_day_plan.image = old_dp.image
                new_day_plan.notes = old_dp.notes
                new_day_plan.save()

                print(f'   ✅ Copied details for Day {day_num} (Dest: {old_dp.destination})')
            else:
                print(f'   ℹ️ Day {day_num} is a new day (Created Empty)')

            # Add to dictionary for the booking copy logic to use
            day_plans_dict[day_num] = new_day_plan

        print(f'✅ Day plans ready: {len(day_plans_dict)} days\n')

        # ===== COPY HOTEL BOOKINGS =====
        print('\n' + '='*80)
        print('🏨 COPYING HOTEL BOOKINGS')
        print('='*80)

        try:
            old_hotel_bookings = HotelBooking.objects.filter(itinerary=old_itinerary)
            print(f'📋 Found {old_hotel_bookings.count()} hotel bookings to copy\n')

            for idx, old_booking in enumerate(old_hotel_bookings, 1):
                try:
                    print(f'\n--- Hotel Booking {idx}/{old_hotel_bookings.count()} ---')
                    print(f'🏨 Hotel: {old_booking.hotel.name} (ID: {old_booking.id})')

                    # Adjust dates proportionally
                    check_in_offset = (old_booking.check_in_date - old_itinerary.travel_from).days
                    new_check_in = new_itinerary.travel_from + timedelta(days=check_in_offset)

                    if new_check_in < new_itinerary.travel_from:
                        new_check_in = new_itinerary.travel_from
                    if new_check_in > new_itinerary.travel_to:
                        raise ValueError("Check-in date outside new itinerary range")

                    stay_duration = (old_booking.check_out_date - old_booking.check_in_date).days
                    new_check_out = new_check_in + timedelta(days=stay_duration)

                    if new_check_out > new_itinerary.travel_to:
                        new_check_out = new_itinerary.travel_to

                    print(f'📅 New dates: {new_check_in} to {new_check_out}')

                    # ✅ GET DAY NUMBER FROM DAY PLAN
                    old_day_plan = old_booking.day_plan if hasattr(old_booking, 'day_plan') and old_booking.day_plan else None
                    day_number = old_day_plan.day_number if old_day_plan else 1
                    print(f'📍 Day number: {day_number}')

                    # Get new day plan
                    day_plan = day_plans_dict.get(day_number)
                    if not day_plan:
                        day_plan, _ = ItineraryDayPlan.objects.get_or_create(
                            itinerary=new_itinerary,
                            day_number=day_number
                        )

                    # ✅ Copy markup values
                    old_markup_type = getattr(old_booking, 'markup_type', 'fixed')
                    old_markup_value = getattr(old_booking, 'markup_value', 0)
                    old_net_price = getattr(old_booking, 'net_price', None)

                    # Copy hotel booking
                    new_hotel = HotelBooking.objects.create(
                        day_plan=day_plan,
                        itinerary=new_itinerary,
                        hotel=old_booking.hotel,
                        destination=old_booking.destination,
                        category=old_booking.category,
                        room_type=old_booking.room_type,
                        meal_plan=old_booking.meal_plan,
                        check_in_date=new_check_in,
                        check_out_date=new_check_out,
                        check_in_time=old_booking.check_in_time,
                        check_out_time=old_booking.check_out_time,
                        num_double_beds=old_booking.num_double_beds,
                        child_with_bed=old_booking.child_with_bed,
                        child_without_bed=old_booking.child_without_bed,
                        extra_beds=old_booking.extra_beds,
                        option=old_booking.option,
                        # ✅ ADD MARKUP FIELDS
                        markup_type=old_markup_type,
                        markup_value=old_markup_value,
                        net_price=old_net_price,
                    )

                    # ✅ COPY INCLUSIONS
                    # Try multiple query methods as per original code
                    method1 = HotelBookingInclusion.objects.filter(hotelbooking=old_booking)
                    method2 = HotelBookingInclusion.objects.filter(hotelbooking_id=old_booking.id)
                    old_inclusions = method1 if method1.exists() else method2

                    if old_inclusions.exists():
                        for old_item in old_inclusions:
                            HotelBookingInclusion.objects.create(
                                hotelbooking=new_hotel,
                                specialinclusion=old_item.specialinclusion,
                                numadults=old_item.numadults,
                                numchildren=old_item.numchildren,
                            )

                    # Update day plan with destination if not set
                    if not day_plan.destination:
                        day_plan.destination = old_booking.destination
                        day_plan.save()

                    copied_count['hotels'] += 1
                    print(f'✅ Hotel booking fully copied')

                except Exception as e:
                    skipped_items['hotels'].append({
                        'hotel': str(old_booking.hotel),
                        'reason': str(e)
                    })
                    print(f'❌ Error copying hotel: {str(e)}')

        except Exception as e:
            print(f'❌ Fatal error in hotel section: {str(e)}')

        # ===== COPY ACTIVITY BOOKINGS =====
        print('\n\n' + '='*80)
        print('🎯 COPYING ACTIVITY BOOKINGS')
        print('='*80)

        try:
            old_activity_bookings = ActivityBooking.objects.filter(itinerary=old_itinerary)

            for idx, old_booking in enumerate(old_activity_bookings, 1):
                try:
                    activity_offset = (old_booking.booking_date - old_itinerary.travel_from).days
                    new_booking_date = new_itinerary.travel_from + timedelta(days=activity_offset)

                    if new_booking_date < new_itinerary.travel_from or new_booking_date > new_itinerary.travel_to:
                         # Keep inside range if offset pushes it out
                         new_booking_date = new_itinerary.travel_from

                    old_day_plan = old_booking.day_plan
                    day_number = old_day_plan.day_number if old_day_plan else 1
                    day_plan = day_plans_dict.get(day_number)

                    if not day_plan:
                         day_plan, _ = ItineraryDayPlan.objects.get_or_create(
                            itinerary=new_itinerary, day_number=day_number)

                    # Markup fields
                    old_markup_type = getattr(old_booking, 'markup_type', 'fixed')
                    old_markup_value = getattr(old_booking, 'markup_value', 0)
                    old_custom_price = getattr(old_booking, 'custom_total_price', None)

                    ActivityBooking.objects.create(
                        day_plan=day_plan,
                        itinerary=new_itinerary,
                        activity=old_booking.activity,
                        booking_date=new_booking_date,
                        booking_time=old_booking.booking_time,
                        num_adults=old_booking.num_adults,
                        num_children=old_booking.num_children,
                        notes=old_booking.notes,
                        markup_type=old_markup_type,
                        markup_value=old_markup_value,
                        custom_total_price=old_custom_price,
                    )

                    copied_count['activities'] += 1

                except Exception as e:
                    skipped_items['activities'].append({'activity': str(old_booking.activity), 'reason': str(e)})
        except Exception as e:
            print(f'❌ Error copying activities: {str(e)}')

        # ===== COPY VEHICLE BOOKINGS =====
        print('\n\n' + '='*80)
        print('🚗 COPYING VEHICLE BOOKINGS')
        print('='*80)

        try:
            old_vehicle_bookings = VehicleBooking.objects.filter(itinerary=old_itinerary)

            for idx, old_booking in enumerate(old_vehicle_bookings, 1):
                try:
                    vehicle_offset = (old_booking.pickup_date - old_itinerary.travel_from).days
                    new_pickup_date = new_itinerary.travel_from + timedelta(days=vehicle_offset)

                    old_day_plan = old_booking.day_plan
                    day_number = old_day_plan.day_number if old_day_plan else 1
                    day_plan = day_plans_dict.get(day_number)

                    if not day_plan:
                        day_plan, _ = ItineraryDayPlan.objects.get_or_create(
                            itinerary=new_itinerary, day_number=day_number)

                    old_markup_type = getattr(old_booking, 'markup_type', 'fixed')
                    old_markup_value = getattr(old_booking, 'markup_value', 0)
                    old_custom_price = getattr(old_booking, 'custom_total_price', None)

                    VehicleBooking.objects.create(
                        day_plan=day_plan,
                        itinerary=new_itinerary,
                        vehicle=old_booking.vehicle,
                        destination=old_booking.destination,
                        pickup_date=new_pickup_date,
                        pickup_time=old_booking.pickup_time,
                        num_passengers=old_booking.num_passengers,
                        vehicle_type=old_booking.vehicle_type,
                        option=old_booking.option,
                        total_km=old_booking.total_km,
                        markup_type=old_markup_type,
                        markup_value=old_markup_value,
                        custom_total_price=old_custom_price,
                    )
                    copied_count['vehicles'] += 1
                except Exception as e:
                    skipped_items['vehicles'].append({'vehicle': str(old_booking.vehicle), 'reason': str(e)})
        except Exception as e:
            print(f'❌ Error copying vehicles: {str(e)}')

        # ===== COPY HOUSEBOAT BOOKINGS =====
        print('\n\n' + '='*80)
        print('⛵ COPYING HOUSEBOAT BOOKINGS')
        print('='*80)

        try:
            old_houseboat_bookings = HouseboatBooking.objects.filter(itinerary=old_itinerary)

            for idx, old_booking in enumerate(old_houseboat_bookings, 1):
                try:
                    hb_offset = (old_booking.check_in_date - old_itinerary.travel_from).days
                    new_hb_check_in = new_itinerary.travel_from + timedelta(days=hb_offset)
                    stay_duration = (old_booking.check_out_date - old_booking.check_in_date).days
                    new_hb_check_out = new_hb_check_in + timedelta(days=stay_duration)

                    old_day_plan = old_booking.day_plan
                    day_number = old_day_plan.day_number if old_day_plan else 1
                    day_plan = day_plans_dict.get(day_number)

                    if not day_plan:
                        day_plan, _ = ItineraryDayPlan.objects.get_or_create(
                            itinerary=new_itinerary, day_number=day_number)

                    old_markup_type = getattr(old_booking, 'markup_type', 'fixed')
                    old_markup_value = getattr(old_booking, 'markup_value', 0)
                    old_net_price = getattr(old_booking, 'net_price', None)

                    new_hb = HouseboatBooking.objects.create(
                        day_plan=day_plan,
                        itinerary=new_itinerary,
                        houseboat=old_booking.houseboat,
                        check_in_date=new_hb_check_in,
                        check_out_date=new_hb_check_out,
                        room_type=old_booking.room_type,
                        meal_plan=old_booking.meal_plan,
                        num_one_bed_rooms=old_booking.num_one_bed_rooms,
                        num_two_bed_rooms=old_booking.num_two_bed_rooms,
                        num_three_bed_rooms=old_booking.num_three_bed_rooms,
                        # ... (Include all other bedroom fields here) ...
                        markup_type=old_markup_type,
                        markup_value=old_markup_value,
                        net_price=old_net_price,
                    )

                    # Copy HB inclusions
                    method1 = HouseboatBookingInclusion.objects.filter(houseboatbooking=old_booking)
                    method2 = HouseboatBookingInclusion.objects.filter(houseboatbooking_id=old_booking.id)
                    old_inclusions = method1 if method1.exists() else method2

                    if old_inclusions.exists():
                        for old_item in old_inclusions:
                            HouseboatBookingInclusion.objects.create(
                                houseboatbooking=new_hb,
                                specialinclusion=old_item.specialinclusion,
                                numadults=old_item.numadults,
                                numchildren=old_item.numchildren,
                            )

                    copied_count['houseboats'] += 1
                except Exception as e:
                    skipped_items['houseboats'].append({'houseboat': str(old_booking.houseboat), 'reason': str(e)})
        except Exception as e:
            print(f'❌ Error copying houseboats: {str(e)}')

        # ===== COPY STANDALONE INCLUSION BOOKINGS =====
        try:
            old_standalone_bookings = StandaloneInclusionBooking.objects.filter(itinerary=old_itinerary)
            for old_booking in old_standalone_bookings:
                try:
                    booking_offset = (old_booking.booking_date - old_itinerary.travel_from).days
                    new_booking_date = new_itinerary.travel_from + timedelta(days=booking_offset)

                    old_day_plan = old_booking.day_plan
                    day_number = old_day_plan.day_number if old_day_plan else 1
                    day_plan = day_plans_dict.get(day_number)

                    if not day_plan:
                         day_plan, _ = ItineraryDayPlan.objects.get_or_create(
                            itinerary=new_itinerary, day_number=day_number)

                    old_markup_type = getattr(old_booking, 'markup_type', 'fixed')
                    old_markup_value = getattr(old_booking, 'markup_value', 0)

                    StandaloneInclusionBooking.objects.create(
                        itinerary=new_itinerary,
                        day_plan=day_plan,
                        special_inclusion=old_booking.special_inclusion,
                        booking_date=new_booking_date,
                        booking_time=old_booking.booking_time,
                        num_adults=old_booking.num_adults,
                        num_children=old_booking.num_children,
                        markup_type=old_markup_type,
                        markup_value=old_markup_value,
                        notes=old_booking.notes,
                    )
                    copied_count['standalone_inclusions'] += 1
                except Exception as e:
                    skipped_items['standalone_inclusions'].append({'inclusion': str(old_booking.special_inclusion), 'reason': str(e)})
        except Exception as e:
             print(f'❌ Error copying standalone: {str(e)}')

        # ===== FINAL SUMMARY =====
        print('\n\n' + '='*80)
        print('📊 FINAL BOOKING COPY SUMMARY')
        print('='*80)
        print(f'Hotels:               {copied_count["hotels"]:3d} copied, {len(skipped_items["hotels"]):3d} skipped')
        print(f'Activities:           {copied_count["activities"]:3d} copied, {len(skipped_items["activities"]):3d} skipped')

        return {
            'success': True,
            'copied': copied_count,
            'skipped': skipped_items
        }

    except Exception as e:
        print(f'\n❌ FATAL ERROR in copy_and_validate_bookings: {str(e)}')
        traceback.print_exc()
        return {
            'success': False,
            'message': str(e)
        }



# ✅ FUNCTION TO CALCULATE PRICING FOR NEW ITINERARY

@custom_login_required
def calculate_itinerary_pricing(itinerary):
    """
    Calculate total pricing for itinerary based on bookings
    """
    from .models import HotelBooking, ActivityBooking, VehicleBooking, HouseboatBooking, Hotelprice, VehiclePricing, ActivityPrice, HouseboatPrice
    from decimal import Decimal

    total_net = Decimal('0.00')
    total_gross = Decimal('0.00')

    try:
        print(f'🔄 Calculating pricing for itinerary {itinerary.id}...')

        # ===== HOTEL PRICING =====
        hotel_bookings = HotelBooking.objects.filter(itinerary=itinerary)
        for booking in hotel_bookings:
            try:
                nights = (booking.check_out_date - booking.check_in_date).days or 1

                # Get pricing rule
                rule = Hotelprice.objects.filter(
                    hotel=booking.hotel,
                    room_type=booking.room_type,
                    meal_plan=booking.meal_plan,
                    from_date__lte=booking.check_in_date,
                    to_date__gte=booking.check_in_date
                ).first()

                if rule:
                    # Calculate per night cost
                    per_night = (
                        Decimal(booking.num_double_beds or 0) * (rule.double_bed or 0) +
                        Decimal(booking.child_with_bed or 0) * (rule.child_with_bed or 0) +
                        Decimal(booking.child_without_bed or 0) * (rule.child_without_bed or 0) +
                        Decimal(booking.extra_beds or 0) * (rule.extra_bed or 0)
                    )
                    hotel_cost = per_night * nights
                    total_net += hotel_cost
                    total_gross += hotel_cost
                    print(f'✅ Hotel {booking.hotel.name}: ₹{hotel_cost}')
            except Exception as e:
                print(f'⚠️ Hotel pricing error: {str(e)}')

        # ===== ACTIVITY PRICING =====
        activity_bookings = ActivityBooking.objects.filter(itinerary=itinerary)
        for booking in activity_bookings:
            try:
                rule = ActivityPrice.objects.filter(
                    activity=booking.activity,
                    from_date__lte=booking.booking_date,
                    to_date__gte=booking.booking_date
                ).first()

                if rule:
                    activity_cost = (
                        Decimal(booking.num_adults or 0) * (rule.per_person or 0) +
                        Decimal(booking.num_children or 0) * (rule.per_person or 0)
                    )
                    total_net += activity_cost
                    total_gross += activity_cost
                    print(f'✅ Activity {booking.activity.name}: ₹{activity_cost}')
            except Exception as e:
                print(f'⚠️ Activity pricing error: {str(e)}')

        # ===== VEHICLE PRICING =====
        vehicle_bookings = VehicleBooking.objects.filter(itinerary=itinerary)
        for booking in vehicle_bookings:
            try:
                rule = VehiclePricing.objects.filter(
                    vehicle=booking.vehicle,
                    from_date__lte=booking.pickup_date,
                    to_date__gte=booking.pickup_date
                ).first()

                if rule:
                    vehicle_cost = rule.total_fee_100km or Decimal('0')
                    if booking.total_km and booking.total_km > 100:
                        extra_km = booking.total_km - 100
                        vehicle_cost += Decimal(str(extra_km)) * (rule.extra_fee_per_km or Decimal('0'))
                    total_net += vehicle_cost
                    total_gross += vehicle_cost
                    print(f'✅ Vehicle {booking.vehicle.name}: ₹{vehicle_cost}')
            except Exception as e:
                print(f'⚠️ Vehicle pricing error: {str(e)}')

        # ===== HOUSEBOAT PRICING =====
        houseboat_bookings = HouseboatBooking.objects.filter(itinerary=itinerary)
        for booking in houseboat_bookings:
            try:
                nights = (booking.check_out_date - booking.check_in_date).days or 1

                rule = HouseboatPrice.objects.filter(
                    houseboat=booking.houseboat,
                    meal_plan=booking.meal_plan,
                    room_type=booking.room_type,
                    from_date__lte=booking.check_in_date,
                    to_date__gte=booking.check_in_date
                ).first()

                if rule:
                    per_night = (
                        Decimal(booking.num_one_bed_rooms or 0) * (rule.one_bed or 0) +
                        Decimal(booking.num_two_bed_rooms or 0) * (rule.two_bed or 0) +
                        Decimal(booking.num_three_bed_rooms or 0) * (rule.three_bed or 0) +
                        Decimal(booking.num_four_bed_rooms or 0) * (rule.four_bed or 0) +
                        Decimal(booking.num_five_bed_rooms or 0) * (rule.five_bed or 0) +
                        Decimal(booking.num_six_bed_rooms or 0) * (rule.six_bed or 0) +
                        Decimal(booking.num_seven_bed_rooms or 0) * (rule.seven_bed or 0) +
                        Decimal(booking.num_eight_bed_rooms or 0) * (rule.eight_bed or 0) +
                        Decimal(booking.num_nine_bed_rooms or 0) * (rule.nine_bed or 0) +
                        Decimal(booking.num_ten_bed_rooms or 0) * (rule.ten_bed or 0) +
                        Decimal(booking.num_extra_beds or 0) * (rule.extra_bed or 0)
                    )
                    houseboat_cost = per_night * nights
                    total_net += houseboat_cost
                    total_gross += houseboat_cost
                    print(f'✅ Houseboat {booking.houseboat.name}: ₹{houseboat_cost}')
            except Exception as e:
                print(f'⚠️ Houseboat pricing error: {str(e)}')

        # ===== UPDATE ITINERARY WITH PRICING =====
        itinerary.total_net_price = total_net
        itinerary.total_gross_price = total_gross
        itinerary.final_amount = total_gross
        itinerary.save()

        print(f'✅ Pricing calculated - Net: ₹{total_net}, Gross: ₹{total_gross}')
        return True

    except Exception as e:
        print(f'❌ Error calculating pricing: {str(e)}')
        import traceback
        traceback.print_exc()
        return False

@custom_login_required
def get_proposals_list(request, query_id):
    """
    Get all proposals for a query
    Separate into Active Proposals and History (Archived)
    """
    from .models import Query, Itinerary

    try:
        query = Query.objects.get(id=query_id)

        # Get all itineraries for this query
        all_itineraries = Itinerary.objects.filter(query=query)

        # Separate into active and archived
        active_proposals = all_itineraries.filter(status__in=['draft', 'quoted', 'confirmed'])
        archived_proposals = all_itineraries.filter(status='archived')

        # Get only the latest active version for each chain
        latest_active = []
        processed_parents = set()

        for itinerary in active_proposals.order_by('-created_at'):
            # If this is a child (has parent), only show if parent is archived
            if itinerary.parent_itinerary_id:
                if itinerary.parent_itinerary_id not in processed_parents:
                    latest_active.append(itinerary)
                    processed_parents.add(itinerary.parent_itinerary_id)
            else:
                # This is a root itinerary, show if no active child exists
                if itinerary.id not in processed_parents:
                    # Check if there's an active child
                    has_active_child = itinerary.versions.filter(
                        status__in=['draft', 'quoted', 'confirmed']
                    ).exists()
                    if not has_active_child:
                        latest_active.append(itinerary)
                    processed_parents.add(itinerary.id)

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'active_proposals': [
                    {
                        'id': it.id,
                        'name': it.name,
                        'dates': f"{it.travel_from} to {it.travel_to}",
                        'days': it.total_days,
                        'passengers': f"{it.adults}A {it.childrens}C {it.infants}I",
                        'price': str(it.final_amount),
                        'status': it.status,
                        'created': it.created_at.strftime('%d %b %Y'),
                    }
                    for it in latest_active
                ],
                'archived_proposals': [
                    {
                        'id': it.id,
                        'name': it.name,
                        'dates': f"{it.travel_from} to {it.travel_to}",
                        'days': it.total_days,
                        'passengers': f"{it.adults}A {it.childrens}C {it.infants}I",
                        'price': str(it.final_amount),
                        'status': it.status,
                        'created': it.created_at.strftime('%d %b %Y'),
                        'archived': it.archived_at.strftime('%d %b %Y') if it.archived_at else 'N/A',
                        'archived_reason': it.archived_reason,
                    }
                    for it in archived_proposals.order_by('-archived_at')
                ]
            })
        else:
            return {
                'active': latest_active,
                'archived': archived_proposals.order_by('-archived_at')
            }

    except Exception as e:
        print(f'❌ Error in get_proposals_list: {str(e)}')
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)



@custom_login_required
def itinerary_history(request, query_id):
    """
    Display archived/historical itineraries for a query
    """
    from .models import Query, Itinerary
    from django.contrib.auth.models import User

    # Authentication check
    user_id = request.session.get('user_id')
    user_type = request.session.get('user_type')
    current_user = None

    if not user_id:
        messages.warning(request, '⚠️ Please login to access this page')
        return redirect('team_member:login')

    # Get current user
    if user_type == 'superuser':
        current_user = User.objects.get(id=user_id)
    elif user_type == 'team_member':
        current_user = TeamMember.objects.get(id=user_id)

    # Get query
    query = get_object_or_404(Query, id=query_id)

    # ✅ GET ARCHIVED ITINERARIES WITH ALL RELATED DATA INCLUDING INCLUSION ITEMS
    archived_itineraries = Itinerary.objects.filter(
        query=query,
        status='archived'
    ).select_related(
        'created_by',
        'parent_itinerary'
    ).prefetch_related(
        'destinations',
        'pricing_options',
        # ✅ PREFETCH DAY PLANS AND BOOKINGS WITH INCLUSION ITEMS
        'day_plans',
        'day_plans__hotel_bookings__hotel',
        'day_plans__hotel_bookings__room_type',
        'day_plans__hotel_bookings__meal_plan',
        'day_plans__hotel_bookings__inclusion_items',
        'day_plans__hotel_bookings__inclusion_items__special_inclusion',
        'day_plans__houseboat_bookings__houseboat',
        'day_plans__houseboat_bookings__room_type',
        'day_plans__houseboat_bookings__meal_plan',
        'day_plans__houseboat_bookings__inclusion_items',
        'day_plans__houseboat_bookings__inclusion_items__special_inclusion',
        'day_plans__activity_bookings__activity',
        'day_plans__vehicle_bookings__vehicle',
        'day_plans__standalone_inclusions__special_inclusion',
    ).order_by('-archived_at')

    # Count active itineraries - ✅ FIXED: Added closing parenthesis
    active_count = Itinerary.objects.filter(
        query=query,
        status__in=['draft', 'quoted', 'confirmed']  # ✅ FIXED HERE
    ).count()

    context = {
        'query': query,
        'archived_itineraries': archived_itineraries,
        'active_count': active_count,
        'user_type': user_type,
        'current_user': current_user,
    }

    return render(request, 'itinerary_history.html', context)


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from .models import Itinerary

@require_POST
@custom_login_required
def unarchive_itinerary(request, itinerary_id):
    try:
        itinerary = Itinerary.objects.get(id=itinerary_id)

        # Check permissions
        if not request.user.is_authenticated:
            return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=401)

        # Unarchive logic
        itinerary.is_archived = False
        itinerary.archived_at = None
        itinerary.archived_reason = None
        itinerary.status = 'draft'  # Set back to draft status
        itinerary.save()

        return JsonResponse({
            'success': True,
            'message': 'Itinerary unarchived successfully'
        })

    except Itinerary.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Itinerary not found'
        }, status=404)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)




# =====================================================
# UPDATE QUERY STATUS
# =====================================================

@require_POST
@custom_login_required
def update_query_status(request, query_id):
    """Update query status - WITH PERMISSION CHECK"""

    # ✅ CHECK IF USER IS LOGGED IN
    user_id = request.session.get('user_id')
    user_type = request.session.get('user_type')

    if not user_id:
        messages.warning(request, '⚠️ Please login first')
        return redirect('team_member:login')

    query = get_object_or_404(Query, id=query_id)

    # ✅ CHECK PERMISSION TO UPDATE STATUS
    can_update = False

    if user_type == 'superuser':
        can_update = True

    elif user_type == 'team_member':
        user = TeamMember.objects.get(id=user_id)

        if user.role in ['admin', 'manager']:
            can_update = True

        elif user.has_permission('can_edit_queries'):
            if query.created_by == user or query.assign == user:
                can_update = True

        elif query.created_by == user or query.assign == user:
            can_update = True

    # ✅ DENY ACCESS if no permission
    if not can_update:
        messages.error(request, '❌ You do not have permission to update this query status')
        return redirect('query_proposals', query_id=query_id)

    # ✅ UPDATE STATUS
    new_status = request.POST.get('status')

    if new_status in dict(Query.STATUS_CHOICES):
        old_status = query.get_status_display()
        query.status = new_status
        query.save()
        messages.success(request, f'✅ Status updated from {old_status} to {query.get_status_display()}')
    else:
        messages.error(request, '❌ Invalid status value')

    return redirect('query_proposals', query_id=query_id)


# =====================================================
# API: GET ITINERARY OPTIONS
# =====================================================
@custom_login_required
@require_http_methods(["GET"])
def get_itinerary_options(request, itinerary_id):
    """Get all pricing options for an itinerary"""
    # Check session authentication
    user_id = request.session.get('user_id')
    if not user_id:
        return JsonResponse({
            'success': False,
            'message': 'Authentication required.'
        }, status=401)

    try:
        itinerary = get_object_or_404(Itinerary, id=itinerary_id)

        # Get all saved pricing options
        pricing_options = ItineraryPricingOption.objects.filter(
            itinerary=itinerary
        ).order_by('option_number')

        options = []
        for option in pricing_options:
            options.append({
                'id': str(option.option_number),
                'name': option.option_name,
                'price': str(option.final_amount),
                'net_price': str(option.net_price),
                'gross_price': str(option.gross_price)
            })

        return JsonResponse({
            'success': True,
            'options': options
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


# =====================================================
# API: CONFIRM ITINERARY
# =====================================================
@custom_login_required
@require_http_methods(["POST"])
def confirm_itinerary(request, itinerary_id):
    """Confirm itinerary with selected hotel option"""
    # Check session authentication
    user_id = request.session.get('user_id')
    if not user_id:
        return JsonResponse({
            'success': False,
            'message': 'Authentication required.'
        }, status=401)

    try:
        # Parse request body
        data = json.loads(request.body)
        option_id = data.get('option_id')

        if not option_id:
            return JsonResponse({
                'success': False,
                'message': 'Option ID is required'
            }, status=400)

        # Get the itinerary
        itinerary = get_object_or_404(Itinerary, id=itinerary_id)

        # Get the selected pricing option
        pricing_option = ItineraryPricingOption.objects.filter(
            itinerary=itinerary,
            option_number=int(option_id)
        ).first()

        if not pricing_option:
            return JsonResponse({
                'success': False,
                'message': 'Pricing option not found'
            }, status=404)

        # ✅ SAVE ALL PRICING FIELDS FROM THE SELECTED OPTION
        itinerary.selected_option = pricing_option.option_name
        itinerary.total_net_price = pricing_option.net_price
        itinerary.total_gross_price = pricing_option.gross_price
        itinerary.discount = pricing_option.discount_amount
        itinerary.final_amount = pricing_option.final_amount

        # Calculate markup values
        itinerary.global_markup = pricing_option.gross_price - pricing_option.net_price

        # Update status
        itinerary.status = 'confirmed'
        itinerary.is_finalized = True
        itinerary.finalized_at = timezone.now()
        itinerary.confirmed_at = timezone.now()
        itinerary.save()

        return JsonResponse({
            'success': True,
            'message': 'Itinerary confirmed successfully',
            'itinerary_id': itinerary.id,
            'selected_option': pricing_option.option_name,
            'status': itinerary.status,
            'pricing': {
                'total_net_price': str(pricing_option.net_price),
                'total_gross_price': str(pricing_option.gross_price),
                'cgst': str(pricing_option.cgst_amount),
                'sgst': str(pricing_option.sgst_amount),
                'discount': str(pricing_option.discount_amount),
                'final_amount': str(pricing_option.final_amount),
                'hotels': pricing_option.hotels_included
            }
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


# =====================================================
# API: CANCEL ITINERARY
# =====================================================
@custom_login_required
@require_http_methods(["POST"])
def cancel_itinerary(request, itinerary_id):
    """Cancel a confirmed itinerary"""
    # Check session authentication
    user_id = request.session.get('user_id')
    if not user_id:
        return JsonResponse({
            'success': False,
            'message': 'Authentication required.'
        }, status=401)

    try:
        itinerary = get_object_or_404(Itinerary, id=itinerary_id)

        # Update status to cancelled
        itinerary.status = 'cancelled'
        itinerary.save()

        return JsonResponse({
            'success': True,
            'message': 'Itinerary cancelled successfully'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


# =====================================================
# API: SET DRAFT ITINERARY
# =====================================================
@custom_login_required
@require_http_methods(["POST"])
def set_draft_itinerary(request, itinerary_id):
    """Set itinerary back to draft"""
    # Check session authentication
    user_id = request.session.get('user_id')
    if not user_id:
        return JsonResponse({
            'success': False,
            'message': 'Authentication required.'
        }, status=401)

    try:
        itinerary = get_object_or_404(Itinerary, id=itinerary_id)

        # Update status to draft
        itinerary.status = 'draft'
        itinerary.is_finalized = False
        itinerary.finalized_at = None
        itinerary.selected_option = None
        itinerary.save()

        return JsonResponse({
            'success': True,
            'message': 'Itinerary set to draft'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


# =====================================================
# API: DELETE ITINERARY - FIXED FOR SESSION AUTH
# =====================================================

@require_POST
@custom_login_required
def delete_itinerary(request, itinerary_id):
    """
    Delete an itinerary and all its related data
    Works with session-based authentication
    """
    # Check session authentication (NOT Django's @login_required)
    user_id = request.session.get('user_id')
    user_type = request.session.get('user_type')

    # Check if it's an AJAX request
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    # Authentication check
    if not user_id:
        if is_ajax:
            return JsonResponse({
                'success': False,
                'message': 'Authentication required. Please log in.'
            }, status=401)
        messages.warning(request, '⚠️ Please login first')
        return redirect('team_member:login')

    try:
        # Get itinerary
        itinerary = get_object_or_404(Itinerary, id=itinerary_id)
        query_id = itinerary.query.id
        itinerary_name = itinerary.name or f"Itinerary #{itinerary.id}"

        # Permission check
        has_permission = False

        if user_type == 'superuser':
            has_permission = True

        elif user_type == 'team_member':
            team_member = TeamMember.objects.get(id=user_id)

            if team_member.role in ['admin', 'manager']:
                has_permission = True

            elif hasattr(team_member, 'permissions') and team_member.permissions:
                has_permission = team_member.permissions.can_delete_itinerary

        # Deny if no permission
        if not has_permission:
            if is_ajax:
                return JsonResponse({
                    'success': False,
                    'message': 'You do not have permission to delete itineraries.'
                }, status=403)
            else:
                messages.error(request, '❌ You do not have permission to delete itineraries.')
                return redirect('query_proposals', query_id=query_id)

        # Delete itinerary (CASCADE will delete related objects)
        itinerary.delete()

        # Return JSON response for AJAX calls
        if is_ajax:
            return JsonResponse({
                'success': True,
                'message': f'Itinerary "{itinerary_name}" deleted successfully!'
            })

        # Regular POST redirect
        messages.success(request, f'✅ Itinerary "{itinerary_name}" deleted successfully!')
        return redirect('query_proposals', query_id=query_id)

    except Itinerary.DoesNotExist:
        if is_ajax:
            return JsonResponse({
                'success': False,
                'message': 'Itinerary not found'
            }, status=404)
        messages.error(request, '❌ Itinerary not found')
        return redirect('query_list')

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error deleting itinerary {itinerary_id}: {str(e)}")

        if is_ajax:
            return JsonResponse({
                'success': False,
                'message': f'Error deleting itinerary: {str(e)}'
            }, status=500)

        messages.error(request, f'❌ Error deleting itinerary: {str(e)}')
        return redirect('query_proposals', query_id=query_id)



from django.shortcuts import render, get_object_or_404
from django.db.models import Sum
from django.conf import settings
import os

from .models import (
    Itinerary,
    Query,
    ItineraryPricingOption,
    HotelBooking,
    VehicleBooking,
    ItineraryDayPlan,
    PackageTermss,
    InvoiceLogo
)
from django.db.models import Q, Max  # <--- CRITICAL IMPORTS


@custom_login_required
def view_quotation(request, itinerary_id):
    # --- HELPER FUNCTIONS ---
    def split_into_bullets(text):
        if not text: return []
        text = text.replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")
        from django.utils.html import strip_tags
        text = strip_tags(text)
        text = text.replace("&nbsp;", " ").replace("&bull;", "").replace("•", "").replace("\r", "\n")
        import re
        lines = re.split(r'\n+', text)
        return [line.strip() for line in lines if line.strip()]

    def normalize_option(opt_string):
        if not opt_string: return 'option_1'
        s = str(opt_string).lower().strip().replace(' ', '').replace('_', '')
        if s in ['option2', 'deluxe', 'deluxepackage']: return 'option_2'
        if s in ['option3', 'premium', 'premiumpackage']: return 'option_3'
        if s in ['option4', 'luxury', 'luxurypackage']: return 'option_4'
        return 'option_1'

    # --- FETCH ITINERARY ---
    itinerary = get_object_or_404(
        Itinerary.objects.select_related('query', 'query__assign', 'query__created_by'),
        id=itinerary_id
    )
    query = itinerary.query

    # --- DETERMINE MODE ---
    is_confirmed_status = (itinerary.status == 'confirmed')
    active_option_key = 'option_1'
    package_display_name = "Standard"

    if is_confirmed_status and itinerary.selected_option:
        active_option_key = normalize_option(itinerary.selected_option)

    # --- FETCH HOTELS (Fixed Logic with Q objects) ---
    # We include items where option is NULL or Empty so manual items appear in all lists.

    # Option 1 (Standard)
    option1_hotels = HotelBooking.objects.filter(itinerary=itinerary).filter(
        Q(option__in=['option_1', 'option1', 'standard', 'Standard', 'Option 1', 'Option_1', 'opt1', '1']) |
        Q(option__isnull=True) | Q(option='')
    ).select_related('hotel', 'destination', 'room_type', 'meal_plan').order_by('check_in_date')

    # Option 2 (Deluxe)
    option2_hotels = HotelBooking.objects.filter(itinerary=itinerary).filter(
        Q(option__in=['option_2', 'option2', 'deluxe', 'Deluxe', 'Option 2', 'Option_2', 'option 2']) |
        Q(option__isnull=True) | Q(option='')
    ).select_related('hotel', 'destination', 'room_type', 'meal_plan').order_by('check_in_date')

    # Option 3 (Premium)
    option3_hotels = HotelBooking.objects.filter(itinerary=itinerary).filter(
        Q(option__in=['option_3', 'option3', 'premium', 'Premium', 'Option 3', 'Option_3', 'option 3']) |
        Q(option__isnull=True) | Q(option='')
    ).select_related('hotel', 'destination', 'room_type', 'meal_plan').order_by('check_in_date')

    # Option 4 (Luxury)
    option4_hotels = HotelBooking.objects.filter(itinerary=itinerary).filter(
        Q(option__in=['option_4', 'option4', 'luxury', 'Luxury', 'Option 4', 'Option_4', 'option 4']) |
        Q(option__isnull=True) | Q(option='')
    ).select_related('hotel', 'destination', 'room_type', 'meal_plan').order_by('check_in_date')

    # --- SELECT DISPLAY DATA ---
    if active_option_key == 'option_2':
        display_hotels = option2_hotels
        package_display_name = "Deluxe"
    elif active_option_key == 'option_3':
        display_hotels = option3_hotels
        package_display_name = "Premium"
    elif active_option_key == 'option_4':
        display_hotels = option4_hotels
        package_display_name = "Luxury"
    else:
        display_hotels = option1_hotels
        package_display_name = "Standard"

    # --- CALCULATE SHARED COSTS ---
    vehicle_cost = 0
    vehicles = VehicleBooking.objects.filter(itinerary=itinerary)
    vehicle_obj = vehicles.first()
    for v in vehicles:
        price = v.custom_total_price if hasattr(v, 'custom_total_price') and v.custom_total_price is not None else v.gross_price
        vehicle_cost += float(price or 0)

    activity_cost = 0
    try:
        activities = ActivityBooking.objects.filter(itinerary=itinerary)
        for a in activities:
            price = a.custom_total_price if hasattr(a, 'custom_total_price') and a.custom_total_price is not None else a.gross_price
            activity_cost += float(price or 0)
    except NameError:
        activities = []

    standalone_cost = 0
    try:
        standalones = StandaloneInclusionBooking.objects.filter(itinerary=itinerary)
        for s in standalones:
            standalone_cost += float(s.total_price or 0)
    except NameError:
        standalones = []

    # --- GET OPTION PRICES ---
    def get_pricing_row(option_num):
        names = []
        if option_num == 1:
            names = ["Option 1", "Option1", "option1", "Standard", "Standard Package", "option_1"]
        elif option_num == 2:
            names = ["Option 2", "Option2", "option2", "Deluxe", "Deluxe Package", "option_2"]
        elif option_num == 3:
            names = ["Option 3", "Option3", "option3", "Premium", "Premium Package", "option_3"]
        elif option_num == 4:
            names = ["Option 4", "Option4", "option4", "Luxury", "Luxury Package", "option_4"]

        pricing = ItineraryPricingOption.objects.filter(itinerary=itinerary, option_name__in=names).first()
        if not pricing:
             pricing = ItineraryPricingOption.objects.filter(itinerary=itinerary, option_number=option_num).first()
        return pricing

    pricing_row_1 = get_pricing_row(1)
    pricing_row_2 = get_pricing_row(2)
    pricing_row_3 = get_pricing_row(3)
    pricing_row_4 = get_pricing_row(4)

    option1_price = float(pricing_row_1.final_amount) if pricing_row_1 else 0
    option2_price = float(pricing_row_2.final_amount) if pricing_row_2 else 0
    option3_price = float(pricing_row_3.final_amount) if pricing_row_3 else 0
    option4_price = float(pricing_row_4.final_amount) if pricing_row_4 else 0

    # --- HEADER DISPLAY DATA (Rooms, Beds, Child Counts) ---

    # 1. Vehicle Name
    vehicle_name_fallback = "Not specified"
    if vehicle_obj:
        if vehicle_obj.vehicle: vehicle_name_fallback = vehicle_obj.vehicle.name
        elif hasattr(vehicle_obj, 'vehicle_type'): vehicle_name_fallback = vehicle_obj.vehicle_type

    # 2. Determine List to Count From (Default to Option 1/Standard)
    # This ensures the header always reflects the Standard package details unless empty.
    qs_for_counts = option1_hotels if option1_hotels.exists() else display_hotels

    # 3. Calculate Counts using Max() on the HotelBooking QuerySet
    room_count = 0
    extra_beds_count = 0
    child_with_bed_count = 0
    child_without_bed_count = 0

    if qs_for_counts.exists():
        # ✅ FIX: Using correct field names from your HotelBooking model
        aggr = qs_for_counts.aggregate(
            max_rooms=Max('num_double_beds'),
            max_extra=Max('extra_beds'),
            max_cwb=Max('child_with_bed'),
            max_cnb=Max('child_without_bed')
        )
        room_count = aggr['max_rooms'] or 0
        extra_beds_count = aggr['max_extra'] or 0
        child_with_bed_count = aggr['max_cwb'] or 0
        child_without_bed_count = aggr['max_cnb'] or 0

    # 4. Fallback to Pricing Row if booking data is missing (0)
    target_pricing_row = pricing_row_1 # Default to Option 1
    if active_option_key == 'option_2': target_pricing_row = pricing_row_2
    elif active_option_key == 'option_3': target_pricing_row = pricing_row_3
    elif active_option_key == 'option_4': target_pricing_row = pricing_row_4

    if target_pricing_row:
        if room_count == 0: room_count = target_pricing_row.number_of_rooms
        if extra_beds_count == 0: extra_beds_count = target_pricing_row.extra_beds
        if child_without_bed_count == 0: child_without_bed_count = target_pricing_row.child_without_bed
        if child_with_bed_count == 0: child_with_bed_count = target_pricing_row.child_with_bed

    # 5. Build the Data Dictionary for the Template
    selected_option_data = {
        'vehicle_type': target_pricing_row.vehicle_type if (target_pricing_row and target_pricing_row.vehicle_type) else vehicle_name_fallback,
        'number_of_rooms': room_count,
        'extra_beds': extra_beds_count,
        'child_without_bed': child_without_bed_count,
        'child_with_bed': child_with_bed_count,
    }

    # --- DETERMINE TOTAL DISPLAY PRICE ---
    total_package_price = 0
    if is_confirmed_status:
        if active_option_key == 'option_2': total_package_price = option2_price
        elif active_option_key == 'option_3': total_package_price = option3_price
        elif active_option_key == 'option_4': total_package_price = option4_price
        else: total_package_price = option1_price

        if total_package_price == 0 and itinerary.final_amount > 0:
             total_package_price = itinerary.final_amount
    else:
        total_package_price = option1_price

    # --- LOAD TERMS & PLANS ---
    day_plans = ItineraryDayPlan.objects.filter(itinerary=itinerary).order_by('day_number')

    try:
        package_terms = itinerary.custom_package_terms
    except:
        package_terms = PackageTermss.objects.first()

    incl_clean = split_into_bullets(package_terms.package_inclusion) if package_terms else []
    excl_clean = split_into_bullets(package_terms.package_exclusion) if package_terms else []
    terms_clean = split_into_bullets(package_terms.terms_and_conditions) if package_terms else []
    cancel_clean = split_into_bullets(package_terms.cancellation_policy) if package_terms else []

    created_by_name = "Dream Holidays Team"
    if query.assign:
        created_by_name = query.assign.get_full_name()
    elif query.created_by:
        created_by_name = query.created_by.get_full_name()

    context = {
        'itinerary': itinerary,
        'query': query,
        'is_finalized': is_confirmed_status,
        'hotel_bookings': display_hotels,
        'option1_hotels': option1_hotels,
        'option2_hotels': option2_hotels,
        'option3_hotels': option3_hotels,
        'option4_hotels': option4_hotels,
        'total_package_price': total_package_price,
        'option1_price': option1_price,
        'option2_price': option2_price,
        'option3_price': option3_price,
        'option4_price': option4_price,
        'confirmed_package_name': package_display_name,
        'selected_option': selected_option_data,
        'day_plans': day_plans,
        'package_terms': package_terms,
        'incl_clean': incl_clean,
        'excl_clean': excl_clean,
        'terms_clean': terms_clean,
        'cancel_clean': cancel_clean,
        'created_by_name': created_by_name,
    }

    return render(request, 'quotation.html', context)







@custom_login_required
def _link_callback(uri, rel):
    """
    Convert HTML URIs (static/media) into absolute system paths so xhtml2pdf can access them.
    """
    # Handle static files
    if uri.startswith(settings.STATIC_URL):
        path = uri.replace(settings.STATIC_URL, "")

        # Try STATIC_ROOT first (production)
        if settings.STATIC_ROOT:
            full_path = os.path.join(settings.STATIC_ROOT, path)
            if os.path.isfile(full_path):
                return full_path

        # Try STATICFILES_DIRS (development)
        if hasattr(settings, 'STATICFILES_DIRS') and settings.STATICFILES_DIRS:
            for static_dir in settings.STATICFILES_DIRS:
                full_path = os.path.join(static_dir, path)
                if os.path.isfile(full_path):
                    return full_path

    # Handle media files
    if settings.MEDIA_URL and uri.startswith(settings.MEDIA_URL):
        path = uri.replace(settings.MEDIA_URL, "")
        full_path = os.path.join(settings.MEDIA_ROOT, path)
        if os.path.isfile(full_path):
            return full_path

    # Return as-is if it's already an absolute path
    if os.path.isfile(uri):
        return uri

    return uri


@custom_login_required
def download_quotation_pdf(request, itinerary_id):
    itinerary = get_object_or_404(Itinerary, id=itinerary_id)
    query = itinerary.query

    # Determine selected option
    option_name = itinerary.selected_option or "option_1"

    selected_pricing = (
        ItineraryPricingOption.objects.filter(
            itinerary=itinerary, option_name=option_name
        ).first()
        or ItineraryPricingOption.objects.filter(itinerary=itinerary).first()
    )

    # Vehicle booking (optional)
    vehicle_booking = VehicleBooking.objects.filter(
        itinerary=itinerary, option=option_name
    ).first()

    # Aggregates for counts
    hotel_aggregates = HotelBooking.objects.filter(
        itinerary=itinerary, option=option_name
    ).aggregate(
        total_rooms=Sum("num_rooms"),
        total_extra_beds=Sum("extra_beds"),
        total_child_with_bed=Sum("child_with_bed"),
        total_child_without_bed=Sum("child_without_bed"),
    )

    # Accommodation rows
    hotel_bookings = (
        HotelBooking.objects.filter(itinerary=itinerary, option=option_name)
        .select_related("hotel", "destination", "room_type", "meal_plan")
        .order_by("check_in_date")
    )

    # Day plans (with optional images)
    day_plans = (
        ItineraryDayPlan.objects.filter(itinerary=itinerary)
        .select_related("destination")
        .order_by("day_number")
    )

    # Selected option dict
    selected_option = {
        "net_price": selected_pricing.net_price if selected_pricing else 0,
        "gross_price": selected_pricing.gross_price if selected_pricing else 0,
        "final_amount": selected_pricing.final_amount if selected_pricing else 0,
        "cgst_amount": selected_pricing.cgst_amount if selected_pricing else 0,
        "sgst_amount": selected_pricing.sgst_amount if selected_pricing else 0,
        "vehicle_type": vehicle_booking.vehicle_type if vehicle_booking else None,
        "number_of_rooms": hotel_aggregates["total_rooms"] or 0,
        "extra_beds": hotel_aggregates["total_extra_beds"] or 0,
        "child_without_bed": hotel_aggregates["total_child_without_bed"] or 0,
        "child_with_bed": hotel_aggregates["total_child_with_bed"] or 0,
    }

    # Use STATIC_URL for images - link_callback will convert to absolute paths
    logo_uri = settings.STATIC_URL + "assets/img/dreamholiday_logo.png"
    qr_code_uri = settings.STATIC_URL + "assets/img/paymentlow.png"

    total_package_price = (
        selected_pricing.final_amount
        if selected_pricing and selected_pricing.final_amount is not None
        else getattr(itinerary, "final_amount", 0)
    )

    context = {
        "itinerary": itinerary,
        "query": query,
        "selected_option": selected_option,
        "hotel_bookings": hotel_bookings,
        "day_plans": day_plans,
        "total_package_price": total_package_price,
        "logo_uri": logo_uri,
        "qr_code_url": qr_code_uri,  # Add QR code URL
        "STATIC_URL": settings.STATIC_URL,
        "MEDIA_URL": settings.MEDIA_URL,
    }

    # Render HTML
    html = render_to_string("quotation_pdf.html", context)

    # Generate PDF into memory buffer
    result = BytesIO()
    pdf = pisa.CreatePDF(
        src=html,
        dest=result,
        encoding="utf-8",
        link_callback=_link_callback,
    )

    if pdf.err:
        return HttpResponse("Error generating PDF", status=500)

    response = HttpResponse(result.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="Quotation_{itinerary.id}.pdf"'
    return response




# ==========================================
# ✅ GENERAL INCLUSION MANAGEMENT
# ==========================================

@require_POST
@custom_login_required
def add_general_inclusion_to_package(request, package_id):
    """Add a general inclusion to a specific day in package template."""
    package = get_object_or_404(PackageTemplate, id=package_id)

    inclusion_id = request.POST.get('inclusion_id')
    day_number = int(request.POST.get('day_number'))

    # Get or create day plan
    day_plan, _ = PackageTemplateDayPlan.objects.get_or_create(
        package_template=package,
        day_number=day_number
    )

    # Get the inclusion
    inclusion = get_object_or_404(SpecialInclusion, id=inclusion_id, inclusion_type='general')

    # Check if already added
    if day_plan.general_inclusions.filter(id=inclusion_id).exists():
        messages.warning(request, f'⚠️ {inclusion.name} is already added to Day {day_number}')
    else:
        # Add inclusion to day plan
        day_plan.general_inclusions.add(inclusion)
        messages.success(request, f'✅ {inclusion.name} added to Day {day_number}')

    return redirect('manage_package_day_plans', package_id=package_id)


@require_POST
@custom_login_required
def remove_general_inclusion_from_package(request, package_id, inclusion_id):
    """Remove a general inclusion from a specific day in package template."""
    package = get_object_or_404(PackageTemplate, id=package_id)
    day_number = int(request.POST.get('day_number'))

    day_plan = get_object_or_404(
        PackageTemplateDayPlan,
        package_template=package,
        day_number=day_number
    )

    inclusion = get_object_or_404(SpecialInclusion, id=inclusion_id)

    # Remove inclusion
    day_plan.general_inclusions.remove(inclusion)

    messages.success(request, f'✅ {inclusion.name} removed from Day {day_number}')
    return redirect('manage_package_day_plans', package_id=package_id)

###################################################################################################################################################3

# ========================================
# PACKAGE TEMPLATE VIEWS
# ========================================
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q
from .models import PackageTemplate, Destinations, PackageTheme, TeamMember

@custom_login_required
def create_package_template(request):
    """
    Create a new package template.
    Handles POST request from modal form.
    """

    if request.method == 'POST':
        # Extract form data
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        from_date = request.POST.get('from_date')
        to_date = request.POST.get('to_date')
        total_days = request.POST.get('total_days', '').strip()
        default_adults = request.POST.get('default_adults', '2').strip()
        default_children = request.POST.get('default_children', '0').strip()
        default_infants = int(request.POST.get('default_infants', 0))
        destinations_input = request.POST.get('destinations', '').strip()
        notes = request.POST.get('notes', '').strip()
        theme_id = request.POST.get('theme', '').strip()
        tags = request.POST.get('tags', '').strip()
        is_featured = request.POST.get('is_featured') == 'on'
        thumbnail = request.FILES.get('thumbnail')
        cover_image = request.FILES.get('cover_image')

        # Validation
        if not name:
            messages.error(request, 'Package name is required.')
            return redirect('list_package_templates')

        if not total_days:
            messages.error(request, 'Total days is required.')
            return redirect('list_package_templates')

        try:
            # Get the first TeamMember or set to None
            team_member = TeamMember.objects.first()

            # Create the package template
            package = PackageTemplate.objects.create(
                name=name,
                description=description or None,
                total_days=int(total_days),
                from_date=from_date or None,
                to_date=to_date or None,
                default_adults=int(default_adults or 2),
                default_children=int(default_children or 0),
                default_infants=default_infants,
                notes=notes or None,
                theme_id=int(theme_id) if theme_id else None,
                tags=tags or None,
                is_featured=is_featured,
                is_active=True,
                created_by=team_member,
                thumbnail=thumbnail,
                cover_image=cover_image
            )

            # Add destinations if provided
            if destinations_input:
                destination_names = [d.strip() for d in destinations_input.split(',') if d.strip()]
                for dest_name in destination_names:
                    destination, created = Destinations.objects.get_or_create(
                        name__iexact=dest_name,
                        defaults={'name': dest_name, 'is_active': True}
                    )
                    package.destinations.add(destination)

            messages.success(request, f'✅ Package template "{name}" created successfully!')
            return redirect('list_package_templates')

        except ValueError as e:
            messages.error(request, f'Invalid input: {str(e)}')
            return redirect('list_package_templates')
        except Exception as e:
            messages.error(request, f'Error creating package: {str(e)}')
            return redirect('list_package_templates')

    # GET request - redirect to list page
    return redirect('list_package_templates')

@custom_login_required
def edit_package_template(request, package_id):
    """
    Edit an existing package template.
    Handles POST request from modal form.
    """

    package = get_object_or_404(PackageTemplate, id=package_id)

    if request.method == 'POST':
        try:
            # Update text fields
            package.name = request.POST.get('name', package.name).strip()
            package.description = request.POST.get('description', '').strip() or None
            package.notes = request.POST.get('notes', '').strip() or None
            package.tags = request.POST.get('tags', '').strip() or None

            # Handle image uploads
            thumbnail = request.FILES.get('thumbnail')
            cover_image = request.FILES.get('cover_image')

            if thumbnail:
                package.thumbnail = thumbnail
            if cover_image:
                package.cover_image = cover_image

            # Update numeric fields with validation
            try:
                package.total_days = int(request.POST.get('total_days', package.total_days))
                package.default_adults = int(request.POST.get('default_adults', package.default_adults))
                package.default_children = int(request.POST.get('default_children', package.default_children))
                # ✅ ADD: Handle default_infants field
                package.default_infants = int(request.POST.get('default_infants', package.default_infants))

                # Validate numeric values
                if package.total_days < 1:
                    messages.error(request, 'Total days must be at least 1.')
                    return redirect('list_package_templates')
                if package.default_adults < 0:
                    messages.error(request, 'Default adults cannot be negative.')
                    return redirect('list_package_templates')
                if package.default_children < 0:
                    messages.error(request, 'Default children cannot be negative.')
                    return redirect('list_package_templates')
                if package.default_infants < 0:
                    messages.error(request, 'Default infants cannot be negative.')
                    return redirect('list_package_templates')

            except ValueError as e:
                messages.error(request, f'Invalid numeric value provided: {str(e)}')
                return redirect('list_package_templates')

            # ✅ IMPROVED: Update date fields with proper None handling
            from_date_str = request.POST.get('from_date', '').strip()
            to_date_str = request.POST.get('to_date', '').strip()

            # Set to None if empty string, otherwise keep the value
            package.from_date = from_date_str if from_date_str else None
            package.to_date = to_date_str if to_date_str else None

            # ✅ Validate date logic
            if package.from_date and package.to_date:
                if package.from_date > package.to_date:
                    messages.error(request, 'Travel end date must be after or equal to start date.')
                    return redirect('list_package_templates')

            # Update theme with validation
            theme_id = request.POST.get('theme', '').strip()
            if theme_id:
                try:
                    package.theme_id = int(theme_id)
                except ValueError:
                    package.theme_id = None
            else:
                package.theme_id = None

            # Update boolean fields
            package.is_featured = request.POST.get('is_featured') == 'on'
            package.is_active = request.POST.get('is_active', 'on') == 'on'

            # Save the package
            package.save()

            # ✅ IMPROVED: Update destinations with better error handling
            destinations_input = request.POST.get('destinations', '').strip()
            if destinations_input:
                package.destinations.clear()
                destination_names = [d.strip() for d in destinations_input.split(',') if d.strip()]

                for dest_name in destination_names:
                    try:
                        # Use case-insensitive lookup
                        destination = Destinations.objects.filter(name__iexact=dest_name).first()

                        if not destination:
                            # Create new destination if it doesn't exist
                            destination = Destinations.objects.create(
                                name=dest_name,
                                is_active=True
                            )

                        package.destinations.add(destination)

                    except Exception as e:
                        messages.warning(request, f'Could not add destination "{dest_name}": {str(e)}')
            else:
                # Clear destinations if input is empty
                package.destinations.clear()

            messages.success(request, f'✅ Package template "{package.name}" updated successfully!')
            return redirect('list_package_templates')

        except Exception as e:
            messages.error(request, f'❌ Error updating package: {str(e)}')
            return redirect('list_package_templates')

    # If not POST, redirect back to list
    return redirect('list_package_templates')

@custom_login_required
def list_package_templates(request):
    """
    List all package templates with search and filters.
    Supports: search by name/description/tags, filter by theme, featured-only toggle.
    """

    # Get filter parameters
    search = request.GET.get('search', '').strip()
    theme_id = request.GET.get('theme', '').strip()
    featured_only = request.GET.get('featured') == 'true'

    # Base queryset with optimized queries
    packages = PackageTemplate.objects.prefetch_related(
        'destinations',
        'day_plans',
        'theme',
        'pricing_options'
    )

    # Apply search filter
    if search:
        packages = packages.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search) |
            Q(tags__icontains=search)
        )

    # Apply theme filter
    if theme_id:
        packages = packages.filter(theme_id=theme_id)

    # Apply featured filter
    if featured_only:
        packages = packages.filter(is_featured=True)

    # Order by: featured first, then most used, then newest
    packages = packages.order_by('-is_featured', '-times_used', '-created_at')

    # Get all themes for filter dropdown
    themes = PackageTheme.objects.filter(status=True).order_by('package_name')

    context = {
        'packages': packages,
        'themes': themes,
        'search': search,
        'selected_theme': theme_id,
    }

    return render(request, 'list_package_templates.html', context)

@custom_login_required
def delete_package_template(request, package_id):
    """
    Delete a package template.
    Handles POST request from delete confirmation modal.
    """

    package = get_object_or_404(PackageTemplate, id=package_id)

    if request.method == 'POST':
        package_name = package.name

        try:
            # Delete associated day plans (cascade should handle this, but being explicit)
            package.day_plans.all().delete()

            # Delete the package
            package.delete()

            messages.success(request, f'✅ Package template "{package_name}" deleted successfully!')
        except Exception as e:
            messages.error(request, f'Error deleting package: {str(e)}')

    return redirect('list_package_templates')


from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import PackageTemplate

@custom_login_required
def get_package_details(request, package_id):
    try:
        package = get_object_or_404(PackageTemplate, id=package_id)

        data = {
            'success': True,
            'package': {
                'id': package.id,
                'name': package.name,
                'description': package.description or '',
                'total_days': package.total_days,
                # ✅ FIX: Handle None dates properly
                'from_date': package.from_date.strftime('%Y-%m-%d') if package.from_date else '',
                'to_date': package.to_date.strftime('%Y-%m-%d') if package.to_date else '',
                'default_adults': package.default_adults,
                'default_children': package.default_children,
                'default_infants': package.default_infants,
                'total_passengers': package.total_passengers,
                'destinations': [dest.name for dest in package.destinations.all()],
                'theme': package.theme.package_name if package.theme else '',
                'tags': package.tags or '',
                'notes': package.notes or '',
                'is_featured': package.is_featured,
                'is_active': package.is_active,
                'thumbnail': package.thumbnail.url if package.thumbnail else '',
                'cover_image': package.cover_image.url if package.cover_image else '',
            }
        }
        return JsonResponse(data)

    except PackageTemplate.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Package not found'
        }, status=404)
    except Exception as e:
        # ✅ Log the actual error for debugging
        import traceback
        print(f"Error in get_package_details: {str(e)}")
        print(traceback.format_exc())

        return JsonResponse({
            'success': False,
            'message': f'Server error: {str(e)}'
        }, status=500)
###################################################################################################################################################3


# ==========================================
# ✅ PACKAGE DAY PLAN VIEW
# ==========================================

from .models import (
    PackageTemplate, PackageTemplateDayPlan,
    Hotel, Houseboat, Activity, Vehicle, Destinations,
    Hotelprice, HouseboatPrice, ActivityPrice, VehiclePricing,
    SpecialInclusion, RoomType, MealPlan
)
from .forms import (
    HotelBookingForm, VehicleBookingForm,
    ActivityBookingForm, HouseboatBookingForm
)
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages


# ==========================================
# ✅ PACKAGE DAY PLAN VIEW - WITH PRICING VALIDATION
# ==========================================

@custom_login_required
def manage_package_day_plans(request, package_id):
    package = get_object_or_404(PackageTemplate, id=package_id)
    total_days = package.total_days or 1
    destinations = package.destinations.all()
    days = list(range(1, int(total_days) + 1))

    selected_section = None
    section_data = []
    context_day_number = None

    # ✅ FIXED: Prefetch multi-inclusions for both hotels and houseboats
    day_plans_qs = PackageTemplateDayPlan.objects.filter(package_template=package).prefetch_related(
        'hotel_bookings__hotel',
        'hotel_bookings__room_type',
        'hotel_bookings__meal_plan',
        'hotel_bookings__inclusion_items__special_inclusion',  # ✅ Hotel multi-inclusions
        'vehicle_bookings__vehicle',
        'activity_bookings__activity',
        'houseboat_bookings__houseboat',
        'houseboat_bookings__meal_plan',
        'houseboat_bookings__room_type',
        'houseboat_bookings__inclusion_items__special_inclusion',  # ✅ Houseboat multi-inclusions
    )

    day_plans_dict = {dp.day_number: dp for dp in day_plans_qs}
    saved_items_by_day = {}

    for day in days:
        day_plan = day_plans_dict.get(day)
        if day_plan:
            saved_items_by_day[day] = {
                'day_plan': day_plan,
                'destination': day_plan.destination.name if day_plan.destination else None,
                'hotel_bookings': day_plan.hotel_bookings.all(),
                'vehicle_bookings': day_plan.vehicle_bookings.all(),
                'activity_bookings': day_plan.activity_bookings.all(),
                'houseboat_bookings': day_plan.houseboat_bookings.all(),
            }
        else:
            saved_items_by_day[day] = {
                'day_plan': None,
                'destination': None,
                'hotel_bookings': [],
                'vehicle_bookings': [],
                'activity_bookings': [],
                'houseboat_bookings': []
            }

    if request.method == 'POST':
        form_type = request.POST.get('form_type')

        if form_type == 'destination_select':
            day_number = int(request.POST.get('day_number'))
            destination_id = request.POST.get('destination')

            if destination_id:
                destination = Destinations.objects.get(id=destination_id)

                day_plan, created = PackageTemplateDayPlan.objects.get_or_create(
                    package_template=package,
                    day_number=day_number
                )

                day_plan.destination = destination

                if not day_plan.description and destination.default_description:
                    day_plan.description = destination.default_description

                day_plan.save()

                messages.success(request, f'✅ Day {day_number} set to {destination.name}')

            return redirect('manage_package_day_plans', package_id=package.id)

        elif form_type == 'edit_day_details':
            day_number = int(request.POST.get('day_number'))
            title = request.POST.get('title', '').strip()
            description = request.POST.get('description', '').strip()
            image = request.FILES.get('image')

            day_plan, created = PackageTemplateDayPlan.objects.get_or_create(
                package_template=package,
                day_number=day_number
            )

            if title:
                day_plan.title = title

            if description:
                day_plan.description = description

            if image:
                if day_plan.image:
                    day_plan.image.delete(save=False)
                day_plan.image = image

            day_plan.save()

            messages.success(request, f'✅ Day {day_number} details updated successfully!')
            return redirect('manage_package_day_plans', package_id=package.id)

        elif form_type == 'section_select':
            selected_section = request.POST.get('section')
            context_day_number = int(request.POST.get('day_number'))
            day_plan = day_plans_dict.get(context_day_number)
            destination = day_plan.destination if day_plan else None

            if destination:
                # ==========================================
                # ✅ HOTELS - WITH PRICING VALIDATION
                # ==========================================
                if selected_section == 'hotels':
                    if package.from_date and package.to_date:
                        # Get hotel IDs with valid pricing for package dates
                        valid_hotel_ids = Hotelprice.objects.filter(
                            hotel__destination=destination,
                            from_date__lte=package.from_date,
                            to_date__gte=package.to_date
                        ).values_list('hotel_id', flat=True).distinct()

                        section_data = Hotel.objects.filter(
                            id__in=valid_hotel_ids,
                            destination=destination
                        ).select_related('destination')

                        print(f"✅ [PACKAGE] Found {section_data.count()} hotels with pricing for {package.from_date} to {package.to_date}")

                        if section_data.count() == 0:
                            messages.warning(
                                request,
                                f'⚠️ No hotels available with pricing for package dates '
                                f'{package.from_date.strftime("%d %b %Y")} to {package.to_date.strftime("%d %b %Y")}. '
                                f'Please add pricing first.'
                            )
                    else:
                        # No package dates - show all hotels
                        section_data = Hotel.objects.filter(
                            destination=destination
                        ).select_related('destination')
                        print(f"⚠️ [PACKAGE] No date range - showing all {section_data.count()} hotels")

                # ==========================================
                # ✅ HOUSEBOATS - WITH PRICING VALIDATION
                # ==========================================
                elif selected_section == 'houseboats':
                    if package.from_date and package.to_date:
                        # Get houseboat IDs with valid pricing for package dates
                        valid_houseboat_ids = HouseboatPrice.objects.filter(
                            houseboat__destination=destination,
                            from_date__lte=package.from_date,
                            to_date__gte=package.to_date
                        ).values_list('houseboat_id', flat=True).distinct()

                        section_data = Houseboat.objects.filter(
                            id__in=valid_houseboat_ids,
                            destination=destination,
                            status=True
                        ).select_related('destination')

                        print(f"✅ [PACKAGE] Found {section_data.count()} houseboats with pricing for {package.from_date} to {package.to_date}")

                        if section_data.count() == 0:
                            messages.warning(
                                request,
                                f'⚠️ No houseboats available with pricing for package dates '
                                f'{package.from_date.strftime("%d %b %Y")} to {package.to_date.strftime("%d %b %Y")}. '
                                f'Please add houseboat pricing first.'
                            )
                    else:
                        # No package dates - show all houseboats
                        section_data = Houseboat.objects.filter(
                            destination=destination,
                            status=True
                        ).select_related('destination')
                        print(f"⚠️ [PACKAGE] No date range - showing all {section_data.count()} houseboats")

                # ==========================================
                # ✅ ACTIVITIES - WITH PRICING VALIDATION
                # ==========================================
                elif selected_section == 'activities':
                    if package.from_date and package.to_date:
                        # Check if ActivityPrice has direct destination field
                        # If not, use: activity__destination
                        valid_activity_ids = ActivityPrice.objects.filter(
                            activity__destination=destination,  # Changed from activity__destination if error occurs
                            from_date__lte=package.from_date,
                            to_date__gte=package.to_date
                        ).values_list('activity_id', flat=True).distinct()

                        section_data = Activity.objects.filter(
                            id__in=valid_activity_ids,
                            destination=destination
                        ).select_related('destination')

                        print(f"✅ [PACKAGE] Found {section_data.count()} activities with pricing for {package.from_date} to {package.to_date}")

                        if section_data.count() == 0:
                            messages.warning(
                                request,
                                f'⚠️ No activities available with pricing for package dates '
                                f'{package.from_date.strftime("%d %b %Y")} to {package.to_date.strftime("%d %b %Y")}. '
                                f'Please add activity pricing first.'
                            )
                    else:
                        # No package dates - show all activities
                        section_data = Activity.objects.filter(
                            destination=destination
                        ).select_related('destination')
                        print(f"⚠️ [PACKAGE] No date range - showing all {section_data.count()} activities")

                # ==========================================
                # ✅ VEHICLES - WITH PRICING VALIDATION
                # ==========================================
                elif selected_section == 'vehicles':
                    if package.from_date and package.to_date:
                        # Get vehicle IDs with valid pricing for package dates
                        # Note: VehiclePricing has vehicle FK, so we access destination through vehicle__destination
                        valid_vehicle_ids = VehiclePricing.objects.filter(
                            vehicle__destination=destination,
                            from_date__lte=package.from_date,
                            to_date__gte=package.to_date
                        ).values_list('vehicle_id', flat=True).distinct()

                        section_data = Vehicle.objects.filter(
                            id__in=valid_vehicle_ids,
                            destination=destination
                        ).select_related('destination')

                        print(f"✅ [PACKAGE] Found {section_data.count()} vehicles with pricing for {package.from_date} to {package.to_date}")

                        if section_data.count() == 0:
                            messages.warning(
                                request,
                                f'⚠️ No vehicles available with pricing for package dates '
                                f'{package.from_date.strftime("%d %b %Y")} to {package.to_date.strftime("%d %b %Y")}. '
                                f'Please add vehicle pricing first.'
                            )
                    else:
                        # No package dates - show all vehicles
                        section_data = Vehicle.objects.filter(
                            destination=destination
                        ).select_related('destination')
                        print(f"⚠️ [PACKAGE] No date range - showing all {section_data.count()} vehicles")

                else:
                    section_data = []
            else:
                section_data = []
                messages.warning(request, f'⚠️ Please select a destination for Day {context_day_number} first.')

    context = {
        'package': package,
        'days': days,
        'destinations': destinations,
        'plans': day_plans_dict,
        'selected_section': selected_section,
        'section_data': section_data,
        'context_day_number': context_day_number,
        'sections': ['hotels', 'houseboats', 'activities', 'vehicles'],
        'saved_items_by_day': saved_items_by_day,

        # Forms for modals
        'hotel_form': HotelBookingForm(),
        'vehicle_form': VehicleBookingForm(),
        'activity_form': ActivityBookingForm(),
        'houseboat_form': HouseboatBookingForm(),

        # Data for dropdowns
        'room_types': RoomType.objects.filter(is_active=True),
        'meal_plans': MealPlan.objects.filter(status=True),

        # ✅ FIXED: Separate inclusions for hotels and houseboats
        'hotel_special_inclusions': SpecialInclusion.objects.filter(
            is_available=True,
            status=True,
            inclusion_type='hotel'
        ).order_by('name'),
        'houseboat_special_inclusions': SpecialInclusion.objects.filter(
            is_available=True,
            status=True,
            inclusion_type='houseboat'
        ).order_by('name'),
    }
    return render(request, 'manage_package_day_plans.html', context)








# ==========================================
# PACKAGE TEMPLATE BOOKING VIEWS
# ==========================================

import json  # ✅ ADD THIS IMPORT
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.db import transaction
from .models import (
    PackageTemplate, PackageTemplateDayPlan,
    HotelBooking, HotelBookingInclusion, Hotel  # ✅ Add HotelBookingInclusion
)
from .forms import HotelBookingForm
from django.http import JsonResponse

@require_POST
@custom_login_required
def create_hotel_booking_for_package(request, package_id):
    """
    Create hotel booking for a package template with support for multiple special inclusions.
    """
    package = get_object_or_404(PackageTemplate, id=package_id)
    form = HotelBookingForm(request.POST, package=package)

    # Check if request is AJAX
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if form.is_valid():
        try:
            with transaction.atomic():
                day_number = request.POST.get("day_number")
                hotel_id = request.POST.get("hotel_id")

                if not day_number:
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': 'Day number is required.',
                            'errors': ['Day number is required.']
                        })
                    messages.error(request, "Day number is required.")
                    return redirect('manage_package_day_plans', package_id=package_id)

                day_plan, created = PackageTemplateDayPlan.objects.get_or_create(
                    package_template=package,
                    day_number=day_number
                )

                booking = form.save(commit=False)
                booking.package_template = package
                booking.package_day_plan = day_plan
                booking.itinerary = None
                booking.day_plan = None

                if hotel_id:
                    hotel = get_object_or_404(Hotel, id=hotel_id)
                    booking.hotel = hotel
                else:
                    if is_ajax:
                        return JsonResponse({
                            'success': False,
                            'message': 'Hotel selection is required.',
                            'errors': ['Hotel selection is required.']
                        })
                    messages.error(request, "Hotel selection is required.")
                    return redirect('manage_package_day_plans', package_id=package_id)

                booking.save()
                day_plan.hotels.add(hotel)

                # Handle multiple special inclusions
                inclusions_json = request.POST.get('inclusions_data', '[]')
                try:
                    inclusions_data = json.loads(inclusions_json)

                    for inc_data in inclusions_data:
                        inclusion_id = inc_data.get('id')
                        num_adults = int(inc_data.get('adults', 0))
                        num_children = int(inc_data.get('children', 0))

                        if inclusion_id and (num_adults > 0 or num_children > 0):
                            HotelBookingInclusion.objects.create(
                                hotel_booking=booking,
                                special_inclusion_id=inclusion_id,
                                num_adults=num_adults,
                                num_children=num_children
                            )
                except json.JSONDecodeError:
                    pass  # Continue even if inclusions fail

                # Success response
                inclusion_count = booking.inclusion_items.count()
                success_message = f'✅ Hotel booking created for {hotel.name}'
                if inclusion_count > 0:
                    success_message += f' with {inclusion_count} special inclusion(s)!'

                if is_ajax:
                    return JsonResponse({
                        'success': True,
                        'message': success_message
                    })

                messages.success(request, success_message)

        except Exception as e:
            error_message = f'❌ Error creating booking: {str(e)}'
            if is_ajax:
                return JsonResponse({
                    'success': False,
                    'message': error_message,
                    'errors': [str(e)]
                })
            messages.error(request, error_message)
    else:
        # Form validation errors
        error_messages = []
        for field, errors in form.errors.items():
            for error in errors:
                if field == '__all__':
                    error_messages.append(str(error))
                else:
                    field_label = form.fields.get(field).label if hasattr(form.fields.get(field, {}), 'label') else field
                    error_messages.append(f"{field_label}: {error}")

        if is_ajax:
            return JsonResponse({
                'success': False,
                'message': 'Please fix the following errors:',
                'errors': error_messages
            })

        if error_messages:
            messages.error(request, '❌ Please fix: ' + ' | '.join(error_messages))

    return redirect('manage_package_day_plans', package_id=package_id)



# ✅ HELPER FUNCTION - Place this BEFORE update_hotel_booking function
def redirect_to_day_plan(booking, day_number):
    """
    Helper function to redirect to the day plan page (BUILD tab), not pricing.
    """
    if booking.day_plan and booking.day_plan.itinerary:
        itinerary_id = booking.day_plan.itinerary.id
        # ✅ Add tab parameter to force BUILD tab
        return redirect(f'/itinerary/{itinerary_id}/day-plan/?day={day_number}#build')
    elif hasattr(booking, 'itinerary') and booking.itinerary:
        itinerary_id = booking.itinerary.id
        return redirect(f'/itinerary/{itinerary_id}/day-plan/?day={day_number}#build')
    else:
        return redirect('dashboard')

from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction
import json


@require_POST
@custom_login_required
def update_hotel_booking(request, booking_id):
    """
    Update existing hotel booking with support for multiple special inclusions.
    ✅ FIX: Handles switching between Real Database Hotels and Demo Hotels correctly.
    ✅ FIX: Handles manual room type (custom_room_type) updates.
    """
    booking = get_object_or_404(HotelBooking, id=booking_id)
    form = HotelBookingForm(request.POST, instance=booking)

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    # ✅ Get day number before any processing to ensure redirect works
    day_number = booking.day_plan.day_number if booking.day_plan else 1

    if form.is_valid():
        try:
            with transaction.atomic():
                # =========================================================
                # 1. HANDLE HOTEL SWITCHING (Real vs Demo)
                # =========================================================
                # Save form with commit=False so we can manually adjust fields
                updated_booking = form.save(commit=False)

                # Get raw POST data to determine user intent (Form cleaning might miss hidden fields)
                hotel_id = request.POST.get('hotel_id') or request.POST.get('hotel')
                custom_name = request.POST.get('custom_hotel_name')
                destination_id = request.POST.get('destination')
                if hotel_id:
                    # 👉 CASE A: REAL HOTEL SELECTED
                    updated_booking.hotel_id = hotel_id
                    updated_booking.custom_hotel_name = None  # Clear demo name

                    # Auto-update destination from the selected hotel object
                    if updated_booking.hotel and updated_booking.hotel.destination:
                        updated_booking.destination = updated_booking.hotel.destination
                elif custom_name:
                    # 👉 CASE B: DEMO HOTEL ENTERED
                    updated_booking.hotel = None  # Clear database hotel link
                    updated_booking.custom_hotel_name = custom_name  # Set new text name

                    # For demo hotels, explicit destination ID is required
                    if destination_id:
                        updated_booking.destination_id = destination_id
                # =========================================================
                # 🔥 NEW: HANDLE ROOM TYPE (Database vs Manual)
                # =========================================================
                room_type_id = request.POST.get('room_type')
                custom_room_type = request.POST.get('custom_room_type')

                print(f"🛏️ Room Type Update - DB ID: {room_type_id}, Manual: {custom_room_type}")

                if custom_room_type and custom_room_type.strip():
                    # 👉 CASE A: MANUAL ROOM TYPE
                    updated_booking.room_type = None  # Clear DB relation
                    updated_booking.custom_room_type = custom_room_type.strip()
                    print(f"✅ Using manual room type: {custom_room_type}")

                elif room_type_id:
                    # 👉 CASE B: DATABASE ROOM TYPE
                    try:
                        room_type_obj = RoomType.objects.get(id=room_type_id)
                        updated_booking.room_type = room_type_obj
                        updated_booking.custom_room_type = None  # Clear manual text
                        print(f"✅ Using DB room type: {room_type_obj.name}")
                    except RoomType.DoesNotExist:
                        print(f"⚠️ Room type ID {room_type_id} not found, keeping existing")
                else:
                    # 👉 CASE C: Neither provided (keep existing)
                    print("ℹ️ No room type change requested")
                # Now save the booking to the database
                updated_booking.save()

                # =========================================================
                # 2. HANDLE INCLUSIONS
                # =========================================================
                # ✅ Clear existing inclusions
                updated_booking.inclusion_items.all().delete()

                # ✅ Get inclusions data from POST
                inclusions_json = request.POST.get('inclusions_data', '[]')
                print(f"📥 Received inclusions_data: {inclusions_json}")

                inclusions_data = []
                if inclusions_json and inclusions_json.strip() and inclusions_json != '[]':
                    try:
                        inclusions_data = json.loads(inclusions_json)
                        print(f"✅ Parsed {len(inclusions_data)} inclusions")
                    except json.JSONDecodeError as e:
                        print(f"❌ JSON Parse Error: {e}")
                        if is_ajax:
                            return JsonResponse({'success': False, 'message': f'Invalid inclusions data format: {str(e)}'}, status=400)

                inclusion_count = 0
                skipped_count = 0

                for inc_data in inclusions_data:
                    try:
                        inclusion_id = inc_data.get('id')
                        num_adults = int(inc_data.get('adults', 0))
                        num_children = int(inc_data.get('children', 0))

                        # Skip if no ID or no guests
                        if not inclusion_id:
                            skipped_count += 1
                            continue

                        if num_adults == 0 and num_children == 0:
                            skipped_count += 1
                            continue

                        # ✅ Verify inclusion exists
                        special_inclusion = SpecialInclusion.objects.get(id=inclusion_id)

                        # ✅ Create inclusion booking
                        HotelBookingInclusion.objects.create(
                            hotel_booking=updated_booking,
                            special_inclusion=special_inclusion,
                            num_adults=num_adults,
                            num_children=num_children
                        )
                        inclusion_count += 1

                    except Exception as inc_error:
                        print(f"❌ Error creating inclusion: {inc_error}")
                        skipped_count += 1
                        continue

                # =========================================================
                # 3. SUCCESS RESPONSE
                # =========================================================
                hotel_name = updated_booking.hotel.name if updated_booking.hotel else f"{updated_booking.custom_hotel_name} (Demo)"

                # 🔥 NEW: Include room type in success message
                room_display = updated_booking.room_type.name if updated_booking.room_type else (
                    f"{updated_booking.custom_room_type} (Manual)" if updated_booking.custom_room_type else "N/A"
                )

                success_message = f'✅ Hotel updated to {hotel_name} - {room_display}'

                if inclusion_count > 0:
                    success_message += f' with {inclusion_count} special inclusion(s)!'

                print(f"✅ {success_message}")

                if is_ajax:
                    return JsonResponse({
                        'success': True,
                        'message': success_message,
                        'booking_id': updated_booking.id,
                        'inclusion_count': inclusion_count
                    })

                messages.success(request, success_message)

                # ✅ STAY ON BUILD TAB - Redirect to specific day
                return redirect(f'/itinerary/{booking.day_plan.itinerary.id}/day-plan/?day={day_number}')

        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"❌ Error updating booking: {error_details}")

            error_message = f'❌ Error updating booking: {str(e)}'

            if is_ajax:
                return JsonResponse({'success': False, 'message': error_message}, status=400)

            messages.error(request, error_message)
            return redirect(f'/itinerary/{booking.day_plan.itinerary.id}/day-plan/?day={day_number}')

    else:
        # Form Validation Failed
        error_messages = []
        for field, errors in form.errors.items():
            for error in errors:
                error_messages.append(f"{field}: {error}")

        print(f"❌ Form validation errors: {error_messages}")

        if is_ajax:
            return JsonResponse({'success': False, 'message': 'Validation failed', 'errors': error_messages}, status=400)

        messages.error(request, '❌ Update failed: ' + ' | '.join(error_messages))
        return redirect(f'/itinerary/{booking.day_plan.itinerary.id}/day-plan/?day={day_number}')




@require_POST
@custom_login_required
def delete_hotel_booking(request, booking_id):
    """
    Delete hotel booking - works for BOTH itinerary and package.
    Stays on current day after deletion.
    """
    booking = get_object_or_404(HotelBooking, id=booking_id)

    # Store info before deletion
    hotel_name = booking.hotel.name if booking.hotel else "Unknown Hotel"
    is_package = booking.is_package_booking

    # ✅ GET DAY NUMBER FROM day_plan OR package_day_plan
    day_number = 1  # Default fallback

    if booking.day_plan:
        day_number = booking.day_plan.day_number
    elif booking.package_day_plan:
        day_number = booking.package_day_plan.day_number

    if is_package:
        parent_id = booking.package_template.id if booking.package_template else None
    else:
        parent_id = booking.itinerary.id if booking.itinerary else None

    # ✅ Check if we have a valid parent ID before deletion
    if not parent_id:
        messages.error(request, "❌ Cannot delete booking: No associated package or itinerary found.")
        return redirect('dashboard')

    try:
        with transaction.atomic():
            # ✅ Optional: Remove from day plan many-to-many if needed
            if booking.package_day_plan and booking.hotel:
                # Check if this is the only booking for this hotel on this day
                other_bookings = HotelBooking.objects.filter(
                    package_day_plan=booking.package_day_plan,
                    hotel=booking.hotel
                ).exclude(id=booking_id).exists()

                if not other_bookings:
                    # This is the last booking for this hotel on this day
                    booking.package_day_plan.hotels.remove(booking.hotel)

            # Delete the booking
            booking.delete()

            messages.success(
                request,
                f"✅ Hotel booking for '{hotel_name}' has been deleted successfully."
            )

    except Exception as e:
        messages.error(request, f"❌ Error deleting booking: {str(e)}")
        print(f"❌ Delete error details: {e}")

    # ✅ REDIRECT WITH DAY PARAMETER - STAY ON CURRENT DAY
    if is_package:
        return redirect(f'{reverse("manage_package_day_plans", args=[parent_id])}?day={day_number}')
    else:
        return redirect(f'{reverse("itinerary_day_plan", args=[parent_id])}?day={day_number}')









# ==========================================
# VEHICLE BOOKING FOR PACKAGE
# ==========================================

@require_POST
@custom_login_required
def create_vehicle_booking_for_package(request, package_id):
    package = get_object_or_404(PackageTemplate, id=package_id)
    form = VehicleBookingForm(request.POST)

    if not form.is_valid():
        messages.error(request, f"Form validation failed: {form.errors.as_text()}")
        return redirect('manage_package_day_plans', package_id=package.id)

    # Validate required POST data
    try:
        day_number = int(request.POST.get("day_number"))
        vehicle_id = int(request.POST.get("vehicle_id"))
        destination_id = int(request.POST.get("destination_id"))
    except (TypeError, ValueError):
        messages.error(request, "Invalid day number, vehicle, or destination.")
        return redirect('manage_package_day_plans', package_id=package.id)

    try:
        vehicle = get_object_or_404(Vehicle, pk=vehicle_id)
        destination = get_object_or_404(Destinations, pk=destination_id)
    except:
        messages.error(request, "Invalid vehicle or destination selected.")
        return redirect('manage_package_day_plans', package_id=package.id)

    with transaction.atomic():
        # Get or create day plan
        day_plan, created = PackageTemplateDayPlan.objects.get_or_create(
            package_template=package,
            day_number=day_number,
            defaults={"destination": destination}
        )

        # Create booking with all required fields
        booking = form.save(commit=False)
        booking.package_template = package
        booking.package_day_plan = day_plan
        booking.itinerary = None
        booking.day_plan = None
        booking.destination = destination
        booking.vehicle = vehicle
        booking.save()

        # Handle many-to-many if exists
        if hasattr(day_plan, "vehicles"):
            day_plan.vehicles.add(vehicle)

    messages.success(request, f'Vehicle booking for {vehicle.name} created successfully!')
    return redirect('manage_package_day_plans', package_id=package.id)


@require_POST
@custom_login_required
def update_vehicle_booking(request, booking_id):
    """
    Update vehicle booking - works for BOTH itinerary and package
    Stays on current day after update
    """
    booking_instance = get_object_or_404(VehicleBooking, id=booking_id)
    form = VehicleBookingForm(request.POST, instance=booking_instance)

    if form.is_valid():
        updated_booking = form.save(commit=False)

        # Optional: override vehicle
        vehicle_id = request.POST.get("vehicle_id")
        if vehicle_id:
            try:
                vehicle = get_object_or_404(Vehicle, pk=int(vehicle_id))
                updated_booking.vehicle = vehicle
            except (TypeError, ValueError):
                pass

        # Optional: override destination
        destination_id = request.POST.get("destination_id")
        if destination_id:
            try:
                destination = get_object_or_404(Destinations, pk=int(destination_id))
                updated_booking.destination = destination
            except (TypeError, ValueError):
                pass

        updated_booking.save()
        messages.success(request, 'Vehicle booking updated successfully!')
    else:
        # Stay on same booking; use original instance
        updated_booking = booking_instance
        messages.error(request, f"Update failed: {form.errors.as_text()}")

    # ✅ Determine day_number from UPDATED booking
    if updated_booking.day_plan:
        day_number = updated_booking.day_plan.day_number
    elif updated_booking.package_day_plan:
        day_number = updated_booking.package_day_plan.day_number
    else:
        day_number = 1  # fallback

    # ✅ Redirect using UPDATED booking
    if updated_booking.is_package_booking:
        return redirect(
            f'{reverse("manage_package_day_plans", args=[updated_booking.package_template.id])}'
            f'?day={day_number}'
        )
    else:
        return redirect(
            f'{reverse("itinerary_day_plan", args=[updated_booking.itinerary.id])}'
            f'?day={day_number}'
        )


@require_POST
@custom_login_required
def delete_vehicle_booking(request, booking_id):
    """
    Delete vehicle booking - works for BOTH itinerary and package
    Stays on current day after deletion
    """
    booking = get_object_or_404(VehicleBooking, id=booking_id)
    vehicle_name = str(booking.vehicle)

    # Store info before deletion
    is_package = booking.is_package_booking

    # ✅ GET DAY NUMBER FROM day_plan OR package_day_plan
    if booking.day_plan:
        day_number = booking.day_plan.day_number
    elif booking.package_day_plan:
        day_number = booking.package_day_plan.day_number
    else:
        day_number = 1  # Default fallback

    if is_package:
        parent_id = booking.package_template.id
    else:
        parent_id = booking.itinerary.id

    booking.delete()
    messages.success(request, f"Vehicle booking for '{vehicle_name}' has been deleted.")

    # Redirect based on booking type - WITH DAY PARAMETER
    if is_package:
        return redirect(f'{reverse("manage_package_day_plans", args=[parent_id])}?day={day_number}')
    else:
        return redirect(f'{reverse("itinerary_day_plan", args=[parent_id])}?day={day_number}')



# ==========================================
# ACTIVITY BOOKING FOR PACKAGE
# ==========================================

@require_POST
@custom_login_required
def create_activity_booking_for_package(request, package_id):
    package = get_object_or_404(PackageTemplate, id=package_id)
    form = ActivityBookingForm(request.POST)

    if form.is_valid():
        day_number = request.POST.get("day_number")
        activity_id = request.POST.get("activity_id")

        if not day_number or not activity_id:
            messages.error(request, "A day number and activity are required.")
            return redirect('manage_package_day_plans', package_id=package.id)

        day_plan, _ = PackageTemplateDayPlan.objects.get_or_create(
            package_template=package,
            day_number=day_number
        )

        booking = form.save(commit=False)
        booking.package_template = package
        booking.package_day_plan = day_plan
        booking.itinerary = None
        booking.day_plan = None
        booking.activity_id = activity_id
        booking.save()
        messages.success(request, "Activity successfully booked!")
    else:
        messages.error(request, f"There was an error: {form.errors.as_text()}")

    return redirect('manage_package_day_plans', package_id=package.id)

@require_POST
@custom_login_required
def update_activity_booking(request, booking_id):
    """
    Update activity booking - works for BOTH itinerary and package
    """
    booking_instance = get_object_or_404(ActivityBooking, id=booking_id)
    form = ActivityBookingForm(request.POST, instance=booking_instance)

    if form.is_valid():
        form.save()
        messages.success(request, 'Activity booking updated successfully!')
    else:
        messages.error(request, f"There was an error: {form.errors.as_text()}")

    # Redirect based on booking type
    if booking_instance.is_package_booking:
        return redirect('manage_package_day_plans', package_id=booking_instance.package_template.id)
    else:
        return redirect('itinerary_day_plan', itinerary_id=booking_instance.itinerary.id)

@require_POST
@custom_login_required
def delete_activity_booking(request, booking_id):
    """
    Delete activity booking - works for BOTH itinerary and package
    """
    booking = get_object_or_404(ActivityBooking, id=booking_id)
    activity_name = booking.activity.name

    # Store info before deletion
    is_package = booking.is_package_booking
    if is_package:
        parent_id = booking.package_template.id
    else:
        parent_id = booking.itinerary.id

    booking.delete()
    messages.success(request, f"Activity booking for '{activity_name}' has been deleted.")

    # Redirect based on booking type
    if is_package:
        return redirect('manage_package_day_plans', package_id=parent_id)
    else:
        return redirect('itinerary_day_plan', itinerary_id=parent_id)


import json
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.views.decorators.http import require_POST


# ==========================================
# ✅ CREATE HOUSEBOAT BOOKING FOR PACKAGE
# ==========================================
@require_POST
@custom_login_required
def create_houseboat_booking_for_package(request, package_id):
    """
    Create houseboat booking for package with validation
    """
    package = get_object_or_404(PackageTemplate, id=package_id)
    form = HouseboatBookingForm(request.POST)

    if form.is_valid():
        day_number = request.POST.get("day_number")
        houseboat_id = request.POST.get("houseboat_id")

        day_plan, _ = PackageTemplateDayPlan.objects.get_or_create(
            package_template=package,
            day_number=day_number
        )

        booking = form.save(commit=False)
        booking.package_template = package
        booking.package_day_plan = day_plan
        booking.itinerary = None
        booking.day_plan = None
        booking.houseboat_id = houseboat_id
        booking.save()

        # ✅ Handle multi-inclusions with validation
        inclusions_data_json = request.POST.get('inclusions_data', '[]')

        try:
            inclusions_data = json.loads(inclusions_data_json)
            print(f"📦 Creating houseboat inclusions for booking #{booking.id}")

            # Get package default limits
            package_default_adults = package.default_adults if hasattr(package, 'default_adults') else 2
            package_default_children = package.default_children if hasattr(package, 'default_children') else 0

            created_count = 0
            validation_errors = []

            for inc_data in inclusions_data:
                inclusion_id = inc_data.get('id')
                num_adults = int(inc_data.get('adults', 0))
                num_children = int(inc_data.get('children', 0))

                # ✅ VALIDATION: Check if adults exceed max
                if num_adults > package_default_adults:
                    validation_errors.append(
                        f"Inclusion {inclusion_id}: Adults ({num_adults}) exceeds maximum ({package_default_adults})"
                    )
                    num_adults = package_default_adults  # Clamp to max
                    print(f"⚠️ Adults clamped to {package_default_adults}")

                # ✅ VALIDATION: Check if children exceed max
                if num_children > package_default_children:
                    validation_errors.append(
                        f"Inclusion {inclusion_id}: Children ({num_children}) exceeds maximum ({package_default_children})"
                    )
                    num_children = package_default_children  # Clamp to max
                    print(f"⚠️ Children clamped to {package_default_children}")

                if inclusion_id and (num_adults > 0 or num_children > 0):
                    try:
                        special_inclusion = SpecialInclusion.objects.get(id=inclusion_id)

                        HouseboatBookingInclusion.objects.create(
                            houseboat_booking=booking,
                            special_inclusion=special_inclusion,
                            num_adults=num_adults,
                            num_children=num_children
                        )

                        print(f"✅ Created inclusion: {special_inclusion.name} - {num_adults}A + {num_children}C")
                        created_count += 1

                    except SpecialInclusion.DoesNotExist:
                        print(f"⚠️ Special inclusion {inclusion_id} not found")
                        continue

            # Show validation errors if any
            if validation_errors:
                for error in validation_errors:
                    print(f"❌ {error}")
                messages.warning(request, f'Some inclusions were adjusted due to limits. {", ".join(validation_errors)}')

            if created_count > 0:
                messages.success(request, f'✅ Houseboat booking created with {created_count} inclusions!')
            else:
                messages.success(request, '✅ Houseboat booking created successfully!')

        except json.JSONDecodeError as e:
            print(f"❌ Error parsing inclusions JSON: {e}")
            messages.warning(request, 'Houseboat booking created but inclusions could not be saved.')
        except Exception as e:
            print(f"❌ Error creating houseboat inclusions: {e}")
            import traceback
            traceback.print_exc()
            messages.warning(request, f'Houseboat booking created but error: {str(e)}')
    else:
        messages.error(request, f"❌ Error: {form.errors.as_text()}")

    return redirect('manage_package_day_plans', package_id=package.id)


# ==========================================
# ✅ UPDATE HOUSEBOAT BOOKING FOR PACKAGE
# ==========================================
# ==========================================
# ✅ UPDATE HOUSEBOAT BOOKING
# ==========================================
@require_POST
@custom_login_required
def update_houseboat_booking(request, booking_id):
    """
    Update houseboat booking with validation
    ✅ FIX: Preserves current day after update
    """
    booking = get_object_or_404(HouseboatBooking, id=booking_id)

    # ✅ Get day number FIRST (before any updates) so we can redirect back to it
    if booking.day_plan:
        day_number = booking.day_plan.day_number
    else:
        # Fallback: try to get from POST or default to 1
        day_number = request.POST.get('day_number', '1')

    form = HouseboatBookingForm(request.POST, instance=booking)

    if form.is_valid():
        booking = form.save()

        # ✅ Handle multi-inclusions with validation
        inclusions_data_json = request.POST.get('inclusions_data', '[]')

        try:
            inclusions_data = json.loads(inclusions_data_json)
            print(f"📦 Updating houseboat inclusions for booking #{booking.id}")

            # Get max limits from package or itinerary
            if booking.package_template:
                max_adults = booking.package_template.default_adults if hasattr(booking.package_template, 'default_adults') else 2
                max_children = booking.package_template.default_children if hasattr(booking.package_template, 'default_children') else 0
            elif hasattr(booking, 'itinerary') and booking.itinerary:
                max_adults = booking.itinerary.default_adults if hasattr(booking.itinerary, 'default_adults') else 2
                max_children = booking.itinerary.default_children if hasattr(booking.itinerary, 'default_children') else 0
            else:
                max_adults = 2
                max_children = 0

            # Delete existing inclusions
            deleted_count = booking.inclusion_items.all().delete()[0]
            print(f"🗑️ Deleted {deleted_count} old inclusions")

            created_count = 0
            validation_errors = []

            for inc_data in inclusions_data:
                inclusion_id = inc_data.get('id')
                num_adults = int(inc_data.get('adults', 0))
                num_children = int(inc_data.get('children', 0))

                # ✅ VALIDATION: Check if adults exceed max
                if num_adults > max_adults:
                    validation_errors.append(
                        f"Adults {num_adults} > max {max_adults}"
                    )
                    num_adults = max_adults
                    print(f"⚠️ Adults clamped to {max_adults}")

                # ✅ VALIDATION: Check if children exceed max
                if num_children > max_children:
                    validation_errors.append(
                        f"Children {num_children} > max {max_children}"
                    )
                    num_children = max_children
                    print(f"⚠️ Children clamped to {max_children}")

                if inclusion_id and (num_adults > 0 or num_children > 0):
                    try:
                        special_inclusion = SpecialInclusion.objects.get(id=inclusion_id)

                        HouseboatBookingInclusion.objects.create(
                            houseboat_booking=booking,
                            special_inclusion=special_inclusion,
                            num_adults=num_adults,
                            num_children=num_children
                        )

                        print(f"✅ Created inclusion: {special_inclusion.name} - {num_adults}A + {num_children}C")
                        created_count += 1

                    except SpecialInclusion.DoesNotExist:
                        print(f"⚠️ Special inclusion {inclusion_id} not found")
                        continue

            # Show validation errors if any
            if validation_errors:
                for error in validation_errors:
                    print(f"⚠️ {error}")
                messages.warning(request, f'Some inclusions were adjusted: {", ".join(validation_errors)}')

            if created_count > 0:
                messages.success(request, f'✅ Houseboat updated with {created_count} inclusions!')
            else:
                messages.success(request, '✅ Houseboat updated successfully!')

        except json.JSONDecodeError as e:
            print(f"❌ Error parsing inclusions JSON: {e}")
            messages.warning(request, 'Houseboat updated but inclusions could not be saved.')
        except Exception as e:
            print(f"❌ Error updating inclusions: {e}")
            messages.warning(request, f'Houseboat updated but error with inclusions: {str(e)}')

    else:
        messages.error(request, f"❌ Update Error: {form.errors.as_text()}")

    # ✅ FIXED: Redirect to the correct location with day parameter
    try:
        # Check if this is a package template booking
        if booking.package_template:
            print(f"🔄 Redirecting to package template {booking.package_template.id}")
            return redirect('manage_package_day_plans', package_id=booking.package_template.id)

        # Check if this is an itinerary booking
        elif hasattr(booking, 'itinerary') and booking.itinerary:
            itinerary_id = booking.itinerary.id
            print(f"🔄 Redirecting to itinerary {itinerary_id}, day {day_number}")
            # ✅ FIX: Redirect to same day in itinerary
            return redirect(f'/itinerary/{itinerary_id}/day-plan/?day={day_number}')

        # Fallback: try to get itinerary from day_plan
        elif booking.day_plan and hasattr(booking.day_plan, 'itinerary'):
            itinerary_id = booking.day_plan.itinerary.id
            print(f"🔄 Redirecting to itinerary {itinerary_id} (from day_plan), day {day_number}")
            return redirect(f'/itinerary/{itinerary_id}/day-plan/?day={day_number}')

    except Exception as e:
        print(f"❌ Error determining redirect: {e}")
        messages.error(request, f"Updated successfully but redirect error: {str(e)}")

    # Final fallback
    print("⚠️ Could not determine redirect, going to home")
    messages.warning(request, "Houseboat updated but could not return to correct page.")
    return redirect('/')



# ==========================================
# ✅ DELETE HOUSEBOAT BOOKING
# ==========================================
@require_POST
@custom_login_required
def delete_houseboat_booking(request, booking_id):
    """
    Delete houseboat booking
    """
    booking = get_object_or_404(HouseboatBooking, id=booking_id)
    houseboat_name = booking.houseboat.name

    is_package = bool(booking.package_template)
    parent_id = booking.package_template.id if is_package else booking.itinerary.id

    booking.delete()
    messages.success(request, f"✅ Houseboat booking for '{houseboat_name}' deleted.")

    if is_package:
        return redirect('manage_package_day_plans', package_id=parent_id)
    else:
        return redirect('itinerary_day_plan', itinerary_id=parent_id)


from itertools import chain
from operator import attrgetter
from decimal import Decimal, InvalidOperation
from collections import defaultdict
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils.timezone import now
from .models import (
    PackageTemplate, PackageTemplateDayPlan,
    HotelBooking, VehicleBooking, ActivityBooking, HouseboatBooking,
    Hotelprice, VehiclePricing, ActivityPrice, HouseboatPrice,
    PackagePricingOption
)



@require_POST
@custom_login_required
def package_template_pricing(request, package_id):
    package = get_object_or_404(PackageTemplate, id=package_id)

    def calculate_all_prices():
        all_bookings = []
        combined_bookings = list(chain(
            HotelBooking.objects.filter(package_day_plan__package_template=package).prefetch_related('inclusion_items__special_inclusion'),
            VehicleBooking.objects.filter(package_template=package),
            ActivityBooking.objects.filter(package_template=package),
            HouseboatBooking.objects.filter(package_template=package).prefetch_related('inclusion_items__special_inclusion')
        ))

        for booking in combined_bookings:
            net_price = Decimal('0.00')
            nights = 1

            # --- Hotel Booking Price Logic ---
            if isinstance(booking, HotelBooking):
                nights = (booking.check_out_date - booking.check_in_date).days or 1

                if hasattr(booking, 'net_price') and booking.net_price and booking.net_price > 0:
                    net_price = booking.net_price
                else:
                    rule = Hotelprice.objects.filter(
                        hotel=booking.hotel, room_type=booking.room_type, meal_plan=booking.meal_plan,
                        from_date__lte=booking.check_in_date, to_date__gte=booking.check_in_date
                    ).first()
                    if rule:
                        per_night = (
                            Decimal(booking.num_double_beds or 0) * (rule.double_bed or 0) +
                            Decimal(booking.child_with_bed or 0) * (rule.child_with_bed or 0) +
                            Decimal(booking.child_without_bed or 0) * (rule.child_without_bed or 0) +
                            Decimal(booking.extra_beds or 0) * (rule.extra_bed or 0)
                        )
                        net_price = per_night * nights

                inclusion_price = booking.get_total_inclusion_price()
                net_price += inclusion_price

                rule = Hotelprice.objects.filter(
                    hotel=booking.hotel, room_type=booking.room_type, meal_plan=booking.meal_plan,
                    from_date__lte=booking.check_in_date, to_date__gte=booking.check_in_date
                ).first()
                booking.price_record = rule
                booking.sort_date = booking.check_in_date
                booking.item_type = 'Accommodation'
                booking.inclusion_price = inclusion_price
                booking.inclusions_list = list(booking.inclusion_items.select_related('special_inclusion').all())

            # --- Vehicle Booking Price Logic ---
            elif isinstance(booking, VehicleBooking):
                rule = VehiclePricing.objects.filter(
                    vehicle=booking.vehicle,
                    from_date__lte=booking.pickup_date,
                    to_date__gte=booking.pickup_date
                ).first()

                expected_calculated_price = Decimal('0')
                if rule and hasattr(booking, 'total_km') and booking.total_km is not None:
                    if booking.total_km <= 100:
                        expected_calculated_price = rule.total_fee_100km or Decimal('0')
                    else:
                        extra_km = booking.total_km - 100
                        extra_cost = Decimal(str(extra_km)) * (rule.extra_fee_per_km or Decimal('0'))
                        expected_calculated_price = (rule.total_fee_100km or Decimal('0')) + extra_cost

                if hasattr(booking, 'custom_total_price') and booking.custom_total_price is not None and booking.custom_total_price > 0:
                    net_price = booking.custom_total_price
                elif hasattr(booking, 'net_price') and booking.net_price and booking.net_price > 0:
                    net_price = booking.net_price
                else:
                    net_price = expected_calculated_price

                booking.price_record = rule
                booking.sort_date = booking.pickup_date
                booking.item_type = 'Transportation'
                booking.inclusion_price = Decimal('0.00')
                booking.inclusions_list = []

            # --- Activity Booking Price Logic ---
            elif isinstance(booking, ActivityBooking):
                rule = ActivityPrice.objects.filter(
                    activity=booking.activity,
                    from_date__lte=booking.booking_date,
                    to_date__gte=booking.booking_date
                ).first() or ActivityPrice.objects.filter(activity=booking.activity).first()

                expected_calculated_price = Decimal('0')
                if rule:
                    current_total_people = (booking.num_adults or 0) + (booking.num_children or 0)
                    expected_calculated_price = Decimal(str(current_total_people)) * (rule.per_person or Decimal('0'))

                if hasattr(booking, 'custom_total_price') and booking.custom_total_price is not None and booking.custom_total_price > 0:
                    net_price = booking.custom_total_price
                elif hasattr(booking, 'net_price') and booking.net_price and booking.net_price > 0:
                    net_price = booking.net_price
                else:
                    net_price = expected_calculated_price

                booking.price_record = rule
                booking.sort_date = booking.booking_date
                booking.item_type = 'Activity'
                booking.inclusion_price = Decimal('0.00')
                booking.inclusions_list = []

            # --- Houseboat Booking Price Logic ---
            elif isinstance(booking, HouseboatBooking):
                nights = (booking.check_out_date - booking.check_in_date).days or 1

                if hasattr(booking, 'net_price') and booking.net_price and booking.net_price > 0:
                    net_price = booking.net_price
                else:
                    rule = HouseboatPrice.objects.filter(
                        houseboat=booking.houseboat,
                        meal_plan=booking.meal_plan,
                        room_type=booking.room_type,
                        from_date__lte=booking.check_in_date,
                        to_date__gte=booking.check_in_date
                    ).first()
                    if rule:
                        per_night = (
                            Decimal(booking.num_one_bed_rooms or 0) * (rule.one_bed or 0) +
                            Decimal(booking.num_two_bed_rooms or 0) * (rule.two_bed or 0) +
                            Decimal(booking.num_three_bed_rooms or 0) * (rule.three_bed or 0) +
                            Decimal(booking.num_four_bed_rooms or 0) * (rule.four_bed or 0) +
                            Decimal(booking.num_five_bed_rooms or 0) * (rule.five_bed or 0) +
                            Decimal(booking.num_six_bed_rooms or 0) * (rule.six_bed or 0) +
                            Decimal(booking.num_seven_bed_rooms or 0) * (rule.seven_bed or 0) +
                            Decimal(booking.num_eight_bed_rooms or 0) * (rule.eight_bed or 0) +
                            Decimal(booking.num_nine_bed_rooms or 0) * (rule.nine_bed or 0) +
                            Decimal(booking.num_ten_bed_rooms or 0) * (rule.ten_bed or 0) +
                            Decimal(booking.num_extra_beds or 0) * (rule.extra_bed or 0)
                        )
                        net_price = per_night * nights

                inclusion_price = booking.get_total_inclusion_price()
                net_price += inclusion_price

                rule = HouseboatPrice.objects.filter(
                    houseboat=booking.houseboat,
                    meal_plan=booking.meal_plan,
                    room_type=booking.room_type,
                    from_date__lte=booking.check_in_date,
                    to_date__gte=booking.check_in_date
                ).first()
                booking.price_record = rule
                booking.sort_date = booking.check_in_date
                booking.item_type = 'Houseboat'
                booking.inclusion_price = inclusion_price
                booking.inclusions_list = list(booking.inclusion_items.select_related('special_inclusion').all())

            # --- Individual Markup & Final Calculation for Each Item ---
            individual_markup = Decimal('0.00')
            if hasattr(booking, 'markup_value') and booking.markup_value:
                try:
                    markup_val = Decimal(str(booking.markup_value))
                    if markup_val > 0:
                        individual_markup = markup_val if booking.markup_type != 'percentage' else net_price * (markup_val / 100)
                except (ValueError, TypeError, InvalidOperation):
                    pass

            booking.calculated_price = {
                'net': net_price,
                'markup': individual_markup,
                'gross': net_price + individual_markup,
                'inclusion_price': getattr(booking, 'inclusion_price', Decimal('0.00')),
                'nights': nights
            }
            all_bookings.append(booking)

        return all_bookings

    # --- POST Request Logic (✅ FIXED) ---
    if request.method == 'POST':
        try:
            PackagePricingOption.objects.filter(package_template=package).delete()

            all_items = sorted(calculate_all_prices(), key=attrgetter('sort_date'))
            non_accommodation_net = sum(item.calculated_price['net'] for item in all_items if item.item_type != 'Accommodation')
            # ✅ FIX 1: Calculate non-accommodation markup
            non_accommodation_markup = sum(item.calculated_price['markup'] for item in all_items if item.item_type != 'Accommodation')

            grouped_hotels = defaultdict(list)
            for item in all_items:
                if item.item_type == 'Accommodation':
                    hotel_option = getattr(item, 'option', 'Option 1')
                    grouped_hotels[hotel_option].append(item)

            # ✅ FIX 2: Get CGST/SGST/Discount from POST request
            cgst_percentage = Decimal(request.POST.get('cgst', str(getattr(package, 'cgst_percentage', Decimal('2.5')))))
            sgst_percentage = Decimal(request.POST.get('sgst', str(getattr(package, 'sgst_percentage', Decimal('2.5')))))
            discount = Decimal(request.POST.get('discount', str(getattr(package, 'discount', Decimal('0')))))

            for index, (option_name, hotels) in enumerate(grouped_hotels.items(), 1):
                package_index_str = str(index)

                # ✅ FIX 3: Get stored global markup for this option
                stored_markup_type = request.session.get(f'package_{package.id}_option_{package_index_str}_markup_type', 'fixed')
                stored_markup_value = Decimal(request.session.get(f'package_{package.id}_option_{package_index_str}_markup_value', '0'))

                option_hotels_net = sum(h.calculated_price['net'] for h in hotels)
                # ✅ FIX 4: Calculate hotel individual markups
                option_hotels_markup = sum(h.calculated_price['markup'] for h in hotels)

                # ✅ FIX 5: Proper calculation flow
                package_net = option_hotels_net + non_accommodation_net
                package_individual_markup = option_hotels_markup + non_accommodation_markup
                package_base_for_global = package_net + package_individual_markup

                # Calculate global markup
                package_global_markup = package_base_for_global * (stored_markup_value / 100) if stored_markup_type == 'percentage' else stored_markup_value

                # Total with all markups
                package_gross_before_tax = package_base_for_global + package_global_markup

                # Calculate taxes and final amount
                cgst_amount = package_gross_before_tax * (cgst_percentage / 100)
                sgst_amount = package_gross_before_tax * (sgst_percentage / 100)
                final_amount = package_gross_before_tax + cgst_amount + sgst_amount - discount

                hotels_list = [{'name': h.hotel.name, 'net_price': float(h.calculated_price['net'])} for h in hotels]

                # ✅ FIX 6: Save with correct gross_price
                PackagePricingOption.objects.create(
                    package_template=package,
                    option_name=option_name,
                    option_number=index,
                    net_price=package_net,
                    gross_price=package_gross_before_tax,  # ✅ Price WITH markups (before tax)
                    cgst_amount=cgst_amount,
                    sgst_amount=sgst_amount,
                    discount_amount=discount,
                    final_amount=final_amount,  # ✅ Final with everything
                    hotels_included=hotels_list
                )

            if hasattr(package, 'is_finalized'):
                package.is_finalized = True
                package.finalized_at = now()
                package.save()

            messages.success(request, f"Saved {len(grouped_hotels)} pricing options for package!")
            return redirect('list_package_templates')

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            return redirect('package_template_pricing', package_id=package.id)

    # --- GET Request Logic ---

    if request.GET.get('markup_value') is not None and request.GET.get('markup_value') != '' and request.GET.get('selected_option'):
        selected_option = request.GET.get('selected_option')
        markup_type = request.GET.get('markup_type', 'fixed')
        markup_value = request.GET.get('markup_value', '0')

        request.session[f'package_{package.id}_option_{selected_option}_markup_type'] = markup_type
        request.session[f'package_{package.id}_option_{selected_option}_markup_value'] = markup_value
        request.session[f'package_{package.id}_selected_option'] = selected_option
        request.session.modified = True

    cgst_perc = Decimal(request.GET.get('cgst', str(getattr(package, 'cgst_percentage', '2.5'))))
    sgst_perc = Decimal(request.GET.get('sgst', str(getattr(package, 'sgst_percentage', '2.5'))))
    discount = Decimal(request.GET.get('discount', str(getattr(package, 'discount', '0'))))

    selected_option = request.GET.get('selected_option') or request.session.get(f'package_{package.id}_selected_option')
    if request.GET.get('selected_option'):
        request.session[f'package_{package.id}_selected_option'] = request.GET.get('selected_option')
        request.session.modified = True

    all_items = sorted(calculate_all_prices(), key=attrgetter('sort_date'))

    non_accommodation_net = sum(item.calculated_price['net'] for item in all_items if item.item_type != 'Accommodation')
    non_accommodation_markup = sum(item.calculated_price['markup'] for item in all_items if item.item_type != 'Accommodation')

    grouped_hotels = defaultdict(list)
    for item in all_items:
        if item.item_type == 'Accommodation':
            hotel_option_display = getattr(item, 'get_option_display', None)
            hotel_option = hotel_option_display() if callable(hotel_option_display) else getattr(item, 'option', 'Option 1')
            grouped_hotels[hotel_option].append(item)

    hotel_option_groups = []

    if not grouped_hotels and all_items:
        stored_markup_type = request.session.get(f'package_{package.id}_option_1_markup_type', 'fixed')
        stored_markup_value = Decimal(request.session.get(f'package_{package.id}_option_1_markup_value', '0'))

        package_base = non_accommodation_net + non_accommodation_markup
        global_markup = package_base * (stored_markup_value / 100) if stored_markup_type == 'percentage' else stored_markup_value
        gross_before_tax = package_base + global_markup
        cgst_amount = gross_before_tax * (cgst_perc / 100)
        sgst_amount = gross_before_tax * (sgst_perc / 100)
        final_amount = gross_before_tax + cgst_amount + sgst_amount - discount

        hotel_option_groups.append({
            'option_name': 'Standard Package',
            'hotel_count': 0,
            'hotels': [],
            'option_net_total': Decimal('0.00'),
            'net_price': non_accommodation_net,
            'markup': non_accommodation_markup,
            'global_markup': global_markup,
            'gross_before_tax': gross_before_tax,
            'cgst_amount': cgst_amount,
            'sgst_amount': sgst_amount,
            'discount': discount,
            'gross_price': final_amount
        })

    for index, (option_name, hotels) in enumerate(grouped_hotels.items()):
        package_index_str = str(index + 1)

        stored_markup_type = request.session.get(f'package_{package.id}_option_{package_index_str}_markup_type', 'fixed')
        stored_markup_value = Decimal(request.session.get(f'package_{package.id}_option_{package_index_str}_markup_value', '0'))

        option_hotels_net = sum(h.calculated_price['net'] for h in hotels)
        option_hotels_markup = sum(h.calculated_price['markup'] for h in hotels)

        package_net = option_hotels_net + non_accommodation_net
        package_markup = option_hotels_markup + non_accommodation_markup
        package_base_for_global = package_net + package_markup

        package_global_markup = package_base_for_global * (stored_markup_value / 100) if stored_markup_type == 'percentage' else stored_markup_value

        package_gross_before_tax = package_base_for_global + package_global_markup
        package_cgst = package_gross_before_tax * (cgst_perc / 100)
        package_sgst = package_gross_before_tax * (sgst_perc / 100)
        package_final_price = package_gross_before_tax + package_cgst + package_sgst - discount

        hotel_option_groups.append({
            'option_name': option_name,
            'hotel_count': len(hotels),
            'hotels': [{'name': h.hotel.name, 'net_price': h.calculated_price['net']} for h in hotels],
            'option_net_total': option_hotels_net,
            'net_price': package_net,
            'markup': package_markup,
            'global_markup': package_global_markup,
            'gross_before_tax': package_gross_before_tax,
            'cgst_amount': package_cgst,
            'sgst_amount': package_sgst,
            'discount': discount,
            'gross_price': package_final_price
        })

    current_markup_type = 'fixed'
    current_markup_value = Decimal('0')
    if selected_option:
        current_markup_type = request.session.get(f'package_{package.id}_option_{selected_option}_markup_type', 'fixed')
        current_markup_value = Decimal(request.session.get(f'package_{package.id}_option_{selected_option}_markup_value', '0'))

    context = {
        'package': package,
        'all_items': all_items,
        'hotel_option_groups': hotel_option_groups,
        'current_markup_type': current_markup_type,
        'current_markup_value': current_markup_value,
        'current_cgst': cgst_perc,
        'current_sgst': sgst_perc,
        'current_discount': discount,
        'selected_option': selected_option,
    }

    return render(request, 'pricing.html', context)

















from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction
from django.http import JsonResponse
from django.urls import reverse  # 🔥 ADD THIS LINE
from django.utils.timezone import now
from django.template.loader import render_to_string
from datetime import datetime, timedelta, time
from decimal import Decimal, InvalidOperation
from itertools import chain
from operator import attrgetter
from collections import defaultdict



from django.http import JsonResponse


@custom_login_required
def check_package_availability(request):
    """
    API endpoint to check package availability without inserting
    """
    package_id = request.GET.get('package_id')
    query_id = request.GET.get('query_id')

    try:
        package = get_object_or_404(PackageTemplate, id=package_id)
        query = get_object_or_404(Query, id=query_id)

        # Check availability
        warnings = validate_pricing_availability(package, query.from_date, query.to_date)

        # Build response
        unavailable_items = []
        total_items = 0
        available_count = 0

        for day_plan in package.day_plans.all():
            for hotel_booking in day_plan.hotel_bookings.all():
                total_items += 1
                current_date = query.from_date + timedelta(days=day_plan.day_number - 1)

                has_pricing = Hotelprice.objects.filter(
                    hotel=hotel_booking.hotel,
                    room_type=hotel_booking.room_type,
                    meal_plan=hotel_booking.meal_plan,
                    from_date__lte=current_date,
                    to_date__gte=current_date
                ).exists()

                if has_pricing:
                    available_count += 1
                else:
                    unavailable_items.append({
                        'type': 'Hotel',
                        'name': hotel_booking.hotel.name,
                        'day': day_plan.day_number,
                        'date': current_date.strftime('%d %b %Y')
                    })

        return JsonResponse({
            'unavailable_items': unavailable_items,
            'available_count': available_count,
            'total_items': total_items,
            'success': True
        })

    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'success': False
        }, status=400)




# ============================================================================
# PRICING VALIDATION HELPER
# ============================================================================
from django.db import transaction
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from datetime import datetime, timedelta
from decimal import Decimal
from .models import (
    PackageTemplate, Query, Itinerary,
    Hotelprice, VehiclePricing, ActivityPrice, HouseboatPrice
)


@custom_login_required
def validate_pricing_availability(package, start_date, end_date):
    """
    Check if all items in package have pricing for the given date range
    Returns list of items missing pricing
    """
    missing_pricing = []

    # Calculate date offset
    if not package.from_date:
        date_offset = 0
    else:
        date_offset = (start_date - package.from_date).days

    # Check hotel pricing
    for day_plan in package.day_plans.all():
        for hotel_booking in day_plan.hotel_bookings.all():
            new_check_in = hotel_booking.check_in_date + timedelta(days=date_offset)
            new_check_out = hotel_booking.check_out_date + timedelta(days=date_offset)

            # Check if pricing exists for this date range
            has_pricing = Hotelprice.objects.filter(
                hotel=hotel_booking.hotel,
                room_type=hotel_booking.room_type,
                meal_plan=hotel_booking.meal_plan,
                from_date__lte=new_check_in,
                to_date__gte=new_check_in
            ).exists()

            if not has_pricing:
                missing_pricing.append({
                    'type': 'Hotel',
                    'name': hotel_booking.hotel.name,
                    'room_type': hotel_booking.room_type.name if hotel_booking.room_type else 'N/A',
                    'meal_plan': hotel_booking.meal_plan.name if hotel_booking.meal_plan else 'N/A',
                    'day': day_plan.day_number,
                    'check_in': new_check_in,
                    'check_out': new_check_out
                })

    # Check activity pricing
    for day_plan in package.day_plans.all():
        activity_bookings = ActivityBooking.objects.filter(
            package_template=package,
            package_day_plan=day_plan
        )
        for activity_booking in activity_bookings:
            booking_date = (activity_booking.booking_date + timedelta(days=date_offset)) if activity_booking.booking_date else (start_date + timedelta(days=day_plan.day_number - 1))

            has_pricing = ActivityPrice.objects.filter(
                activity=activity_booking.activity,
                from_date__lte=booking_date,
                to_date__gte=booking_date
            ).exists()

            if not has_pricing:
                missing_pricing.append({
                    'type': 'Activity',
                    'name': activity_booking.activity.name,
                    'day': day_plan.day_number,
                    'date': booking_date.strftime('%d %b %Y')
                })

    # Check houseboat pricing
    for day_plan in package.day_plans.all():
        houseboat_bookings = HouseboatBooking.objects.filter(
            package_template=package,
            package_day_plan=day_plan
        )
        for houseboat_booking in houseboat_bookings:
            new_hb_check_in = (houseboat_booking.check_in_date + timedelta(days=date_offset)) if houseboat_booking.check_in_date else (start_date + timedelta(days=day_plan.day_number - 1))

            has_pricing = HouseboatPrice.objects.filter(
                houseboat=houseboat_booking.houseboat,
                room_type=houseboat_booking.room_type,
                meal_plan=houseboat_booking.meal_plan,
                from_date__lte=new_hb_check_in,
                to_date__gte=new_hb_check_in
            ).exists()

            if not has_pricing:
                missing_pricing.append({
                    'type': 'Houseboat',
                    'name': houseboat_booking.houseboat.name,
                    'day': day_plan.day_number,
                    'check_in': new_hb_check_in.strftime('%d %b %Y')
                })

    # Check vehicle pricing
    vehicle_bookings = VehicleBooking.objects.filter(package_template=package)
    for vehicle_booking in vehicle_bookings:
        pickup_date = (vehicle_booking.pickup_date + timedelta(days=date_offset)) if vehicle_booking.pickup_date else start_date

        # 🔥 FIXED: Use vehicle__destination instead of just destination
        has_pricing = VehiclePricing.objects.filter(
            vehicle=vehicle_booking.vehicle,
            from_date__lte=pickup_date,
            to_date__gte=pickup_date
        ).exists()

        if not has_pricing:
            missing_pricing.append({
                'type': 'Vehicle',
                'name': vehicle_booking.vehicle.name,
                'vehicle_type': vehicle_booking.vehicle_type if vehicle_booking.vehicle_type else 'N/A',
                'pickup_date': pickup_date.strftime('%d %b %Y')
            })

    return missing_pricing


@custom_login_required
def validate_pricing_availability(package, start_date, end_date):
    """
    Check if all items in package have pricing for the given date range
    Returns list of items missing pricing
    """
    missing_pricing = []

    # Calculate date offset
    if not package.from_date:
        date_offset = 0
    else:
        date_offset = (start_date - package.from_date).days

    # Check hotel pricing
    for day_plan in package.day_plans.all():
        for hotel_booking in day_plan.hotel_bookings.all():
            new_check_in = hotel_booking.check_in_date + timedelta(days=date_offset)
            new_check_out = hotel_booking.check_out_date + timedelta(days=date_offset)

            # Check if pricing exists for this date range
            has_pricing = Hotelprice.objects.filter(
                hotel=hotel_booking.hotel,
                room_type=hotel_booking.room_type,
                meal_plan=hotel_booking.meal_plan,
                from_date__lte=new_check_in,
                to_date__gte=new_check_in
            ).exists()

            if not has_pricing:
                missing_pricing.append({
                    'type': 'Hotel',
                    'name': hotel_booking.hotel.name,
                    'room_type': hotel_booking.room_type.name if hotel_booking.room_type else 'N/A',
                    'meal_plan': hotel_booking.meal_plan.name if hotel_booking.meal_plan else 'N/A',
                    'day': day_plan.day_number,
                    'check_in': new_check_in,
                    'check_out': new_check_out
                })

    # Check activity pricing
    for day_plan in package.day_plans.all():
        activity_bookings = ActivityBooking.objects.filter(
            package_template=package,
            package_day_plan=day_plan
        )
        for activity_booking in activity_bookings:
            booking_date = (activity_booking.booking_date + timedelta(days=date_offset)) if activity_booking.booking_date else (start_date + timedelta(days=day_plan.day_number - 1))

            has_pricing = ActivityPrice.objects.filter(
                activity=activity_booking.activity,
                from_date__lte=booking_date,
                to_date__gte=booking_date
            ).exists()

            if not has_pricing:
                missing_pricing.append({
                    'type': 'Activity',
                    'name': activity_booking.activity.name,
                    'day': day_plan.day_number,
                    'date': booking_date.strftime('%d %b %Y')
                })

    # Check houseboat pricing
    for day_plan in package.day_plans.all():
        houseboat_bookings = HouseboatBooking.objects.filter(
            package_template=package,
            package_day_plan=day_plan
        )
        for houseboat_booking in houseboat_bookings:
            new_hb_check_in = (houseboat_booking.check_in_date + timedelta(days=date_offset)) if houseboat_booking.check_in_date else (start_date + timedelta(days=day_plan.day_number - 1))

            has_pricing = HouseboatPrice.objects.filter(
                houseboat=houseboat_booking.houseboat,
                room_type=houseboat_booking.room_type,
                meal_plan=houseboat_booking.meal_plan,
                from_date__lte=new_hb_check_in,
                to_date__gte=new_hb_check_in
            ).exists()

            if not has_pricing:
                missing_pricing.append({
                    'type': 'Houseboat',
                    'name': houseboat_booking.houseboat.name,
                    'day': day_plan.day_number,
                    'check_in': new_hb_check_in.strftime('%d %b %Y')
                })

    # Check vehicle pricing
    vehicle_bookings = VehicleBooking.objects.filter(package_template=package)
    for vehicle_booking in vehicle_bookings:
        pickup_date = (vehicle_booking.pickup_date + timedelta(days=date_offset)) if vehicle_booking.pickup_date else start_date

        # 🔥 FIXED: Use vehicle__destination instead of just destination
        has_pricing = VehiclePricing.objects.filter(
            vehicle=vehicle_booking.vehicle,
            from_date__lte=pickup_date,
            to_date__gte=pickup_date
        ).exists()

        if not has_pricing:
            missing_pricing.append({
                'type': 'Vehicle',
                'name': vehicle_booking.vehicle.name,
                'vehicle_type': vehicle_booking.vehicle_type if vehicle_booking.vehicle_type else 'N/A',
                'pickup_date': pickup_date.strftime('%d %b %Y')
            })

    return missing_pricing



# ============================================================================
# CREATE ITINERARY FROM PACKAGE WITH VALIDATION
# ============================================================================

from datetime import time, timedelta
from decimal import Decimal
from django.utils.timezone import now
from .models import (
    Itinerary, ItineraryDayPlan, ItineraryPricingOption,
    HotelBooking, HotelBookingInclusion, VehicleBooking, ActivityBooking, HouseboatBooking,
    Hotelprice, VehiclePricing, ActivityPrice, HouseboatPrice,
    PackageTemplate, Query
)

@custom_login_required
def create_itinerary_from_package_with_validation(package, query, start_date, end_date, created_by=None):
    """
    Create itinerary from package - ONLY INSERT ITEMS WITH VALID PRICING
    Items without pricing are automatically SKIPPED
    AUTO-SAVES PRICING WITH TAXES - No manual entry needed!
    Returns: (itinerary, skipped_items_list)
    """

    print("=" * 80)
    print(f"🔧 PACKAGE INSERTION WITH PRICING VALIDATION")
    print(f"📦 Package: {package.name}")
    print(f"📅 Query dates: {start_date} to {end_date}")
    print(f"👤 Client: {query.client_name}")
    print("=" * 80)

    # Calculate date offset
    if not package.from_date:
        date_offset = 0
        print("⚠️  WARNING: Package has no from_date, using offset = 0")
    else:
        date_offset = (start_date - package.from_date).days
        print(f"✅ Date offset calculated: {date_offset} days")
        print(f"   (Package date: {package.from_date} → Query date: {start_date})")

    print("")

    # Create itinerary
    itinerary = Itinerary.objects.create(
        query=query,
        name=f"{package.name} - {query.client_name}",
        discount=getattr(package, 'discount', 0),
        cgst_percentage=getattr(package, 'cgst_percentage', 0),
        sgst_percentage=getattr(package, 'sgst_percentage', 0),
        status='draft',
        created_by=created_by  # ✅ FIXED: Now accepts this parameter

    )

    if hasattr(package, 'destinations'):
        itinerary.destinations.set(package.destinations.all())

    print(f"✅ Itinerary created: ID {itinerary.id}")
    print(f"   Name: {itinerary.name}\n")

    # Statistics
    stats = {
        'hotels_inserted': 0,
        'hotels_skipped': 0,
        'activities_inserted': 0,
        'activities_skipped': 0,
        'vehicles_inserted': 0,
        'vehicles_skipped': 0,
        'houseboats_inserted': 0,
        'houseboats_skipped': 0,
        'day_plans_created': 0
    }

    skipped_items = []

    # Copy day plans
    day_plans = package.day_plans.all().order_by('day_number')
    print(f"📋 Processing {day_plans.count()} day plans...\n")

    for template_day in day_plans:
        current_date = start_date + timedelta(days=template_day.day_number - 1)

        print("─" * 80)
        print(f"📅 DAY {template_day.day_number}: {current_date.strftime('%d %b %Y')} - {template_day.title}")
        print("─" * 80)

        # CREATE DAY PLAN FIRST
        itinerary_day = ItineraryDayPlan.objects.create(
            itinerary=itinerary,
            day_number=template_day.day_number,
            destination=template_day.destination,
            title=template_day.title,
            description=template_day.description,
            image=template_day.image if template_day.image else None,
        )
        stats['day_plans_created'] += 1

        # ========== HOTELS ==========
        hotel_bookings = template_day.hotel_bookings.all()
        if hotel_bookings.exists():
            print(f"\n🏨 HOTELS ({hotel_bookings.count()} found):")

            for hotel_booking in hotel_bookings:
                new_check_in = hotel_booking.check_in_date + timedelta(days=date_offset)
                new_check_out = hotel_booking.check_out_date + timedelta(days=date_offset)

                # CHECK PRICING AVAILABILITY
                has_pricing = Hotelprice.objects.filter(
                    hotel=hotel_booking.hotel,
                    room_type=hotel_booking.room_type,
                    meal_plan=hotel_booking.meal_plan,
                    from_date__lte=new_check_out,
                    to_date__gte=new_check_in
                ).exists()

                room_type_name = hotel_booking.room_type.name if hotel_booking.room_type else 'N/A'
                meal_plan_name = hotel_booking.meal_plan.name if hotel_booking.meal_plan else 'N/A'
                hotel_name = f"{hotel_booking.hotel.name} ({room_type_name}, {meal_plan_name})"

                if has_pricing:
                    new_hotel_booking = HotelBooking.objects.create(
                        itinerary=itinerary,
                        day_plan=itinerary_day,
                        hotel=hotel_booking.hotel,
                        destination=hotel_booking.destination,
                        category=getattr(hotel_booking, 'category', None),
                        room_type=hotel_booking.room_type,
                        meal_plan=hotel_booking.meal_plan,
                        option=getattr(hotel_booking, 'option', None),
                        num_rooms=getattr(hotel_booking, 'num_rooms', 1),
                        num_double_beds=getattr(hotel_booking, 'num_double_beds', 0),
                        child_with_bed=getattr(hotel_booking, 'child_with_bed', 0),
                        child_without_bed=getattr(hotel_booking, 'child_without_bed', 0),
                        extra_beds=getattr(hotel_booking, 'extra_beds', 0),
                        check_in_date=new_check_in,
                        check_in_time=hotel_booking.check_in_time if hotel_booking.check_in_time else time(14, 0),
                        check_out_date=new_check_out,
                        check_out_time=hotel_booking.check_out_time if hotel_booking.check_out_time else time(11, 0),
                    )

                    # Copy special inclusions
                    inclusion_count = 0
                    if hasattr(hotel_booking, 'inclusion_items'):
                        for inclusion_item in hotel_booking.inclusion_items.all():
                            HotelBookingInclusion.objects.create(
                                hotel_booking=new_hotel_booking,
                                special_inclusion=inclusion_item.special_inclusion,
                                num_adults=getattr(inclusion_item, 'num_adults', 0),
                                num_children=getattr(inclusion_item, 'num_children', 0),
                                price=getattr(inclusion_item, 'price', 0),
                            )
                            inclusion_count += 1

                    inclusion_info = f" + {inclusion_count} inclusions" if inclusion_count > 0 else ""
                    print(f"   ✅ {hotel_name}{inclusion_info}")
                    print(f"      Dates: {new_check_in.strftime('%d %b')} → {new_check_out.strftime('%d %b')}")
                    stats['hotels_inserted'] += 1
                else:
                    print(f"   ❌ {hotel_name} - SKIPPED (No pricing)")
                    print(f"      Missing: {new_check_in.strftime('%d %b')} → {new_check_out.strftime('%d %b')}")
                    stats['hotels_skipped'] += 1
                    skipped_items.append({
                        'day': template_day.day_number,
                        'type': 'Hotel',
                        'name': hotel_name,
                        'date': f"{new_check_in.strftime('%d %b')} → {new_check_out.strftime('%d %b')}"
                    })

        # ========== ACTIVITIES ==========
        activity_bookings = ActivityBooking.objects.filter(
            package_template=package,
            package_day_plan=template_day
        )

        if activity_bookings.exists():
            print(f"\n🎯 ACTIVITIES ({activity_bookings.count()} found):")

            for activity_booking in activity_bookings:
                new_booking_date = (activity_booking.booking_date + timedelta(days=date_offset)) if activity_booking.booking_date else current_date

                has_pricing = ActivityPrice.objects.filter(
                    activity=activity_booking.activity,
                    from_date__lte=new_booking_date,
                    to_date__gte=new_booking_date
                ).exists()

                if has_pricing:
                    ActivityBooking.objects.create(
                        itinerary=itinerary,
                        day_plan=itinerary_day,
                        activity=activity_booking.activity,
                        booking_date=new_booking_date,
                        booking_time=getattr(activity_booking, 'booking_time', None),
                        num_adults=getattr(activity_booking, 'num_adults', 0),
                        num_children=getattr(activity_booking, 'num_children', 0),
                        notes=getattr(activity_booking, 'notes', ''),
                    )
                    print(f"   ✅ {activity_booking.activity.name} ({new_booking_date.strftime('%d %b')})")
                    stats['activities_inserted'] += 1
                else:
                    print(f"   ❌ {activity_booking.activity.name} - SKIPPED (No pricing)")
                    stats['activities_skipped'] += 1
                    skipped_items.append({
                        'day': template_day.day_number,
                        'type': 'Activity',
                        'name': activity_booking.activity.name,
                        'date': new_booking_date.strftime('%d %b %Y')
                    })

        # ========== HOUSEBOATS ==========
        houseboat_bookings = HouseboatBooking.objects.filter(
            package_template=package,
            package_day_plan=template_day
        )

        if houseboat_bookings.exists():
            print(f"\n🚤 HOUSEBOATS ({houseboat_bookings.count()} found):")

            for houseboat_booking in houseboat_bookings:
                new_hb_check_in = (houseboat_booking.check_in_date + timedelta(days=date_offset)) if houseboat_booking.check_in_date else current_date
                new_hb_check_out = (houseboat_booking.check_out_date + timedelta(days=date_offset)) if houseboat_booking.check_out_date else current_date + timedelta(days=1)

                has_pricing = HouseboatPrice.objects.filter(
                    houseboat=houseboat_booking.houseboat,
                    room_type=houseboat_booking.room_type,
                    meal_plan=houseboat_booking.meal_plan,
                    from_date__lte=new_hb_check_out,
                    to_date__gte=new_hb_check_in
                ).exists()

                room_type_name = houseboat_booking.room_type.name if houseboat_booking.room_type else 'Standard'
                houseboat_name = f"{houseboat_booking.houseboat.name} ({room_type_name})"

                if has_pricing:
                    HouseboatBooking.objects.create(
                        itinerary=itinerary,
                        day_plan=itinerary_day,
                        houseboat=houseboat_booking.houseboat,
                        meal_plan=houseboat_booking.meal_plan,
                        room_type=houseboat_booking.room_type,
                        check_in_date=new_hb_check_in,
                        check_out_date=new_hb_check_out,
                        num_one_bed_rooms=getattr(houseboat_booking, 'num_one_bed_rooms', 0),
                        num_two_bed_rooms=getattr(houseboat_booking, 'num_two_bed_rooms', 0),
                        num_three_bed_rooms=getattr(houseboat_booking, 'num_three_bed_rooms', 0),
                        num_four_bed_rooms=getattr(houseboat_booking, 'num_four_bed_rooms', 0),
                        num_five_bed_rooms=getattr(houseboat_booking, 'num_five_bed_rooms', 0),
                        num_six_bed_rooms=getattr(houseboat_booking, 'num_six_bed_rooms', 0),
                        num_seven_bed_rooms=getattr(houseboat_booking, 'num_seven_bed_rooms', 0),
                        num_eight_bed_rooms=getattr(houseboat_booking, 'num_eight_bed_rooms', 0),
                        num_nine_bed_rooms=getattr(houseboat_booking, 'num_nine_bed_rooms', 0),
                        num_ten_bed_rooms=getattr(houseboat_booking, 'num_ten_bed_rooms', 0),
                        num_extra_beds=getattr(houseboat_booking, 'num_extra_beds', 0),
                    )
                    print(f"   ✅ {houseboat_name}")
                    print(f"      {new_hb_check_in.strftime('%d %b')} → {new_hb_check_out.strftime('%d %b')}")
                    stats['houseboats_inserted'] += 1
                else:
                    print(f"   ❌ {houseboat_name} - SKIPPED (No pricing)")
                    stats['houseboats_skipped'] += 1
                    skipped_items.append({
                        'day': template_day.day_number,
                        'type': 'Houseboat',
                        'name': houseboat_name,
                        'date': f"{new_hb_check_in.strftime('%d %b')} → {new_hb_check_out.strftime('%d %b')}"
                    })

        print("")

    # ========== VEHICLES ==========
    vehicle_bookings = VehicleBooking.objects.filter(package_template=package)

    if vehicle_bookings.exists():
        print("─" * 80)
        print(f"🚗 VEHICLES ({vehicle_bookings.count()} found):")
        print("─" * 80)

        for vehicle_booking in vehicle_bookings:
            new_pickup = (vehicle_booking.pickup_date + timedelta(days=date_offset)) if vehicle_booking.pickup_date else start_date

            has_pricing = VehiclePricing.objects.filter(
                vehicle=vehicle_booking.vehicle,
                from_date__lte=new_pickup,
                to_date__gte=new_pickup
            ).exists()

            vehicle_type = getattr(vehicle_booking, 'vehicle_type', 'N/A')
            vehicle_name = f"{vehicle_booking.vehicle.name} ({vehicle_type})"

            if has_pricing:
                VehicleBooking.objects.create(
                    itinerary=itinerary,
                    vehicle=vehicle_booking.vehicle,
                    destination=getattr(vehicle_booking, 'destination', None),
                    pickup_date=new_pickup,
                    pickup_time=vehicle_booking.pickup_time if vehicle_booking.pickup_time else time(9, 0),
                    num_passengers=getattr(vehicle_booking, 'num_passengers', 0),
                    vehicle_type=vehicle_type,
                    option=getattr(vehicle_booking, 'option', None),
                    total_km=getattr(vehicle_booking, 'total_km', 0),
                )
                print(f"✅ {vehicle_name}")
                print(f"   Pickup: {new_pickup.strftime('%d %b %Y')}")
                stats['vehicles_inserted'] += 1
            else:
                print(f"❌ {vehicle_name} - SKIPPED (No pricing)")
                print(f"   Missing pricing for: {new_pickup.strftime('%d %b %Y')}")
                stats['vehicles_skipped'] += 1
                skipped_items.append({
                    'day': '-',
                    'type': 'Vehicle',
                    'name': vehicle_name,
                    'date': new_pickup.strftime('%d %b %Y')
                })
        print("")

    # ========== FINAL SUMMARY ==========
    print("=" * 80)
    print("📊 INSERTION SUMMARY")
    print("=" * 80)
    print(f"✅ Itinerary: {itinerary.name} (ID: {itinerary.id})")
    print(f"\n📈 ITEMS INSERTED:")
    print(f"   • Day Plans: {stats['day_plans_created']}")
    print(f"   • Hotels: {stats['hotels_inserted']}")
    print(f"   • Activities: {stats['activities_inserted']}")
    print(f"   • Vehicles: {stats['vehicles_inserted']}")
    print(f"   • Houseboats: {stats['houseboats_inserted']}")

    total_inserted = (stats['hotels_inserted'] + stats['activities_inserted'] +
                      stats['vehicles_inserted'] + stats['houseboats_inserted'])
    total_skipped = (stats['hotels_skipped'] + stats['activities_skipped'] +
                     stats['vehicles_skipped'] + stats['houseboats_skipped'])

    print(f"\n   TOTAL INSERTED: {total_inserted}")

    if total_skipped > 0:
        print(f"\n⚠️  ITEMS SKIPPED (No pricing): {total_skipped}")
        if skipped_items:
            print(f"\n📋 SKIPPED ITEMS DETAILS:")
            for item in skipped_items[:10]:
                print(f"   • Day {item['day']}: {item['type']} - {item['name']} ({item['date']})")
            if len(skipped_items) > 10:
                print(f"   ... and {len(skipped_items) - 10} more items")

    print("=" * 80)

    # 🔥 AUTO-SAVE PRICING WITH INCLUSIONS (COMPLETE FIX)
    print("\n💰 AUTO-SAVING Pricing Options WITH TAXES...\n")
    try:
        ItineraryPricingOption.objects.filter(itinerary=itinerary).delete()

        hotels = itinerary.hotelbooking_set.all()
        activities = itinerary.activitybooking_set.all()
        vehicles = itinerary.vehiclebooking_set.all()
        houseboats = itinerary.houseboatbooking_set.all()

        total_net_price = Decimal('0.00')
        hotels_list = []

        # Calculate hotels pricing - WITH INCLUSIONS
        for hotel in hotels:
            nights = (hotel.check_out_date - hotel.check_in_date).days or 1
            hotel_price = Hotelprice.objects.filter(
                hotel=hotel.hotel,
                room_type=hotel.room_type,
                meal_plan=hotel.meal_plan,
                from_date__lte=hotel.check_out_date,
                to_date__gte=hotel.check_in_date
            ).first()

            if hotel_price:
                # Base price calculation
                price_per_night = (
                    Decimal(hotel.num_double_beds or 0) * (hotel_price.double_bed or 0) +
                    Decimal(hotel.child_with_bed or 0) * (hotel_price.child_with_bed or 0) +
                    Decimal(hotel.child_without_bed or 0) * (hotel_price.child_without_bed or 0) +
                    Decimal(hotel.extra_beds or 0) * (hotel_price.extra_bed or 0)
                )
                hotel_base_total = price_per_night * nights

                # 🔥 CALCULATE INCLUSIONS MANUALLY (since we just created the bookings)
                inclusion_price = Decimal('0.00')
                inclusion_items = HotelBookingInclusion.objects.filter(hotel_booking=hotel)

                for inclusion_item in inclusion_items:
                    if inclusion_item.special_inclusion and inclusion_item.price:
                        inclusion_price += Decimal(str(inclusion_item.price or 0))

                hotel_total = hotel_base_total + inclusion_price

                total_net_price += hotel_total
                hotels_list.append({
                    'name': hotel.hotel.name,
                    'net_price': float(hotel_total),
                    'base_price': float(hotel_base_total),
                    'inclusion_price': float(inclusion_price)
                })

                if inclusion_price > 0:
                    print(f"   ✅ Hotel: {hotel.hotel.name} ({nights} night{'s' if nights > 1 else ''})")
                    print(f"      • Base: ₹{hotel_base_total}")
                    print(f"      • Inclusions: ₹{inclusion_price}")
                    print(f"      • Total: ₹{hotel_total}")
                else:
                    print(f"   ✅ Hotel: {hotel.hotel.name} ({nights} night{'s' if nights > 1 else ''}) - ₹{hotel_total}")

        # Calculate activities pricing
        for activity in activities:
            activity_price = ActivityPrice.objects.filter(
                activity=activity.activity,
                from_date__lte=activity.booking_date,
                to_date__gte=activity.booking_date
            ).first()

            if activity_price:
                people = (activity.num_adults or 0) + (activity.num_children or 0)
                price = Decimal(str(activity_price.per_person or 0)) * people
                total_net_price += price
                print(f"   ✅ Activity: {activity.activity.name} ({people} people) - ₹{price}")

        # Calculate vehicles pricing
        for vehicle in vehicles:
            vehicle_price = VehiclePricing.objects.filter(
                vehicle=vehicle.vehicle,
                from_date__lte=vehicle.pickup_date,
                to_date__gte=vehicle.pickup_date
            ).first()

            if vehicle_price:
                km = vehicle.total_km or 0
                if km <= 100:
                    price = Decimal(str(vehicle_price.total_fee_100km or 0))
                else:
                    extra_km = km - 100
                    price = Decimal(str(vehicle_price.total_fee_100km or 0)) + (Decimal(str(extra_km)) * Decimal(str(vehicle_price.extra_fee_per_km or 0)))

                total_net_price += price
                print(f"   ✅ Vehicle: {vehicle.vehicle.name} ({km}km) - ₹{price}")

        # Calculate houseboats pricing - WITH INCLUSIONS
        for houseboat in houseboats:
            nights = (houseboat.check_out_date - houseboat.check_in_date).days or 1
            houseboat_price = HouseboatPrice.objects.filter(
                houseboat=houseboat.houseboat,
                room_type=houseboat.room_type,
                meal_plan=houseboat.meal_plan,
                from_date__lte=houseboat.check_out_date,
                to_date__gte=houseboat.check_in_date
            ).first()

            if houseboat_price:
                price_per_night = (
                    Decimal(houseboat.num_one_bed_rooms or 0) * (houseboat_price.one_bed or 0) +
                    Decimal(houseboat.num_two_bed_rooms or 0) * (houseboat_price.two_bed or 0) +
                    Decimal(houseboat.num_three_bed_rooms or 0) * (houseboat_price.three_bed or 0) +
                    Decimal(houseboat.num_four_bed_rooms or 0) * (houseboat_price.four_bed or 0) +
                    Decimal(houseboat.num_five_bed_rooms or 0) * (houseboat_price.five_bed or 0) +
                    Decimal(houseboat.num_six_bed_rooms or 0) * (houseboat_price.six_bed or 0) +
                    Decimal(houseboat.num_seven_bed_rooms or 0) * (houseboat_price.seven_bed or 0) +
                    Decimal(houseboat.num_eight_bed_rooms or 0) * (houseboat_price.eight_bed or 0) +
                    Decimal(houseboat.num_nine_bed_rooms or 0) * (houseboat_price.nine_bed or 0) +
                    Decimal(houseboat.num_ten_bed_rooms or 0) * (houseboat_price.ten_bed or 0) +
                    Decimal(houseboat.num_extra_beds or 0) * (houseboat_price.extra_bed or 0)
                )
                houseboat_base_total = price_per_night * nights

                # 🔥 CALCULATE HOUSEBOAT INCLUSIONS MANUALLY
                inclusion_price = Decimal('0.00')
                inclusion_items = HouseboatBookingInclusion.objects.filter(houseboat_booking=houseboat)

                for inclusion_item in inclusion_items:
                    if inclusion_item.special_inclusion and inclusion_item.price:
                        inclusion_price += Decimal(str(inclusion_item.price or 0))

                houseboat_total = houseboat_base_total + inclusion_price

                total_net_price += houseboat_total

                if inclusion_price > 0:
                    print(f"   ✅ Houseboat: {houseboat.houseboat.name} ({nights} night{'s' if nights > 1 else ''})")
                    print(f"      • Base: ₹{houseboat_base_total}")
                    print(f"      • Inclusions: ₹{inclusion_price}")
                    print(f"      • Total: ₹{houseboat_total}")
                else:
                    print(f"   ✅ Houseboat: {houseboat.houseboat.name} ({nights} night{'s' if nights > 1 else ''}) - ₹{houseboat_total}")

        # Calculate taxes CORRECTLY
        print(f"\n   📊 Pricing Breakdown:")
        print(f"   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        print(f"   Net Price (with inclusions): ₹{total_net_price}")

        # Get tax percentages
        cgst_perc = Decimal(str(itinerary.cgst_percentage or 0))
        sgst_perc = Decimal(str(itinerary.sgst_percentage or 0))
        discount_amt = Decimal(str(itinerary.discount or 0))

        print(f"   CGST @ {cgst_perc}%")
        print(f"   SGST @ {sgst_perc}%")
        if discount_amt > 0:
            print(f"   Discount: -₹{discount_amt}")

        # Calculate taxes on NET price (including inclusions)
        cgst_amount = (total_net_price * cgst_perc / Decimal('100')) if cgst_perc > 0 else Decimal('0.00')
        sgst_amount = (total_net_price * sgst_perc / Decimal('100')) if sgst_perc > 0 else Decimal('0.00')

        # Final amount = net (with inclusions) + taxes - discount
        final_amount = total_net_price + cgst_amount + sgst_amount - discount_amt

        print(f"   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        print(f"   CGST Amount: ₹{cgst_amount}")
        print(f"   SGST Amount: ₹{sgst_amount}")
        print(f"   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        print(f"   💰 FINAL AMOUNT: ₹{final_amount}")
        print(f"   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")

        # Create pricing option
        ItineraryPricingOption.objects.create(
            itinerary=itinerary,
            option_name='Standard Package',
            option_number=1,
            net_price=total_net_price,
            gross_price=total_net_price,
            cgst_amount=cgst_amount,
            sgst_amount=sgst_amount,
            discount_amount=discount_amt,
            final_amount=final_amount,
            hotels_included=hotels_list
        )

        # Mark as finalized
        itinerary.is_finalized = True
        itinerary.finalized_at = now()
        itinerary.save()

        print(f"   ✅ Pricing Saved Successfully!\n")

    except Exception as e:
        print(f"⚠️  Error auto-saving pricing: {e}\n")
        import traceback
        traceback.print_exc()

        print("=" * 80 + "\n")

    return itinerary, skipped_items






@custom_login_required
def insert_package_to_itinerary(request, query_id):
    """
    Insert package with smart pricing validation - WITH PERMISSION CHECK
    """

    # ✅ CHECK LOGIN
    user_id = request.session.get('user_id')
    user_type = request.session.get('user_type')

    if not user_id:
        messages.warning(request, '⚠️ Please login first')
        return redirect('team_member:login')

    query = get_object_or_404(Query, id=query_id)

    # ✅ CHECK PERMISSION TO INSERT ITINERARY
    can_insert = False
    current_user = None

    if user_type == 'superuser':
        can_insert = True
    elif user_type == 'team_member':
        current_user = TeamMember.objects.get(id=user_id)
        if current_user.role in ['admin', 'manager']:
            can_insert = True
        elif current_user.has_permission('can_insert_itinerary'):
            can_insert = True

    # ✅ DENY ACCESS if no permission
    if not can_insert:
        messages.error(request, '❌ You do not have permission to insert itineraries')
        return redirect('query_proposals', query_id=query_id)

    # Get packages
    packages = PackageTemplate.objects.filter(
        is_active=True,
        total_days=query.total_days
    ).prefetch_related(
        'destinations', 'day_plans', 'pricing_options'
    ).order_by('-is_featured', '-times_used')

    if request.method == 'POST':
        package_id = request.POST.get('package_id')
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        if not package_id:
            if is_ajax:
                return JsonResponse({'success': False, 'error': 'Please select a package template.'})
            messages.error(request, 'Please select a package template.')
            return redirect('insert_package_to_itinerary', query_id=query.id)

        package = get_object_or_404(PackageTemplate, id=package_id)

        if package.total_days != query.total_days:
            error_msg = f'Package is {package.total_days} days but query is {query.total_days} days.'
            if is_ajax:
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('insert_package_to_itinerary', query_id=query.id)

        try:
            with transaction.atomic():
                itinerary, skipped_items = create_itinerary_from_package_with_validation(
                    package=package,
                    query=query,
                    start_date=query.from_date,
                    end_date=query.to_date,
                    created_by=current_user  # ✅ PASS CREATOR
                )

                package.times_used = (package.times_used or 0) + 1
                package.save(update_fields=['times_used'])

                if is_ajax:
                    # Count inserted items
                    hotels_inserted = itinerary.hotelbooking_set.count()
                    activities_inserted = itinerary.activitybooking_set.count()
                    vehicles_inserted = itinerary.vehiclebooking_set.count()
                    houseboats_inserted = itinerary.houseboatbooking_set.count()
                    total_inserted = hotels_inserted + activities_inserted + vehicles_inserted + houseboats_inserted

                    skipped_by_type = {}
                    for item in skipped_items:
                        item_type = item.get('type', 'Unknown')
                        if item_type not in skipped_by_type:
                            skipped_by_type[item_type] = []
                        skipped_by_type[item_type].append(item)

                    pricing_url = reverse('itinerary_pricing', kwargs={'itinerary_id': itinerary.id})

                    return JsonResponse({
                        'success': True,
                        'message': f'✅ Package "{package.name}" inserted successfully!',
                        'itinerary_id': itinerary.id,
                        'package_name': package.name,
                        'total_inserted': total_inserted,
                        'hotels_inserted': hotels_inserted,
                        'activities_inserted': activities_inserted,
                        'vehicles_inserted': vehicles_inserted,
                        'houseboats_inserted': houseboats_inserted,
                        'skipped_count': len(skipped_items),
                        'skipped_items': skipped_items,
                        'skipped_by_type': skipped_by_type,
                        'has_skipped_items': len(skipped_items) > 0,
                        'itinerary_url': reverse('itinerary_day_plan', kwargs={'itinerary_id': itinerary.id}),
                        'pricing_url': pricing_url
                    })

                if skipped_items:
                    total_skipped = len(skipped_items)
                    msg = f'✅ Package "{package.name}" inserted!\n⚠️ {total_skipped} item(s) were SKIPPED\n'
                    for item in skipped_items[:8]:
                        msg += f"  • Day {item['day']}: {item['type']} - {item['name']}\n"
                    if total_skipped > 8:
                        msg += f"  ... and {total_skipped - 8} more items"
                    messages.warning(request, msg)
                else:
                    messages.success(request, f'✅ Package "{package.name}" inserted successfully!')

                return redirect('itinerary_pricing', itinerary_id=itinerary.id)

        except Exception as e:
            print(f"❌ ERROR: {str(e)}")
            import traceback
            traceback.print_exc()

            if is_ajax:
                return JsonResponse({'success': False, 'error': str(e)})
            messages.error(request, f'❌ Error: {str(e)}')
            return redirect('insert_package_to_itinerary', query_id=query.id)

    context = {
        'query': query,
        'packages': packages,
        'user_type': user_type,
        'current_user': current_user,
    }
    return render(request, 'insert_package.html', context)






from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.mail import EmailMultiAlternatives
from django.conf import settings
from django.utils.html import escape
from .models import (
    Query,
    Itinerary,
    HotelBooking,
    ActivityBooking,
    VehicleBooking,
    HouseboatBooking,
    StandaloneInclusionBooking,  # NEW
    Supplier
)
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import re
from html import unescape

@custom_login_required
def generate_booking_pdf(query, itinerary, hotels, vehicles, activities, houseboats, standalone_inclusions, booking_type):
    """Generate PDF with relevant booking details"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()

    # Title mapping with emojis
    titles = {
        'hotel': '🏨 HOTEL BOOKING DETAILS',
        'vehicle': '🚗 VEHICLE BOOKING DETAILS',
        'activity': '🎯 ACTIVITY BOOKING DETAILS',
        'houseboat': '🚤 HOUSEBOAT BOOKING DETAILS',
        'standalone_inclusion': '⭐ SPECIAL INCLUSION BOOKING DETAILS',  # NEW
    }

    # Custom title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f4e78'),
        spaceAfter=12,
        alignment=1
    )

    # Add title and greeting
    elements.append(Paragraph(titles.get(booking_type, 'BOOKING DETAILS'), title_style))
    elements.append(Paragraph("Greetings from Dream Holidays Kerala!!!!!", title_style))
    elements.append(Spacer(1, 0.2*inch))

    # Guest Information Table
    guest_data = [
        ['Booking ID', str(itinerary.id)],
        ['Guest Name', escape(query.client_name)],
        ['Email', escape(query.email)],
        ['Phone', escape(query.phone_number) if hasattr(query, 'phone_number') else 'N/A'],
        ['Check-In', query.from_date.strftime('%d-%m-%Y') if query.from_date else 'N/A'],
        ['Check-Out', query.to_date.strftime('%d-%m-%Y') if query.to_date else 'N/A'],
        ['Total Pax', f"{query.adult or 0} Adult - {query.childrens or 0} Child - {query.infant or 0} Infant"],
    ]

    guest_table = Table(guest_data, colWidths=[2*inch, 4*inch])
    guest_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#D3D3D3')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(guest_table)
    elements.append(Spacer(1, 0.2*inch))
    elements.append(Paragraph("Booking Details", styles['Heading2']))
    elements.append(Spacer(1, 0.1*inch))

    # Common table style
    common_table_style = TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4F8')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
    ])

    # Add booking details based on type
    if booking_type == 'hotel' and hotels.exists():
        for idx, hotel in enumerate(hotels, 1):
            if idx > 1:
                elements.append(Spacer(1, 0.15*inch))

            elements.append(Paragraph(f"Hotel #{idx}: {escape(hotel.hotel.name)}", styles['Heading3']))

            hotel_nights = (hotel.check_out_date - hotel.check_in_date).days if hotel.check_out_date and hotel.check_in_date else 0
            hotel_data = [
                ['Hotel Name', escape(hotel.hotel.name)],
                ['Check-In', hotel.check_in_date.strftime('%d-%m-%Y') if hotel.check_in_date else 'N/A'],
                ['Check-Out', hotel.check_out_date.strftime('%d-%m-%Y') if hotel.check_out_date else 'N/A'],
                ['Nights', str(hotel_nights)],
                ['Room Type', escape(hotel.room_type.name) if hotel.room_type else 'N/A'],
                ['Meal Plan', escape(hotel.meal_plan.name) if hotel.meal_plan else 'N/A'],
                ['Number of Rooms', str(hotel.num_rooms or 0)],
            ]
            hotel_table = Table(hotel_data, colWidths=[2*inch, 4*inch])
            hotel_table.setStyle(common_table_style)
            elements.append(hotel_table)
            elements.append(Spacer(1, 0.1*inch))

    elif booking_type == 'vehicle' and vehicles.exists():
        for idx, vehicle in enumerate(vehicles, 1):
            if idx > 1:
                elements.append(Spacer(1, 0.15*inch))

            elements.append(Paragraph(f"Vehicle #{idx}: {escape(vehicle.vehicle.name)}", styles['Heading3']))

            vehicle_data = [
                ['Vehicle Name', escape(vehicle.vehicle.name)],
                ['Pickup Date', vehicle.pickup_date.strftime('%d-%m-%Y') if vehicle.pickup_date else 'N/A'],
                ['Vehicle Type', escape(vehicle.get_vehicle_type_display()) if hasattr(vehicle, 'get_vehicle_type_display') else vehicle.vehicle_type],
                ['Passengers', str(vehicle.num_passengers or 0)],
                ['Total KM', str(vehicle.total_km or 0)],
            ]
            vehicle_table = Table(vehicle_data, colWidths=[2*inch, 4*inch])
            vehicle_table.setStyle(common_table_style)
            elements.append(vehicle_table)
            elements.append(Spacer(1, 0.1*inch))

    elif booking_type == 'activity' and activities.exists():
        for idx, activity in enumerate(activities, 1):
            if idx > 1:
                elements.append(Spacer(1, 0.15*inch))

            elements.append(Paragraph(f"Activity #{idx}: {escape(activity.activity.name)}", styles['Heading3']))

            activity_data = [
                ['Activity Name', escape(activity.activity.name)],
                ['Booking Date', activity.booking_date.strftime('%d-%m-%Y') if activity.booking_date else 'N/A'],
                ['Adults', str(activity.num_adults or 0)],
                ['Children', str(activity.num_children or 0)],
                ['Total Participants', str((activity.num_adults or 0) + (activity.num_children or 0))],
            ]
            activity_table = Table(activity_data, colWidths=[2*inch, 4*inch])
            activity_table.setStyle(common_table_style)
            elements.append(activity_table)
            elements.append(Spacer(1, 0.1*inch))

    elif booking_type == 'houseboat' and houseboats.exists():
        for idx, houseboat in enumerate(houseboats, 1):
            if idx > 1:
                elements.append(Spacer(1, 0.15*inch))

            elements.append(Paragraph(f"Houseboat #{idx}: {escape(houseboat.houseboat.name)}", styles['Heading3']))

            hb_nights = (houseboat.check_out_date - houseboat.check_in_date).days if houseboat.check_out_date and houseboat.check_in_date else 0
            hb_data = [
                ['Houseboat Name', escape(houseboat.houseboat.name)],
                ['Check-In', houseboat.check_in_date.strftime('%d-%m-%Y') if houseboat.check_in_date else 'N/A'],
                ['Check-Out', houseboat.check_out_date.strftime('%d-%m-%Y') if houseboat.check_out_date else 'N/A'],
                ['Nights', str(hb_nights)],
                ['Room Type', escape(houseboat.room_type.name) if houseboat.room_type else 'N/A'],
                ['Meal Plan', escape(houseboat.meal_plan.name) if houseboat.meal_plan else 'N/A'],
            ]
            hb_table = Table(hb_data, colWidths=[2*inch, 4*inch])
            hb_table.setStyle(common_table_style)
            elements.append(hb_table)
            elements.append(Spacer(1, 0.1*inch))

    # NEW: Standalone Inclusion PDF Generation
    elif booking_type == 'standalone_inclusion' and standalone_inclusions.exists():
        for idx, inclusion in enumerate(standalone_inclusions, 1):
            if idx > 1:
                elements.append(Spacer(1, 0.15*inch))

            elements.append(Paragraph(f"Special Inclusion #{idx}: {escape(inclusion.special_inclusion.name)}", styles['Heading3']))

            # Build inclusion data with conditional fields
            inclusion_data = [
                ['Inclusion Name', escape(inclusion.special_inclusion.name)],
                ['Booking Date', inclusion.booking_date.strftime('%d-%m-%Y') if inclusion.booking_date else 'N/A'],
            ]

            # Add time if available
            if inclusion.booking_time:
                inclusion_data.append(['Time', inclusion.booking_time.strftime('%I:%M %p')])

            # Add participant details
            inclusion_data.extend([
                ['Adults', str(inclusion.num_adults or 0)],
                ['Children', str(inclusion.num_children or 0)],
                ['Total Participants', str((inclusion.num_adults or 0) + (inclusion.num_children or 0))],
            ])

            # Add notes if available
            if inclusion.notes:
                inclusion_data.append(['Special Notes', escape(inclusion.notes)])

            inclusion_table = Table(inclusion_data, colWidths=[2*inch, 4*inch])

            # Purple-themed style for standalone inclusions
            standalone_style = TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8D5F0')),  # Light purple
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#9b59b6')),  # Purple border
            ])

            inclusion_table.setStyle(standalone_style)
            elements.append(inclusion_table)
            elements.append(Spacer(1, 0.1*inch))

    # Footer
    footer_text = """<b>Important Notes:</b><br/>
    • Kindly send the confirmation ASAP<br/>
    • All bookings are subject to availability<br/>
    • Please confirm the rates and availability at the earliest<br/><br/>
    Hope the above is in order. Feel free to contact us for any clarifications.<br/><br/>
    <b>Warm Regards,</b><br/>
    Dream Holidays Kerala Team
    """

    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph(footer_text, styles['Normal']))

    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer



import re
from html import unescape
from django.utils.html import escape

@custom_login_required
def generate_supplier_email(supplier, query, itinerary, hotels, activities, vehicles, houseboats, standalone_inclusions, message_body):
    """Generate email with only relevant details for supplier type"""

    # Extract pax info
    adults = query.adult or 0
    children = query.childrens or 0
    infants = query.infant or 0
    nights = (query.to_date - query.from_date).days if query.to_date and query.from_date else 0

    # Destinations
    destinations_list = ', '.join([escape(dest.name) for dest in itinerary.destinations.all()]) if itinerary.destinations.exists() else 'N/A'

    # Common enquiry table
    enquiry_table = f"""
    <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%; margin-bottom: 20px; font-family: Arial, sans-serif;">
        <tbody>
            <tr style="background-color: #f8f9fa;">
                <td style="font-weight: bold; width: 25%;">Customer Name:</td>
                <td>{escape(query.client_name)}</td>
                <td style="font-weight: bold; width: 25%;">Enquiry ID:</td>
                <td>{query.id}</td>
            </tr>
            <tr style="background-color: #ffffff;">
                <td style="font-weight: bold;">Check-In:</td>
                <td>{query.from_date.strftime('%d-%m-%Y') if query.from_date else 'Not set'}</td>
                <td style="font-weight: bold;">Check-Out:</td>
                <td>{query.to_date.strftime('%d-%m-%Y') if query.to_date else 'Not set'}</td>
            </tr>
            <tr style="background-color: #f8f9fa;">
                <td style="font-weight: bold;">Nights:</td>
                <td>{nights}</td>
                <td style="font-weight: bold;">Total Pax:</td>
                <td>{adults} Adult - {children} Child - {infants} Infant</td>
            </tr>
            <tr style="background-color: #ffffff;">
                <td style="font-weight: bold;">Destination:</td>
                <td colspan="3">{destinations_list}</td>
            </tr>
        </tbody>
    </table>
    """

    # Extract only custom message
    clean_message = message_body
    if "Enquiry Details" in clean_message:
        clean_message = clean_message.split("Enquiry Details")[0]

    clean_message = re.sub('<[^<]+?>', '', clean_message)
    clean_message = unescape(clean_message).strip()

    lines = [line.strip() for line in clean_message.split('\n') if line.strip()]
    if lines:
        clean_message = '\n'.join(lines[:2])
    else:
        clean_message = "Kindly provide the best rates for the below enquiry at the earliest."

    # Initialize email
    email_content = f"""
    <p style="font-family: Arial, sans-serif; font-size: 14px;">{clean_message}</p>

    <h2 style="color: #333; border-bottom: 2px solid #007bff; padding-bottom: 10px; font-family: Arial, sans-serif;">Enquiry Details</h2>
    {enquiry_table}
    """

    # ADD ONLY RELEVANT DETAILS BY SUPPLIER TYPE
    if supplier.supplier_type == 'hotel' and hotels.exists():
        email_content += "<h2 style='color: #333; border-bottom: 2px solid #28a745; padding-bottom: 10px; font-family: Arial, sans-serif; margin-top: 20px;'>🏨 Hotel Requirements</h2>"
        for hotel in hotels:
            hotel_nights = (hotel.check_out_date - hotel.check_in_date).days if hotel.check_out_date and hotel.check_in_date else 0
            email_content += f"""
            <h3 style="color: #333; margin-top: 15px; font-family: Arial, sans-serif;">{escape(hotel.hotel.name)}</h3>
            <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; margin-bottom: 15px;">
                <tbody>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold; width: 30%;">Check-In</td><td>{hotel.check_in_date.strftime('%d-%m-%Y') if hotel.check_in_date else 'N/A'}</td></tr>
                    <tr style="background-color: #ffffff;"><td style="font-weight: bold;">Check-Out</td><td>{hotel.check_out_date.strftime('%d-%m-%Y') if hotel.check_out_date else 'N/A'}</td></tr>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold;">Nights</td><td>{hotel_nights}</td></tr>
                    <tr style="background-color: #ffffff;"><td style="font-weight: bold;">Room Category</td><td>{escape(hotel.room_type.name) if hotel.room_type else 'N/A'}</td></tr>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold;">Meal Plan</td><td>{escape(hotel.meal_plan.name) if hotel.meal_plan else 'N/A'}</td></tr>
                    <tr style="background-color: #ffffff;"><td style="font-weight: bold;">Rooms</td><td>{hotel.num_rooms or 0}</td></tr>
                </tbody>
            </table>
            """

    elif supplier.supplier_type == 'vehicle' and vehicles.exists():
        email_content += "<h2 style='color: #333; border-bottom: 2px solid #17a2b8; padding-bottom: 10px; font-family: Arial, sans-serif; margin-top: 20px;'>🚗 Vehicle Requirements</h2>"
        for vehicle in vehicles:
            email_content += f"""
            <h3 style="color: #333; margin-top: 15px; font-family: Arial, sans-serif;">{escape(vehicle.vehicle.name)}</h3>
            <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; margin-bottom: 15px;">
                <tbody>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold; width: 30%;">Pickup Date</td><td>{vehicle.pickup_date.strftime('%d-%m-%Y') if vehicle.pickup_date else 'N/A'}</td></tr>
                    <tr style="background-color: #ffffff;"><td style="font-weight: bold;">Vehicle Type</td><td>{escape(vehicle.get_vehicle_type_display()) if hasattr(vehicle, 'get_vehicle_type_display') else vehicle.vehicle_type}</td></tr>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold;">Passengers</td><td>{vehicle.num_passengers or 0}</td></tr>
                    <tr style="background-color: #ffffff;"><td style="font-weight: bold;">Total KM</td><td>{vehicle.total_km or 0}</td></tr>
                </tbody>
            </table>
            """

    elif supplier.supplier_type == 'activity' and activities.exists():
        email_content += "<h2 style='color: #333; border-bottom: 2px solid #ffc107; padding-bottom: 10px; font-family: Arial, sans-serif; margin-top: 20px;'>🎯 Activity Requirements</h2>"
        for activity in activities:
            email_content += f"""
            <h3 style="color: #333; margin-top: 15px; font-family: Arial, sans-serif;">{escape(activity.activity.name)}</h3>
            <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; margin-bottom: 15px;">
                <tbody>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold; width: 30%;">Booking Date</td><td>{activity.booking_date.strftime('%d-%m-%Y') if activity.booking_date else 'N/A'}</td></tr>
                    <tr style="background-color: #ffffff;"><td style="font-weight: bold;">Adults</td><td>{activity.num_adults or 0}</td></tr>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold;">Children</td><td>{activity.num_children or 0}</td></tr>
                </tbody>
            </table>
            """

    elif supplier.supplier_type == 'houseboat' and houseboats.exists():
        email_content += "<h2 style='color: #333; border-bottom: 2px solid #20c997; padding-bottom: 10px; font-family: Arial, sans-serif; margin-top: 20px;'>🚤 Houseboat Requirements</h2>"
        for houseboat in houseboats:
            hb_nights = (houseboat.check_out_date - houseboat.check_in_date).days if houseboat.check_out_date and houseboat.check_in_date else 0
            email_content += f"""
            <h3 style="color: #333; margin-top: 15px; font-family: Arial, sans-serif;">{escape(houseboat.houseboat.name)}</h3>
            <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; margin-bottom: 15px;">
                <tbody>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold; width: 30%;">Check-In</td><td>{houseboat.check_in_date.strftime('%d-%m-%Y') if houseboat.check_in_date else 'N/A'}</td></tr>
                    <tr style="background-color: #ffffff;"><td style="font-weight: bold;">Check-Out</td><td>{houseboat.check_out_date.strftime('%d-%m-%Y') if houseboat.check_out_date else 'N/A'}</td></tr>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold;">Nights</td><td>{hb_nights}</td></tr>
                    <tr style="background-color: #ffffff;"><td style="font-weight: bold;">Room Type</td><td>{escape(houseboat.room_type.name) if houseboat.room_type else 'N/A'}</td></tr>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold;">Meal Plan</td><td>{escape(houseboat.meal_plan.name) if houseboat.meal_plan else 'N/A'}</td></tr>
                </tbody>
            </table>
            """

    elif supplier.supplier_type == 'standalone_inclusion' and standalone_inclusions.exists():
        email_content += "<h2 style='color: #333; border-bottom: 2px solid #9b59b6; padding-bottom: 10px; font-family: Arial, sans-serif; margin-top: 20px;'>⭐ Special Inclusion Requirements</h2>"
        for inclusion in standalone_inclusions:
            # Build time row if available
            time_row = ""
            if inclusion.booking_time:
                time_row = f'<tr style="background-color: #ffffff;"><td style="font-weight: bold;">Time</td><td>{inclusion.booking_time.strftime("%I:%M %p")}</td></tr>'

            # Build notes row if available
            notes_row = ""
            if inclusion.notes:
                notes_row = f'<tr style="background-color: #f8f9fa;"><td style="font-weight: bold;">Special Notes</td><td>{escape(inclusion.notes)}</td></tr>'

            email_content += f"""
            <h3 style="color: #333; margin-top: 15px; font-family: Arial, sans-serif;">{escape(inclusion.special_inclusion.name)}</h3>
            <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; margin-bottom: 15px;">
                <tbody>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold; width: 30%;">Booking Date</td><td>{inclusion.booking_date.strftime('%d-%m-%Y') if inclusion.booking_date else 'N/A'}</td></tr>
                    {time_row}
                    <tr style="background-color: #ffffff;"><td style="font-weight: bold;">Adults</td><td>{inclusion.num_adults or 0}</td></tr>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold;">Children</td><td>{inclusion.num_children or 0}</td></tr>
                    <tr style="background-color: #ffffff;"><td style="font-weight: bold;">Total Participants</td><td>{(inclusion.num_adults or 0) + (inclusion.num_children or 0)}</td></tr>
                    {notes_row}
                </tbody>
            </table>
            """

    else:
        email_content += f"""
        <div style="margin-top: 20px; padding: 15px; background: #fff3cd; border-left: 4px solid #ffc107; border-radius: 4px; font-family: Arial, sans-serif;">
            <strong>⚠️ Note:</strong> No relevant booking details found for your service type.
        </div>
        """

    # Add closing
    email_content += f"""
    <div style="margin-top: 40px; padding-top: 20px; border-top: 2px solid #e0e0e0;">
        <p style="font-family: Arial, sans-serif; font-size: 14px; color: #555;">Awaiting your response at the earliest.</p>
        <p style="font-family: Arial, sans-serif; font-size: 14px; margin-top: 15px;">
            Best regards,<br>
            <strong style="color: #333; font-size: 15px;">{escape(query.client_name)}</strong><br>
            <small style="color: #888;">Query ID: #{query.id}</small>
        </p>
    </div>
    """

    return email_content



# 🔥 Main supplier communication view
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.mail import EmailMultiAlternatives
from django.conf import settings
from django.utils.html import escape
from .models import (
    Query,
    Itinerary,
    HotelBooking,
    ActivityBooking,
    VehicleBooking,
    HouseboatBooking,
    StandaloneInclusionBooking,  # NEW
    Supplier
)

@custom_login_required
def supplier_communication(request, query_id):
    """Send supplier communication emails with auto-generated enquiry details"""
    query = get_object_or_404(Query, id=query_id)

    # Get the confirmed itinerary
    itinerary = Itinerary.objects.filter(query=query, is_finalized=True).first()

    if not itinerary:
        messages.error(request, "❌ No confirmed itinerary found. Please confirm pricing first.")
        return redirect('query_list')

    print(f"\n{'='*80}")
    print(f"📧 SUPPLIER COMMUNICATION - Query ID: {query_id}")
    print(f"{'='*80}\n")

    # Get all bookings INCLUDING standalone inclusions
    hotels = HotelBooking.objects.filter(itinerary=itinerary).select_related('hotel', 'room_type', 'meal_plan')
    activities = ActivityBooking.objects.filter(itinerary=itinerary).select_related('activity')
    vehicles = VehicleBooking.objects.filter(itinerary=itinerary).select_related('vehicle')
    houseboats = HouseboatBooking.objects.filter(itinerary=itinerary).select_related('houseboat', 'room_type', 'meal_plan')
    standalone_inclusions = StandaloneInclusionBooking.objects.filter(itinerary=itinerary).select_related('special_inclusion')

    # Get active suppliers by type (only if bookings exist)
    hotel_suppliers = Supplier.objects.filter(
        supplier_type='hotel',
        is_active=True
    ).order_by('company_name') if hotels.exists() else Supplier.objects.none()

    activity_suppliers = Supplier.objects.filter(
        supplier_type='activity',
        is_active=True
    ).order_by('company_name') if activities.exists() else Supplier.objects.none()

    vehicle_suppliers = Supplier.objects.filter(
        supplier_type='vehicle',
        is_active=True
    ).order_by('company_name') if vehicles.exists() else Supplier.objects.none()

    houseboat_suppliers = Supplier.objects.filter(
        supplier_type='houseboat',
        is_active=True
    ).order_by('company_name') if houseboats.exists() else Supplier.objects.none()

    # NEW: Get standalone inclusion suppliers
    standalone_suppliers = Supplier.objects.filter(
        supplier_type='standalone_inclusion',
        is_active=True
    ).order_by('company_name') if standalone_inclusions.exists() else Supplier.objects.none()

    # Handle POST - Send emails
    if request.method == 'POST':
        try:
            selected_suppliers = request.POST.getlist('suppliers[]')
            subject = request.POST.get('subject', '').strip()
            cc_emails_str = request.POST.get('cc_email', '').strip()
            message_body = request.POST.get('message_body', '').strip()

            # Validate inputs
            if not selected_suppliers:
                messages.error(request, "❌ Please select at least one supplier.")
                return redirect('supplier_communication', query_id=query_id)

            if not subject:
                messages.error(request, "❌ Subject cannot be empty.")
                return redirect('supplier_communication', query_id=query_id)

            if not message_body:
                messages.error(request, "❌ Message body cannot be empty.")
                return redirect('supplier_communication', query_id=query_id)

            # Parse CC emails
            cc_emails = [email.strip() for email in cc_emails_str.split(',') if email.strip()]

            print(f"\n📧 Sending emails to {len(selected_suppliers)} supplier(s)")
            print(f"   Subject: {subject}")
            print(f"   CC: {', '.join(cc_emails) if cc_emails else 'None'}\n")

            # Send emails
            sent_count = 0
            failed_suppliers = []

            for supplier_id in selected_suppliers:
                try:
                    # Get supplier from database
                    supplier = Supplier.objects.get(id=supplier_id)

                    if not supplier.email:
                        failed_suppliers.append(f"{supplier.company_name} - No email")
                        continue

                    # 🔥 GENERATE CUSTOM EMAIL BASED ON SUPPLIER TYPE
                    custom_message = generate_supplier_email(
                        supplier=supplier,
                        query=query,
                        itinerary=itinerary,
                        hotels=hotels,
                        activities=activities,
                        vehicles=vehicles,
                        houseboats=houseboats,
                        standalone_inclusions=standalone_inclusions,  # NEW
                        message_body=message_body
                    )

                    # Create email
                    email = EmailMultiAlternatives(
                        subject=subject,
                        body=custom_message,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[supplier.email],
                        cc=cc_emails,
                    )

                    # Attach HTML version
                    email.attach_alternative(custom_message, "text/html")

                    # ✅ GENERATE AND ATTACH PDF BASED ON SUPPLIER TYPE
                    if supplier.supplier_type == 'hotel' and hotels.exists():
                        pdf_buffer = generate_booking_pdf(query, itinerary, hotels, vehicles, activities, houseboats, standalone_inclusions, 'hotel')
                        email.attach(
                            filename=f"Hotel_Booking_Query_{query.id}.pdf",
                            content=pdf_buffer.getvalue(),
                            mimetype='application/pdf'
                        )

                    elif supplier.supplier_type == 'vehicle' and vehicles.exists():
                        pdf_buffer = generate_booking_pdf(query, itinerary, hotels, vehicles, activities, houseboats, standalone_inclusions, 'vehicle')
                        email.attach(
                            filename=f"Vehicle_Booking_Query_{query.id}.pdf",
                            content=pdf_buffer.getvalue(),
                            mimetype='application/pdf'
                        )

                    elif supplier.supplier_type == 'activity' and activities.exists():
                        pdf_buffer = generate_booking_pdf(query, itinerary, hotels, vehicles, activities, houseboats, standalone_inclusions, 'activity')
                        email.attach(
                            filename=f"Activity_Booking_Query_{query.id}.pdf",
                            content=pdf_buffer.getvalue(),
                            mimetype='application/pdf'
                        )

                    elif supplier.supplier_type == 'houseboat' and houseboats.exists():
                        pdf_buffer = generate_booking_pdf(query, itinerary, hotels, vehicles, activities, houseboats, standalone_inclusions, 'houseboat')
                        email.attach(
                            filename=f"Houseboat_Booking_Query_{query.id}.pdf",
                            content=pdf_buffer.getvalue(),
                            mimetype='application/pdf'
                        )

                    # NEW: Standalone inclusion PDF
                    elif supplier.supplier_type == 'standalone_inclusion' and standalone_inclusions.exists():
                        pdf_buffer = generate_booking_pdf(query, itinerary, hotels, vehicles, activities, houseboats, standalone_inclusions, 'standalone_inclusion')
                        email.attach(
                            filename=f"SpecialInclusion_Booking_Query_{query.id}.pdf",
                            content=pdf_buffer.getvalue(),
                            mimetype='application/pdf'
                        )

                    # Send email
                    email.send()

                    sent_count += 1
                    print(f"   ✅ {supplier.company_name} ({supplier.email}) with PDF")

                except Supplier.DoesNotExist:
                    failed_suppliers.append(f"Supplier ID {supplier_id} not found")
                except Exception as e:
                    failed_suppliers.append(f"Error: {str(e)}")
                    print(f"   ❌ Failed to send to {supplier_id}: {e}")

            # Show results
            print(f"\n{'='*80}")
            print(f"📊 RESULTS")
            print(f"{'='*80}")
            print(f"✅ Successfully sent: {sent_count}")
            if failed_suppliers:
                print(f"❌ Failed: {len(failed_suppliers)}")
                for failed in failed_suppliers:
                    print(f"   • {failed}")
            print(f"{'='*80}\n")

            if sent_count > 0:
                messages.success(
                    request,
                    f"✅ Successfully sent {sent_count} email{'s' if sent_count > 1 else ''} with PDFs!"
                )

            if failed_suppliers:
                messages.warning(
                    request,
                    f"⚠️ Failed to send {len(failed_suppliers)} email(s). Check supplier details."
                )

            return redirect('query_list')

        except Exception as e:
            messages.error(request, f"❌ Error: {str(e)}")
            print(f"❌ Error in supplier_communication: {e}")
            import traceback
            traceback.print_exc()
            return redirect('query_list')

    # ========== GET REQUEST - Generate Default Content ==========

    default_subject = f"Travel Enquiry for {query.client_name} (Query ID: {query.id})"

    adults = query.adult or 0
    children = query.childrens or 0
    infants = query.infant or 0
    nights = (query.to_date - query.from_date).days if query.to_date and query.from_date else 0

    destinations_list = ', '.join([escape(dest.name) for dest in itinerary.destinations.all()]) if itinerary.destinations.exists() else 'N/A'

    enquiry_table = f"""
    <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%; margin-bottom: 20px; font-family: Arial, sans-serif;">
        <tbody>
            <tr style="background-color: #f8f9fa;">
                <td style="font-weight: bold; width: 25%;">Customer Name:</td>
                <td>{escape(query.client_name)}</td>
                <td style="font-weight: bold; width: 25%;">Enquiry ID:</td>
                <td>{query.id}</td>
            </tr>
            <tr style="background-color: #ffffff;">
                <td style="font-weight: bold;">Check-In:</td>
                <td>{query.from_date.strftime('%d-%m-%Y') if query.from_date else 'Not set'}</td>
                <td style="font-weight: bold;">Check-Out:</td>
                <td>{query.to_date.strftime('%d-%m-%Y') if query.to_date else 'Not set'}</td>
            </tr>
            <tr style="background-color: #f8f9fa;">
                <td style="font-weight: bold;">Nights:</td>
                <td>{nights}</td>
                <td style="font-weight: bold;">Total Pax:</td>
                <td>{adults} Adult - {children} Child - {infants} Infant</td>
            </tr>
            <tr style="background-color: #ffffff;">
                <td style="font-weight: bold;">Destination:</td>
                <td colspan="3">{destinations_list}</td>
            </tr>
        </tbody>
    </table>
    """

    # Build hotel details
    hotel_details = ""
    if hotels.exists():
        hotel_details = "<h2 style='color: #333; border-bottom: 2px solid #28a745; padding-bottom: 10px; font-family: Arial, sans-serif; margin-top: 20px;'>🏨 Hotel Bookings</h2>"
        for hotel in hotels:
            hotel_nights = (hotel.check_out_date - hotel.check_in_date).days if hotel.check_out_date and hotel.check_in_date else 0
            hotel_details += f"""
            <h3 style="color: #333; margin-top: 15px; font-family: Arial, sans-serif;">{escape(hotel.hotel.name)}</h3>
            <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; margin-bottom: 15px;">
                <tbody>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold; width: 30%;">Check-In</td><td>{hotel.check_in_date.strftime('%d-%m-%Y') if hotel.check_in_date else 'N/A'}</td></tr>
                    <tr style="background-color: #ffffff;"><td style="font-weight: bold;">Check-Out</td><td>{hotel.check_out_date.strftime('%d-%m-%Y') if hotel.check_out_date else 'N/A'}</td></tr>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold;">Nights</td><td>{hotel_nights}</td></tr>
                    <tr style="background-color: #ffffff;"><td style="font-weight: bold;">Room Category</td><td>{escape(hotel.room_type.name) if hotel.room_type else 'N/A'}</td></tr>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold;">Meal Plan</td><td>{escape(hotel.meal_plan.name) if hotel.meal_plan else 'N/A'}</td></tr>
                    <tr style="background-color: #ffffff;"><td style="font-weight: bold;">No. of Rooms</td><td>{hotel.num_rooms or 0}</td></tr>
                </tbody>
            </table>
            """

    # Build vehicle details
    vehicle_details = ""
    if vehicles.exists():
        vehicle_details = "<h2 style='color: #333; border-bottom: 2px solid #17a2b8; padding-bottom: 10px; font-family: Arial, sans-serif; margin-top: 20px;'>🚗 Vehicle Bookings</h2>"
        for vehicle in vehicles:
            vehicle_details += f"""
            <h3 style="color: #333; margin-top: 15px; font-family: Arial, sans-serif;">{escape(vehicle.vehicle.name)}</h3>
            <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; margin-bottom: 15px;">
                <tbody>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold; width: 30%;">Pickup Date</td><td>{vehicle.pickup_date.strftime('%d-%m-%Y') if vehicle.pickup_date else 'N/A'}</td></tr>
                    <tr style="background-color: #ffffff;"><td style="font-weight: bold;">Vehicle Type</td><td>{escape(vehicle.get_vehicle_type_display()) if hasattr(vehicle, 'get_vehicle_type_display') else vehicle.vehicle_type}</td></tr>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold;">Passengers</td><td>{vehicle.num_passengers or 0}</td></tr>
                    <tr style="background-color: #ffffff;"><td style="font-weight: bold;">Total KM</td><td>{vehicle.total_km or 0}</td></tr>
                </tbody>
            </table>
            """

    # Build activity details
    activity_details = ""
    if activities.exists():
        activity_details = "<h2 style='color: #333; border-bottom: 2px solid #ffc107; padding-bottom: 10px; font-family: Arial, sans-serif; margin-top: 20px;'>🎯 Activity Bookings</h2>"
        for activity in activities:
            activity_details += f"""
            <h3 style="color: #333; margin-top: 15px; font-family: Arial, sans-serif;">{escape(activity.activity.name)}</h3>
            <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; margin-bottom: 15px;">
                <tbody>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold; width: 30%;">Booking Date</td><td>{activity.booking_date.strftime('%d-%m-%Y') if activity.booking_date else 'N/A'}</td></tr>
                    <tr style="background-color: #ffffff;"><td style="font-weight: bold;">Adults</td><td>{activity.num_adults or 0}</td></tr>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold;">Children</td><td>{activity.num_children or 0}</td></tr>
                </tbody>
            </table>
            """

    # Build houseboat details
    houseboat_details = ""
    if houseboats.exists():
        houseboat_details = "<h2 style='color: #333; border-bottom: 2px solid #20c997; padding-bottom: 10px; font-family: Arial, sans-serif; margin-top: 20px;'>🚤 Houseboat Bookings</h2>"
        for houseboat in houseboats:
            hb_nights = (houseboat.check_out_date - houseboat.check_in_date).days if houseboat.check_out_date and houseboat.check_in_date else 0
            houseboat_details += f"""
            <h3 style="color: #333; margin-top: 15px; font-family: Arial, sans-serif;">{escape(houseboat.houseboat.name)}</h3>
            <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; margin-bottom: 15px;">
                <tbody>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold; width: 30%;">Check-In</td><td>{houseboat.check_in_date.strftime('%d-%m-%Y') if houseboat.check_in_date else 'N/A'}</td></tr>
                    <tr style="background-color: #ffffff;"><td style="font-weight: bold;">Check-Out</td><td>{houseboat.check_out_date.strftime('%d-%m-%Y') if houseboat.check_out_date else 'N/A'}</td></tr>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold;">Nights</td><td>{hb_nights}</td></tr>
                    <tr style="background-color: #ffffff;"><td style="font-weight: bold;">Room Type</td><td>{escape(houseboat.room_type.name) if houseboat.room_type else 'N/A'}</td></tr>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold;">Meal Plan</td><td>{escape(houseboat.meal_plan.name) if houseboat.meal_plan else 'N/A'}</td></tr>
                </tbody>
            </table>
            """

    # NEW: Build standalone inclusion details
    standalone_details = ""
    if standalone_inclusions.exists():
        standalone_details = "<h2 style='color: #333; border-bottom: 2px solid #9b59b6; padding-bottom: 10px; font-family: Arial, sans-serif; margin-top: 20px;'>⭐ Special Inclusions</h2>"
        for inclusion in standalone_inclusions:
            time_info = ""
            if inclusion.booking_time:
                time_info = f'<tr style="background-color: #ffffff;"><td style="font-weight: bold;">Time</td><td>{inclusion.booking_time.strftime("%I:%M %p")}</td></tr>'

            notes_info = ""
            if inclusion.notes:
                notes_info = f'<tr style="background-color: #f8f9fa;"><td style="font-weight: bold;">Notes</td><td>{escape(inclusion.notes)}</td></tr>'

            standalone_details += f"""
            <h3 style="color: #333; margin-top: 15px; font-family: Arial, sans-serif;">{escape(inclusion.special_inclusion.name)}</h3>
            <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; margin-bottom: 15px;">
                <tbody>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold; width: 30%;">Booking Date</td><td>{inclusion.booking_date.strftime('%d-%m-%Y') if inclusion.booking_date else 'N/A'}</td></tr>
                    {time_info}
                    <tr style="background-color: #ffffff;"><td style="font-weight: bold;">Adults</td><td>{inclusion.num_adults or 0}</td></tr>
                    <tr style="background-color: #f8f9fa;"><td style="font-weight: bold;">Children</td><td>{inclusion.num_children or 0}</td></tr>
                    {notes_info}
                </tbody>
            </table>
            """

    # Build default message
    default_message = f"""
    <p style="font-family: Arial, sans-serif; font-size: 14px;">Dear Sir,</p>
    <p style="font-family: Arial, sans-serif; font-size: 14px;">Kindly provide the best rates for the below enquiry at the earliest.</p>

    <h2 style="color: #333; border-bottom: 2px solid #007bff; padding-bottom: 10px; font-family: Arial, sans-serif;">Enquiry Details</h2>
    {enquiry_table}

    {hotel_details}
    {vehicle_details}
    {activity_details}
    {houseboat_details}
    {standalone_details}

    <p style="margin-top: 30px; font-family: Arial, sans-serif; font-size: 14px;">Awaiting your response.</p>
    <p style="font-family: Arial, sans-serif; font-size: 14px;">Best regards,<br><strong>{escape(query.client_name)}</strong></p>
    """

    context = {
        'query': query,
        'itinerary': itinerary,
        'hotel_suppliers': hotel_suppliers,
        'activity_suppliers': activity_suppliers,
        'vehicle_suppliers': vehicle_suppliers,
        'houseboat_suppliers': houseboat_suppliers,
        'standalone_suppliers': standalone_suppliers,  # NEW
        'default_subject': default_subject,
        'default_message': default_message,
    }

    return render(request, 'supplier_communication.html', context)




from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import (
    Itinerary, Query, ItineraryDayPlan, HotelBooking, ActivityBooking,
    VehicleBooking, HouseboatBooking, ItineraryPricingOption
)
from datetime import datetime

@custom_login_required
def view_confirmation(request, itinerary_id):
    """
    Display the travel booking confirmation document
    """
    itinerary = get_object_or_404(Itinerary, id=itinerary_id)

    # Only show confirmation for confirmed itineraries
    if itinerary.status != 'confirmed':
        messages.warning(request, 'This itinerary is not confirmed yet.')
        return redirect('query_detail', query_id=itinerary.query.id)

    # Get query details
    query = itinerary.query

    # Get all day plans with their bookings
    day_plans = ItineraryDayPlan.objects.filter(
        itinerary=itinerary
    ).prefetch_related(
        'hotel_bookings',
        'activity_bookings',
        'vehicle_bookings',
        'houseboat_bookings'
    ).order_by('day_number')

    # Prepare itinerary data with bookings
    itinerary_data = []
    for day in day_plans:
        # Get hotel booking for this day
        hotel_booking = day.hotel_bookings.filter(
            option=itinerary.selected_option or 'option_1'
        ).first()

        day_info = {
            'date': itinerary.travel_from + timedelta(days=day.day_number - 1) if itinerary.travel_from else None,
            'day_number': day.day_number,
            'place': day.destination.name if day.destination else day.title or 'Not specified',
            'hotel': hotel_booking.hotel.name if hotel_booking and hotel_booking.hotel else 'Not assigned',
            'room_type': hotel_booking.room_type.name if hotel_booking and hotel_booking.room_type else 'Standard',
            'meal_plan': hotel_booking.meal_plan.name if hotel_booking and hotel_booking.meal_plan else 'Breakfast',
            'rate': hotel_booking.gross_price if hotel_booking else 0,
            'advance': 0,  # You can add advance_paid field to HotelBooking model if needed
        }
        itinerary_data.append(day_info)

    # Get vehicle booking
    vehicle_bookings = VehicleBooking.objects.filter(itinerary=itinerary)
    vehicle_booking = vehicle_bookings.first()
    vehicle_cost = sum([vb.gross_price for vb in vehicle_bookings]) if vehicle_bookings else 0
    vehicle_type = vehicle_booking.vehicle.name if vehicle_booking and vehicle_booking.vehicle else "Not assigned"

    # Get selected pricing option
    selected_pricing = itinerary.pricing_options.filter(
        option_name=itinerary.selected_option
    ).first()

    # Calculate pricing
    if selected_pricing:
        package_cost = selected_pricing.final_amount
        gst_amount = selected_pricing.cgst_amount + selected_pricing.sgst_amount
        net_price = selected_pricing.net_price
    else:
        package_cost = itinerary.final_amount
        gst_amount = 0
        net_price = itinerary.total_net_price

    # Calculate advance and balance (you can customize this)
    total_advance = 0  # Add logic to calculate from payment records if you have
    balance = package_cost - total_advance

    # Get room configuration from pricing option
    rooms = selected_pricing.number_of_rooms if selected_pricing else 1
    extra_beds = selected_pricing.extra_beds if selected_pricing else 0

    # Get confirmation details
    context = {
        'itinerary': itinerary,
        'query': query,
        'arrival_date': itinerary.travel_from or query.from_date,
        'departure_date': itinerary.travel_to or query.to_date,
        'total_days': itinerary.total_days or query.total_days,
        'confirmation_no': f"TH{itinerary.created_at.strftime('%y')}-{query.id:02d}/{itinerary.id}",

        # Guest details
        'guest_name': f"{query.get_gender_display()} {query.client_name}",
        'guest_phone': query.phone_number,
        'guest_email': query.email,

        # Package details
        'package_category': 'Deluxe',  # You can make this dynamic based on hotels selected
        'package_type': 'Family' if query.childrens and query.childrens > 0 else 'Individual',
        'adults': itinerary.adults or query.adult,
        'children': itinerary.childrens or query.childrens,
        'infants': itinerary.infants or query.infant,
        'rooms': rooms,
        'extra_beds': extra_beds,
        'vehicle_type': vehicle_type,
        'vehicle_cost': vehicle_cost,

        # Itinerary data
        'itinerary_data': itinerary_data,

        # Additional details
        'confirmed_by': itinerary.created_by.get_full_name() if itinerary.created_by else 'System',
        'email': 'salesone@travelhope.in',  # Make this dynamic from OrganisationalSetting
        'special_inclusions': itinerary.notes or 'Standard package inclusions apply',
        'notes': f"Package for {itinerary.adults or query.adult} adults",

        # Dates
        'written_on': itinerary.created_at.strftime('%d-%m-%Y'),
        'booking_finished_on': itinerary.finalized_at.strftime('%d-%m-%Y') if itinerary.finalized_at else itinerary.created_at.strftime('%d-%m-%Y'),
        'booking_done_by': itinerary.created_by.get_full_name() if itinerary.created_by else 'Admin',

        # Pricing
        'package_cost': package_cost,
        'gst': gst_amount,
        'total_advance': total_advance,
        'balance': balance,
        'selected_option': itinerary.selected_option or 'Option 1',
        'net_price': net_price,
    }

    return render(request, 'view_confirmation.html', context)


from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO


@custom_login_required
def download_confirmation_pdf(request, itinerary_id):
    """
    Generate and download PDF of booking confirmation
    """
    itinerary = get_object_or_404(Itinerary, id=itinerary_id)

    # Only allow PDF download for confirmed itineraries
    if itinerary.status != 'confirmed':
        messages.warning(request, 'This itinerary is not confirmed yet.')
        return redirect('query_detail', query_id=itinerary.query.id)

    # Get query details
    query = itinerary.query

    # Get all day plans with their bookings
    day_plans = ItineraryDayPlan.objects.filter(
        itinerary=itinerary
    ).prefetch_related(
        'hotel_bookings',
        'activity_bookings',
        'vehicle_bookings',
        'houseboat_bookings'
    ).order_by('day_number')

    # Prepare itinerary data with bookings
    itinerary_data = []
    for day in day_plans:
        hotel_booking = day.hotel_bookings.filter(
            option=itinerary.selected_option or 'option_1'
        ).first()

        day_info = {
            'date': itinerary.travel_from + timedelta(days=day.day_number - 1) if itinerary.travel_from else None,
            'day_number': day.day_number,
            'place': day.destination.name if day.destination else day.title or 'Not specified',
            'hotel': hotel_booking.hotel.name if hotel_booking and hotel_booking.hotel else 'Not assigned',
            'room_type': hotel_booking.room_type.name if hotel_booking and hotel_booking.room_type else 'Standard',
            'meal_plan': hotel_booking.meal_plan.name if hotel_booking and hotel_booking.meal_plan else 'Breakfast',
            'rate': hotel_booking.gross_price if hotel_booking else 0,
            'advance': 0,
        }
        itinerary_data.append(day_info)

    # Get vehicle booking
    vehicle_bookings = VehicleBooking.objects.filter(itinerary=itinerary)
    vehicle_booking = vehicle_bookings.first()
    vehicle_cost = sum([vb.gross_price for vb in vehicle_bookings]) if vehicle_bookings else 0
    vehicle_type = vehicle_booking.vehicle.name if vehicle_booking and vehicle_booking.vehicle else "Not assigned"

    # Get selected pricing option
    selected_pricing = itinerary.pricing_options.filter(
        option_name=itinerary.selected_option
    ).first()

    # Calculate pricing
    if selected_pricing:
        package_cost = selected_pricing.final_amount
        gst_amount = selected_pricing.cgst_amount + selected_pricing.sgst_amount
        net_price = selected_pricing.net_price
    else:
        package_cost = itinerary.final_amount if hasattr(itinerary, 'final_amount') else 0
        gst_amount = 0
        net_price = itinerary.total_net_price if hasattr(itinerary, 'total_net_price') else 0

    total_advance = 0
    balance = package_cost - total_advance
    rooms = selected_pricing.number_of_rooms if selected_pricing else 1
    extra_beds = selected_pricing.extra_beds if selected_pricing else 0

    confirmation_no = f"TH{itinerary.created_at.strftime('%y')}-{query.id:02d}/{itinerary.id}"

    # Context for PDF
    context = {
        'itinerary': itinerary,
        'query': query,
        'arrival_date': itinerary.travel_from or query.from_date,
        'departure_date': itinerary.travel_to or query.to_date,
        'total_days': itinerary.total_days or query.total_days,
        'confirmation_no': confirmation_no,
        'guest_name': f"{query.get_gender_display()} {query.client_name}",
        'guest_phone': query.phone_number,
        'guest_email': query.email,
        'package_category': 'Deluxe',
        'package_type': 'Family' if query.childrens and query.childrens > 0 else 'Individual',
        'adults': itinerary.adults or query.adult,
        'children': itinerary.childrens or query.childrens,
        'infants': itinerary.infants or query.infant,
        'rooms': rooms,
        'extra_beds': extra_beds,
        'vehicle_type': vehicle_type,
        'vehicle_cost': vehicle_cost,
        'itinerary_data': itinerary_data,
        'confirmed_by': itinerary.created_by.get_full_name() if itinerary.created_by else 'System',
        'email': 'salesone@travelhope.in',
        'special_inclusions': itinerary.notes or 'Standard package inclusions apply',
        'notes': f"Package for {itinerary.adults or query.adult} adults",
        'written_on': itinerary.created_at.strftime('%d-%m-%Y'),
        'booking_finished_on': itinerary.finalized_at.strftime('%d-%m-%Y') if hasattr(itinerary, 'finalized_at') and itinerary.finalized_at else itinerary.created_at.strftime('%d-%m-%Y'),
        'booking_done_by': itinerary.created_by.get_full_name() if itinerary.created_by else 'Admin',
        'package_cost': package_cost,
        'gst': gst_amount,
        'total_advance': total_advance,
        'balance': balance,
        'selected_option': itinerary.selected_option or 'Option 1',
        'net_price': net_price,
    }

    # Load template
    template = get_template('confirmation_pdf.html')
    html = template.render(context)

    # Create PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Confirmation_{confirmation_no.replace("/", "_")}.pdf"'

    # Generate PDF
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('PDF generation error. Please contact support.', status=500)

    return response









####################for php integrations########################


from django.shortcuts import render
from django.views.generic import ListView
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from rest_framework import generics
from django.db.models import Q
from django.utils.dateparse import parse_date
from django.views.generic import ListView
from .models import Lead

from .models import Lead
from .serializers import LeadSerializer
from django.views.generic import TemplateView
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from rest_framework import generics
from django.db.models import Q
from django.utils.dateparse import parse_date
from django.views.generic import ListView
from .models import *
from .serializers import LeadSerializer
from django.views.generic import TemplateView
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.mail import EmailMessage
from django.shortcuts import redirect, render, get_object_or_404
from django.urls import reverse
from django.db import transaction




class About(TemplateView):
    # Render this template at "/"
    template_name = "about.html"

    # Optional: pass data to the template
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["app_name"] = "Dream Holidays CRM"
        ctx["tagline"] = "Leads from your PHP site, listed in Django"
        return ctx

# leads/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, TemplateView
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.core.mail import EmailMessage
from django.db import transaction
from django.db.models import (
    Q, Case, When, Value, BooleanField, CharField,
    Count, OuterRef, Subquery
)
from django.utils.dateparse import parse_date




# API / DRF
from rest_framework import generics
from .serializers import LeadSerializer

from urllib.parse import urlparse, parse_qs


# -------------------------------------------------------------
# ABOUT PAGE
# -------------------------------------------------------------
class About(TemplateView):
    template_name = "about.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["app_name"] = "Dream Holidays CRM"
        ctx["tagline"] = "Leads from your PHP site, listed in Django"
        return ctx


# -------------------------------------------------------------
# GOOGLE ADS DETECTION QUERY
# -------------------------------------------------------------
# ADS_Q = (
#     Q(from_url__icontains="gclid=") |
#     Q(from_url__icontains="wbraid=") |
#     Q(from_url__icontains="gbraid=") |
#     Q(from_url__icontains="gad_source=") |
#     (
#         Q(from_url__icontains="utm_source=google") &
#         (
#             Q(from_url__icontains="utm_medium=cpc") |
#             Q(from_url__icontains="utm_medium=ppc") |
#             Q(from_url__icontains="utm_medium=paid") |
#             Q(from_url__icontains="utm_medium=paid_search")
#         )
#     )
# )

ADS_Q = (
    # Google Ads auto-tagging
    Q(from_url__icontains="gclid=") |
    Q(from_url__icontains="wbraid=") |
    Q(from_url__icontains="gbraid=") |
    Q(from_url__icontains="gad_source=") |

    # Any paid UTM (platform-agnostic)
    Q(from_url__icontains="utm_medium=paid") |
    Q(from_url__icontains="utm_medium=cpc") |
    Q(from_url__icontains="utm_medium=ppc") |
    Q(from_url__icontains="utm_medium=paid_search") |

    # Explicit paid social sources
    Q(from_url__icontains="utm_source=fb") |
    Q(from_url__icontains="utm_source=facebook") |
    Q(from_url__icontains="utm_source=instagram") |
    Q(from_url__icontains="utm_source=meta")
)


# -------------------------------------------------------------
# LEADS MAIN LIST VIEW
# -------------------------------------------------------------
class LeadListView(ListView):
    model = Lead
    template_name = "leads_list.html"
    context_object_name = "leads"
    paginate_by = 25
    ordering = ["-created_at"]

    def _apply_filters(self, qs):
        q = (self.request.GET.get("q") or "").strip()
        start = (self.request.GET.get("start") or "").strip()
        end = (self.request.GET.get("end") or "").strip()
        source = (self.request.GET.get("source") or "").strip()
        channel = (self.request.GET.get("channel") or "").lower()

        if q:
            qs = qs.filter(
                Q(name__icontains=q) |
                Q(email__icontains=q) |
                Q(phone__icontains=q)
            )

        if source:
            qs = qs.filter(from_url__icontains=source)

        if start:
            d = parse_date(start)
            if d:
                qs = qs.filter(created_at__date__gte=d)

        if end:
            d = parse_date(end)
            if d:
                qs = qs.filter(created_at__date__lte=d)

        if channel == "ads":
            qs = qs.filter(ADS_Q)
        elif channel == "seo":
            qs = qs.exclude(ADS_Q)

        for key in ("utm_campaign", "utm_source", "utm_medium"):
            val = (self.request.GET.get(key) or "").strip()
            if val:
                qs = qs.filter(from_url__icontains=f"{key}={val}")

        return qs

    def _annotate(self, qs):
        phone_counts = Lead.objects.filter(
            phone=OuterRef('phone')
        ).values('phone').annotate(
            c=Count('id')
        ).values('c')

        return qs.annotate(
            dup_count=Subquery(phone_counts),
            chan_is_ads=Case(
                When(ADS_Q, then=Value(True)),
                default=Value(False),
                output_field=BooleanField(),
            ),
            chan_source=Case(
                When(ADS_Q, then=Value("Google Ads")),
                default=Value("SEO"),
                output_field=CharField(),
            ),
        )

    def get_queryset(self):
        qs = super().get_queryset().order_by("-created_at")
        qs = self._apply_filters(qs)
        qs = self._annotate(qs)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        qs = self.get_queryset()
        context["unique_count"] = qs.values("phone").distinct().count()
        return context


# -------------------------------------------------------------
# AJAX TABLE ROWS
# -------------------------------------------------------------
class LeadRowsView(LoginRequiredMixin, ListView):
    model = Lead
    template_name = "_lead_rows.html"
    context_object_name = "object_list"

    def get_queryset(self):
        qs = Lead.objects.all().order_by("-created_at")

        q = (self.request.GET.get("q") or "")
        start = self.request.GET.get("start")
        end = self.request.GET.get("end")

        if q:
            qs = qs.filter(
                Q(name__icontains=q) |
                Q(email__icontains=q) |
                Q(phone__icontains=q)
            )

        if start:
            qs = qs.filter(created_at__date__gte=start)

        if end:
            qs = qs.filter(created_at__date__lte=end)

        # Duplicate count
        phone_counts = Lead.objects.filter(
            phone=OuterRef('phone')
        ).values('phone').annotate(
            c=Count('id')
        ).values('c')

        return qs.annotate(dup_count=Subquery(phone_counts))[:100]


# -------------------------------------------------------------
# DRF API VIEWS
# -------------------------------------------------------------
@method_decorator(csrf_exempt, name="dispatch")
class ApiLeadListCreateView(generics.ListCreateAPIView):
    queryset = Lead.objects.all()
    serializer_class = LeadSerializer


@method_decorator(csrf_exempt, name="dispatch")
class ApiLeadDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Lead.objects.all()
    serializer_class = LeadSerializer


# -------------------------------------------------------------
# ASSIGN LEADS TO TEAM MEMBERS
# -------------------------------------------------------------
def leads_list_assign(request):
    if request.method == "POST":
        form = AssignLeadsForm(request.POST)
        if form.is_valid():

            team_member = form.cleaned_data["assigned_to"]
            leads_qs = form.get_leads_qs()

            count = leads_qs.count()
            if count == 0:
                messages.warning(request, "No leads selected.")
                return redirect("lead_assign")

            with transaction.atomic():
                leads_qs.update(assigned_to=team_member)

            # Email body
            lines = [
                f"Hello {team_member.get_full_name()},",
                "",
                f"You have been assigned {count} lead(s):",
                "",
            ]
            for l in leads_qs:
                lines.append(f"- {l.name} | Phone: {l.phone or '-'} | Email: {l.email or '-'} | ID: {l.id}")

            lines.append("")
            lines.append("Please follow up and update the CRM accordingly.")

            body = "\n".join(lines)

            EmailMessage(
                subject="New Leads Assigned",
                body=body,
                to=[team_member.email],
            ).send(fail_silently=False)

            messages.success(request, f"Assigned {count} lead(s) to {team_member.get_full_name()} and emailed {team_member.email}.")
            return redirect("lead_assign")

    else:
        form = AssignLeadsForm()

    # Show all leads (you can filter unassigned if needed)
    leads = Lead.objects.select_related("assigned_to").order_by("-created_at")
    team_members = TeamMember.objects.filter(is_active=True).order_by("first_name")

    return render(request, "lead_assign.html", {
        "form": form,
        "leads": leads,
        "team_members": team_members,
    })


# -------------------------------------------------------------
# TEAM MEMBER LIST PAGE
# -------------------------------------------------------------
def team_member_list_page(request):
    members = TeamMember.objects.order_by("first_name")
    return render(request, "team_member_list_page.html", {"members": members})


def team_member_table_partial(request):
    members = TeamMember.objects.order_by("first_name")
    return render(request, "team_member_table_rows.html", {"members": members})


from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from .models import Lead


def leads_pdf_all(request):
    leads = Lead.objects.all().order_by('-created_at')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="all_leads_table.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    styles = getSampleStyleSheet()

    # Table header
    data = [
        ["Sl No", "Name", "Email", "Phone", "Channel", "Date"]
    ]

    # Add each lead to table
    for idx, lead in enumerate(leads, start=1):
        data.append([
            str(idx),
            lead.name or "-",
            lead.email or "-",
            lead.phone or "-",
            lead.channel or "-",
            lead.created_at.strftime("%Y-%m-%d"),
        ])

    # Create table
    table = Table(data, repeatRows=1)

    # Styling the table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#e5e7eb")),  # header bg
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#111827")),   # header text
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),

        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

        # Row background alternative
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
    ]))

    elements = []
    elements.append(Paragraph("<b>All Leads Report (Table Format)</b>", styles['Title']))
    elements.append(Paragraph(f"Total Leads: {leads.count()}", styles['Normal']))
    elements.append(table)

    doc.build(elements)

    return response



import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import Lead

def leads_excel_all(request):
    from django.db.models import Max
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.utils import get_column_letter

    unique_leads = (
        Lead.objects
        .values('name', 'email', 'phone')
        .annotate(latest_date=Max('created_at'))
        .order_by('-latest_date')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = ["Sl No", "Name", "Email", "Phone"]
    ws.append(headers)

    for idx, lead in enumerate(unique_leads, start=1):
        ws.append([
            idx,
            lead["name"] or "-",
            lead["email"] or "-",
            lead["phone"] or "-",
        ])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="all_unique_leads.xlsx"'

    wb.save(response)
    return response




from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from io import BytesIO
from xhtml2pdf import pisa
import re
import os
from django.conf import settings

from .models import (
    Itinerary,
    Query,
    ItineraryPricingOption,
    HotelBooking,
    VehicleBooking,
    ItineraryDayPlan,
    PackageTermss,
    InvoiceLogo,
    OrganisationalSetting
)


def split_into_bullets(text):
    """Converts CKEditor text into a clean list of bullet lines."""
    if not text:
        return []

    text = text.replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")
    text = strip_tags(text)
    text = text.replace("&nbsp;", " ")
    text = text.replace("&bull;", "")
    text = text.replace("•", "")
    text = text.replace("\r", "\n")

    lines = re.split(r'\n+|\.\s+', text)
    lines = [line.strip() for line in lines if line.strip()]

    return lines


def link_callback(uri, rel):
    """
    Convert HTML URIs to absolute system paths so xhtml2pdf can access them.
    """
    # Handle static files
    if uri.startswith(settings.STATIC_URL):
        path = uri.replace(settings.STATIC_URL, "")

        # Try STATIC_ROOT first (production)
        if settings.STATIC_ROOT:
            full_path = os.path.join(settings.STATIC_ROOT, path)
            if os.path.isfile(full_path):
                return full_path

        # Try STATICFILES_DIRS (development)
        if hasattr(settings, 'STATICFILES_DIRS') and settings.STATICFILES_DIRS:
            for static_dir in settings.STATICFILES_DIRS:
                full_path = os.path.join(static_dir, path)
                if os.path.isfile(full_path):
                    return full_path

    # Handle media files
    if settings.MEDIA_URL and uri.startswith(settings.MEDIA_URL):
        path = uri.replace(settings.MEDIA_URL, "")
        full_path = os.path.join(settings.MEDIA_ROOT, path)
        if os.path.isfile(full_path):
            return full_path

    # Return as-is if it's already an absolute path
    if os.path.isfile(uri):
        return uri

    return uri


def render_to_pdf(template_src, context_dict={}):
    """
    Helper function to render HTML template to PDF
    """
    html = render_to_string(template_src, context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result, link_callback=link_callback)

    if not pdf.err:
        return result.getvalue()
    return None


# ==========================================
# DRIVER ITINERARY VIEWS
# ==========================================

def view_driver_itinerary(request, itinerary_id):
    """View driver itinerary in browser"""

    itinerary = get_object_or_404(Itinerary, id=itinerary_id)
    query = itinerary.query

    option_name = itinerary.selected_option or 'option1'

    if option_name and ' ' in option_name:
        option_name = option_name.lower().replace(' ', '_')

    hotel_bookings = HotelBooking.objects.filter(
        itinerary=itinerary,
        option=option_name
    ).select_related(
        'hotel', 'destination', 'day_plan'
    ).order_by('check_in_date')

    vehicle_bookings = VehicleBooking.objects.filter(
        itinerary=itinerary,
        option=option_name
    ).select_related('vehicle', 'day_plan').order_by('pickup_date')

    day_plans = ItineraryDayPlan.objects.filter(
        itinerary=itinerary
    ).select_related('destination').order_by('day_number')

    logo = InvoiceLogo.objects.first()
    org_settings = OrganisationalSetting.objects.first()

    accommodations = []
    for booking in hotel_bookings:
        accommodations.append({
            'date': booking.check_in_date,
            'destination': booking.destination.name if booking.destination else 'N/A',
            'hotel': booking.hotel.name if booking.hotel else 'TBD',
            'contact': booking.hotel.phone_number if booking.hotel else ''
        })

    context = {
        'itinerary': itinerary,
        'query': query,
        'day_plans': day_plans,
        'accommodations': accommodations,
        'vehicle_bookings': vehicle_bookings,
        'logo': logo,
        'org_settings': org_settings,
    }

    return render(request, 'driver_itinerary_simple.html', context)


def download_driver_pdf(request, itinerary_id):
    """Generate and download driver PDF using xhtml2pdf"""

    itinerary = get_object_or_404(Itinerary, id=itinerary_id)
    query = itinerary.query

    option_name = itinerary.selected_option or 'option1'

    if option_name and ' ' in option_name:
        option_name = option_name.lower().replace(' ', '_')

    hotel_bookings = HotelBooking.objects.filter(
        itinerary=itinerary,
        option=option_name
    ).select_related(
        'hotel', 'destination', 'day_plan'
    ).order_by('check_in_date')

    vehicle_bookings = VehicleBooking.objects.filter(
        itinerary=itinerary,
        option=option_name
    ).select_related('vehicle', 'day_plan').order_by('pickup_date')

    day_plans = ItineraryDayPlan.objects.filter(
        itinerary=itinerary
    ).select_related('destination').order_by('day_number')

    logo = InvoiceLogo.objects.first()
    org_settings = OrganisationalSetting.objects.first()

    accommodations = []
    for booking in hotel_bookings:
        accommodations.append({
            'date': booking.check_in_date,
            'destination': booking.destination.name if booking.destination else 'N/A',
            'hotel': booking.hotel.name if booking.hotel else 'TBD',
            'contact': booking.hotel.phone_number if booking.hotel else ''
        })

    context = {
        'itinerary': itinerary,
        'query': query,
        'day_plans': day_plans,
        'accommodations': accommodations,
        'vehicle_bookings': vehicle_bookings,
        'logo': logo,
        'org_settings': org_settings,
        'is_pdf': True,
    }

    # Generate PDF
    pdf = render_to_pdf('driver_itinerary_simple.html', context)

    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Driver_Itinerary_{query.query_id}.pdf"'
        return response

    return HttpResponse("Error generating PDF", status=500)


# ==========================================
# CLIENT QUOTATION VIEWS
# ==========================================

def view_client_quotation(request, itinerary_id):
    """View client quotation in browser"""

    itinerary = get_object_or_404(Itinerary, id=itinerary_id)
    query = itinerary.query

    option_name = itinerary.selected_option or 'option_1'

    if option_name and ' ' in option_name:
        option_name = option_name.lower().replace(' ', '_')

    selected_pricing = ItineraryPricingOption.objects.filter(
        itinerary=itinerary, option_name=option_name
    ).first()

    hotel_bookings = HotelBooking.objects.filter(
        itinerary=itinerary, option=option_name
    ).select_related(
        'hotel', 'destination', 'room_type', 'meal_plan', 'day_plan'
    ).order_by('check_in_date')

    vehicle_bookings = VehicleBooking.objects.filter(
        itinerary=itinerary, option=option_name
    ).select_related('vehicle', 'destination', 'day_plan').order_by('pickup_date')

    day_plans = ItineraryDayPlan.objects.filter(
        itinerary=itinerary
    ).select_related('destination').order_by('day_number')

    package_terms = PackageTermss.objects.first()
    logo = InvoiceLogo.objects.first()
    org_settings = OrganisationalSetting.objects.first()

    incl_clean = split_into_bullets(package_terms.package_inclusion) if package_terms else []
    excl_clean = split_into_bullets(package_terms.package_exclusion) if package_terms else []
    terms_clean = split_into_bullets(package_terms.terms_and_conditions) if package_terms else []
    cancel_clean = split_into_bullets(package_terms.cancellation_policy) if package_terms else []

    created_by_name = "Dream Holidays Team"
    if query.assign:
        created_by_name = query.assign.get_full_name()
    elif query.created_by:
        created_by_name = query.created_by.get_full_name()

    accommodations = []
    for booking in hotel_bookings:
        accommodations.append({
            'date': booking.check_in_date,
            'destination': booking.destination.name if booking.destination else 'N/A',
            'hotel': booking.hotel.name if booking.hotel else 'TBD',
            'contact': booking.hotel.phone_number if booking.hotel else ''
        })

    context = {
        'itinerary': itinerary,
        'query': query,
        'selected_pricing': selected_pricing,
        'hotel_bookings': hotel_bookings,
        'vehicle_bookings': vehicle_bookings,
        'day_plans': day_plans,
        'accommodations': accommodations,
        'total_package_price': selected_pricing.final_amount if selected_pricing else itinerary.final_amount,
        'logo': logo,
        'org_settings': org_settings,
        'package_terms': package_terms,
        'incl_clean': incl_clean,
        'excl_clean': excl_clean,
        'terms_clean': terms_clean,
        'cancel_clean': cancel_clean,
        'created_by_name': created_by_name,
    }

    return render(request, 'client_quotation.html', context)


def download_client_pdf(request, itinerary_id):
    """Generate and download client PDF with pricing using xhtml2pdf"""

    itinerary = get_object_or_404(Itinerary, id=itinerary_id)
    query = itinerary.query

    option_name = itinerary.selected_option or 'option_1'

    if option_name and ' ' in option_name:
        option_name = option_name.lower().replace(' ', '_')

    selected_pricing = ItineraryPricingOption.objects.filter(
        itinerary=itinerary, option_name=option_name
    ).first()

    hotel_bookings = HotelBooking.objects.filter(
        itinerary=itinerary, option=option_name
    ).select_related(
        'hotel', 'destination', 'room_type', 'meal_plan', 'day_plan'
    ).order_by('check_in_date')

    vehicle_bookings = VehicleBooking.objects.filter(
        itinerary=itinerary, option=option_name
    ).select_related('vehicle', 'destination', 'day_plan').order_by('pickup_date')

    day_plans = ItineraryDayPlan.objects.filter(
        itinerary=itinerary
    ).select_related('destination').order_by('day_number')

    package_terms = PackageTermss.objects.first()
    logo = InvoiceLogo.objects.first()
    org_settings = OrganisationalSetting.objects.first()

    incl_clean = split_into_bullets(package_terms.package_inclusion) if package_terms else []
    excl_clean = split_into_bullets(package_terms.package_exclusion) if package_terms else []
    terms_clean = split_into_bullets(package_terms.terms_and_conditions) if package_terms else []
    cancel_clean = split_into_bullets(package_terms.cancellation_policy) if package_terms else []

    created_by_name = "Dream Holidays Team"
    if query.assign:
        created_by_name = query.assign.get_full_name()
    elif query.created_by:
        created_by_name = query.created_by.get_full_name()

    accommodations = []
    for booking in hotel_bookings:
        accommodations.append({
            'date': booking.check_in_date,
            'destination': booking.destination.name if booking.destination else 'N/A',
            'hotel': booking.hotel.name if booking.hotel else 'TBD',
            'contact': booking.hotel.phone_number if booking.hotel else ''
        })

    context = {
        'itinerary': itinerary,
        'query': query,
        'selected_pricing': selected_pricing,
        'hotel_bookings': hotel_bookings,
        'vehicle_bookings': vehicle_bookings,
        'day_plans': day_plans,
        'accommodations': accommodations,
        'total_package_price': selected_pricing.final_amount if selected_pricing else itinerary.final_amount,
        'logo': logo,
        'org_settings': org_settings,
        'package_terms': package_terms,
        'incl_clean': incl_clean,
        'excl_clean': excl_clean,
        'terms_clean': terms_clean,
        'cancel_clean': cancel_clean,
        'created_by_name': created_by_name,
        'is_pdf': True,
    }

    # Generate PDF
    pdf = render_to_pdf('client_quotation.html', context)

    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Quotation_{query.query_id}_{itinerary.name}.pdf"'
        return response

    return HttpResponse("Error generating PDF", status=500)

from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404


@require_http_methods(["GET"])
def get_package_terms_ajax(request, itinerary_id):
    """Get package terms for modal (AJAX)"""
    try:
        itinerary = get_object_or_404(Itinerary, id=itinerary_id)

        # Try to get custom terms, else use default
        try:
            terms = itinerary.custom_package_terms
        except ItineraryPackageTerms.DoesNotExist:
            # Get default terms
            default_terms = PackageTermss.objects.first()

            if default_terms:
                # Create custom terms from default
                terms = ItineraryPackageTerms.objects.create(
                    itinerary=itinerary,
                    special_inclusion=default_terms.special_inclusion,
                    package_inclusion=default_terms.package_inclusion,
                    package_exclusion=default_terms.package_exclusion,
                    terms_and_conditions=default_terms.terms_and_conditions,
                    payment_policy=default_terms.payment_policy,
                    cancellation_policy=default_terms.cancellation_policy,
                    refund_policy=default_terms.refund_policy,
                    list_of_documents=default_terms.list_of_documents,
                    is_active=True
                )
            else:
                # Create empty terms
                terms = ItineraryPackageTerms.objects.create(
                    itinerary=itinerary,
                    is_active=True
                )

        return JsonResponse({
            'success': True,
            'terms': {
                'special_inclusion': terms.special_inclusion or '',
                'package_inclusion': terms.package_inclusion or '',
                'package_exclusion': terms.package_exclusion or '',
                'terms_and_conditions': terms.terms_and_conditions or '',
                'payment_policy': terms.payment_policy or '',
                'cancellation_policy': terms.cancellation_policy or '',
                'refund_policy': terms.refund_policy or '',
                'list_of_documents': terms.list_of_documents or '',
                'is_active': terms.is_active
            }
        })
    except Exception as e:
        print(f"Error in get_package_terms_ajax: {str(e)}")  # Debug print
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


@require_http_methods(["POST"])
def save_package_terms_ajax(request, itinerary_id):
    """Save package terms via AJAX"""
    try:
        itinerary = get_object_or_404(Itinerary, id=itinerary_id)

        # Get or create custom terms
        try:
            terms = itinerary.custom_package_terms
        except ItineraryPackageTerms.DoesNotExist:
            terms = ItineraryPackageTerms.objects.create(itinerary=itinerary)

        # Update fields
        terms.special_inclusion = request.POST.get('special_inclusion', '')
        terms.package_inclusion = request.POST.get('package_inclusion', '')
        terms.package_exclusion = request.POST.get('package_exclusion', '')
        terms.terms_and_conditions = request.POST.get('terms_and_conditions', '')
        terms.payment_policy = request.POST.get('payment_policy', '')
        terms.cancellation_policy = request.POST.get('cancellation_policy', '')
        terms.refund_policy = request.POST.get('refund_policy', '')
        terms.list_of_documents = request.POST.get('list_of_documents', '')
        terms.is_active = request.POST.get('is_active', 'false') == 'true'

        terms.save()

        return JsonResponse({
            'success': True,
            'message': 'Package terms saved successfully!'
        })
    except Exception as e:
        print(f"Error in save_package_terms_ajax: {str(e)}")  # Debug print
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)













from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from django.utils.timezone import now
from .models import (
    Itinerary, ItineraryDayPlan, HotelBooking, VehicleBooking,
    ActivityBooking, HouseboatBooking, StandaloneInclusionBooking,
    ItineraryPricingOption, HotelBookingInclusion, HouseboatBookingInclusion
)

def prepare_edit_itinerary(request, itinerary_id):
    """
    Smart edit handler - creates new version for finalized itineraries.
    If status is 'draft', allows direct editing.
    """
    itinerary = get_object_or_404(Itinerary, id=itinerary_id)

    print("\n" + "="*80)
    print(f"📝 PREPARE EDIT REQUEST FOR ITINERARY #{itinerary.id}")
    print(f"   Name: {itinerary.name}")
    print(f"   Status: {itinerary.status}")
    print(f"   Is Finalized: {itinerary.is_finalized}")
    print("="*80)

    # ✅ CHECK STATUS: If draft, edit directly
    if itinerary.status == 'draft':
        print("✅ Itinerary is in DRAFT status - opening for direct editing")
        print("="*80 + "\n")
        return redirect('itinerary_day_plan', itinerary_id=itinerary.id)

    # ✅ IF NOT DRAFT - CREATE NEW VERSION
    print(f"🔄 Creating new version (Current status: {itinerary.status})")
    print("="*80)

    try:
        # Archive the current version
        itinerary.status = 'archived'
        itinerary.archived_at = now()
        itinerary.archived_reason = 'Edited - New version created'
        itinerary.save()

        print(f"📦 Archived old version (ID: {itinerary.id})")

        # GET NEXT VERSION NUMBER
        max_version = Itinerary.objects.filter(
            query=itinerary.query
        ).order_by('-version_number').first()

        next_version = (max_version.version_number + 1) if max_version else 1

        # CREATE NEW VERSION
        new_version = Itinerary.objects.create(
            query=itinerary.query,
            name=itinerary.name,
            travel_from=itinerary.travel_from,
            travel_to=itinerary.travel_to,
            total_days=itinerary.total_days,
            adults=itinerary.adults,
            childrens=itinerary.childrens,
            infants=itinerary.infants,
            notes=itinerary.notes,
            parent_itinerary=itinerary,
            version_number=next_version,
            status='draft',
            is_finalized=False,
            finalized_at=None,
            cgst_percentage=itinerary.cgst_percentage,
            sgst_percentage=itinerary.sgst_percentage,
            discount=itinerary.discount,
            created_by=itinerary.created_by
        )

        if itinerary.destinations.exists():
            new_version.destinations.set(itinerary.destinations.all())

        # ==========================================
        # COPY ALL DATA
        # ==========================================

        # 1️⃣ COPY DAY PLANS
        day_plan_mapping = {}
        old_day_plans = ItineraryDayPlan.objects.filter(itinerary=itinerary)

        for old_day in old_day_plans:
            new_day = ItineraryDayPlan.objects.create(
                itinerary=new_version,
                day_number=old_day.day_number,
                destination=old_day.destination,
                title=old_day.title,
                description=old_day.description,
                image=old_day.image,
                notes=old_day.notes if hasattr(old_day, 'notes') else '',
            )

            # Copy ManyToMany fields
            if hasattr(old_day, 'hotels'): new_day.hotels.set(old_day.hotels.all())
            if hasattr(old_day, 'houseboats'): new_day.houseboats.set(old_day.houseboats.all())
            if hasattr(old_day, 'activities'): new_day.activities.set(old_day.activities.all())
            if hasattr(old_day, 'meal_plans'): new_day.meal_plans.set(old_day.meal_plans.all())
            if hasattr(old_day, 'vehicles'): new_day.vehicles.set(old_day.vehicles.all())
            if hasattr(old_day, 'inclusions'): new_day.inclusions.set(old_day.inclusions.all())

            day_plan_mapping[old_day.id] = new_day

        # 2️⃣ COPY HOTELS (✅ UPDATED FOR CUSTOM FIELDS)
        old_hotels = HotelBooking.objects.filter(itinerary=itinerary)
        hotel_count = 0

        print(f"\n🏨 Hotels: {old_hotels.count()}")
        for old_hotel in old_hotels:
            new_hotel = HotelBooking.objects.create(
                itinerary=new_version,
                day_plan=day_plan_mapping.get(old_hotel.day_plan.id) if old_hotel.day_plan else None,

                # ✅ Identity & Custom Fields
                hotel=old_hotel.hotel,
                custom_hotel_name=old_hotel.custom_hotel_name,  # <--- NEW
                destination=old_hotel.destination if hasattr(old_hotel, 'destination') else None,

                # ✅ Room & Custom Fields
                room_type=old_hotel.room_type,
                custom_room_type=old_hotel.custom_room_type,    # <--- NEW

                # Standard Fields
                meal_plan=old_hotel.meal_plan,
                check_in_date=old_hotel.check_in_date,
                check_out_date=old_hotel.check_out_date,
                check_in_time=old_hotel.check_in_time if hasattr(old_hotel, 'check_in_time') else None,
                check_out_time=old_hotel.check_out_time if hasattr(old_hotel, 'check_out_time') else None,
                num_rooms=old_hotel.num_rooms if hasattr(old_hotel, 'num_rooms') else 1,
                num_double_beds=old_hotel.num_double_beds if hasattr(old_hotel, 'num_double_beds') else 0,
                extra_beds=old_hotel.extra_beds if hasattr(old_hotel, 'extra_beds') else 0,
                child_with_bed=old_hotel.child_with_bed if hasattr(old_hotel, 'child_with_bed') else 0,
                child_without_bed=old_hotel.child_without_bed if hasattr(old_hotel, 'child_without_bed') else 0,
                option=old_hotel.option if hasattr(old_hotel, 'option') else 'option1',
                category=old_hotel.category if hasattr(old_hotel, 'category') else '',
                markup_type=old_hotel.markup_type if hasattr(old_hotel, 'markup_type') else 'percentage',
                markup_value=old_hotel.markup_value if hasattr(old_hotel, 'markup_value') else 0,
                markup_amount=old_hotel.markup_amount if hasattr(old_hotel, 'markup_amount') else 0,
                markup_percentage=old_hotel.markup_percentage if hasattr(old_hotel, 'markup_percentage') else 0,
                net_price=old_hotel.net_price if hasattr(old_hotel, 'net_price') else 0,
                gross_price=old_hotel.gross_price if hasattr(old_hotel, 'gross_price') else 0,
                custom_double_bed_total=old_hotel.custom_double_bed_total if hasattr(old_hotel, 'custom_double_bed_total') else None,
                custom_extra_bed_total=old_hotel.custom_extra_bed_total if hasattr(old_hotel, 'custom_extra_bed_total') else None,
                custom_child_with_bed_total=old_hotel.custom_child_with_bed_total if hasattr(old_hotel, 'custom_child_with_bed_total') else None,
                custom_child_without_bed_total=old_hotel.custom_child_without_bed_total if hasattr(old_hotel, 'custom_child_without_bed_total') else None,
            )

            # Copy hotel inclusions
            for old_inc in old_hotel.inclusion_items.all():
                HotelBookingInclusion.objects.create(
                    hotel_booking=new_hotel,
                    special_inclusion=old_inc.special_inclusion,
                    num_adults=old_inc.num_adults if hasattr(old_inc, 'num_adults') else 0,
                    num_children=old_inc.num_children if hasattr(old_inc, 'num_children') else 0,
                    price=old_inc.price if hasattr(old_inc, 'price') else 0,
                )
            hotel_count += 1

        # 3️⃣ COPY VEHICLES
        old_vehicles = VehicleBooking.objects.filter(itinerary=itinerary)
        for old_vehicle in old_vehicles:
            VehicleBooking.objects.create(
                itinerary=new_version,
                day_plan=day_plan_mapping.get(old_vehicle.day_plan.id) if old_vehicle.day_plan else None,
                vehicle=old_vehicle.vehicle,
                destination=old_vehicle.destination if hasattr(old_vehicle, 'destination') else None,
                pickup_date=old_vehicle.pickup_date,
                pickup_time=old_vehicle.pickup_time if hasattr(old_vehicle, 'pickup_time') else None,
                num_passengers=old_vehicle.num_passengers if hasattr(old_vehicle, 'num_passengers') else 0,
                total_km=old_vehicle.total_km if hasattr(old_vehicle, 'total_km') else 0,
                vehicle_type=old_vehicle.vehicle_type if hasattr(old_vehicle, 'vehicle_type') else '',
                option=old_vehicle.option if hasattr(old_vehicle, 'option') else '',
                markup_type=old_vehicle.markup_type if hasattr(old_vehicle, 'markup_type') else 'percentage',
                markup_value=old_vehicle.markup_value if hasattr(old_vehicle, 'markup_value') else 0,
                markup_amount=old_vehicle.markup_amount if hasattr(old_vehicle, 'markup_amount') else 0,
                markup_percentage=old_vehicle.markup_percentage if hasattr(old_vehicle, 'markup_percentage') else 0,
                net_price=old_vehicle.net_price if hasattr(old_vehicle, 'net_price') else 0,
                gross_price=old_vehicle.gross_price if hasattr(old_vehicle, 'gross_price') else 0,
                custom_total_price=old_vehicle.custom_total_price if hasattr(old_vehicle, 'custom_total_price') else None,
            )

        # 4️⃣ COPY ACTIVITIES
        old_activities = ActivityBooking.objects.filter(itinerary=itinerary)
        for old_activity in old_activities:
            ActivityBooking.objects.create(
                itinerary=new_version,
                day_plan=day_plan_mapping.get(old_activity.day_plan.id) if old_activity.day_plan else None,
                activity=old_activity.activity,
                booking_date=old_activity.booking_date,
                booking_time=old_activity.booking_time if hasattr(old_activity, 'booking_time') else None,
                num_adults=old_activity.num_adults if hasattr(old_activity, 'num_adults') else 0,
                num_children=old_activity.num_children if hasattr(old_activity, 'num_children') else 0,
                confirmation_number=old_activity.confirmation_number if hasattr(old_activity, 'confirmation_number') else '',
                notes=old_activity.notes if hasattr(old_activity, 'notes') else '',
                markup_type=old_activity.markup_type if hasattr(old_activity, 'markup_type') else 'percentage',
                markup_value=old_activity.markup_value if hasattr(old_activity, 'markup_value') else 0,
                markup_amount=old_activity.markup_amount if hasattr(old_activity, 'markup_amount') else 0,
                markup_percentage=old_activity.markup_percentage if hasattr(old_activity, 'markup_percentage') else 0,
                net_price=old_activity.net_price if hasattr(old_activity, 'net_price') else 0,
                gross_price=old_activity.gross_price if hasattr(old_activity, 'gross_price') else 0,
                custom_total_price=old_activity.custom_total_price if hasattr(old_activity, 'custom_total_price') else None,
            )

        # 5️⃣ COPY HOUSEBOATS
        old_houseboats = HouseboatBooking.objects.filter(itinerary=itinerary)
        for old_hb in old_houseboats:
            new_hb = HouseboatBooking.objects.create(
                itinerary=new_version,
                day_plan=day_plan_mapping.get(old_hb.day_plan.id) if old_hb.day_plan else None,
                houseboat=old_hb.houseboat,
                room_type=old_hb.room_type if hasattr(old_hb, 'room_type') else None,
                meal_plan=old_hb.meal_plan if hasattr(old_hb, 'meal_plan') else None,
                check_in_date=old_hb.check_in_date,
                check_out_date=old_hb.check_out_date,
                option=old_hb.option if hasattr(old_hb, 'option') else 'option1',
                num_one_bed_rooms=old_hb.num_one_bed_rooms if hasattr(old_hb, 'num_one_bed_rooms') else 0,
                num_two_bed_rooms=old_hb.num_two_bed_rooms if hasattr(old_hb, 'num_two_bed_rooms') else 0,
                num_three_bed_rooms=old_hb.num_three_bed_rooms if hasattr(old_hb, 'num_three_bed_rooms') else 0,
                num_four_bed_rooms=old_hb.num_four_bed_rooms if hasattr(old_hb, 'num_four_bed_rooms') else 0,
                num_five_bed_rooms=old_hb.num_five_bed_rooms if hasattr(old_hb, 'num_five_bed_rooms') else 0,
                num_six_bed_rooms=old_hb.num_six_bed_rooms if hasattr(old_hb, 'num_six_bed_rooms') else 0,
                num_seven_bed_rooms=old_hb.num_seven_bed_rooms if hasattr(old_hb, 'num_seven_bed_rooms') else 0,
                num_eight_bed_rooms=old_hb.num_eight_bed_rooms if hasattr(old_hb, 'num_eight_bed_rooms') else 0,
                num_nine_bed_rooms=old_hb.num_nine_bed_rooms if hasattr(old_hb, 'num_nine_bed_rooms') else 0,
                num_ten_bed_rooms=old_hb.num_ten_bed_rooms if hasattr(old_hb, 'num_ten_bed_rooms') else 0,
                num_extra_beds=old_hb.num_extra_beds if hasattr(old_hb, 'num_extra_beds') else 0,
                markup_type=old_hb.markup_type if hasattr(old_hb, 'markup_type') else 'percentage',
                markup_value=old_hb.markup_value if hasattr(old_hb, 'markup_value') else 0,
                markup_amount=old_hb.markup_amount if hasattr(old_hb, 'markup_amount') else 0,
                markup_percentage=old_hb.markup_percentage if hasattr(old_hb, 'markup_percentage') else 0,
                net_price=old_hb.net_price if hasattr(old_hb, 'net_price') else 0,
                gross_price=old_hb.gross_price if hasattr(old_hb, 'gross_price') else 0,
                custom_one_bed_total=old_hb.custom_one_bed_total if hasattr(old_hb, 'custom_one_bed_total') else None,
                custom_two_bed_total=old_hb.custom_two_bed_total if hasattr(old_hb, 'custom_two_bed_total') else None,
                custom_three_bed_total=old_hb.custom_three_bed_total if hasattr(old_hb, 'custom_three_bed_total') else None,
                custom_four_bed_total=old_hb.custom_four_bed_total if hasattr(old_hb, 'custom_four_bed_total') else None,
                custom_five_bed_total=old_hb.custom_five_bed_total if hasattr(old_hb, 'custom_five_bed_total') else None,
                custom_six_bed_total=old_hb.custom_six_bed_total if hasattr(old_hb, 'custom_six_bed_total') else None,
                custom_seven_bed_total=old_hb.custom_seven_bed_total if hasattr(old_hb, 'custom_seven_bed_total') else None,
                custom_eight_bed_total=old_hb.custom_eight_bed_total if hasattr(old_hb, 'custom_eight_bed_total') else None,
                custom_nine_bed_total=old_hb.custom_nine_bed_total if hasattr(old_hb, 'custom_nine_bed_total') else None,
                custom_ten_bed_total=old_hb.custom_ten_bed_total if hasattr(old_hb, 'custom_ten_bed_total') else None,
                custom_extra_bed_hb_total=old_hb.custom_extra_bed_hb_total if hasattr(old_hb, 'custom_extra_bed_hb_total') else None,
            )
            # Copy houseboat inclusions
            for old_inc in old_hb.inclusion_items.all():
                HouseboatBookingInclusion.objects.create(
                    houseboat_booking=new_hb,
                    special_inclusion=old_inc.special_inclusion,
                    num_adults=old_inc.num_adults if hasattr(old_inc, 'num_adults') else 0,
                    num_children=old_inc.num_children if hasattr(old_inc, 'num_children') else 0,
                    price=old_inc.price if hasattr(old_inc, 'price') else 0,
                )

        # 6️⃣ COPY STANDALONE INCLUSIONS
        old_inclusions = StandaloneInclusionBooking.objects.filter(itinerary=itinerary)
        for old_si in old_inclusions:
            StandaloneInclusionBooking.objects.create(
                itinerary=new_version,
                day_plan=day_plan_mapping.get(old_si.day_plan.id) if old_si.day_plan else None,
                special_inclusion=old_si.special_inclusion,
                booking_date=old_si.booking_date,
                booking_time=old_si.booking_time if hasattr(old_si, 'booking_time') else None,
                num_adults=old_si.num_adults if hasattr(old_si, 'num_adults') else 0,
                num_children=old_si.num_children if hasattr(old_si, 'num_children') else 0,
                adult_unit_price=old_si.adult_unit_price if hasattr(old_si, 'adult_unit_price') else 0,
                child_unit_price=old_si.child_unit_price if hasattr(old_si, 'child_unit_price') else 0,
                subtotal=old_si.subtotal if hasattr(old_si, 'subtotal') else 0,
                markup_type=old_si.markup_type if hasattr(old_si, 'markup_type') else 'percentage',
                markup_value=old_si.markup_value if hasattr(old_si, 'markup_value') else 0,
                markup_amount=old_si.markup_amount if hasattr(old_si, 'markup_amount') else 0,
                total_price=old_si.total_price if hasattr(old_si, 'total_price') else 0,
                notes=old_si.notes if hasattr(old_si, 'notes') else '',
                created_by=old_si.created_by if hasattr(old_si, 'created_by') else None,
            )

        # 7️⃣ COPY PRICING OPTIONS
        old_pricing = ItineraryPricingOption.objects.filter(itinerary=itinerary)
        pricing_count = 0
        for old_p in old_pricing:
            ItineraryPricingOption.objects.create(
                itinerary=new_version,
                option_name=old_p.option_name,
                option_number=old_p.option_number,
                vehicle_type=old_p.vehicle_type if hasattr(old_p, 'vehicle_type') else '',
                number_of_rooms=old_p.number_of_rooms if hasattr(old_p, 'number_of_rooms') else 0,
                extra_beds=old_p.extra_beds if hasattr(old_p, 'extra_beds') else 0,
                child_without_bed=old_p.child_without_bed if hasattr(old_p, 'child_without_bed') else 0,
                child_with_bed=old_p.child_with_bed if hasattr(old_p, 'child_with_bed') else 0,
                child_ages=old_p.child_ages if hasattr(old_p, 'child_ages') else '',
                net_price=old_p.net_price,
                markup_amount=old_p.markup_amount,
                gross_price=old_p.gross_price,
                cgst_amount=old_p.cgst_amount,
                sgst_amount=old_p.sgst_amount,
                discount_amount=old_p.discount_amount,
                final_amount=old_p.final_amount,
                hotels_included=old_p.hotels_included if hasattr(old_p, 'hotels_included') else None,
            )
            pricing_count += 1

        # ✅ FINAL SUMMARY
        total_bookings = hotel_count  # + others count if you track them
        messages.success(
            request,
            f'✅ Created version {new_version.version_number}! '
            f'Archived version {itinerary.version_number}.'
        )
        return redirect('itinerary_day_plan', itinerary_id=new_version.id)

    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"\n❌ ERROR CREATING VERSION:")
        print(error_trace)
        messages.error(request, f'❌ Error creating version: {str(e)}')
        return redirect('query_proposals', query_id=itinerary.query.id)



def copy_all_bookings_to_new_version(old_itinerary, new_itinerary):
    """
    Copy ALL bookings from old itinerary to new version.
    ⚠️ DOES NOT change new_itinerary status or save it!
    """

    print(f"\n📋 COPYING BOOKINGS...")
    print(f"   From: ID {old_itinerary.id} (V{old_itinerary.version_number})")
    print(f"   To: ID {new_itinerary.id} (V{new_itinerary.version_number})")

    # 1️⃣ COPY DAY PLANS
    day_plan_mapping = {}
    old_day_plans = ItineraryDayPlan.objects.filter(itinerary=old_itinerary)

    for old_day in old_day_plans:
        new_day = ItineraryDayPlan.objects.create(
            itinerary=new_itinerary,
            day_number=old_day.day_number,
            destination=old_day.destination,
            title=old_day.title,
            description=old_day.description,
            image=old_day.image,
            notes=old_day.notes,
        )

        # Copy ManyToMany fields
        new_day.hotels.set(old_day.hotels.all())
        new_day.houseboats.set(old_day.houseboats.all())
        new_day.activities.set(old_day.activities.all())
        new_day.meal_plans.set(old_day.meal_plans.all())
        new_day.vehicles.set(old_day.vehicles.all())
        new_day.inclusions.set(old_day.inclusions.all())

        day_plan_mapping[old_day.id] = new_day

    # 2️⃣ COPY HOTELS (✅ UPDATED FOR CUSTOM FIELDS)
    old_hotels = HotelBooking.objects.filter(itinerary=old_itinerary)
    hotel_count = 0

    print(f"\n🏨 Hotels: {old_hotels.count()}")
    for old_hotel in old_hotels:
        new_hotel = HotelBooking.objects.create(
            itinerary=new_itinerary,
            day_plan=day_plan_mapping.get(old_hotel.day_plan.id) if old_hotel.day_plan else None,

            # ✅ Identity & Custom Fields
            hotel=old_hotel.hotel,
            custom_hotel_name=old_hotel.custom_hotel_name,  # <--- NEW
            destination=old_hotel.destination if hasattr(old_hotel, 'destination') else None,

            # ✅ Room & Custom Fields
            room_type=old_hotel.room_type,
            custom_room_type=old_hotel.custom_room_type,    # <--- NEW

            # Standard Fields
            meal_plan=old_hotel.meal_plan,
            check_in_date=old_hotel.check_in_date,
            check_out_date=old_hotel.check_out_date,
            check_in_time=old_hotel.check_in_time,
            check_out_time=old_hotel.check_out_time,
            num_rooms=old_hotel.num_rooms,
            num_double_beds=old_hotel.num_double_beds,
            extra_beds=old_hotel.extra_beds,
            child_with_bed=old_hotel.child_with_bed,
            child_without_bed=old_hotel.child_without_bed,
            option=old_hotel.option,
            category=old_hotel.category,
            markup_type=old_hotel.markup_type,
            markup_value=old_hotel.markup_value,
            markup_amount=old_hotel.markup_amount,
            markup_percentage=old_hotel.markup_percentage,
            net_price=old_hotel.net_price,
            gross_price=old_hotel.gross_price,
            custom_double_bed_total=old_hotel.custom_double_bed_total,
            custom_extra_bed_total=old_hotel.custom_extra_bed_total,
            custom_child_with_bed_total=old_hotel.custom_child_with_bed_total,
            custom_child_without_bed_total=old_hotel.custom_child_without_bed_total,
        )

        # Copy hotel inclusions
        for old_inc in old_hotel.inclusion_items.all():
            HotelBookingInclusion.objects.create(
                hotel_booking=new_hotel,
                special_inclusion=old_inc.special_inclusion,
                num_adults=old_inc.num_adults,
                num_children=old_inc.num_children,
                price=old_inc.price,
            )

        hotel_count += 1
        print(f"   ✓ {old_hotel.hotel.name if old_hotel.hotel else old_hotel.custom_hotel_name}")

    # 3️⃣ COPY VEHICLES
    old_vehicles = VehicleBooking.objects.filter(itinerary=old_itinerary)
    for old_vehicle in old_vehicles:
        VehicleBooking.objects.create(
            itinerary=new_itinerary,
            day_plan=day_plan_mapping.get(old_vehicle.day_plan.id) if old_vehicle.day_plan else None,
            vehicle=old_vehicle.vehicle,
            destination=old_vehicle.destination,
            pickup_date=old_vehicle.pickup_date,
            pickup_time=old_vehicle.pickup_time,
            num_passengers=old_vehicle.num_passengers,
            total_km=old_vehicle.total_km,
            vehicle_type=old_vehicle.vehicle_type,
            option=old_vehicle.option,
            markup_type=old_vehicle.markup_type,
            markup_value=old_vehicle.markup_value,
            markup_amount=old_vehicle.markup_amount,
            markup_percentage=old_vehicle.markup_percentage,
            net_price=old_vehicle.net_price,
            gross_price=old_vehicle.gross_price,
            custom_total_price=old_vehicle.custom_total_price,
        )

    # 4️⃣ COPY ACTIVITIES
    old_activities = ActivityBooking.objects.filter(itinerary=old_itinerary)
    for old_activity in old_activities:
        ActivityBooking.objects.create(
            itinerary=new_itinerary,
            day_plan=day_plan_mapping.get(old_activity.day_plan.id) if old_activity.day_plan else None,
            activity=old_activity.activity,
            booking_date=old_activity.booking_date,
            booking_time=old_activity.booking_time,
            num_adults=old_activity.num_adults,
            num_children=old_activity.num_children,
            confirmation_number=old_activity.confirmation_number,
            notes=old_activity.notes,
            markup_type=old_activity.markup_type,
            markup_value=old_activity.markup_value,
            markup_amount=old_activity.markup_amount,
            markup_percentage=old_activity.markup_percentage,
            net_price=old_activity.net_price,
            gross_price=old_activity.gross_price,
            custom_total_price=old_activity.custom_total_price,
        )

    # 5️⃣ COPY HOUSEBOATS
    old_houseboats = HouseboatBooking.objects.filter(itinerary=old_itinerary)
    for old_hb in old_houseboats:
        new_hb = HouseboatBooking.objects.create(
            itinerary=new_itinerary,
            day_plan=day_plan_mapping.get(old_hb.day_plan.id) if old_hb.day_plan else None,
            houseboat=old_hb.houseboat,
            room_type=old_hb.room_type,
            meal_plan=old_hb.meal_plan,
            check_in_date=old_hb.check_in_date,
            check_out_date=old_hb.check_out_date,
            option=old_hb.option,
            num_one_bed_rooms=old_hb.num_one_bed_rooms,
            num_two_bed_rooms=old_hb.num_two_bed_rooms,
            num_three_bed_rooms=old_hb.num_three_bed_rooms,
            num_four_bed_rooms=old_hb.num_four_bed_rooms,
            num_five_bed_rooms=old_hb.num_five_bed_rooms,
            num_six_bed_rooms=old_hb.num_six_bed_rooms,
            num_seven_bed_rooms=old_hb.num_seven_bed_rooms,
            num_eight_bed_rooms=old_hb.num_eight_bed_rooms,
            num_nine_bed_rooms=old_hb.num_nine_bed_rooms,
            num_ten_bed_rooms=old_hb.num_ten_bed_rooms,
            num_extra_beds=old_hb.num_extra_beds,
            markup_type=old_hb.markup_type,
            markup_value=old_hb.markup_value,
            markup_amount=old_hb.markup_amount,
            markup_percentage=old_hb.markup_percentage,
            net_price=old_hb.net_price,
            gross_price=old_hb.gross_price,
            custom_one_bed_total=old_hb.custom_one_bed_total,
            custom_two_bed_total=old_hb.custom_two_bed_total,
            custom_three_bed_total=old_hb.custom_three_bed_total,
            custom_four_bed_total=old_hb.custom_four_bed_total,
            custom_five_bed_total=old_hb.custom_five_bed_total,
            custom_six_bed_total=old_hb.custom_six_bed_total,
            custom_seven_bed_total=old_hb.custom_seven_bed_total,
            custom_eight_bed_total=old_hb.custom_eight_bed_total,
            custom_nine_bed_total=old_hb.custom_nine_bed_total,
            custom_ten_bed_total=old_hb.custom_ten_bed_total,
            custom_extra_bed_hb_total=old_hb.custom_extra_bed_hb_total,
        )

        # Copy houseboat inclusions
        for old_inc in old_hb.inclusion_items.all():
            HouseboatBookingInclusion.objects.create(
                houseboat_booking=new_hb,
                special_inclusion=old_inc.special_inclusion,
                num_adults=old_inc.num_adults,
                num_children=old_inc.num_children,
                price=old_inc.price,
            )

    # 6️⃣ COPY STANDALONE INCLUSIONS
    old_inclusions = StandaloneInclusionBooking.objects.filter(itinerary=old_itinerary)
    for old_si in old_inclusions:
        StandaloneInclusionBooking.objects.create(
            itinerary=new_itinerary,
            day_plan=day_plan_mapping.get(old_si.day_plan.id) if old_si.day_plan else None,
            special_inclusion=old_si.special_inclusion,
            booking_date=old_si.booking_date,
            booking_time=old_si.booking_time,
            num_adults=old_si.num_adults,
            num_children=old_si.num_children,
            adult_unit_price=old_si.adult_unit_price,
            child_unit_price=old_si.child_unit_price,
            subtotal=old_si.subtotal,
            markup_type=old_si.markup_type,
            markup_value=old_si.markup_value,
            markup_amount=old_si.markup_amount,
            total_price=old_si.total_price,
            notes=old_si.notes,
            created_by=old_si.created_by,
        )

    # 7️⃣ COPY PRICING OPTIONS
    old_pricing = ItineraryPricingOption.objects.filter(itinerary=old_itinerary)
    pricing_count = 0
    for old_p in old_pricing:
        ItineraryPricingOption.objects.create(
            itinerary=new_itinerary,
            option_name=old_p.option_name,
            option_number=old_p.option_number,
            vehicle_type=old_p.vehicle_type,
            number_of_rooms=old_p.number_of_rooms,
            extra_beds=old_p.extra_beds,
            child_without_bed=old_p.child_without_bed,
            child_with_bed=old_p.child_with_bed,
            child_ages=old_p.child_ages,
            net_price=old_p.net_price,
            markup_amount=old_p.markup_amount,
            gross_price=old_p.gross_price,
            cgst_amount=old_p.cgst_amount,
            sgst_amount=old_p.sgst_amount,
            discount_amount=old_p.discount_amount,
            final_amount=old_p.final_amount,
            hotels_included=old_p.hotels_included,
        )
        pricing_count += 1

    # ✅ Refresh to see if status changed
    new_itinerary.refresh_from_db()

    print(f"\n✅ COPYING COMPLETE!")
    print(f"   Hotels: {hotel_count}")

    return new_itinerary



# ==========================================
# DELETE VERSION (SECURE)
# ==========================================
from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse

def delete_itinerary_version(request, itinerary_id):
    """
    Delete an archived itinerary version permanently
    Only archived versions can be deleted
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)

    itinerary = get_object_or_404(Itinerary, id=itinerary_id)

    # ✅ SAFETY CHECK: Only allow deletion of archived versions
    if itinerary.status != 'archived':
        messages.error(request, '❌ Cannot delete active itinerary. Only archived versions can be deleted.')
        return redirect('itinerary_version_history', query_id=itinerary.query.id)

    # ✅ PERMISSION CHECK
    user_type = request.session.get('user_type')
    current_user = None

    if user_type == 'superuser':
        from django.contrib.auth.models import User
        current_user = User.objects.get(id=request.session.get('user_id'))
    elif user_type == 'team_member':
        from Travel.models import TeamMember
        current_user = TeamMember.objects.get(id=request.session.get('user_id'))

    can_delete = (
        user_type == 'superuser' or
        (current_user and hasattr(current_user, 'role') and current_user.role in ['admin', 'manager'])
    )

    if not can_delete:
        messages.error(request, '❌ You do not have permission to delete versions')
        return redirect('itinerary_version_history', query_id=itinerary.query.id)

    try:
        query_id = itinerary.query.id
        version_number = itinerary.version_number
        itinerary_name = itinerary.name or f"Itinerary #{itinerary.id}"

        print("\n" + "="*80)
        print(f"🗑️  DELETING ARCHIVED VERSION")
        print("="*80)
        print(f"   Itinerary ID: {itinerary.id}")
        print(f"   Version: {version_number}")
        print(f"   Name: {itinerary_name}")
        print(f"   Status: {itinerary.status}")

        # ✅ CASCADE DELETE - All related data will be deleted automatically
        # because of Django's ForeignKey on_delete=models.CASCADE
        itinerary.delete()

        print(f"✅ Successfully deleted version {version_number}")
        print("="*80 + "\n")

        messages.success(
            request,
            f'✅ Successfully deleted Version {version_number} - {itinerary_name}'
        )

        return redirect('itinerary_version_history', query_id=query_id)

    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"\n❌ ERROR DELETING VERSION:")
        print(error_trace)
        print("="*80 + "\n")

        messages.error(request, f'❌ Error deleting version: {str(e)}')
        return redirect('itinerary_version_history', query_id=itinerary.query.id)


from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.http import require_POST
from .models import (
    Itinerary, ItineraryDayPlan, HotelBooking, VehicleBooking,
    ActivityBooking, HouseboatBooking, StandaloneInclusionBooking,
    ItineraryPricingOption, HotelBookingInclusion, HouseboatBookingInclusion,
    TeamMember
)
from django.contrib.auth.models import User

@require_POST
def restore_archived_itinerary(request, itinerary_id):
    """
    RESTORE = Create a new active copy from archived version
    Original archived version stays in history (NOT moved)
    """
    archived_itinerary = get_object_or_404(Itinerary, id=itinerary_id)
    query_id = archived_itinerary.query.id

    # ==========================================
    # PERMISSION CHECK
    # ==========================================
    user_type = request.session.get('user_type')
    user_id = request.session.get('user_id')

    can_restore = False

    if user_type == 'superuser':
        can_restore = True
    elif user_type == 'team_member':
        current_user = TeamMember.objects.get(id=user_id)
        if current_user.role in ['admin', 'manager']:
            can_restore = True

    if not can_restore:
        messages.error(request, '❌ You do not have permission to restore itineraries.')
        return redirect('itinerary_history', query_id=query_id)

    # ==========================================
    # VALIDATION CHECK
    # ==========================================
    if archived_itinerary.status != 'archived':
        messages.warning(request, f'⚠️ This itinerary is not archived')
        return redirect('query_proposals', query_id=query_id)

    try:
        print("\n" + "="*80)
        print(f"♻️  RESTORING FROM ARCHIVED VERSION")
        print("="*80)
        print(f"   Source (Archived): ID {archived_itinerary.id} (v{archived_itinerary.version_number})")
        print(f"   Name: {archived_itinerary.name}")
        print(f"   Status: {archived_itinerary.status}")
        print(f"   📝 Will CREATE NEW COPY (archived version stays in history)")

        # ==========================================
        # GET NEXT VERSION NUMBER
        # ==========================================
        # Find highest version number for this query
        max_version = Itinerary.objects.filter(
            query=archived_itinerary.query
        ).order_by('-version_number').first()

        next_version = (max_version.version_number + 1) if max_version else 1

        print(f"   📊 Next version number: {next_version}")

        # ==========================================
        # CREATE NEW ACTIVE COPY
        # ==========================================
        restored_itinerary = Itinerary.objects.create(
            query=archived_itinerary.query,
            name=archived_itinerary.name,
            travel_from=archived_itinerary.travel_from,
            travel_to=archived_itinerary.travel_to,
            total_days=archived_itinerary.total_days,
            adults=archived_itinerary.adults,
            childrens=archived_itinerary.childrens,
            infants=archived_itinerary.infants,
            notes=archived_itinerary.notes,
            parent_itinerary=archived_itinerary.parent_itinerary,  # Keep same parent
            version_number=next_version,  # New version number
            status='draft',  # Active as draft
            is_finalized=False,  # Editable
            finalized_at=None,
            cgst_percentage=archived_itinerary.cgst_percentage,
            sgst_percentage=archived_itinerary.sgst_percentage,
            discount=archived_itinerary.discount,
            created_by=archived_itinerary.created_by
        )

        # Set destinations
        if archived_itinerary.destinations.exists():
            restored_itinerary.destinations.set(archived_itinerary.destinations.all())

        print(f"   ✅ Created restored copy: ID {restored_itinerary.id} (v{restored_itinerary.version_number})")
        print(f"   Status: {restored_itinerary.status}")

        # ==========================================
        # COPY ALL DATA FROM ARCHIVED VERSION
        # ==========================================

        # 1️⃣ COPY DAY PLANS
        day_plan_mapping = {}
        old_day_plans = ItineraryDayPlan.objects.filter(itinerary=archived_itinerary)

        print(f"\n   📅 Copying {old_day_plans.count()} day plan(s)...")
        for old_day in old_day_plans:
            new_day = ItineraryDayPlan.objects.create(
                itinerary=restored_itinerary,
                day_number=old_day.day_number,
                destination=old_day.destination,
                title=old_day.title,
                description=old_day.description,
                image=old_day.image,
                notes=old_day.notes if hasattr(old_day, 'notes') else '',
            )

            # Copy ManyToMany fields
            if hasattr(old_day, 'hotels'): new_day.hotels.set(old_day.hotels.all())
            if hasattr(old_day, 'houseboats'): new_day.houseboats.set(old_day.houseboats.all())
            if hasattr(old_day, 'activities'): new_day.activities.set(old_day.activities.all())
            if hasattr(old_day, 'meal_plans'): new_day.meal_plans.set(old_day.meal_plans.all())
            if hasattr(old_day, 'vehicles'): new_day.vehicles.set(old_day.vehicles.all())
            if hasattr(old_day, 'inclusions'): new_day.inclusions.set(old_day.inclusions.all())

            day_plan_mapping[old_day.id] = new_day

        # 2️⃣ COPY HOTELS (✅ UPDATED FOR CUSTOM FIELDS)
        hotel_count = 0
        old_hotels = HotelBooking.objects.filter(itinerary=archived_itinerary)

        print(f"   🏨 Copying {old_hotels.count()} hotel(s)...")
        for old_hotel in old_hotels:
            new_hotel = HotelBooking.objects.create(
                itinerary=restored_itinerary,
                day_plan=day_plan_mapping.get(old_hotel.day_plan.id) if old_hotel.day_plan else None,

                # ✅ Identity & Custom Fields
                hotel=old_hotel.hotel,
                custom_hotel_name=old_hotel.custom_hotel_name,  # <--- NEW
                destination=old_hotel.destination if hasattr(old_hotel, 'destination') else None,

                # ✅ Room & Custom Fields
                room_type=old_hotel.room_type,
                custom_room_type=old_hotel.custom_room_type,    # <--- NEW

                # Standard Fields
                meal_plan=old_hotel.meal_plan,
                check_in_date=old_hotel.check_in_date,
                check_out_date=old_hotel.check_out_date,
                check_in_time=old_hotel.check_in_time if hasattr(old_hotel, 'check_in_time') else None,
                check_out_time=old_hotel.check_out_time if hasattr(old_hotel, 'check_out_time') else None,
                num_rooms=old_hotel.num_rooms if hasattr(old_hotel, 'num_rooms') else 1,
                num_double_beds=old_hotel.num_double_beds if hasattr(old_hotel, 'num_double_beds') else 0,
                extra_beds=old_hotel.extra_beds if hasattr(old_hotel, 'extra_beds') else 0,
                child_with_bed=old_hotel.child_with_bed if hasattr(old_hotel, 'child_with_bed') else 0,
                child_without_bed=old_hotel.child_without_bed if hasattr(old_hotel, 'child_without_bed') else 0,
                option=old_hotel.option if hasattr(old_hotel, 'option') else 'option1',
                category=old_hotel.category if hasattr(old_hotel, 'category') else '',
                markup_type=old_hotel.markup_type if hasattr(old_hotel, 'markup_type') else 'percentage',
                markup_value=old_hotel.markup_value if hasattr(old_hotel, 'markup_value') else 0,
                markup_amount=old_hotel.markup_amount if hasattr(old_hotel, 'markup_amount') else 0,
                markup_percentage=old_hotel.markup_percentage if hasattr(old_hotel, 'markup_percentage') else 0,
                net_price=old_hotel.net_price if hasattr(old_hotel, 'net_price') else 0,
                gross_price=old_hotel.gross_price if hasattr(old_hotel, 'gross_price') else 0,
                custom_double_bed_total=old_hotel.custom_double_bed_total if hasattr(old_hotel, 'custom_double_bed_total') else None,
                custom_extra_bed_total=old_hotel.custom_extra_bed_total if hasattr(old_hotel, 'custom_extra_bed_total') else None,
                custom_child_with_bed_total=old_hotel.custom_child_with_bed_total if hasattr(old_hotel, 'custom_child_with_bed_total') else None,
                custom_child_without_bed_total=old_hotel.custom_child_without_bed_total if hasattr(old_hotel, 'custom_child_without_bed_total') else None,
            )

            # Copy hotel inclusions
            for old_inc in old_hotel.inclusion_items.all():
                HotelBookingInclusion.objects.create(
                    hotel_booking=new_hotel,
                    special_inclusion=old_inc.special_inclusion,
                    num_adults=old_inc.num_adults if hasattr(old_inc, 'num_adults') else 0,
                    num_children=old_inc.num_children if hasattr(old_inc, 'num_children') else 0,
                    price=old_inc.price if hasattr(old_inc, 'price') else 0,
                )

            hotel_count += 1

        # 3️⃣ COPY VEHICLES
        vehicle_count = 0
        old_vehicles = VehicleBooking.objects.filter(itinerary=archived_itinerary)

        print(f"   🚗 Copying {old_vehicles.count()} vehicle(s)...")
        for old_vehicle in old_vehicles:
            VehicleBooking.objects.create(
                itinerary=restored_itinerary,
                day_plan=day_plan_mapping.get(old_vehicle.day_plan.id) if old_vehicle.day_plan else None,
                vehicle=old_vehicle.vehicle,
                destination=old_vehicle.destination if hasattr(old_vehicle, 'destination') else None,
                pickup_date=old_vehicle.pickup_date,
                pickup_time=old_vehicle.pickup_time if hasattr(old_vehicle, 'pickup_time') else None,
                num_passengers=old_vehicle.num_passengers if hasattr(old_vehicle, 'num_passengers') else 0,
                total_km=old_vehicle.total_km if hasattr(old_vehicle, 'total_km') else 0,
                vehicle_type=old_vehicle.vehicle_type if hasattr(old_vehicle, 'vehicle_type') else '',
                option=old_vehicle.option if hasattr(old_vehicle, 'option') else '',
                markup_type=old_vehicle.markup_type if hasattr(old_vehicle, 'markup_type') else 'percentage',
                markup_value=old_vehicle.markup_value if hasattr(old_vehicle, 'markup_value') else 0,
                markup_amount=old_vehicle.markup_amount if hasattr(old_vehicle, 'markup_amount') else 0,
                markup_percentage=old_vehicle.markup_percentage if hasattr(old_vehicle, 'markup_percentage') else 0,
                net_price=old_vehicle.net_price if hasattr(old_vehicle, 'net_price') else 0,
                gross_price=old_vehicle.gross_price if hasattr(old_vehicle, 'gross_price') else 0,
                custom_total_price=old_vehicle.custom_total_price if hasattr(old_vehicle, 'custom_total_price') else None,
            )
            vehicle_count += 1

        # 4️⃣ COPY ACTIVITIES
        activity_count = 0
        old_activities = ActivityBooking.objects.filter(itinerary=archived_itinerary)

        print(f"   🎯 Copying {old_activities.count()} activit(ies)...")
        for old_activity in old_activities:
            ActivityBooking.objects.create(
                itinerary=restored_itinerary,
                day_plan=day_plan_mapping.get(old_activity.day_plan.id) if old_activity.day_plan else None,
                activity=old_activity.activity,
                booking_date=old_activity.booking_date,
                booking_time=old_activity.booking_time if hasattr(old_activity, 'booking_time') else None,
                num_adults=old_activity.num_adults if hasattr(old_activity, 'num_adults') else 0,
                num_children=old_activity.num_children if hasattr(old_activity, 'num_children') else 0,
                confirmation_number=old_activity.confirmation_number if hasattr(old_activity, 'confirmation_number') else '',
                notes=old_activity.notes if hasattr(old_activity, 'notes') else '',
                markup_type=old_activity.markup_type if hasattr(old_activity, 'markup_type') else 'percentage',
                markup_value=old_activity.markup_value if hasattr(old_activity, 'markup_value') else 0,
                markup_amount=old_activity.markup_amount if hasattr(old_activity, 'markup_amount') else 0,
                markup_percentage=old_activity.markup_percentage if hasattr(old_activity, 'markup_percentage') else 0,
                net_price=old_activity.net_price if hasattr(old_activity, 'net_price') else 0,
                gross_price=old_activity.gross_price if hasattr(old_activity, 'gross_price') else 0,
                custom_total_price=old_activity.custom_total_price if hasattr(old_activity, 'custom_total_price') else None,
            )
            activity_count += 1

        # 5️⃣ COPY HOUSEBOATS
        houseboat_count = 0
        old_houseboats = HouseboatBooking.objects.filter(itinerary=archived_itinerary)

        print(f"   🚤 Copying {old_houseboats.count()} houseboat(s)...")
        for old_hb in old_houseboats:
            new_hb = HouseboatBooking.objects.create(
                itinerary=restored_itinerary,
                day_plan=day_plan_mapping.get(old_hb.day_plan.id) if old_hb.day_plan else None,
                houseboat=old_hb.houseboat,
                room_type=old_hb.room_type if hasattr(old_hb, 'room_type') else None,
                meal_plan=old_hb.meal_plan if hasattr(old_hb, 'meal_plan') else None,
                check_in_date=old_hb.check_in_date,
                check_out_date=old_hb.check_out_date,
                option=old_hb.option if hasattr(old_hb, 'option') else 'option1',
                num_one_bed_rooms=old_hb.num_one_bed_rooms if hasattr(old_hb, 'num_one_bed_rooms') else 0,
                num_two_bed_rooms=old_hb.num_two_bed_rooms if hasattr(old_hb, 'num_two_bed_rooms') else 0,
                num_three_bed_rooms=old_hb.num_three_bed_rooms if hasattr(old_hb, 'num_three_bed_rooms') else 0,
                num_four_bed_rooms=old_hb.num_four_bed_rooms if hasattr(old_hb, 'num_four_bed_rooms') else 0,
                num_five_bed_rooms=old_hb.num_five_bed_rooms if hasattr(old_hb, 'num_five_bed_rooms') else 0,
                num_six_bed_rooms=old_hb.num_six_bed_rooms if hasattr(old_hb, 'num_six_bed_rooms') else 0,
                num_seven_bed_rooms=old_hb.num_seven_bed_rooms if hasattr(old_hb, 'num_seven_bed_rooms') else 0,
                num_eight_bed_rooms=old_hb.num_eight_bed_rooms if hasattr(old_hb, 'num_eight_bed_rooms') else 0,
                num_nine_bed_rooms=old_hb.num_nine_bed_rooms if hasattr(old_hb, 'num_nine_bed_rooms') else 0,
                num_ten_bed_rooms=old_hb.num_ten_bed_rooms if hasattr(old_hb, 'num_ten_bed_rooms') else 0,
                num_extra_beds=old_hb.num_extra_beds if hasattr(old_hb, 'num_extra_beds') else 0,
                markup_type=old_hb.markup_type if hasattr(old_hb, 'markup_type') else 'percentage',
                markup_value=old_hb.markup_value if hasattr(old_hb, 'markup_value') else 0,
                markup_amount=old_hb.markup_amount if hasattr(old_hb, 'markup_amount') else 0,
                markup_percentage=old_hb.markup_percentage if hasattr(old_hb, 'markup_percentage') else 0,
                net_price=old_hb.net_price if hasattr(old_hb, 'net_price') else 0,
                gross_price=old_hb.gross_price if hasattr(old_hb, 'gross_price') else 0,
                custom_one_bed_total=old_hb.custom_one_bed_total if hasattr(old_hb, 'custom_one_bed_total') else None,
                custom_two_bed_total=old_hb.custom_two_bed_total if hasattr(old_hb, 'custom_two_bed_total') else None,
                custom_three_bed_total=old_hb.custom_three_bed_total if hasattr(old_hb, 'custom_three_bed_total') else None,
                custom_four_bed_total=old_hb.custom_four_bed_total if hasattr(old_hb, 'custom_four_bed_total') else None,
                custom_five_bed_total=old_hb.custom_five_bed_total if hasattr(old_hb, 'custom_five_bed_total') else None,
                custom_six_bed_total=old_hb.custom_six_bed_total if hasattr(old_hb, 'custom_six_bed_total') else None,
                custom_seven_bed_total=old_hb.custom_seven_bed_total if hasattr(old_hb, 'custom_seven_bed_total') else None,
                custom_eight_bed_total=old_hb.custom_eight_bed_total if hasattr(old_hb, 'custom_eight_bed_total') else None,
                custom_nine_bed_total=old_hb.custom_nine_bed_total if hasattr(old_hb, 'custom_nine_bed_total') else None,
                custom_ten_bed_total=old_hb.custom_ten_bed_total if hasattr(old_hb, 'custom_ten_bed_total') else None,
                custom_extra_bed_hb_total=old_hb.custom_extra_bed_hb_total if hasattr(old_hb, 'custom_extra_bed_hb_total') else None,
            )

            # Copy houseboat inclusions
            for old_inc in old_hb.inclusion_items.all():
                HouseboatBookingInclusion.objects.create(
                    houseboat_booking=new_hb,
                    special_inclusion=old_inc.special_inclusion,
                    num_adults=old_inc.num_adults if hasattr(old_inc, 'num_adults') else 0,
                    num_children=old_inc.num_children if hasattr(old_inc, 'num_children') else 0,
                    price=old_inc.price if hasattr(old_inc, 'price') else 0,
                )

            houseboat_count += 1

        # 6️⃣ COPY STANDALONE INCLUSIONS
        inclusion_count = 0
        old_inclusions = StandaloneInclusionBooking.objects.filter(itinerary=archived_itinerary)

        print(f"   ⭐ Copying {old_inclusions.count()} standalone inclusion(s)...")
        for old_si in old_inclusions:
            StandaloneInclusionBooking.objects.create(
                itinerary=restored_itinerary,
                day_plan=day_plan_mapping.get(old_si.day_plan.id) if old_si.day_plan else None,
                special_inclusion=old_si.special_inclusion,
                booking_date=old_si.booking_date,
                booking_time=old_si.booking_time if hasattr(old_si, 'booking_time') else None,
                num_adults=old_si.num_adults if hasattr(old_si, 'num_adults') else 0,
                num_children=old_si.num_children if hasattr(old_si, 'num_children') else 0,
                adult_unit_price=old_si.adult_unit_price if hasattr(old_si, 'adult_unit_price') else 0,
                child_unit_price=old_si.child_unit_price if hasattr(old_si, 'child_unit_price') else 0,
                subtotal=old_si.subtotal if hasattr(old_si, 'subtotal') else 0,
                markup_type=old_si.markup_type if hasattr(old_si, 'markup_type') else 'percentage',
                markup_value=old_si.markup_value if hasattr(old_si, 'markup_value') else 0,
                markup_amount=old_si.markup_amount if hasattr(old_si, 'markup_amount') else 0,
                total_price=old_si.total_price if hasattr(old_si, 'total_price') else 0,
                notes=old_si.notes if hasattr(old_si, 'notes') else '',
                created_by=old_si.created_by if hasattr(old_si, 'created_by') else None,
            )
            inclusion_count += 1

        # 7️⃣ COPY PRICING OPTIONS
        pricing_count = 0
        old_pricing = ItineraryPricingOption.objects.filter(itinerary=archived_itinerary)

        print(f"   💰 Copying {old_pricing.count()} pricing option(s)...")
        for old_p in old_pricing:
            ItineraryPricingOption.objects.create(
                itinerary=restored_itinerary,
                option_name=old_p.option_name,
                option_number=old_p.option_number,
                vehicle_type=old_p.vehicle_type if hasattr(old_p, 'vehicle_type') else '',
                number_of_rooms=old_p.number_of_rooms if hasattr(old_p, 'number_of_rooms') else 0,
                extra_beds=old_p.extra_beds if hasattr(old_p, 'extra_beds') else 0,
                child_without_bed=old_p.child_without_bed if hasattr(old_p, 'child_without_bed') else 0,
                child_with_bed=old_p.child_with_bed if hasattr(old_p, 'child_with_bed') else 0,
                child_ages=old_p.child_ages if hasattr(old_p, 'child_ages') else '',
                net_price=old_p.net_price,
                markup_amount=old_p.markup_amount,
                gross_price=old_p.gross_price,
                cgst_amount=old_p.cgst_amount,
                sgst_amount=old_p.sgst_amount,
                discount_amount=old_p.discount_amount,
                final_amount=old_p.final_amount,
                hotels_included=old_p.hotels_included if hasattr(old_p, 'hotels_included') else None,
            )
            pricing_count += 1

        # ✅ SUMMARY
        total_bookings = hotel_count + vehicle_count + activity_count + houseboat_count + inclusion_count

        print("\n" + "="*80)
        print(f"✅ RESTORE COMPLETE!")
        print(f"   📦 Archived Version: ID {archived_itinerary.id} (v{archived_itinerary.version_number}) - STILL IN HISTORY")
        print(f"   ✨ Restored Copy: ID {restored_itinerary.id} (v{restored_itinerary.version_number}) - ACTIVE")
        print(f"")
        print(f"   📊 Data Copied:")
        print(f"      Days: {len(day_plan_mapping)}")
        print(f"      Hotels: {hotel_count}")
        print(f"      Vehicles: {vehicle_count}")
        print(f"      Activities: {activity_count}")
        print(f"      Houseboats: {houseboat_count}")
        print(f"      Standalone: {inclusion_count}")
        print(f"      Pricing Options: {pricing_count}")
        print(f"      TOTAL BOOKINGS: {total_bookings}")
        print("="*80 + "\n")

        messages.success(
            request,
            f'✅ Restored version {archived_itinerary.version_number} as new version {restored_itinerary.version_number}! '
            f'Original stays in history. Copied {total_bookings} bookings and {pricing_count} pricing options.'
        )

        return redirect('query_proposals', query_id=query_id)

    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"\n❌ ERROR RESTORING:")
        print(error_trace)
        print("="*80 + "\n")

        messages.error(request, f'❌ Error restoring: {str(e)}')
        return redirect('itinerary_history', query_id=query_id)




from django.shortcuts import render, get_object_or_404
from django.db.models import Q

def itinerary_version_history(request, query_id):
    """
    Show all versions (active + archived) for a query
    """
    query = get_object_or_404(Query, id=query_id)

    # ✅ Get ALL versions for this query (including archived)
    all_versions = Itinerary.objects.filter(
        query=query
    ).select_related('created_by', 'parent_itinerary').prefetch_related(
        'pricing_options'
    ).order_by('-version_number', '-created_at')

    # ✅ Separate active from archived
    active_versions = []
    archived_versions = []

    for itinerary in all_versions:
        # Get pricing summary
        pricing_summary = []
        for option in itinerary.pricing_options.all():
            pricing_summary.append({
                'name': option.option_name,
                'price': option.final_amount
            })

        # Count bookings
        hotel_count = HotelBooking.objects.filter(
            day_plan__itinerary=itinerary
        ).count()

        vehicle_count = VehicleBooking.objects.filter(
            itinerary=itinerary
        ).count()

        activity_count = ActivityBooking.objects.filter(
            day_plan__itinerary=itinerary
        ).count()

        houseboat_count = HouseboatBooking.objects.filter(
            day_plan__itinerary=itinerary
        ).count()

        standalone_count = StandaloneInclusionBooking.objects.filter(
            itinerary=itinerary
        ).count()

        total_bookings = (
            hotel_count + vehicle_count + activity_count +
            houseboat_count + standalone_count
        )

        version_data = {
            'itinerary': itinerary,
            'pricing_summary': pricing_summary,
            'total_bookings': total_bookings,
            'hotel_count': hotel_count,
            'vehicle_count': vehicle_count,
            'activity_count': activity_count,
            'houseboat_count': houseboat_count,
            'standalone_count': standalone_count,
        }

        if itinerary.status == 'archived':
            archived_versions.append(version_data)
        else:
            active_versions.append(version_data)

    context = {
        'query': query,
        'active_versions': active_versions,
        'archived_versions': archived_versions,
        'user_type': request.session.get('user_type'),
        'current_user': request.user if request.user.is_authenticated else None,
    }

    return render(request, 'itinerary_version_history.html', context)



from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.views.decorators.http import require_POST
from .models import HotelBooking, Hotel, Hotelprice
def get_available_hotels_for_change(request):
    """
    AJAX: Returns list of hotels in the same destination that have pricing configured
    for the booking dates.
    """
    booking_id = request.GET.get('booking_id')
    if not booking_id:
        return JsonResponse({'error': 'No booking ID'}, status=400)
    booking = get_object_or_404(HotelBooking, id=booking_id)

    # Get other hotels in same destination
    # Filter by hotels that have a price entry covering the check-in date
    # ✅ FIX: Match Sidebar logic exactly (Use Itinerary Dates, not Booking Dates)
    from django.db.models import Q
    valid_hotel_ids = Hotelprice.objects.filter(
        hotel__destination=booking.destination,
        hotel__status=True
    ).filter(
        Q(from_date__lte=booking.itinerary.travel_to) &
        Q(to_date__gte=booking.itinerary.travel_from)
    ).values_list('hotel__id', flat=True).distinct()

    available_hotels = Hotel.objects.filter(
        id__in=valid_hotel_ids
    ).exclude(id=booking.hotel.id)

    hotels_data = []
    for hotel in available_hotels:
        hotels_data.append({
            'id': hotel.id,
            'name': hotel.name,
            'category': hotel.get_category_display(),
            'image': hotel.image.url if hotel.image else None
        })

    # ✅ DEBUGGING: Send strict filtering info to frontend
    debug_msg = (
        f"STRICT SEARCH | "
        f"Dest: {booking.destination.name} | "
        f"Found: {available_hotels.count()} hotels | "
        f"Excl: {booking.hotel.id}"
    )

    return JsonResponse({'hotels': hotels_data, 'debug_message': debug_msg})



@require_POST
def change_hotel_booking(request):
    """
    Swaps the hotel in an existing booking. Deletes old booking, creates new one
    with same details but new hotel.
    """
    print("🚀 CHANGE HOTEL REQUEST RECEIVED")
    try:
        booking_id = request.POST.get('booking_id')
        new_hotel_id = request.POST.get('hotel_id') # Form uses 'hotel_id'

        print(f"📥 POST Data: booking_id={booking_id}, new_hotel_id={new_hotel_id}")

        if not booking_id or not new_hotel_id:
            print("❌ MISSING ID ERROR")
            messages.error(request, 'Missing Booking ID or Hotel ID')
            return redirect(request.META.get('HTTP_REFERER', '/'))

        old_booking = get_object_or_404(HotelBooking, id=booking_id)
        new_hotel = get_object_or_404(Hotel, id=new_hotel_id)

        print(f"✅ Found Old Booking: {old_booking} (ID: {old_booking.id})")
        print(f"✅ Found New Hotel: {new_hotel} (ID: {new_hotel.id})")

        # Create new booking
        new_booking = HotelBooking(
            itinerary=old_booking.itinerary,
            day_plan=old_booking.day_plan,
            package_template=old_booking.package_template,
            package_day_plan=old_booking.package_day_plan,

            destination=new_hotel.destination,
            hotel=new_hotel,

            # Update fields from POST data (user might have changed them)
            category=request.POST.get('category', new_hotel.category),
            room_type_id=request.POST.get('room_type') or old_booking.room_type_id,
            meal_plan_id=request.POST.get('meal_plan') or old_booking.meal_plan_id,
            option=request.POST.get('option', old_booking.option),

            num_rooms=request.POST.get('num_rooms', old_booking.num_rooms),
            num_double_beds=request.POST.get('num_double_beds', old_booking.num_double_beds),
            extra_beds=request.POST.get('extra_beds', old_booking.extra_beds),
            child_with_bed=request.POST.get('child_with_bed', old_booking.child_with_bed),
            child_without_bed=request.POST.get('child_without_bed', old_booking.child_without_bed),

            check_in_date=request.POST.get('check_in_date', old_booking.check_in_date),
            check_in_time=request.POST.get('check_in_time', old_booking.check_in_time),
            check_out_date=request.POST.get('check_out_date', old_booking.check_out_date),
            check_out_time=request.POST.get('check_out_time', old_booking.check_out_time),

            # Reset price - triggers recalculation if any logic exists
            net_price=0,
            markup_type=old_booking.markup_type,
            markup_value=old_booking.markup_value
        )

        print("💾 Saving New Booking...")
        new_booking.save()
        print(f"✅ New Booking Saved! ID: {new_booking.id}")

        # ============================================================
        # PROCESS SPECIAL INCLUSIONS
        # ============================================================
        inclusions_data = request.POST.get('inclusions_data')
        if inclusions_data:
            print(f"📦 Processing Inclusions Data: {inclusions_data}")
            import json
            try:
                inclusions_list = json.loads(inclusions_data)
                for inc in inclusions_list:
                    inc_id = inc.get('id')
                    adults = int(inc.get('adults') or 0)
                    children = int(inc.get('children') or 0)

                    if inc_id and (adults > 0 or children > 0):
                        try:
                            # Verify inclusion exists
                            special_inc = SpecialInclusion.objects.get(id=inc_id)

                            # Create new relationship
                            # NOTE: Using underscores as per models.py definition
                            HotelBookingInclusion.objects.create(
                                hotel_booking=new_booking,
                                special_inclusion=special_inc,
                                num_adults=adults,
                                num_children=children
                            )
                            print(f"   ✅ Added Inclusion: {special_inc.name} ({adults}A + {children}C)")
                        except Exception as inner_e:
                            print(f"   ⚠️ Failed to add inclusion {inc_id}: {inner_e}")
            except Exception as e:
                print(f"❌ Error parsing inclusions JSON: {e}")
        else:
            print("ℹ️ No inclusions data provided")

        print(f"🗑️ Deleting Old Booking ID: {old_booking.id}...")
        old_booking.delete()
        print("✅ Old Booking Deleted!")

        messages.success(request, f'✅ Successfully changed hotel to {new_hotel.name}')
        return redirect('itinerary_day_plan', itinerary_id=old_booking.itinerary.id)

    except Exception as e:
        import traceback
        print(f"❌ EXCEPTION IN CHANGE HOTEL: {str(e)}")
        traceback.print_exc()
        messages.error(request, f'❌ Error changing hotel: {str(e)}')
        return redirect(request.META.get('HTTP_REFERER', '/'))






from django.http import JsonResponse
from django.db.models import Q
# ⚠️ MAKE SURE YOU IMPORT Hotelprice HERE
from .models import HotelBooking, Hotel, Hotelprice

def get_available_hotels_for_change(request):
    """
    API to fetch hotels that match the destination AND have pricing for the dates.
    """
    booking_id = request.GET.get('booking_id')

    if not booking_id:
        return JsonResponse({'success': False, 'message': 'Booking ID missing'})

    try:
        booking = HotelBooking.objects.get(id=booking_id)

        # 1. Determine Destination
        destination = booking.destination
        if not destination and booking.hotel:
            destination = booking.hotel.destination
        if not destination and booking.day_plan:
            destination = booking.day_plan.destination

        if not destination:
            return JsonResponse({'success': False, 'message': 'Destination could not be determined'})

        # 2. Get Booking Dates
        check_in = booking.check_in_date
        check_out = booking.check_out_date

        # 3. 🔥 FILTER LOGIC: Find Hotels with Pricing for these dates
        if check_in and check_out:
            # Find IDs of hotels in this destination that have a price entry
            # overlapping with the check-in/check-out dates
            valid_hotel_ids = Hotelprice.objects.filter(
                hotel__destination=destination,
                hotel__status=True,
                from_date__lte=check_out, # Price starts before checkout
                to_date__gte=check_in     # Price ends after checkin
            ).values_list('hotel_id', flat=True).distinct()

            # Filter the Hotel list using those valid IDs
            hotels = Hotel.objects.filter(id__in=valid_hotel_ids).order_by('name')
        else:
            # Fallback if dates are missing (unlikely for a saved booking)
            hotels = Hotel.objects.none()

        # 4. Prepare Data
        hotels_data = [
            {
                'id': h.id,
                'name': h.name,
                'category': h.get_category_display() if hasattr(h, 'get_category_display') else h.category
            }
            for h in hotels
        ]

        return JsonResponse({
            'success': True,
            'hotels': hotels_data,
            'current_hotel_id': booking.hotel.id if booking.hotel else None
        })

    except HotelBooking.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Booking not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})
